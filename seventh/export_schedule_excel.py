# export_schedule_excel.py
# ---------------------------------------------------------------------
# Builds an Excel with three sheets from Schedule.yaml + EnvConfig.yaml
# SHEET 1  -> "Tasks x Dates"     (module/fab/region/customer, per-day cells with start/end highlights)
# SHEET 2  -> "Employees x Dates" (company | employee | skills, per-day text, workdays/hours)
# SHEET 3  -> "Dashboard"         (KPIs + tables + CHARTS: line + top-20 bars)
# ---------------------------------------------------------------------

import os
import math
import yaml
from argparse import ArgumentParser
from datetime import date, datetime, timedelta
from collections import defaultdict, Counter

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# ------------------------------- CONFIG -------------------------------

LIGHT_BLUE = "ADD8E6"  # module start highlight
RED        = "FF9999"  # phase/task end highlight

# If True: required hours derived from EnvConfig using workload_days * max(work_hours) * min_worker_num
# If False: required hours = actually assigned hours
REQUIRED_HOURS_MODE = True

# Overtime threshold (hours/day)
OT_THRESHOLD = 8

# Cap threshold for "cap breach %" KPI (soft check only, for dashboard)
CAP_HOURS = 12

# ------------------------------ UTILITIES -----------------------------

def _d(s):
    """Parse 'YYYY/MM/DD' or 'YYYY-MM-DD' to date."""
    if isinstance(s, date):
        return s
    s = str(s).replace("-", "/")
    return datetime.strptime(s, "%Y/%m/%d").date()

def _hfmt(h):
    return f"{int(h)}H"

def _safe(lst, key):
    """Return lst[key] if exists."""
    try:
        return lst[key]
    except Exception:
        return None

# ----------------------------- LOADERS --------------------------------

def load_env(env_path):
    with open(env_path, "r", encoding="utf-8") as f:
        env = yaml.safe_load(f)

    env = env.get("environment", {})

    # maps
    workflows = {w["id"]: w for w in env.get("workflow_list", [])}
    fabs      = {f["id"]: f for f in env.get("fab_list", [])}
    regions   = {r["id"]: r for r in env.get("region_list", [])}
    customers = {c["id"]: c for c in env.get("customer_company_list", [])}
    wcompanies= {c["id"]: c for c in env.get("worker_company_list", [])}
    workers   = {w["id"]: w for w in env.get("worker_list", [])}

    # operation meta: workflow_id -> {phase_id -> {op_id -> meta}}
    op_meta = defaultdict(lambda: defaultdict(dict))
    for wf in env.get("workflow_list", []):
        for ph in wf.get("phase_list", []):
            for op in ph.get("operation_list", []):
                op_meta[wf["id"]][ph["id"]][op["id"]] = {
                    "name": op.get("name", op["id"]),
                    "work_hours": op.get("work_hours", [8,10,12]),
                    "min_worker_num": op.get("min_worker_num", 1),
                    "max_worker_num": op.get("max_worker_num", 99),
                }

    return {
        "workflows": workflows,
        "fabs": fabs,
        "regions": regions,
        "customers": customers,
        "workers": workers,
        "worker_companies": wcompanies,
        "op_meta": op_meta,
    }

def load_schedule(path):
    with open(path, "r", encoding="utf-8") as f:
        s = yaml.safe_load(f)
    s = s.get("schedule", s)

    plan_start = _d(s["planrange"]["start_date"])
    plan_end   = _d(s["planrange"]["end_date"])

    modules = s.get("workflow_task_list", [])
    asg_raw = s.get("assignment_list", [])

    # Normalize assignments and expand per-day rows
    assignments = []
    for a in asg_raw:
        wd_key = "work_date_lsit" if "work_date_lsit" in a else "work_date_list"
        for ditem in a.get(wd_key, []):
            assignments.append({
                "worker": a["worker"],
                "operation_task": a["operation_task"],  # e.g., e1p3o2
                "date": _d(ditem["date"]),
                "hours": int(ditem["hour"]),
            })

    return plan_start, plan_end, modules, assignments

# -------------------------- DATA AGGREGATION --------------------------

def build_maps(env, modules, assignments):
    workers = env["workers"]
    fabs = env["fabs"]
    regions = env["regions"]
    customers = env["customers"]
    wcompanies = env["worker_companies"]

    # module map by id
    mod_map = {m["id"]: m for m in modules}

    # per module: earliest phase start (for start highlight)
    module_start = {}
    # per module + phase: phase end (for deadline highlight)
    phase_end = {}  # (module_id, phase_id) -> date

    # useful lookup for op->phase for each module
    op_phase_of_module = {}  # (module_id, op_id) -> phase_id

    for m in modules:
        starts = []
        for ph in m.get("phase_task_list", []):
            p_start = _d(ph["start_date"])
            p_end   = _d(ph["end_date"])
            starts.append(p_start)
            phase_end[(m["id"], ph["phase"])] = p_end
            for ot in ph.get("operation_task_list", []):
                op_phase_of_module[(m["id"], ot["operation"])] = ph["phase"]
        module_start[m["id"]] = min(starts) if starts else None

    # worker info maps
    worker_name = {wid: workers[wid].get("name", wid) for wid in workers}
    worker_company_name = {
        wid: wcompanies.get(workers[wid].get("worker_company"), {}).get("name", workers[wid].get("worker_company"))
        for wid in workers
    }
    worker_skills = {wid: workers[wid].get("skill_map", {}) for wid in workers}

    # module metadata columns for sheet 1
    mod_meta_cols = {}
    for m in modules:
        fab_id = m.get("fab", "")
        fab = fabs.get(fab_id, {})
        region = regions.get(fab.get("region"), {})
        cust   = customers.get(fab.get("customer_company"), {})
        mod_meta_cols[m["id"]] = {
            "module": m["id"],
            "module_name": m.get("name",""),
            "fab_id": fab_id,
            "fab_name": fab.get("name",""),
            "region": region.get("name",""),     # header 'region'
            "customer": cust.get("name",""),     # header 'customer'
        }

    # explode module task catalog: all (module, op) that exist in phases
    module_ops = []
    for m in modules:
        for ph in m.get("phase_task_list", []):
            for ot in ph.get("operation_task_list", []):
                module_ops.append((m["id"], ot["operation"], ot.get("name", "")))

    # aggregate strings for Tasks x Dates
    tde = defaultdict(list)  # (module_id, op_id, date) -> [ "AA(12H)" ... ]
    for a in assignments:
        idx = a["operation_task"].find("p")
        m_id = a["operation_task"][:idx] if idx > 0 else a["operation_task"]
        op_id = a["operation_task"][idx:] if idx > 0 else ""
        tde[(m_id, op_id, a["date"])].append(f'{worker_name.get(a["worker"], a["worker"])}({_hfmt(a["hours"])})')

    # Employees x Dates aggregation + stats
    edt = defaultdict(list)  # (company, worker_name, date) -> ["p1o2 (12H)"]
    emp_total_hours = Counter()
    emp_workdays = Counter()
    per_day_total = Counter()
    day_op_heads = defaultdict(int)  # <-- IMPORTANT: int default (fixes TypeError)

    for a in assignments:
        wid = a["worker"]
        wname = worker_name.get(wid, wid)
        comp  = worker_company_name.get(wid, "")
        idx = a["operation_task"].find("p")
        m_id = a["operation_task"][:idx] if idx > 0 else a["operation_task"]
        op_id = a["operation_task"][idx:] if idx > 0 else ""
        edt[(comp, wname, a["date"])].append(f'{op_id} ({_hfmt(a["hours"])})')
        emp_total_hours[wname] += a["hours"]
        per_day_total[a["date"]] += a["hours"]
        day_op_heads[(m_id, op_id, a["date"])] += 1

    # employee workdays (distinct days with >=1 assignment)
    for (comp, wname, d), items in edt.items():
        if items:
            emp_workdays[wname] += 1

    return {
        "mod_map": mod_map,
        "module_start": module_start,
        "phase_end": phase_end,
        "op_phase_of_module": op_phase_of_module,
        "mod_meta_cols": mod_meta_cols,
        "module_ops": module_ops,
        "tde": tde,
        "edt": edt,
        "worker_name": worker_name,
        "worker_company_name": worker_company_name,
        "worker_skills": worker_skills,
        "emp_total_hours": emp_total_hours,
        "emp_workdays": emp_workdays,
        "per_day_total": per_day_total,
        "day_op_heads": day_op_heads,
    }

# --------------------------- REQUIRED HOURS ---------------------------

def compute_required_hours(env, modules):
    """
    Estimate required hours from EnvConfig + schedule structure.
    required = sum over operation_tasks ( workload_days * max(work_hours) * min_worker_num )
    """
    op_meta = env["op_meta"]
    required_by_module = defaultdict(int)
    for m in modules:
        wf_id = m.get("workflow")
        for ph in m.get("phase_task_list", []):
            ph_id = ph["phase"]
            for ot in ph.get("operation_task_list", []):
                op_id = ot["operation"]
                meta = _safe(_safe(op_meta, wf_id), ph_id)
                meta = _safe(meta, op_id) or {}
                wh = max(meta.get("work_hours", [8,10,12]))
                min_workers = int(meta.get("min_worker_num", 1))
                days = int(ot.get("workload_days", 0))
                required_by_module[m["id"]] += days * wh * min_workers
    return required_by_module

# ----------------------------- WRITERS --------------------------------

# ======================================================================
# ========================== SHEET 1 (Tasks x Dates) ====================
# ======================================================================
def write_sheet_tasks_dates(wb, plan_start, plan_end, env, maps):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    fill_deadline = PatternFill(start_color=RED,        end_color=RED,        fill_type="solid")
    fill_start    = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    thin = Side(style="thin", color="999999")

    ws = wb.create_sheet("Tasks x Dates")

    headers = ["module", "module_name", "fab_id", "fab_name", "region", "customer", "task"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        c.alignment = center

    dates = []
    d = plan_start
    while d <= plan_end:
        dates.append(d)
        d += timedelta(days=1)
    for j, dt in enumerate(dates, start=len(headers)+1):
        c = ws.cell(row=1, column=j, value=dt.isoformat())
        c.font = bold
        c.alignment = center
        ws.column_dimensions[get_column_letter(j)].width = 26

    widths = [8, 18, 8, 14, 12, 10, 12]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = get_column_letter(len(headers)+1) + "2"

    def mod_sort_key(m_id):
        try:
            return int(m_id[1:])
        except Exception:
            return 999
    rows = sorted(maps["module_ops"], key=lambda x: (mod_sort_key(x[0]), x[1]))

    for r_idx, (m_id, op_id, op_name) in enumerate(rows, start=2):
        meta = maps["mod_meta_cols"][m_id]
        ws.cell(row=r_idx, column=1, value=meta["module"]).font = bold
        ws.cell(row=r_idx, column=2, value=meta["module_name"]).font = bold
        ws.cell(row=r_idx, column=3, value=meta["fab_id"]).font = bold
        ws.cell(row=r_idx, column=4, value=meta["fab_name"]).font = bold
        ws.cell(row=r_idx, column=5, value=meta["region"]).font = bold
        ws.cell(row=r_idx, column=6, value=meta["customer"]).font = bold
        ws.cell(row=r_idx, column=7, value=f"{op_id} {op_name}").font = bold

        m_start = maps["module_start"].get(m_id)
        ph_id = maps["op_phase_of_module"].get((m_id, op_id))
        p_end = maps["phase_end"].get((m_id, ph_id))
        for j, dt in enumerate(dates, start=len(headers)+1):
            text = " | ".join(sorted(maps["tde"].get((m_id, op_id, dt), [])))
            c = ws.cell(row=r_idx, column=j, value=text)
            c.alignment = center
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            if m_start and dt == m_start:
                c.fill = fill_start
            if p_end and dt == p_end:
                c.fill = fill_deadline

    ws.row_dimensions[1].height = 22
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 36

# ======================================================================
# ======================= SHEET 2 (Employees x Dates) ==================
# ======================================================================
def write_sheet_employees_dates(wb, plan_start, plan_end, env, maps):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    ws = wb.create_sheet("Employees x Dates")

    base_headers = ["company", "employee", "skills"]
    for i, h in enumerate(base_headers, start=1):
        c = ws.cell(row=1, column=i, value=h); c.font = bold

    dates = []
    d = plan_start
    while d <= plan_end:
        dates.append(d); d += timedelta(days=1)
    for j, dt in enumerate(dates, start=len(base_headers)+1):
        c = ws.cell(row=1, column=j, value=dt.isoformat()); c.font = bold
        ws.column_dimensions[get_column_letter(j)].width = 30

    workdays_col  = len(base_headers) + len(dates) + 1
    workhours_col = len(base_headers) + len(dates) + 2
    ws.cell(row=1, column=workdays_col,  value="Workdays").font  = bold
    ws.cell(row=1, column=workhours_col, value="WorkHours").font = bold

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 54
    ws.column_dimensions[get_column_letter(workdays_col)].width  = 11
    ws.column_dimensions[get_column_letter(workhours_col)].width = 11

    ws.freeze_panes = "D2"

    roster = []
    seen = set()
    for (comp, wname, d), _ in maps["edt"].items():
        if (comp, wname) not in seen:
            roster.append((comp, wname)); seen.add((comp, wname))
    roster.sort(key=lambda t: (t[0] or "", t[1] or ""))

    def skills_text(wname):
        wid = None
        for k, v in env["workers"].items():
            if v.get("name") == wname:
                wid = k; break
        if wid is None:
            return ""
        items = list(maps["worker_skills"].get(wid, {}).items())
        def keyer(k):
            try:
                p, o = k.split("o")
                return (int(p[1:]), int(o))
            except Exception:
                return (999, 999)
        items.sort(key=lambda kv: keyer(kv[0]))
        return ", ".join(f"{k}:{v}" for k, v in items)

    for i, (comp, wname) in enumerate(roster, start=2):
        ws.row_dimensions[i].height = 34
        ws.cell(row=i, column=1, value=comp or "").alignment = left
        ws.cell(row=i, column=2, value=wname or "").alignment = left
        ws.cell(row=i, column=3, value=skills_text(wname)).alignment = left

        for j, dt in enumerate(dates, start=len(base_headers)+1):
            txt = " | ".join(sorted(maps["edt"].get((comp, wname, dt), [])))
            ws.cell(row=i, column=j, value=txt).alignment = center

        ws.cell(row=i, column=workdays_col,  value=maps["emp_workdays"].get(wname, 0)).alignment = center
        ws.cell(row=i, column=workhours_col, value=maps["emp_total_hours"].get(wname, 0)).alignment = center

# ======================================================================
# ===================== SHEET 3 (Dashboard / Analytics) =================
# ======================================================================
def write_sheet_dashboard(wb, plan_start, plan_end, env, maps, modules, assignments):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws = wb.create_sheet("Dashboard")

    for col in range(1, 80):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # ================= KPIs =================
    unique_workers = len({a["worker"] for a in assignments})
    avg_hours = 0.0
    if maps["emp_workdays"]:
        total_hours = sum(maps["emp_total_hours"].values())
        total_days = sum(maps["emp_workdays"].values())
        if total_days:
            avg_hours = total_hours / total_days

    ot_hours = 0
    per_emp_day = defaultdict(int)
    for (comp, wname, d), items in maps["edt"].items():
        day_hours = 0
        for s in items:
            h = int(s.split("(")[-1].rstrip("H)"))
            day_hours += h
        per_emp_day[(wname, d)] += day_hours
    for (w, d), h in per_emp_day.items():
        if h > OT_THRESHOLD:
            ot_hours += (h - OT_THRESHOLD)

    cap_breach_cnt = sum(1 for h in per_emp_day.values() if h > CAP_HOURS)
    cap_breach_pct = (100.0 * cap_breach_cnt / max(1, len(per_emp_day))) if per_emp_day else 0.0

    # Staffing validations
    breaches = []
    required_by_module = compute_required_hours(env, modules) if REQUIRED_HOURS_MODE else defaultdict(int)
    op_meta = env["op_meta"]
    minmax_breach_count = 0
    for (m_id, op_id, dt), heads in maps["day_op_heads"].items():
        wf = modules[0].get("workflow")
        ph = maps["op_phase_of_module"].get((m_id, op_id))
        meta = _safe(_safe(op_meta, wf), ph)
        meta = _safe(meta, op_id) or {}
        min_w = int(meta.get("min_worker_num", 1))
        max_w = int(meta.get("max_worker_num", 999))
        status = ""
        if heads < min_w:
            status = f"below min ({heads}<{min_w})"
        elif heads > max_w:
            status = f"above max ({heads}>{max_w})"
        if status:
            breaches.append((m_id, op_id, dt.isoformat(), heads, min_w, max_w, status))
            minmax_breach_count += 1

    # Completion %
    assigned_by_module = defaultdict(int)
    for a in assignments:
        idx = a["operation_task"].find("p")
        m_id = a["operation_task"][:idx] if idx > 0 else a["operation_task"]
        assigned_by_module[m_id] += a["hours"]
    if not REQUIRED_HOURS_MODE:
        for m_id, hrs in assigned_by_module.items():
            required_by_module[m_id] = hrs

    completion_pct = 0.0
    total_req = sum(required_by_module.values())
    if total_req:
        completion_pct = 100.0 * sum(assigned_by_module.values()) / total_req

    # KPI block
    ws["A1"].value = "KPIs"; ws["A1"].font = bold
    kpi_rows = [
        ("Unique workers", unique_workers),
        ("Avg hours/worker-day", round(avg_hours, 2)),
        ("Overtime hours(>8)", ot_hours),
        ("Cap breach(>12h) days", cap_breach_cnt),
        ("Cap breach(%)", f"{cap_breach_pct:.1f}%"),
        ("Staffing violations", minmax_breach_count),
        ("Completion", f"{completion_pct:.1f}%"),
    ]
    for i, (k, v) in enumerate(kpi_rows, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

    # -------- Progress by module (table only; chart removed per request) --------
    start_row = 10
    ws.cell(row=start_row, column=1, value="Progress by module").font = bold
    hdr = ["module", "required_hours", "assigned_hours", "%complete", "planned_end", "last_assigned", "delay(vs plan)"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=start_row+1, column=j, value=h).font = bold

    prog_rows = []
    for m in modules:
        m_id = m["id"]
        req = required_by_module.get(m_id, 0)
        asg = assigned_by_module.get(m_id, 0)
        pct = (100.0*asg/req) if req else 100.0
        planned_end = max(_d(ph["end_date"]) for ph in m.get("phase_task_list", [])) if m.get("phase_task_list") else None
        last_asg = max((a["date"] for a in assignments if a["operation_task"].startswith(m_id)), default=None)
        delay = None
        if planned_end and last_asg:
            delay = (last_asg - planned_end).days
        prog_rows.append((m_id, req, asg, pct, planned_end.isoformat() if planned_end else "",
                          last_asg.isoformat() if last_asg else "", delay if delay is not None else ""))

    for i, row in enumerate(prog_rows, start=start_row+2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    # -------- Staffing min/max breaches (table) --------
    b0 = start_row + 2 + len(prog_rows) + 2
    ws.cell(row=b0, column=1, value="Staffing min/max breaches (by op/date)").font = bold
    hdr2 = ["module", "op_id", "date", "heads", "min", "max", "status"]
    for j, h in enumerate(hdr2, start=1):
        ws.cell(row=b0+1, column=j, value=h).font = bold
    for i, row in enumerate(sorted(breaches, key=lambda r:(r[2], r[0], r[1])), start=b0+2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    # -------- Overtime (by employee) & Daily totals (tables) --------
    # Overtime by employee (sorted desc)
    ot_by_emp = Counter()
    for (wname, d), h in per_emp_day.items():
        if h > OT_THRESHOLD:
            ot_by_emp[wname] += (h - OT_THRESHOLD)
    ot_table = sorted(ot_by_emp.items(), key=lambda kv: kv[1], reverse=True)

    # Daily totals (now for EVERY day in the plan range so ticks are consistent)
    # Also store label in MM/DD to keep narrow
    day_rows = []
    d = plan_start
    while d <= plan_end:
        day_rows.append((d.strftime("%m/%d"), maps["per_day_total"].get(d, 0)))
        d += timedelta(days=1)

    c0 = b0 + 2 + len(breaches) + 2
    ws.cell(row=c0, column=1, value="Overtime & capacity").font = bold

    # Left subtable: overtime by employee
    ws.cell(row=c0+1, column=1, value="employee").font = bold
    ws.cell(row=c0+1, column=2, value="total_ot_hours").font = bold
    for i, (name, hrs) in enumerate(ot_table, start=c0+2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=int(hrs))

    # Right subtable: daily total hours
    ws.cell(row=c0+1, column=5, value="date").font = bold
    ws.cell(row=c0+1, column=6, value="total_hours").font = bold
    for i, (dd, hrs) in enumerate(day_rows, start=c0+2):
        ws.cell(row=i, column=5, value=dd)
        ws.cell(row=i, column=6, value=int(hrs))

    # -------- Workload balance by employee (table + helper top-20) --------
    d0 = c0 + 2 + max(len(ot_table), len(day_rows)) + 2
    ws.cell(row=d0, column=1, value="Workload balance (per employee)").font = bold
    hdr3 = ["employee", "workdays", "total_hours", "avg/day", "stdev/day", "CoV"]
    for j, h in enumerate(hdr3, start=1):
        ws.cell(row=d0+1, column=j, value=h).font = bold

    per_emp_day_list = defaultdict(list)
    for (w, d), h in per_emp_day.items():
        per_emp_day_list[w].append(h)

    balance_rows = []
    for wname in sorted(maps["emp_total_hours"].keys()):
        days = maps["emp_workdays"].get(wname, 0)
        tot  = maps["emp_total_hours"].get(wname, 0)
        avg = (tot / days) if days else 0.0
        lst = per_emp_day_list.get(wname, [])
        stdev = 0.0
        if len(lst) >= 2:
            m = sum(lst) / len(lst)
            stdev = math.sqrt(sum((x - m) ** 2 for x in lst) / (len(lst) - 1))
        cov = (stdev / avg * 100.0) if avg else 0.0
        balance_rows.append((wname, days, int(tot), round(avg,2), round(stdev,2), f"{cov:.1f}%"))

    for i, row in enumerate(balance_rows, start=d0+2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    # Helper table for TOP-20 total hours (contiguous so charts look right)
    top_col = 10  # column J
    ws.cell(row=d0, column=top_col, value="Top 20 total hours").font = bold
    ws.cell(row=d0+1, column=top_col,   value="employee").font = bold
    ws.cell(row=d0+1, column=top_col+1, value="total_hours").font = bold
    balance_rows_top20 = sorted(balance_rows, key=lambda r: r[2], reverse=True)[:20]
    for i, row in enumerate(balance_rows_top20, start=d0+2):
        ws.cell(row=i, column=top_col,   value=row[0])  # employee
        ws.cell(row=i, column=top_col+1, value=row[2])  # total_hours

    # ==================== CHARTS ====================
    # Anchors spaced so they don't overlap; adjust to taste.
    anchor_line   = "N2"   # Total hours per day (line)
    anchor_ot20   = "N22"  # Top overtime (top 20)
    anchor_tot20  = "N44"  # Total hours by employee (top 20)

    # Chart: Total hours per day (LINE, label every ~10 days)
    if day_rows:
        chart2 = LineChart()
        chart2.title = "Total hours per day"
        chart2.y_axis.title = "hours"
        chart2.x_axis.title = "date"
        # show fewer date labels:
        chart2.x_axis.tickLblSkip = 9  # roughly every 10th label
        dr0 = c0 + 1
        mrows = len(day_rows)
        cats = Reference(ws, min_col=5, min_row=dr0+1, max_row=dr0 + mrows)
        vals = Reference(ws, min_col=6, min_row=dr0,   max_row=dr0 + mrows)
        chart2.add_data(vals, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.height = 12; chart2.width = 24
        ws.add_chart(chart2, anchor_line)

    # Chart: Top overtime by employee (HORIZONTAL BAR, TOP-20)
    if ot_table:
        chart3 = BarChart()
        chart3.type = "bar"  # horizontal
        chart3.title = "Top overtime (hours) by employee"
        chart3.x_axis.title = "hours"
        chart3.y_axis.title = "employee"
        or0 = c0 + 1
        mrows = min(20, len(ot_table))  # <-- top 20
        cats = Reference(ws, min_col=1, min_row=or0+1, max_row=or0 + mrows)
        vals = Reference(ws, min_col=2, min_row=or0,   max_row=or0 + mrows)
        chart3.add_data(vals, titles_from_data=True)
        chart3.set_categories(cats)
        chart3.height = 12; chart3.width = 24
        ws.add_chart(chart3, anchor_ot20)

    # Chart: Total hours by employee (HORIZONTAL BAR, TOP-20)
    #     Uses the helper table (employee, total_hours) we wrote in columns J-K
    if balance_rows_top20:
        chart4 = BarChart()
        chart4.type = "bar"  # horizontal
        chart4.title = "Total hours by employee (Top 20)"
        chart4.x_axis.title = "hours"
        chart4.y_axis.title = "employee"
        bal0 = d0 + 1
        mrows = len(balance_rows_top20)
        cats = Reference(ws, min_col=top_col,   min_row=bal0+1, max_row=bal0 + mrows)
        vals = Reference(ws, min_col=top_col+1, min_row=bal0,   max_row=bal0 + mrows)
        chart4.add_data(vals, titles_from_data=True)
        chart4.set_categories(cats)
        chart4.height = 12; chart4.width = 24
        ws.add_chart(chart4, anchor_tot20)

# ------------------------------ MAIN ----------------------------------

def main():
    ap = ArgumentParser(description="Read Schedule.yaml + EnvConfig.yaml and export Excel (3 sheets).")
    ap.add_argument("--schedule", default="Schedule.yaml")
    ap.add_argument("--env",      default="EnvConfig.yaml")
    ap.add_argument("--out",      default="schedule_export.xlsx")
    args = ap.parse_args()

    plan_start, plan_end, modules, assignments = load_schedule(args.schedule)
    env = load_env(args.env)
    maps = build_maps(env, modules, assignments)

    wb = Workbook()
    # remove default empty sheet created by Workbook
    del wb[wb.sheetnames[0]]

    # ===== SHEET 1 =====
    write_sheet_tasks_dates(wb, plan_start, plan_end, env, maps)

    # ===== SHEET 2 =====
    write_sheet_employees_dates(wb, plan_start, plan_end, env, maps)

    # ===== SHEET 3 =====
    write_sheet_dashboard(wb, plan_start, plan_end, env, maps, modules, assignments)

    wb.save(args.out)
    print(f"Wrote Excel: {os.path.abspath(args.out)}")

if __name__ == "__main__":
    main()
