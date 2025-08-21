# export_matrix_capacity.py
import os
from argparse import ArgumentParser
from datetime import timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

from employee_scheduler_third import solve_from_config  # or employee_scheduler_by_capacity

LIGHT_BLUE = "ADD8E6"

def add_task_sheet(wb, solution, start_day, deadline_by_task, dates):
    """Sheet 1: rows=task_code, cols=dates, cells=comma-joined team names; deadline highlighted."""
    ws = wb.create_sheet(title="Schedule by Task") if wb.worksheets else wb.active
    if ws.title != "Schedule by Task":
        ws.title = "Schedule by Task"

    # Build team per task (same names all active days)
    team_by_task = defaultdict(list)
    for seat in solution.seats:
        if seat.employee:
            team_by_task[seat.task_code].append(seat.employee.name)

    # Windows per task (start/end day idx)
    window_by_task = {}
    for w in solution.windows:
        window_by_task[w.task_code] = (w.start_day_id, w.end_day_id)

    # Deadline dates
    dl_date_for = {code: start_day + timedelta(days=idx) for code, idx in deadline_by_task.items()}

    # Task order
    tasks = sorted({s.task_code for s in solution.seats},
                   key=lambda code: (int(code.split("-")[0][1:]), code.split("-")[1]))

    deadline_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    # Header
    ws.cell(row=1, column=1, value="task \\ date").font = bold
    for j, d in enumerate(dates, start=2):
        c = ws.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws.column_dimensions[get_column_letter(j)].width = 22
    ws.column_dimensions["A"].width = 12
    ws.freeze_panes = "B2"

    # Body
    for i, code in enumerate(tasks, start=2):
        ws.cell(row=i, column=1, value=code).font = bold
        ws.row_dimensions[i].height = 32
        team = ", ".join(sorted(team_by_task.get(code, [])))
        start_idx, end_idx = window_by_task.get(code, (99999, -1))

        for j, d in enumerate(dates, start=2):
            day_idx = j - 2
            text = team if start_idx <= day_idx <= end_idx else ""
            c = ws.cell(row=i, column=j, value=text)
            c.alignment = center
            if dl_date_for.get(code) == d:
                c.fill = deadline_fill

def add_employee_sheet(wb, solution, dates):
    """Sheet 2: rows=employee, cols=dates, cells=task_code(s) (no deadline highlight)."""
    ws = wb.create_sheet(title="Schedule by Employee")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    # Build (employee -> {date -> set(task_codes)})
    emp_day_tasks = defaultdict(lambda: defaultdict(set))

    # For each seat, fill the employee’s tasks across the window
    # (Double-booking overlap is hard-forbidden, but we guard with set just in case.)
    windows_by_task = {w.task_code: (w.start_day_id, w.end_day_id) for w in solution.windows}

    # Map day_idx -> actual date for quick lookup
    idx_to_date = {i: d.d for i, d in enumerate(solution.days)}

    for seat in solution.seats:
        if not seat.employee:
            continue
        start_idx, end_idx = windows_by_task.get(seat.task_code, (None, None))
        if start_idx is None:
            continue
        for day_idx in range(start_idx, end_idx + 1):
            emp_day_tasks[seat.employee.name][idx_to_date[day_idx]].add(seat.task_code)

    # Sort employees by name
    employee_names = sorted([e.name for e in solution.employees])

    # Header
    ws.cell(row=1, column=1, value="employee \\ date").font = bold
    for j, d in enumerate(dates, start=2):
        c = ws.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws.column_dimensions[get_column_letter(j)].width = 18
    ws.column_dimensions["A"].width = 18
    ws.freeze_panes = "B2"

    # Body
    for i, name in enumerate(employee_names, start=2):
        ws.cell(row=i, column=1, value=name).font = bold
        ws.row_dimensions[i].height = 24
        day_map = emp_day_tasks.get(name, {})
        for j, d in enumerate(dates, start=2):
            tasks_today = sorted(day_map.get(d, []))
            text = ", ".join(tasks_today)
            c = ws.cell(row=i, column=j, value=text)
            c.alignment = center

def write_excel_matrix(solution, start_day, deadline_by_task, out_path="schedule_matrix.xlsx"):
    dates = [d.d for d in solution.days]

    wb = Workbook()
    # Sheet 1
    add_task_sheet(wb, solution, start_day, deadline_by_task, dates)
    # Sheet 2
    add_employee_sheet(wb, solution, dates)

    wb.save(out_path)
    print(f"Wrote Excel: {os.path.abspath(out_path)}")

def main():
    ap = ArgumentParser(description="Export schedule matrix for capacity/team model (two sheets)")
    ap.add_argument("--config", default="config.yaml")
    ap.add_argument("--out", default="schedule_matrix.xlsx")
    args = ap.parse_args()

    solution, start_day, deadline_by_task = solve_from_config(args.config)
    print(f"Best score: {solution.score}")
    write_excel_matrix(solution, start_day, deadline_by_task, args.out)

if __name__ == "__main__":
    main()
