# export_csv.py
import os
from argparse import ArgumentParser
from collections import defaultdict
from datetime import timedelta
from typing import Dict, Tuple, List

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# <<< match the minimal mock solver you’re using >>>
from employee_scheduler_mock_tf import solve_from_config

WEEKEND_FILL = PatternFill(start_color="FFEFEF", end_color="FFEFEF", fill_type="solid")
HEADER_BOLD  = Font(bold=True)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT         = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def write_excel(solution, start_day, out_path: str = "schedule_mock.xlsx"):
    """
    Minimal 2-sheet export.
      - Sheet 1: unchanged.
      - Sheet 2: add per-employee country totals after the date columns.
    """
    wb = Workbook()

    # Collect dates from the solved solution
    dates = [d.d for d in solution.days]  # list[date]
    day_id_to_date = {i: d for i, d in enumerate(dates)}

    # Build quick maps from the solved assignments
    # (task_id, day_id) -> [employee_name, ...]
    task_day_to_emps: Dict[Tuple[str, int], List[str]] = defaultdict(list)
    # (employee_name, day_id) -> task_id
    emp_day_to_task: Dict[Tuple[str, int], str] = {}

    # We also want to know each task's country so we can display it in Sheet 1.
    task_country: Dict[str, str] = {}

    # --- NEW: accumulate per-employee per-country day counts ---
    emp_country_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    all_countries: set = set()

    for r in solution.reqs:
        if r.employee is None or r.day is None:
            continue
        tid = r.task_id
        did = r.day.id
        ename = r.employee.name
        ctry = r.country

        task_day_to_emps[(tid, did)].append(ename)
        emp_day_to_task[(ename, did)] = tid
        if tid not in task_country:
            task_country[tid] = ctry

        # count one working day for this employee in this country
        emp_country_counts[ename][ctry] += 1
        all_countries.add(ctry)

    countries_sorted = sorted(all_countries)

    # ---------------- Sheet 1: Tasks x Dates (unchanged) ----------------
    ws1 = wb.active
    ws1.title = "Tasks x Dates"

    ws1.cell(row=1, column=1, value="task_id").font = HEADER_BOLD
    ws1.cell(row=1, column=2, value="country").font = HEADER_BOLD

    for j, d in enumerate(dates, start=3):
        c = ws1.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = HEADER_BOLD
        ws1.column_dimensions[get_column_letter(j)].width = 18
        if d.weekday() >= 5:
            c.fill = WEEKEND_FILL

    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 10
    ws1.freeze_panes = "C2"

    all_task_ids = sorted({r.task_id for r in solution.reqs})

    for i, tid in enumerate(all_task_ids, start=2):
        ws1.cell(row=i, column=1, value=tid).font = HEADER_BOLD
        ws1.cell(row=i, column=1).alignment = LEFT
        ws1.cell(row=i, column=2, value=task_country.get(tid, "")).alignment = LEFT

        for j, d in enumerate(dates, start=3):
            did = j - 3
            emps = sorted(task_day_to_emps.get((tid, did), []))
            txt = f"{', '.join(emps)} ({len(emps)})" if emps else ""
            cell = ws1.cell(row=i, column=j, value=txt)
            cell.alignment = CENTER
            if d.weekday() >= 5:
                cell.fill = WEEKEND_FILL

    # ---------------- Sheet 2: Employees x Dates (+ country totals) ----------------
    ws2 = wb.create_sheet("Employees x Dates")
    ws2.cell(row=1, column=1, value="employee").font = HEADER_BOLD
    ws2.column_dimensions["A"].width = 16

    # date columns
    for j, d in enumerate(dates, start=2):
        c = ws2.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = HEADER_BOLD
        ws2.column_dimensions[get_column_letter(j)].width = 16
        if d.weekday() >= 5:
            c.fill = WEEKEND_FILL

    # --- NEW: country totals on Sheet 2 (after the date columns) ---
    first_ctry_col = 2 + len(dates)
    for idx, ctry in enumerate(countries_sorted):
        col = first_ctry_col + idx
        cell = ws2.cell(row=1, column=col, value=f"{ctry}")
        cell.font = HEADER_BOLD
        ws2.column_dimensions[get_column_letter(col)].width = 10

    total_col = first_ctry_col + len(countries_sorted)
    ws2.cell(row=1, column=total_col, value="Total").font = HEADER_BOLD
    ws2.column_dimensions[get_column_letter(total_col)].width = 10

    ws2.freeze_panes = "B2"

    employees_sorted = sorted([e.name for e in solution.employees])

    for i, ename in enumerate(employees_sorted, start=2):
        ws2.cell(row=i, column=1, value=ename).font = HEADER_BOLD
        ws2.cell(row=i, column=1).alignment = LEFT

        # per-day task
        for j, d in enumerate(dates, start=2):
            did = j - 2
            tid = emp_day_to_task.get((ename, did), "")
            c = ws2.cell(row=i, column=j, value=tid)
            c.alignment = CENTER
            if d.weekday() >= 5:
                c.fill = WEEKEND_FILL

        # --- NEW: fill per-country counts and row total ---
        row_total = 0
        for k, ctry in enumerate(countries_sorted):
            val = int(emp_country_counts.get(ename, {}).get(ctry, 0))
            ws2.cell(row=i, column=first_ctry_col + k, value=val).alignment = CENTER
            row_total += val
        ws2.cell(row=i, column=total_col, value=row_total).alignment = CENTER

    wb.save(out_path)
    print(f"Wrote Excel: {os.path.abspath(out_path)}")



def main():
    ap = ArgumentParser(description="Export mock Timefold schedule to a simple Excel workbook")
    ap.add_argument("--config", default="config_mock.yaml", help="Path to mock YAML config")
    ap.add_argument("--out", default="schedule_mock.xlsx", help="Output .xlsx path")
    args = ap.parse_args()

    # Solve using your mock Timefold solver
    solution, start_day = solve_from_config(args.config)
    print(f"Best score: {solution.score}")

    # Write a simple, readable workbook
    write_excel(solution, start_day, out_path=args.out)


if __name__ == "__main__":
    main()
