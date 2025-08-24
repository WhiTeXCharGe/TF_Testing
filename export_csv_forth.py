# export_csv_forth.py
import os
from argparse import ArgumentParser
from collections import defaultdict
from datetime import timedelta, datetime
from typing import Dict, Tuple

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import yaml

# Import solver from the 4th model:
from employee_scheduler_forth import solve_from_config

LIGHT_BLUE = "ADD8E6"  # module start highlight
RED       = "FF9999"   # task end highlight

def _parse_modules_config(config_path: str):
    """
    Parse the modules YAML to reconstruct:
      - module start day index (per module)
      - per-task (Sx-Py-Z) end day index and workload
      - skill key mapping (Sx-Py-Z -> P?-?)
    Note: Dates are converted to indices relative to global start_day.
    Returns:
      start_day            : date
      horizon_days         : int
      module_start_idx     : Dict[str, int]
      task_end_idx         : Dict[str, int]
      task_workload        : Dict[str, int]
      skill_key_for_task   : Dict[str, str]   # "S1-P3-A" -> "P3-A"
    """
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    start_day = datetime.strptime(str(cfg["start_day"]), "%Y-%m-%d").date()
    horizon_days = int(cfg.get("horizon_days", 30))

    def to_idx(datestr: str) -> int:
        d = datetime.strptime(str(datestr), "%Y-%m-%d").date()
        return max(0, min(horizon_days - 1, (d - start_day).days))

    module_start_idx: Dict[str, int] = {}
    task_end_idx: Dict[str, int] = {}
    task_workload: Dict[str, int] = {}
    skill_key_for_task: Dict[str, str] = {}

    modules = cfg["modules"]
    for m in modules:
        mcode = str(m["code"]).strip()
        m_start = to_idx(m.get("start_date", cfg["start_day"]))
        module_start_idx[mcode] = m_start

        for proc in m["processes"]:
            pid = int(proc["id"])
            p_end = to_idx(proc["end_date"])
            for t in proc["tasks"]:
                full_code = str(t["code"]).strip().upper()  # e.g., "S1-P3-A"
                # Extract 'P?-?' from the middle part:
                parts = full_code.split("-")
                if len(parts) != 3 or not parts[1].startswith("P"):
                    raise ValueError(f"Bad task code '{full_code}' (expected 'Sx-Py-Z').")
                skill_key = f"{parts[1]}-{parts[2]}"  # "P3-A"
                t_end = to_idx(t.get("end_date", proc["end_date"]))
                end_idx = min(p_end, t_end)

                task_end_idx[full_code] = end_idx
                task_workload[full_code] = int(t["workload"])
                skill_key_for_task[full_code] = skill_key

    return start_day, horizon_days, module_start_idx, task_end_idx, task_workload, skill_key_for_task


def write_excel(solution, start_day, config_path: str, out_path: str = "schedule_matrix.xlsx",
                show_skills: bool = True):
    """
    Create an Excel with 2 sheets:

    Sheet 1: "Tasks x Dates"
      - Rows: task_code (Sx-Py-Z) with workload
      - Columns: dates across the horizon
      - Cells: comma-joined employee names (optionally include skill label P?-?:level)
      - Module start column (light blue), task end column (red)

    Sheet 2: "Employees x Dates"
      - Rows: employee names
      - Columns: dates
      - Cells: the task worked that day (optionally 'P?-?:level')
    """
    # Parse config to reconstruct module/task timing and workloads
    (cfg_start_day, horizon_days, module_start_idx, task_end_idx,
     task_workload, skill_key_for_task) = _parse_modules_config(config_path)

    # Dates from solution
    dates = [d.d for d in solution.days]  # list[date]
    date_to_col = {d: j for j, d in enumerate(dates, start=2)}  # for quick lookup

        # Build (task_code, date) -> list of cell strings for Sheet 1
    # And (employee_name, date) -> task code (no level) for Sheet 2
    task_day_to_entries = defaultdict(list)
    emp_day_to_task: Dict[Tuple[str, datetime], str] = {}

    for r in solution.reqs:
        if r.employee is None or r.day is None:
            continue
        full_task_code = f"{r.module}-P{r.process_id}-{r.task_letter}"  # e.g., S1-P3-A
        d = r.day.d
        emp_name = r.employee.name

        # For Sheet 1 (optionally with skills)
        if show_skills:
            sk = f"P{r.process_id}-{r.task_letter}"
            lvl = r.employee.skills.get(sk, "")
            task_day_to_entries[(full_task_code, d)].append(f"{emp_name} ({sk}:{lvl})")
        else:
            task_day_to_entries[(full_task_code, d)].append(emp_name)

        # For Sheet 2 (always show task only, no level)
        emp_day_to_task[(emp_name, d)] = f"P{r.process_id}-{r.task_letter}"


    # Order tasks by module, process, letter
    def task_sort_key(code: str):
        # code = Sx-Py-Z
        s, p, z = code.split("-")
        return (int(s[1:]) if s[0] == "S" and s[1:].isdigit() else s, int(p[1:]), z)

    all_task_codes = sorted(set([t for (t, _) in task_day_to_entries.keys()] + list(task_end_idx.keys())),
                            key=task_sort_key)

    # Styles
    deadline_fill = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
    start_fill    = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    # Workbook
    wb = Workbook()

    # ---------------- Sheet 1: Tasks x Dates ----------------
    ws1 = wb.active
    ws1.title = "Tasks x Dates"

    LABEL_W = 20
    DATE_W  = 24
    ROW_H   = 36

    ws1.cell(row=1, column=1, value="task \\ date").font = bold
    for j, d in enumerate(dates, start=2):
        c = ws1.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws1.column_dimensions[get_column_letter(j)].width = DATE_W
    ws1.column_dimensions["A"].width = LABEL_W
    ws1.freeze_panes = "B2"

    for i, code in enumerate(all_task_codes, start=2):
        # Row label includes workload if known
        wl = task_workload.get(code, None)
        row_label = f"{code}" + (f" (workload {wl})" if wl is not None else "")
        ws1.cell(row=i, column=1, value=row_label).font = bold
        ws1.row_dimensions[i].height = ROW_H

        # Highlight module start column and task end column
        parts = code.split("-")
        mcode = parts[0]
        start_idx = module_start_idx.get(mcode, None)
        end_idx = task_end_idx.get(code, None)
        for j, d in enumerate(dates, start=2):
            text = ", ".join(sorted(task_day_to_entries.get((code, d), [])))
            cell = ws1.cell(row=i, column=j, value=text)
            cell.alignment = center
            # Color start
            if start_idx is not None and (d == (start_day + timedelta(days=start_idx))):
                cell.fill = start_fill
            # Color end
            if end_idx is not None and (d == (start_day + timedelta(days=end_idx))):
                cell.fill = deadline_fill

        # ---------------- Sheet 2: Employees x Dates ----------------
    ws2 = wb.create_sheet("Employees x Dates")

    # Headers: A=employee, B=skills, then dates from C...
    ws2.cell(row=1, column=1, value="employee").font = bold
    ws2.cell(row=1, column=2, value="skills").font = bold
    for j, d in enumerate(dates, start=3):
        c = ws2.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws2.column_dimensions[get_column_letter(j)].width = 20

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 48
    ws2.freeze_panes = "C2"
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Pre-compute skills summary per employee
    skills_by_employee = {}
    for e in solution.employees:
        # Sort skills for a stable, readable string: by process number then task letter
        def _sort_key(k: str):
            # k like "P2-A"
            try:
                p, t = k.split("-")
                return (int(p[1:]), t)
            except Exception:
                return (999, k)
        items = sorted(e.skills.items(), key=lambda kv: _sort_key(kv[0]))
        skills_by_employee[e.name] = ", ".join(f"{k}:{v}" for k, v in items)

    # stable order of employees by name
    employees_sorted = sorted([e.name for e in solution.employees])
    for i, name in enumerate(employees_sorted, start=2):
        # Column A: name
        ws2.cell(row=i, column=1, value=name).font = bold
        ws2.cell(row=i, column=1).alignment = left

        # Column B: skills summary
        ws2.cell(row=i, column=2, value=skills_by_employee.get(name, "")).alignment = left

        # From column C onward: task code only
        for j, d in enumerate(dates, start=3):
            text = emp_day_to_task.get((name, d), "")
            c = ws2.cell(row=i, column=j, value=text)
            c.alignment = center


    # Save
    wb.save(out_path)
    print(f"Wrote Excel: {os.path.abspath(out_path)}")


def main():
    ap = ArgumentParser(description="Export 4th-model schedule to Excel (2 sheets)")
    ap.add_argument("--config", default="config_modules.yaml", help="Path to modules YAML")
    ap.add_argument("--out", default="schedule_matrix.xlsx", help="Output .xlsx path")
    ap.add_argument("--show-skills", default="true", choices=["true","false"], help="Include skill labels")
    args = ap.parse_args()

    show_skills = (args.show_skills.lower() == "true")

    # Solve using the 4th model
    solution, start_day = solve_from_config(args.config)
    print(f"Best score: {solution.score}")

    write_excel(solution, start_day, config_path=args.config, out_path=args.out, show_skills=show_skills)


if __name__ == "__main__":
    main()
