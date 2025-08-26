# export_csv_fifth.py
import os
from argparse import ArgumentParser
from collections import defaultdict
from datetime import timedelta, datetime, date
from typing import Dict, Tuple, List, Set

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import yaml

# >>> CHANGE HERE if your solver module name differs
from employee_scheduler_fifth_1 import solve_from_config  # adjust if needed

LIGHT_BLUE = "ADD8E6"   # module start highlight
RED        = "FF9999"   # task end highlight
PINK       = "FFC0CB"   # weekend / org-unavailable highlight
PALE_RED   = "FFCCCC"   # personal unavailable (Sheet 2)

def _parse_modules_config(config_path: str):
    """
    Parse config_modules.yaml (v5-lite) and reconstruct:
      - start_day (date), horizon_days (int)
      - business_days set (weekday ints: Mon=0..Sun=6)
      - module metadata: company/country
      - calendars: company/country/module => sets of CLOSED dates
      - module start index
      - per-task end index (inherits process end) and workload_days
      - skill key mapping: "S1-P3-A" -> "P3-A"
      - quantum_per_day (for Sheet 2 quanta calc)
    Returns:
      start_day              : date
      horizon_days           : int
      business_days          : Set[int]
      module_company         : Dict[str, str]
      module_country         : Dict[str, str]
      company_unavail_dates  : Dict[str, Set[date]]
      country_unavail_dates  : Dict[str, Set[date]]
      module_unavail_dates   : Dict[str, Set[date]]
      module_start_idx       : Dict[str, int]
      task_end_idx           : Dict[str, int]
      task_workload          : Dict[str, int]
      skill_key_for_task     : Dict[str, str]
      quantum_per_day        : int
    """
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    # Dates / horizon
    start_day = datetime.strptime(str(cfg["start_day"]), "%Y-%m-%d").date()
    horizon_days = int(cfg.get("horizon_days", 30))

    # Business days
    wd_map = {"Mon":0,"Tue":1,"Wed":2,"Thu":3,"Fri":4,"Sat":5,"Sun":6}
    business_days: Set[int] = set(wd_map[d] for d in cfg.get("business_days", ["Mon","Tue","Wed","Thu","Fri"]))

    def to_idx(datestr: str) -> int:
        d = datetime.strptime(str(datestr), "%Y-%m-%d").date()
        # Align to business-day boundary (if config used a weekend)
        while d.weekday() not in business_days:
            # move forward to next business day
            d = d + timedelta(days=1)
        return max(0, min(horizon_days - 1, (d - start_day).days))

    def to_date_set(dates: List[str]) -> Set[date]:
        s: Set[date] = set()
        for ds in (dates or []):
            d = datetime.strptime(str(ds), "%Y-%m-%d").date()
            while d.weekday() not in business_days:
                d = d + timedelta(days=1)
            s.add(d)
        return s

    # Calendars
    calendars = cfg.get("calendars", {}) or {}
    company_unavail_dates: Dict[str, Set[date]] = {}
    for comp, obj in (calendars.get("company", {}) or {}).items():
        company_unavail_dates[str(comp)] = to_date_set(obj.get("unavailable", []))
    country_unavail_dates: Dict[str, Set[date]] = {}
    for ctry, obj in (calendars.get("country", {}) or {}).items():
        country_unavail_dates[str(ctry)] = to_date_set(obj.get("unavailable", []))
    module_unavail_dates: Dict[str, Set[date]] = {}
    for mod, obj in (calendars.get("module", {}) or {}).items():
        module_unavail_dates[str(mod)] = to_date_set(obj.get("unavailable", []))

    # Module metadata
    module_company: Dict[str, str] = {}
    module_country: Dict[str, str] = {}

    module_start_idx: Dict[str, int] = {}
    task_end_idx: Dict[str, int] = {}
    task_workload: Dict[str, int] = {}
    skill_key_for_task: Dict[str, str] = {}

    modules = cfg["modules"]
    for m in modules:
        mcode = str(m["code"]).strip()
        module_company[mcode] = str(m.get("company", "")).strip()
        module_country[mcode] = str(m.get("country", "")).strip()

        m_start = to_idx(m.get("start_date", cfg["start_day"]))
        module_start_idx[mcode] = m_start

        for proc in m["processes"]:
            pid = int(proc["id"])
            # In v5, tasks inherit process end date
            p_end = to_idx(proc["end_date"])

            for t in proc["tasks"]:
                full_code = str(t["code"]).strip().upper()  # e.g., "S1-P3-A"
                parts = full_code.split("-")
                if len(parts) != 3 or not parts[1].startswith("P"):
                    raise ValueError(f"Bad task code '{full_code}' (expected 'Sx-Py-Z').")
                skill_key = f"{parts[1]}-{parts[2]}"  # "P3-A"

                # workload_days in v5
                wl = int(t.get("workload_days", 0))
                task_end_idx[full_code] = p_end
                task_workload[full_code] = wl
                skill_key_for_task[full_code] = skill_key

    # Quantum per day (for Sheet 2 quanta calc)
    quantum_per_day = int(((cfg.get("quantum") or {}).get("per_day") or 1))

    return (start_day, horizon_days, business_days,
            module_company, module_country,
            company_unavail_dates, country_unavail_dates, module_unavail_dates,
            module_start_idx, task_end_idx, task_workload, skill_key_for_task,
            quantum_per_day)


def write_excel(solution, start_day, config_path: str, out_path: str = "schedule_matrix.xlsx",
                show_skills: bool = True):
    """
    Create an Excel with 2 sheets:

    Sheet 1: "Tasks x Dates"
      - Col A: module details "Sx | company X | country Y"
      - Col B: task-only label "Py-Z"
      - Columns C..: dates across the horizon
      - Cells: "AA (3), AF (1) | avg=2.0"
      - Module start column (light blue), task end column (red)
      - Weekends and org-unavailable (module/company/country) = pink

    Sheet 2: "Employees x Dates"
      - Col A: employee
      - Col B: skills summary
      - Col C: personal unavailable dates (CSV)
      - Columns D..: dates
      - Cell red if employee is personally unavailable that day
      - Row end: "Workdays", "Quanta" (sum). Bottom row: averages for those two.
    """
    # Parse config to reconstruct timing, calendars, workloads, meta
    (cfg_start_day, horizon_days, business_days,
     module_company, module_country,
     company_unavail_dates, country_unavail_dates, module_unavail_dates,
     module_start_idx, task_end_idx, task_workload, skill_key_for_task,
     quantum_per_day) = _parse_modules_config(config_path)

    # Dates from solution
    dates = [d.d for d in solution.days]  # list[date]

    # Build (task_code, date) -> list of "AA (3)" for Sheet 1
    # And (employee_name, date) -> task code (no level) for Sheet 2
    task_day_to_entries = defaultdict(list)
    emp_day_to_task: Dict[Tuple[str, date], str] = {}

    for r in solution.reqs:
        if r.employee is None or r.day is None:
            continue
        full_task_code = f"{r.module}-P{r.process_id}-{r.task_letter}"  # e.g., S1-P3-A
        d = r.day.d
        emp_name = r.employee.name

        # For Sheet 1 (optionally with skills)
        if show_skills:
            sk = f"P{r.process_id}-{r.task_letter}"
            lvl = r.employee.skills.get(sk, "")
            # show only level in parentheses (no "P?-?:" prefix)
            task_day_to_entries[(full_task_code, d)].append(f"{emp_name} ({lvl})")
        else:
            task_day_to_entries[(full_task_code, d)].append(emp_name)

        # For Sheet 2: task code only
        emp_day_to_task[(emp_name, d)] = f"{r.module}-P{r.process_id}-{r.task_letter}"

    # Order tasks by module, process, letter
    def task_sort_key(code: str):
        # code = Sx-Py-Z
        s, p, z = code.split("-")
        return (int(s[1:]) if s[0] == "S" and s[1:].isdigit() else s, int(p[1:]), z)

    all_task_codes = sorted(
        set([t for (t, _) in task_day_to_entries.keys()] + list(task_end_idx.keys())),
        key=task_sort_key
    )

    # Styles
    fill_deadline = PatternFill(start_color=RED,  end_color=RED,  fill_type="solid")
    fill_start    = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    fill_pink     = PatternFill(start_color=PINK, end_color=PINK, fill_type="solid")
    fill_pale_red = PatternFill(start_color=PALE_RED, end_color=PALE_RED, fill_type="solid")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    bold   = Font(bold=True)

    # Workbook
    wb = Workbook()

    # ---------------- Sheet 1: Tasks x Dates ----------------
    ws1 = wb.active
    ws1.title = "Tasks x Dates"

    LABEL1_W = 26  # module details
    LABEL2_W = 12  # task label
    DATE_W   = 24
    ROW_H    = 36

    ws1.cell(row=1, column=1, value="module | company | country").font = bold  # <<< CHANGED
    ws1.cell(row=1, column=2, value="task").font = bold                         # <<< CHANGED
    for j, d in enumerate(dates, start=3):                                       # dates start at col 3
        c = ws1.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws1.column_dimensions[get_column_letter(j)].width = DATE_W
    ws1.column_dimensions["A"].width = LABEL1_W
    ws1.column_dimensions["B"].width = LABEL2_W
    ws1.freeze_panes = "C2"

    for i, code in enumerate(all_task_codes, start=2):
        # code = Sx-Py-Z
        parts = code.split("-")
        mcode, ppart, letter = parts[0], parts[1], parts[2]
        # Row labels
        detail = f"{mcode} | company {module_company.get(mcode,'')} | country {module_country.get(mcode,'')}"
        ws1.cell(row=i, column=1, value=detail).font = bold
        ws1.cell(row=i, column=1).alignment = left

        ws1.cell(row=i, column=2, value=f"{ppart}-{letter}").font = bold
        ws1.row_dimensions[i].height = ROW_H

        # start/end highlights
        start_idx = module_start_idx.get(mcode, None)
        end_idx = task_end_idx.get(code, None)

        for j, d in enumerate(dates, start=3):
            # weekend/org-unavailable background (pink)
            is_weekend = (d.weekday() not in business_days)
            org_closed = (
                (d in (module_unavail_dates.get(mcode, set()))) or
                (d in (company_unavail_dates.get(module_company.get(mcode,""), set()))) or
                (d in (country_unavail_dates.get(module_country.get(mcode,""), set())))
            )

            # cell contents
            entries = task_day_to_entries.get((code, d), [])
            # compute avg skill in this cell
            avg_str = ""
            if entries:
                # derive levels reliably from employee records
                # find all reqs that match this (code,d) and collect level from employee.skills
                # We reconstruct from entries text by reading levels in parentheses:
                levels = []
                for s in entries:
                    # s like "AA (3)"
                    if "(" in s and s.endswith(")"):
                        try:
                            lvl = int(s.split("(")[-1][:-1])
                            levels.append(lvl)
                        except Exception:
                            pass
                if levels:
                    avg_val = round(sum(levels) / len(levels), 2)
                    avg_str = f" | avg={avg_val}"
            text = ", ".join(sorted(entries)) + (avg_str if avg_str else "")
            cell = ws1.cell(row=i, column=j, value=text)
            cell.alignment = center

            # background for weekend/org-unavailable
            if is_weekend or org_closed:
                cell.fill = fill_pink

            # start/end overlays (draw after base fill)
            if start_idx is not None and (d == (start_day + timedelta(days=start_idx))):
                cell.fill = fill_start
            if end_idx is not None and (d == (start_day + timedelta(days=end_idx))):
                cell.fill = fill_deadline

    # ---------------- Sheet 2: Employees x Dates ----------------
    ws2 = wb.create_sheet("Employees x Dates")

    # Headers: A=employee, B=skills, C=unavailable dates, then dates from D...
    ws2.cell(row=1, column=1, value="employee").font = bold
    ws2.cell(row=1, column=2, value="skills").font = bold
    ws2.cell(row=1, column=3, value="unavailable").font = bold
    for j, d in enumerate(dates, start=4):
        c = ws2.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws2.column_dimensions[get_column_letter(j)].width = 20

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 48
    ws2.column_dimensions["C"].width = 28
    ws2.freeze_panes = "D2"
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # Pre-compute skills and personal unavailable per employee
    skills_by_employee: Dict[str, str] = {}
    personal_unavail: Dict[str, Set[date]] = {}

    for e in solution.employees:
        # Skills summary string (sorted P then letter)
        def _sort_key(k: str):
            try:
                p, t = k.split("-")
                return (int(p[1:]), t)
            except Exception:
                return (999, k)
        items = sorted(e.skills.items(), key=lambda kv: _sort_key(kv[0]))
        skills_by_employee[e.name] = ", ".join(f"{k}:{v}" for k, v in items)

        # Collect unavailable as dates from day ids present in solution.days
        did_to_date = {i: d for i, d in enumerate(dates)}
        udates = set()
        # employee has .unavailable_day_ids (set of indices)
        if hasattr(e, "unavailable_day_ids"):
            for did in e.unavailable_day_ids:
                if did in did_to_date:
                    udates.add(did_to_date[did])
        personal_unavail[e.name] = udates

    # stable order of employees by name
    employees_sorted = sorted([e.name for e in solution.employees])

    # Track per-employee counts
    workday_counts: Dict[str, int] = defaultdict(int)
    quanta_counts: Dict[str, int] = defaultdict(int)

    for i, name in enumerate(employees_sorted, start=2):
        ws2.row_dimensions[i].height = 36

        # Column A: name
        ws2.cell(row=i, column=1, value=name).font = bold
        ws2.cell(row=i, column=1).alignment = left

        # Column B: skills summary
        ws2.cell(row=i, column=2, value=skills_by_employee.get(name, "")).alignment = left

        # Column C: personal unavailable dates (CSV)
        unav_csv = ", ".join(sorted([d.strftime("%Y-%m-%d") for d in personal_unavail.get(name, set())]))
        ws2.cell(row=i, column=3, value=unav_csv).alignment = left

        # From column D onward: task code only
        for j, d in enumerate(dates, start=4):
            text = emp_day_to_task.get((name, d), "")
            c = ws2.cell(row=i, column=j, value=text)
            c.alignment = center
            # highlight personal unavailable day
            if d in personal_unavail.get(name, set()):
                c.fill = fill_pale_red
            # count workdays/quanta if assigned
            if text:
                workday_counts[name] += 1
                quanta_counts[name] += quantum_per_day

    # Append Workdays / Quanta columns at the end
    workdays_col = len(dates) + 4
    quanta_col   = len(dates) + 5
    ws2.cell(row=1, column=workdays_col, value="Workdays").font = bold
    ws2.cell(row=1, column=quanta_col,   value="Quanta").font   = bold
    ws2.column_dimensions[get_column_letter(workdays_col)].width = 12
    ws2.column_dimensions[get_column_letter(quanta_col)].width   = 12

    # Fill totals
    for i, name in enumerate(employees_sorted, start=2):
        ws2.cell(row=i, column=workdays_col, value=workday_counts.get(name, 0)).alignment = center
        ws2.cell(row=i, column=quanta_col,   value=quanta_counts.get(name, 0)).alignment   = center

    # AVERAGE row at the bottom
    avg_row = ws2.max_row + 1
    ws2.cell(row=avg_row, column=1, value="AVERAGE").font = bold
    # skills/unavailable left blank
    # compute averages for Workdays and Quanta
    if employees_sorted:
        avg_work = round(sum(workday_counts.values()) / len(employees_sorted), 2)
        avg_quan = round(sum(quanta_counts.values()) / len(employees_sorted), 2)
    else:
        avg_work = 0
        avg_quan = 0
    ws2.cell(row=avg_row, column=workdays_col, value=avg_work).alignment = center
    ws2.cell(row=avg_row, column=quanta_col,   value=avg_quan).alignment = center

    # Save
    wb.save(out_path)
    print(f"Wrote Excel: {os.path.abspath(out_path)}")


def main():
    ap = ArgumentParser(description="Export v5 schedule to Excel (2 sheets)")
    ap.add_argument("--config", default="config_modules.yaml", help="Path to modules YAML")
    ap.add_argument("--out", default="schedule_matrix.xlsx", help="Output .xlsx path")
    ap.add_argument("--show-skills", default="true", choices=["true","false"], help="Include skill labels")
    args = ap.parse_args()

    show_skills = (args.show_skills.lower() == "true")

    solution, start_day = solve_from_config(args.config)
    print(f"Best score: {solution.score}")

    write_excel(solution, start_day, config_path=args.config, out_path=args.out, show_skills=show_skills)


if __name__ == "__main__":
    main()
