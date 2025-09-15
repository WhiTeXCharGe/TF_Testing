# export_csv_overtime_2.py
import os
from argparse import ArgumentParser
from collections import defaultdict
from datetime import timedelta, datetime
from typing import Dict, Tuple, List

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import csv

from employee_scheduler_overtime_2 import solve_from_config

LIGHT_BLUE = "ADD8E6"   # module start highlight
RED        = "FF9999"   # task end highlight

def _parse_modules_config(config_path: str):
    import yaml
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    start_day = datetime.strptime(str(cfg["start_day"]), "%Y-%m-%d").date()
    horizon_days = int(cfg.get("horizon_days", 30))

    def to_idx(datestr: str) -> int:
        d = datetime.strptime(str(datestr), "%Y-%m-%d").date()
        return max(0, min(horizon_days - 1, (d - start_day).days))

    module_start_idx: Dict[str, int] = {}
    task_end_idx: Dict[str, int] = {}
    module_factory: Dict[str, str] = {}

    for m in cfg["modules"]:
        mcode = str(m["code"]).strip()
        module_factory[mcode] = str(m.get("factory", "")).strip()
        m_start = to_idx(m.get("start_date", cfg["start_day"]))
        module_start_idx[mcode] = m_start
        for proc in m["processes"]:
            p_end = to_idx(proc["end_date"])
            for t in proc["tasks"]:
                full_code = str(t["code"]).strip().upper()  # e.g., S1-P2-A
                task_end_idx[full_code] = p_end

    return start_day, horizon_days, module_start_idx, task_end_idx, module_factory

def _aggregate_from_tokens(solution, start_day):
    tde_hours = defaultdict(int)
    edt_hours = defaultdict(int)

    for u in solution.tokens:
        if u.employee is None or u.day is None or (u.hours or 0) <= 0:
            continue
        code = f"{u.module}-P{u.process_id}-{u.task_letter}"
        d = u.day.d
        tde_hours[(code, d, u.employee.name)] += int(u.hours)
        edt_hours[(u.employee.name, d, code)] += int(u.hours)

    return tde_hours, edt_hours

def write_excel(solution, start_day, config_path: str,
                out_path: str = "schedule_matrix_overtime_units.xlsx",
                show_skills: bool = True):
    cfg_start_day, horizon_days, module_start_idx, task_end_idx, module_factory = _parse_modules_config(config_path)

    dates = [cfg_start_day + timedelta(days=i) for i in range(horizon_days)]
    tde_hours, edt_hours = _aggregate_from_tokens(solution, cfg_start_day)

    wb = Workbook()
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    fill_deadline = PatternFill(start_color=RED,  end_color=RED,  fill_type="solid")
    fill_start    = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    # -------------- Sheet 1: Tasks x Dates --------------
    ws1 = wb.active
    ws1.title = "Tasks x Dates"

    ws1.cell(row=1, column=1, value="module").font = bold
    ws1.cell(row=1, column=2, value="factory").font = bold
    ws1.cell(row=1, column=3, value="task").font = bold
    for j, d in enumerate(dates, start=4):
        c = ws1.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws1.column_dimensions[get_column_letter(j)].width = 28
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 14
    ws1.freeze_panes = "D2"

    all_task_codes = sorted(set([k[0] for k in tde_hours.keys()] + list(task_end_idx.keys())))

    def task_sort_key(code: str):
        try:
            s, p, z = code.split("-")
            sN = int(s[1:]) if s.startswith("S") else 999
            pN = int(p[1:]) if p.startswith("P") else 999
            return (sN, pN, z)
        except Exception:
            return (999, 999, code)

    all_task_codes.sort(key=task_sort_key)

    for i, code in enumerate(all_task_codes, start=2):
        mcode, ppart, letter = code.split("-")
        ws1.cell(row=i, column=1, value=mcode).font = bold
        ws1.cell(row=i, column=2, value=module_factory.get(mcode, "")).font = bold
        ws1.cell(row=i, column=3, value=f"{ppart}-{letter}").font = bold

        start_idx = module_start_idx.get(mcode, None)
        end_idx   = task_end_idx.get(code, None)

        for j, d in enumerate(dates, start=4):
            per_emp = {emp: h for (tcode, day, emp), h in tde_hours.items()
                       if tcode == code and day == d}
            entries: List[str] = []
            for emp, hrs in sorted(per_emp.items()):
                lvl = ""
                if show_skills:
                    for e in solution.employees:
                        if e.name == emp:
                            lvl = e.skills.get(f"{ppart}-{letter}", "")
                            break
                entries.append(f"{emp}({lvl},{hrs}H)" if lvl != "" else f"{emp}({hrs}H)")
            text = " | ".join(entries)
            c = ws1.cell(row=i, column=j, value=text)
            c.alignment = center

            if start_idx is not None and d == (cfg_start_day + timedelta(days=start_idx)):
                c.fill = fill_start
            if end_idx is not None and d == (cfg_start_day + timedelta(days=end_idx)):
                c.fill = fill_deadline

    ws1.row_dimensions[1].height = 24
    for r in range(2, ws1.max_row + 1):
        ws1.row_dimensions[r].height = 36

    # -------------- Sheet 2: Employees x Dates --------------
    ws2 = wb.create_sheet("Employees x Dates")
    ws2.cell(row=1, column=1, value="employee").font = bold
    ws2.cell(row=1, column=2, value="skills").font = bold

    for j, d in enumerate(dates, start=3):
        c = ws2.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws2.column_dimensions[get_column_letter(j)].width = 32

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 48
    ws2.freeze_panes = "C2"

    def skills_text(e) -> str:
        def skey(k: str):
            try:
                p, t = k.split("-")
                return (int(p[1:]), t)
            except Exception:
                return (999, k)
        items = sorted(e.skills.items(), key=lambda kv: skey(kv[0]))
        return ", ".join(f"{k}:{v}" for k, v in items)

    employees_sorted = sorted([e for e in solution.employees if e.id != 0], key=lambda e: e.name)

    workday_counts = defaultdict(int)
    workhour_counts = defaultdict(int)

    for i, e in enumerate(employees_sorted, start=2):
        ws2.row_dimensions[i].height = 36
        ws2.cell(row=i, column=1, value=e.name).font = bold
        ws2.cell(row=i, column=1).alignment = left
        ws2.cell(row=i, column=2, value=skills_text(e)).alignment = left

        for j, d in enumerate(dates, start=3):
            tasks_here = [(t, hrs) for (emp, day, t), hrs in edt_hours.items()
                          if emp == e.name and day == d]
            tasks_here.sort(key=lambda kv: kv[0])

            if tasks_here:
                workday_counts[e.name] += 1
                total_h = sum(int(h) for _, h in tasks_here)
                workhour_counts[e.name] += total_h
                text = " | ".join(f"{t} ({int(h)}H)" for t, h in tasks_here)
            else:
                text = ""
            ws2.cell(row=i, column=j, value=text).alignment = center

    workdays_col  = len(dates) + 3
    workhours_col = len(dates) + 4
    ws2.cell(row=1, column=workdays_col,  value="Workdays").font  = bold
    ws2.cell(row=1, column=workhours_col, value="WorkHours").font = bold
    ws2.column_dimensions[get_column_letter(workdays_col)].width  = 12
    ws2.column_dimensions[get_column_letter(workhours_col)].width = 12

    for i, e in enumerate(employees_sorted, start=2):
        ws2.cell(row=i, column=workdays_col,  value=workday_counts[e.name]).alignment  = center
        ws2.cell(row=i, column=workhours_col, value=workhour_counts[e.name]).alignment = center

    wb.save(out_path)
    print(f"Wrote Excel: {os.path.abspath(out_path)}")

def write_csvs(solution, start_day,
               tasks_csv: str = "tasks_by_date.csv",
               employees_csv: str = "employees_by_date.csv"):
    cfg_start_day, horizon_days, *_ = _parse_modules_config("config_modules.yaml")

    tde_hours, edt_hours = _aggregate_from_tokens(solution, cfg_start_day)

    with open(tasks_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["task_code", "date", "employee", "hours"])
        for (code, d, emp), hrs in sorted(tde_hours.items()):
            w.writerow([code, d.isoformat(), emp, hrs])
    print(f"Wrote CSV: {os.path.abspath(tasks_csv)}")

    with open(employees_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["employee", "date", "task_code", "hours"])
        for (emp, d, code), hrs in sorted(edt_hours.items()):
            w.writerow([emp, d.isoformat(), code, hrs])
    print(f"Wrote CSV: {os.path.abspath(employees_csv)}")

def _unpack_solutions(result_tuple):
    """
    Accepts tuples of variable length:
      (final, start_day)
      (final, start_day, p1)
      (final, start_day, p1, p2)
      (final, start_day, p1, p2, p3)
    Returns: (final, start_day, [p1, p2, p3])
    """
    if isinstance(result_tuple, tuple) and len(result_tuple) >= 2:
        final = result_tuple[0]
        start_day = result_tuple[1]
        snapshots = list(result_tuple[2:])  # pass1..passN-1
        return final, start_day, snapshots
    # fallback
    return result_tuple, None, []

def main():
    ap = ArgumentParser(description="Export overtime schedule to Excel/CSV (Pass1/Pass2/Pass3/Final)")
    ap.add_argument("--config", default="config_modules.yaml", help="Path to modules YAML")
    ap.add_argument("--xlsx",   default="schedule_matrix_overtime_units.xlsx", help="(kept for compatibility; final file path if you want)")
    ap.add_argument("--tasks_csv", default="tasks_by_date_final.csv", help="Final Tasks CSV path")
    ap.add_argument("--emps_csv",  default="employees_by_date_final.csv", help="Final Employees CSV path")
    args = ap.parse_args()

    solved = solve_from_config(args.config)
    final, start_day, snapshots = _unpack_solutions(solved)

    # Write as many pass snapshots as we have, naming stays consistent.
    names = [
        "employee_schedule_overtime_pass1.xlsx",
        "employee_schedule_overtime_pass2.xlsx",
        "employee_schedule_overtime_pass3.xlsx",
    ]
    for i, snap in enumerate(snapshots[:3]):  # support up to 3 passes
        if snap is not None:
            write_excel(snap, start_day, config_path=args.config, out_path=names[i])

    # Always write final
    write_excel(final, start_day, config_path=args.config, out_path="employee_schedule_overtime_final.xlsx")

    # Final CSVs
    write_csvs(final, start_day, tasks_csv=args.tasks_csv, employees_csv=args.emps_csv)
    print("All exports done.")

if __name__ == "__main__":
    main()
