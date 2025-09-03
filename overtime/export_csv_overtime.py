# export_csv_overtime.py
import os
from argparse import ArgumentParser
from collections import defaultdict
from datetime import timedelta, datetime, date
from typing import Dict, Tuple, List, Set

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import yaml

# Uses the overtime solver
from employee_scheduler_overtime import solve_from_config

LIGHT_BLUE = "ADD8E6"   # module start highlight
RED        = "FF9999"   # task end highlight

def _parse_modules_config(config_path: str):
    """
    Parse config_modules.yaml (overtime) and reconstruct minimal timing info:
      - start_day (date), horizon_days (int)
      - module start index
      - per-task end index (inherits process end)
    Returns:
      start_day        : date
      horizon_days     : int
      module_start_idx : Dict[str, int]
      task_end_idx     : Dict[str, int]
    """
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    # Dates / horizon
    start_day = datetime.strptime(str(cfg["start_day"]), "%Y-%m-%d").date()
    horizon_days = int(cfg.get("horizon_days", 30))

    def to_idx(datestr: str) -> int:
        d = datetime.strptime(str(datestr), "%Y-%m-%d").date()
        return max(0, min(horizon_days - 1, (d - start_day).days))

    module_start_idx: Dict[str, int] = {}
    task_end_idx: Dict[str, int] = {}

    modules = cfg["modules"]
    for m in modules:
        mcode = str(m["code"]).strip()
        m_start = to_idx(m.get("start_date", cfg["start_day"]))
        module_start_idx[mcode] = m_start

        for proc in m.get("processes", []):
            pid = int(proc["id"])
            p_end = to_idx(proc["end_date"])
            for t in proc.get("tasks", []):
                full_code = str(t["code"]).strip().upper()  # e.g., "S1-P3-A"
                task_end_idx[full_code] = p_end

    return start_day, horizon_days, module_start_idx, task_end_idx


def write_excel(solution, start_day, config_path: str, out_path: str = "schedule_matrix_overtime.xlsx"):
    """
    Create an Excel with 2 sheets (hour-based, no unavailable/weekend shading):

    Sheet 1: "Tasks x Dates"
      - Col A: module code (Sx)
      - Col B: task-only label "Py-Z"
      - Columns C..: dates across the horizon
      - Cells: entries per employee working that task on that date, formatted: "AA(3,8H)"
               where 3 is skill level, 8H is hours that day on that task by that employee.
      - Module start column (light blue), task end column (red).

    Sheet 2: "Employees x Dates"
      - Col A: employee
      - Columns B..: dates
      - Each cell: list of task codes with hours that day, e.g., "S1-P1-A (6H) | S1-P1-B (2H)"
      - Last two columns: "Workdays" (days with any hours) and "WorkHours" (sum of hours).
    """
    # Parse config for start/end highlighting and task list
    (cfg_start_day, horizon_days, module_start_idx, task_end_idx) = _parse_modules_config(config_path)

    # Dates from solution
    dates = [d.d for d in solution.days]  # list[date]

    # --------- Aggregate hours ---------
    # (task_code, date, employee) -> hours
    tde_hours: Dict[Tuple[str, date, str], int] = defaultdict(int)
    # (employee, date, task_code) -> hours
    edt_hours: Dict[Tuple[str, date, str], int] = defaultdict(int)
    # skill cache: (employee_name, pcode) -> level
    skill_level: Dict[Tuple[str, str], int] = {}

    for r in solution.reqs:
        if r.employee is None or r.day is None:
            continue
        full_task_code = f"{r.module}-P{r.process_id}-{r.task_letter}"  # e.g., S1-P3-A
        d = r.day.d
        emp_name = r.employee.name
        pcode = f"P{r.process_id}-{r.task_letter}"
        lvl = r.employee.skills.get(pcode, 0)
        skill_level[(emp_name, pcode)] = lvl

        tde_hours[(full_task_code, d, emp_name)] += 1
        edt_hours[(emp_name, d, full_task_code)] += 1

    # Build sets for ordering and iteration
    all_task_codes = sorted(set([t for (t, _, _) in tde_hours.keys()] + list(task_end_idx.keys())),
                            key=lambda code: (int(code.split('-')[0][1:]) if code.startswith('S') and code.split('-')[0][1:].isdigit() else 9999,
                                              int(code.split('-')[1][1:]) if code.split('-')[1][1:].isdigit() else 9999,
                                              code.split('-')[2]))

    # Styles
    fill_deadline = PatternFill(start_color=RED,  end_color=RED,  fill_type="solid")
    fill_start    = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    bold   = Font(bold=True)

    # Workbook
    wb = Workbook()

    # ---------------- Sheet 1: Tasks x Dates ----------------
    ws1 = wb.active
    ws1.title = "Tasks x Dates"

    LABEL1_W = 10  # module code
    LABEL2_W = 10  # task label
    DATE_W   = 24
    ROW_H    = 36

    ws1.cell(row=1, column=1, value="module").font = bold
    ws1.cell(row=1, column=2, value="task").font = bold
    for j, d in enumerate(dates, start=3):  # dates start at col 3
        c = ws1.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws1.column_dimensions[get_column_letter(j)].width = DATE_W
    ws1.column_dimensions["A"].width = LABEL1_W
    ws1.column_dimensions["B"].width = LABEL2_W
    ws1.freeze_panes = "C2"

    for i, code in enumerate(all_task_codes, start=2):
        # code = Sx-Py-Z
        parts = code.split("-")
        mcode, ppart, letter = parts[0], parts[1], parts[2]
        # Row labels
        ws1.cell(row=i, column=1, value=mcode).font = bold
        ws1.cell(row=i, column=1).alignment = left

        ws1.cell(row=i, column=2, value=f"{ppart}-{letter}").font = bold
        ws1.row_dimensions[i].height = ROW_H

        # start/end highlights
        start_idx = module_start_idx.get(mcode, None)
        end_idx = task_end_idx.get(code, None)

        for j, d in enumerate(dates, start=3):
            # Gather entries for this task/date
            entries = []
            # We want one line per employee on that day for this task, with hours
            # Sort employees for determinism
            emps_here = [(emp, hrs) for (t, day, emp), hrs in tde_hours.items() if t == code and day == d]
            emps_here.sort(key=lambda kv: kv[0])
            for emp, hrs in emps_here:
                lvl = skill_level.get((emp, f"{ppart}-{letter}"), 0)
                entries.append(f"{emp}({lvl},{hrs}H)")
            text = ", ".join(entries)
            cell = ws1.cell(row=i, column=j, value=text)
            cell.alignment = center

            # start/end overlays
            if start_idx is not None and (d == (start_day + timedelta(days=start_idx))):
                cell.fill = fill_start
            if end_idx is not None and (d == (start_day + timedelta(days=end_idx))):
                cell.fill = fill_deadline

    # ---------------- Sheet 2: Employees x Dates ----------------
    ws2 = wb.create_sheet("Employees x Dates")

    # Headers: A=employee, then dates, then Workdays and WorkHours
    ws2.cell(row=1, column=1, value="employee").font = bold
    for j, d in enumerate(dates, start=2):
        c = ws2.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws2.column_dimensions[get_column_letter(j)].width = 30

    ws2.column_dimensions["A"].width = 16
    ws2.freeze_panes = "B2"
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # stable order of employees by name
    employees_sorted = sorted([e.name for e in solution.employees])

    # Track per-employee totals
    workday_counts: Dict[str, int] = defaultdict(int)
    workhour_counts: Dict[str, int] = defaultdict(int)

    for i, name in enumerate(employees_sorted, start=2):
        ws2.row_dimensions[i].height = 36

        # Column A: name
        ws2.cell(row=i, column=1, value=name).font = bold
        ws2.cell(row=i, column=1).alignment = left

        # From column B onward: for each date, join tasks with hours
        days_with_hours = 0
        total_hours = 0
        for j, d in enumerate(dates, start=2):
            # collect tasks for this (employee, date)
            tasks_here = [(t, hrs) for (emp, day, t), hrs in edt_hours.items() if emp == name and day == d]
            tasks_here.sort(key=lambda kv: kv[0])
            if tasks_here:
                days_with_hours += 1
                sumh = sum(h for _, h in tasks_here)
                total_hours += sumh
                s = " | ".join(f"{t} ({h}H)" for t, h in tasks_here)
            else:
                s = ""
            c = ws2.cell(row=i, column=j, value=s)
            c.alignment = center

        # Totals columns at the end
        workdays_col = len(dates) + 2
        workhours_col = len(dates) + 3
        ws2.cell(row=1, column=workdays_col, value="Workdays").font = bold
        ws2.cell(row=1, column=workhours_col, value="WorkHours").font = bold
        ws2.column_dimensions[get_column_letter(workdays_col)].width = 12
        ws2.column_dimensions[get_column_letter(workhours_col)].width = 12

        ws2.cell(row=i, column=workdays_col, value=days_with_hours).alignment = center
        ws2.cell(row=i, column=workhours_col, value=total_hours).alignment = center

        workday_counts[name] = days_with_hours
        workhour_counts[name] = total_hours

    # Save
    wb.save(out_path)
    print(f"Wrote Excel: {os.path.abspath(out_path)}")


def main():
    ap = ArgumentParser(description="Export overtime schedule to Excel (2 sheets, with hours)")
    ap.add_argument("--config", default="config_modules.yaml", help="Path to modules YAML")
    ap.add_argument("--out", default="schedule_matrix_overtime.xlsx", help="Output .xlsx path")
    args = ap.parse_args()

    solution, start_day = solve_from_config(args.config)
    print(f"Best score: {solution.score}")

    write_excel(solution, start_day, config_path=args.config, out_path=args.out)


if __name__ == "__main__":
    main()
