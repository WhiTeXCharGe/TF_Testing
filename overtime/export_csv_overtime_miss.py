# export_csv_overtime_2.py
import os
from argparse import ArgumentParser
from collections import defaultdict
from datetime import timedelta, datetime
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
import csv

from employee_scheduler_overtime_2 import solve_from_config

LIGHT_BLUE = "ADD8E6"   # module start highlight
RED        = "FF9999"   # task end highlight

def _parse_modules_config(config_path: str):
    import yaml
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    start_day = datetime.strptime(str(cfg["start_day"]), "%Y-%m-%d").date()
    horizon_days = int(cfg.get("horizon_days", 30))
    unit_hours = int(((cfg.get("quantum") or {}).get("unit_hours") or 8))

    def to_idx(datestr: str) -> int:
        d = datetime.strptime(str(datestr), "%Y-%m-%d").date()
        return max(0, min(horizon_days - 1, (d - start_day).days))

    module_start_idx: Dict[str, int] = {}
    task_end_idx: Dict[str, int] = {}

    for m in cfg["modules"]:
        mcode = str(m["code"]).strip()
        m_start = to_idx(m.get("start_date", cfg["start_day"]))
        module_start_idx[mcode] = m_start
        for proc in m["processes"]:
            p_end = to_idx(proc["end_date"])
            for t in proc["tasks"]:
                full_code = str(t["code"]).strip().upper()
                task_end_idx[full_code] = p_end

    return start_day, horizon_days, unit_hours, module_start_idx, task_end_idx

def write_excel(solution, start_day, config_path: str,
                out_path: str = "schedule_matrix_overtime_units.xlsx",
                show_skills: bool = True):
    dates = [d.d for d in solution.days]  # list[date]

    # Aggregations using VARIABLE hours
    # Sheet 1: (task_code, date, employee) -> sum hours
    tde_hours = defaultdict(int)
    # Sheet 2: (employee, date, task_code) -> sum hours
    edt_hours = defaultdict(int)

    for u in solution.units:
        if u.employee is None or u.day is None or (u.hours or 0) <= 0:
            continue
        code = f"{u.module}-P{u.process_id}-{u.task_letter}"
        d = u.day.d
        tde_hours[(code, d, u.employee.name)] += int(u.hours)
        edt_hours[(u.employee.name, d, code)] += int(u.hours)

    cfg_start_day, horizon_days, unit_hours, module_start_idx, task_end_idx = _parse_modules_config(config_path)

    wb = Workbook()
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    fill_deadline = PatternFill(start_color=RED,  end_color=RED,  fill_type="solid")
    fill_start    = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    # -------------- Sheet 1: Tasks x Dates --------------
    ws1 = wb.active
    ws1.title = "Tasks x Dates"

    ws1.cell(row=1, column=1, value="module").font = bold
    ws1.cell(row=1, column=2, value="task").font = bold
    for j, d in enumerate(dates, start=3):
        c = ws1.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws1.column_dimensions[get_column_letter(j)].width = 28
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 14
    ws1.freeze_panes = "C2"

    # Collect all task codes so deadlines appear even if empty
    all_task_codes = sorted(set([k[0] for k in tde_hours.keys()] + list(task_end_idx.keys())))

    def task_sort_key(code: str):
        # code = "Sx-Py-Z"
        try:
            s, p, z = code.split("-")
            sN = int(s[1:]) if s.startswith("S") else 999
            pN = int(p[1:]) if p.startswith("P") else 999
            return (sN, pN, z)
        except Exception:
            return (999, 999, code)

    all_task_codes.sort(key=task_sort_key)

    for i, code in enumerate(all_task_codes, start=2):
        mcode, ppart, letter = code.split("-")
        ws1.cell(row=i, column=1, value=mcode).font = bold
        ws1.cell(row=i, column=2, value=f"{ppart}-{letter}").font = bold

        start_idx = module_start_idx.get(mcode, None)
        end_idx   = task_end_idx.get(code, None)

        for j, d in enumerate(dates, start=3):
            # Per-employee entries "AA(3,8H) | AB(2,6H)"
            per_emp = {emp: h for (tcode, day, emp), h in tde_hours.items()
                       if tcode == code and day == d}
            entries = []
            for emp, hrs in sorted(per_emp.items()):
                lvl = ""
                if show_skills:
                    for e in solution.employees:
                        if e.name == emp:
                            lvl = e.skills.get(f"{ppart}-{letter}", "")
                            break
                entries.append(f"{emp}({lvl},{hrs}H)" if lvl != "" else f"{emp}({hrs}H)")
            text = " | ".join(entries)
            c = ws1.cell(row=i, column=j, value=text)
            c.alignment = center

            if start_idx is not None and (d == (start_day + timedelta(days=start_idx))):
                c.fill = fill_start
            if end_idx is not None and (d == (start_day + timedelta(days=end_idx))):
                c.fill = fill_deadline

    # -------------- Sheet 2: Employees x Dates --------------
    ws2 = wb.create_sheet("Employees x Dates")
    ws2.cell(row=1, column=1, value="employee").font = bold
    ws2.cell(row=1, column=2, value="skills").font = bold
    ws2.cell(row=1, column=3, value="capacity").font = bold

    for j, d in enumerate(dates, start=4):
        c = ws2.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws2.column_dimensions[get_column_letter(j)].width = 32

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 48
    ws2.column_dimensions["C"].width = 12
    ws2.freeze_panes = "D2"

    def skills_text(e) -> str:
        def skey(k: str):
            try:
                p, t = k.split("-")
                return (int(p[1:]), t)
            except Exception:
                return (999, k)
        items = sorted(e.skills.items(), key=lambda kv: skey(kv[0]))
        return ", ".join(f"{k}:{v}" for k, v in items)

    employees_sorted = sorted(solution.employees, key=lambda e: e.name)

    workday_counts = defaultdict(int)
    workhour_counts = defaultdict(int)

    for i, e in enumerate(employees_sorted, start=2):
        ws2.row_dimensions[i].height = 36

        ws2.cell(row=i, column=1, value=e.name).font = bold
        ws2.cell(row=i, column=1).alignment = left
        ws2.cell(row=i, column=2, value=skills_text(e)).alignment = left

        cap = int(e.capacity_hours_per_day) + int(e.overtime_hours_per_day)
        ws2.cell(row=i, column=3, value=cap).alignment = center

        for j, d in enumerate(dates, start=4):
            tasks_here = [(t, hrs) for (emp, day, t), hrs in edt_hours.items()
                          if emp == e.name and day == d]
            tasks_here.sort(key=lambda kv: kv[0])

            if tasks_here:
                workday_counts[e.name] += 1
                total_h = sum(int(h) for _, h in tasks_here)
                workhour_counts[e.name] += total_h
                text = " | ".join(f"{t} ({int(h)}H)" for t, h in tasks_here)
            else:
                text = ""
            ws2.cell(row=i, column=j, value=text).alignment = center

    # Totals at the end
    workdays_col  = len(dates) + 4
    workhours_col = len(dates) + 5
    ws2.cell(row=1, column=workdays_col,  value="Workdays").font  = bold
    ws2.cell(row=1, column=workhours_col, value="WorkHours").font = bold
    ws2.column_dimensions[get_column_letter(workdays_col)].width  = 12
    ws2.column_dimensions[get_column_letter(workhours_col)].width = 12

    for i, e in enumerate(employees_sorted, start=2):
        ws2.cell(row=i, column=workdays_col,  value=workday_counts[e.name]).alignment  = center
        ws2.cell(row=i, column=workhours_col, value=workhour_counts[e.name]).alignment = center

    wb.save(out_path)
    print(f"Wrote Excel: {os.path.abspath(out_path)}")

def write_csvs(solution, start_day,
               tasks_csv: str = "tasks_by_date.csv",
               employees_csv: str = "employees_by_date.csv"):
    dates = [d.d for d in solution.days]

    # Build maps with VARIABLE hours
    tde_hours = defaultdict(int)   # (task_code, date, employee) -> hrs
    edt_hours = defaultdict(int)   # (employee, date, task_code) -> hrs

    for u in solution.units:
        if u.employee is None or u.day is None or (u.hours or 0) <= 0:
            continue
        code = f"{u.module}-P{u.process_id}-{u.task_letter}"
        d = u.day.d
        tde_hours[(code, d, u.employee.name)] += int(u.hours)
        edt_hours[(u.employee.name, d, code)] += int(u.hours)

    # ---- tasks_by_date.csv ----
    with open(tasks_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["task_code", "date", "employee", "hours"])
        for (code, d, emp), hrs in sorted(tde_hours.items()):
            w.writerow([code, d.isoformat(), emp, hrs])
    print(f"Wrote CSV: {os.path.abspath(tasks_csv)}")

    # ---- employees_by_date.csv ----
    with open(employees_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["employee", "date", "task_code", "hours"])
        for (emp, d, code), hrs in sorted(edt_hours.items()):
            w.writerow([emp, d.isoformat(), code, hrs])
    print(f"Wrote CSV: {os.path.abspath(employees_csv)}")

def main():
    ap = ArgumentParser(description="Export overtime schedule (variable-hour units) to Excel/CSV")
    ap.add_argument("--config", default="config_modules.yaml", help="Path to modules YAML")
    ap.add_argument("--xlsx",   default="schedule_matrix_overtime_units.xlsx", help="Output .xlsx")
    ap.add_argument("--tasks_csv", default="tasks_by_date.csv", help="Tasks CSV path")
    ap.add_argument("--emps_csv",  default="employees_by_date.csv", help="Employees CSV path")
    args = ap.parse_args()

    solution, start_day = solve_from_config(args.config)
    print(f"Best score: {solution.score}")

    write_excel(solution, start_day, config_path=args.config, out_path=args.xlsx)
    write_csvs(solution, start_day, tasks_csv=args.tasks_csv, employees_csv=args.emps_csv)

if __name__ == "__main__":
    main()
