# yaml_to_suother_like_schedule.py
# ------------------------------------------------------------
# Convert Timefold output (EnvConfig.yaml + Schedule.yaml)
# into a SU_Others-like "Employees x Dates" Excel.
#
# Differences from export_schedule_excel.py Sheet2:
# - Intersect cells show MODULE "name" (module code) instead of task id.
# - Same base code (before first "_") uses same color.
# - Red is reserved ONLY for unavailable dates.
# - Left columns: company(worker_company) | name | manager | today_task
# - Sheet name: "Schedule"
# - Date headers include YEAR (YYYY/MM/DD)
#
# Usage:
#   python yaml_to_suother_like_schedule.py EnvConfig.yaml Schedule.yaml out.xlsx
#
# Optional env var:
#   EXPORT_TODAY_YMD=2025/10/14   (controls "today_task" column)
# ------------------------------------------------------------

import os
import re
import sys
import yaml
from datetime import datetime, date, timedelta
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


# ---------------- CONFIG (easy to adjust) ----------------
DATE_COL_WIDTH = 6.5   # width for the employee x date cells

META_COL_WIDTHS = {
    "company": 14,
    "name": 16,
    "manager": 10,
    "today_task": 22,
}

HEADER_ROW_HEIGHT = 18
DATA_ROW_HEIGHT = 18

# Filter huge worker_list:
# True  -> include everyone in EnvConfig.worker_list (can be 800+)
# False -> include only workers who have at least one assignment OR one unavailable day within plan range
INCLUDE_ALL_WORKERS = False

# If multiple module assignments in same day, join by newline.
JOIN_MULTIPLE_WITH_NEWLINE = True

# Unavailable date fill (must be RED-ish)
UNAV_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

BLACK = Font(color="000000")
BOLD_BLACK = Font(bold=True, color="000000")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

THIN = Side(style="thin", color="999999")
BORDER_THIN = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

# Light palette (avoid red + avoid too-dark; text must remain black)
PALETTE = [
    "FFF2CC", "D9EAD3", "CFE2F3", "D9D2E9", "FCE5CD", "EAD1DC", "D0E0E3", "E2EFDA",
    "DEEAF6", "E7E6E6", "C9DAF8", "D9E1F2", "D0CECE", "E2F0D9", "DDEBF7", "F8CBAD",
]


# ---------------- Utilities ----------------
def _d(s) -> date:
    """Parse 'YYYY/MM/DD' (or 'YYYY-MM-DD') to date."""
    if isinstance(s, date):
        return s
    s = str(s).strip().replace("-", "/")
    return datetime.strptime(s, "%Y/%m/%d").date()


def daterange(start: date, end: date):
    d = start
    while d <= end:
        yield d
        d += timedelta(days=1)


def load_yaml(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def normalize_module_code(name: str) -> str:
    """
    Same module code may appear with different suffix name:
      '530N02621A_JP_Rapidus_新規' and '530N02621A'
    -> base is text before first '_'
    """
    if not name:
        return ""
    s = str(name).strip()
    return s.split("_", 1)[0]


def color_for_key(key: str, key_to_color: dict) -> str:
    """Stable mapping key -> color."""
    if key in key_to_color:
        return key_to_color[key]
    idx = abs(hash(key)) % len(PALETTE)
    col = PALETTE[idx]
    key_to_color[key] = col
    return col


def parse_unavailable_dates(raw, plan_start: date, plan_end: date) -> set:
    """
    Matches the schema used in export_schedule_excel.py:
      - scalar date
      - list of scalar dates
      - {single:{days:[...]}}
      - {weekly:{weekdays:[sat,sun,...]}}
      - list of those dicts
    """
    off = set()
    if raw is None:
        return off

    def add_single(obj):
        try:
            dd = _d(obj)
        except Exception:
            return
        if plan_start <= dd <= plan_end:
            off.add(dd)

    def weekday_to_int(w):
        m = {
            "mon": 0, "monday": 0,
            "tue": 1, "tues": 1, "tuesday": 1,
            "wed": 2, "weds": 2, "wednesday": 2,
            "thu": 3, "thur": 3, "thurs": 3, "thursday": 3,
            "fri": 4, "friday": 4,
            "sat": 5, "saturday": 5,
            "sun": 6, "sunday": 6,
        }
        return m.get(str(w).strip().lower())

    items = raw if isinstance(raw, list) else [raw]
    weekly = set()

    for item in items:
        if item is None:
            continue

        if isinstance(item, dict):
            single = item.get("single")
            if isinstance(single, dict):
                days = single.get("days")
                if isinstance(days, list):
                    for x in days:
                        add_single(x)

            weekly_obj = item.get("weekly")
            if isinstance(weekly_obj, dict):
                wds = weekly_obj.get("weekdays")
                if isinstance(wds, list):
                    for w in wds:
                        wd = weekday_to_int(w)
                        if wd is not None:
                            weekly.add(wd)
        else:
            add_single(item)

    if weekly:
        for dd in daterange(plan_start, plan_end):
            if dd.weekday() in weekly:
                off.add(dd)

    return off


def build_op_task_index(workflow_task_list):
    """
    Build mapping: operation_task_list.id -> module_id
    Uses Schedule.yaml task structure (no guessing).
    """
    idx = {}
    for m in workflow_task_list or []:
        mid = m["id"]
        for ph in (m.get("phase_task_list") or []):
            for ot in (ph.get("operation_task_list") or []):
                idx[ot["id"]] = mid
    return idx


# ---------------- Main exporter ----------------
def main(env_path: str, sched_path: str, out_path: str):
    env = load_yaml(env_path)
    sched = load_yaml(sched_path)

    env_root = env.get("environment", env)
    sched_root = sched.get("schedule", sched)

    # plan range (FULL YEAR RANGE)
    plan_start = _d(sched_root["plan_range"]["start_date"])
    plan_end = _d(sched_root["plan_range"]["end_date"])
    dates = list(daterange(plan_start, plan_end))

    # env lookups
    worker_companies = {c["id"]: c for c in env_root.get("worker_company_list", [])}
    workers = {w["id"]: w for w in env_root.get("worker_list", [])}

    # schedule lookups
    modules = sched_root.get("workflow_task_list", []) or []
    module_name_by_id = {m["id"]: (m.get("name") or m["id"]) for m in modules}

    op_task_to_module = build_op_task_index(modules)

    # assignments expanded to per-day:
    # (wid, date) -> [module_id...]
    cell_modules = defaultdict(list)
    assigned_workers = set()
    for a in (sched_root.get("assignment_list") or []):
        wid = a.get("worker")
        op_task = a.get("operation_task")
        mid = op_task_to_module.get(op_task)

        # IMPORTANT: If not found, keep it visible but DON'T break.
        # However, in correct data this should be found.
        if not mid:
            # fallback: do NOT invent "OTHER"; just keep op_task as key
            mid = f"UNKNOWN::{op_task}"

        wd_key = "work_date_lsit" if "work_date_lsit" in a else "work_date_list"
        for ditem in (a.get(wd_key) or []):
            dd = _d(ditem.get("date"))
            if plan_start <= dd <= plan_end:
                cell_modules[(wid, dd)].append(mid)
                assigned_workers.add(wid)

    # worker unavailable map (and filter set)
    worker_off = {}
    workers_with_off = set()
    for wid, w in workers.items():
        off = parse_unavailable_dates(w.get("unavailable_dates"), plan_start, plan_end)
        worker_off[wid] = off
        if off:
            workers_with_off.add(wid)

    # decide which workers to include
    if INCLUDE_ALL_WORKERS:
        worker_ids = list(workers.keys())
    else:
        worker_ids = sorted(assigned_workers.union(workers_with_off))

    # stable sort: company then name
    def company_name_of(wid):
        w = workers.get(wid, {})
        wc = w.get("worker_company")
        return (worker_companies.get(wc, {}) or {}).get("name", wc or "")

    def worker_name_of(wid):
        return (workers.get(wid, {}) or {}).get("name", wid)

    worker_ids.sort(key=lambda wid: (company_name_of(wid), worker_name_of(wid)))

    # today for today_task column
    today = date.today()
    if os.environ.get("EXPORT_TODAY_YMD"):
        try:
            today = _d(os.environ["EXPORT_TODAY_YMD"])
        except Exception:
            pass

    # workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    headers = ["company", "name", "manager", "today_task"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = BOLD_BLACK
        c.alignment = CENTER
        c.border = BORDER_THIN

    start_col_dates = len(headers) + 1


    # date headers: store real date value (with year) but DISPLAY only MM/DD
    for k, dd in enumerate(dates):
        col = start_col_dates + k

        # store as an Excel date serial (has year internally)
        c = ws.cell(row=1, column=col, value=datetime(dd.year, dd.month, dd.day))

        # display format (what you see without clicking)
        c.number_format = "mm/dd"

        c.font = BOLD_BLACK
        c.alignment = CENTER
        c.border = BORDER_THIN
        ws.column_dimensions[get_column_letter(col)].width = DATE_COL_WIDTH


    # meta widths
    ws.column_dimensions["A"].width = META_COL_WIDTHS["company"]
    ws.column_dimensions["B"].width = META_COL_WIDTHS["name"]
    ws.column_dimensions["C"].width = META_COL_WIDTHS["manager"]
    ws.column_dimensions["D"].width = META_COL_WIDTHS["today_task"]

    ws.freeze_panes = get_column_letter(start_col_dates) + "2"
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    # color maps
    company_color = {}
    module_color = {}

    for r, wid in enumerate(worker_ids, start=2):
        wcfg = workers.get(wid, {}) or {}
        wname = wcfg.get("name", wid)
        is_mgr = bool(wcfg.get("is_manager", False))

        wc_id = wcfg.get("worker_company")
        wc_name = (worker_companies.get(wc_id, {}) or {}).get("name", wc_id or "")

        # company label color (consistent per company)
        comp_key = wc_name or wc_id or ""
        comp_fill = None
        if comp_key:
            col = color_for_key(comp_key, company_color)
            comp_fill = PatternFill(start_color=col, end_color=col, fill_type="solid")

        # left columns
        cA = ws.cell(row=r, column=1, value=wc_name)
        cB = ws.cell(row=r, column=2, value=wname)
        cC = ws.cell(row=r, column=3, value="Yes" if is_mgr else "")
        cD = ws.cell(row=r, column=4, value="")

        for c in (cA, cB, cC, cD):
            c.border = BORDER_THIN
            c.font = BOLD_BLACK if c in (cA, cB) else BLACK
            c.alignment = CENTER if c in (cA, cC) else LEFT
            if comp_fill:
                c.fill = comp_fill
                # text black always
                c.font = Font(bold=c.font.bold, color="000000")

        # grid cells
        todays_text = ""
        for k, dd in enumerate(dates):
            col = start_col_dates + k
            cell = ws.cell(row=r, column=col, value="")
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.font = BLACK

            # unavailable -> RED
            if dd in worker_off.get(wid, set()):
                cell.fill = UNAV_FILL
                continue

            mids = cell_modules.get((wid, dd), [])
            if not mids:
                continue

            # convert module_id -> module_name (module code)
            names = []
            for mid in mids:
                if mid.startswith("UNKNOWN::"):
                    # show something visible if YAML mismatch exists
                    names.append(mid.replace("UNKNOWN::", ""))
                else:
                    names.append(module_name_by_id.get(mid, mid))

            text = names[0] if JOIN_MULTIPLE_WITH_NEWLINE else " | ".join(names)
            cell.value = text

            # color key uses normalized code base from module NAME (not id)
            base = normalize_module_code(names[0])
            mcol = color_for_key(base, module_color)
            cell.fill = PatternFill(start_color=mcol, end_color=mcol, fill_type="solid")

            if dd == today:
                todays_text = text

        if todays_text:
            cD.value = todays_text

        ws.row_dimensions[r].height = DATA_ROW_HEIGHT
    last_row = ws.max_row
    last_col = start_col_dates + len(dates) - 1
    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"

    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python yaml_to_suother_like_schedule.py EnvConfig.yaml Schedule.yaml out.xlsx")
        sys.exit(2)

    main(sys.argv[1], sys.argv[2], sys.argv[3])
