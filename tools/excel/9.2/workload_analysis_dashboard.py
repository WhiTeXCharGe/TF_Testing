# python workload_analysis_dashboard.py EnvConfig.yaml Schedule.yaml workload_analysis.xlsx
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Workload Analysis Dashboard (wf_tool only)

Reads:
  - EnvConfig.yaml (optional, currently not required for metrics)
  - Schedule.yaml

Outputs:
  - Excel workbook with:
      Overview
      Operation Summary
      Phase Summary
      Module Summary
      Distributions
      Data (op tasks)

Filters:
  - workflow == "wf_tool" only
  - excludes dummy/other tasks (operation == "other_op", or id/name contains "dummy", or starts with "other")

Metrics:
  - workload_days (from Schedule.yaml)
  - duration_days (end-start+1)
  - intensity = workload_days / duration_days
"""

import sys
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional

import yaml
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


# ---------------------------
# YAML
# ---------------------------
def load_yaml(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# ---------------------------
# Date helpers
# ---------------------------
def parse_ymd(x: Any) -> Optional[date]:
    """Parse 'YYYY/MM/DD' or 'YYYY-MM-DD' into date."""
    if x is None:
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, date):
        return x
    s = str(x).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    # also accept YYYY/M/D
    m = re.match(r"(\d{4})/(\d{1,2})/(\d{1,2})", s)
    if m:
        y, mo, d = map(int, m.groups())
        return date(y, mo, d)
    raise ValueError(f"Unrecognized date: {x!r}")


# ---------------------------
# Filters
# ---------------------------
def is_dummy_op_task(op_task: Dict[str, Any]) -> bool:
    tid = (op_task.get("id") or "").lower()
    nm = (op_task.get("name") or "").lower()
    op = (op_task.get("operation") or "").lower()

    if op == "other_op":
        return True
    if "dummy" in tid or "dummy" in nm:
        return True
    if tid.startswith("other") or nm.startswith("other"):
        return True
    return False


# ---------------------------
# Excel helpers
# ---------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")


def add_df_sheet(
    wb: Workbook,
    name: str,
    df: pd.DataFrame,
    freeze: str = "A2",
    table_name: Optional[str] = None,
) -> None:
    ws = wb.create_sheet(title=name)

    # write cells
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for c_idx in range(1, len(row) + 1):
                cell = ws.cell(r_idx, c_idx)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        else:
            for c_idx in range(1, len(row) + 1):
                ws.cell(r_idx, c_idx).alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = freeze

    # basic column widths
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col[:2000]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

    # add Excel table (for filtering/sorting)
    if table_name is None:
        table_name = re.sub(r"[^A-Za-z0-9_]", "_", name)[:20] + "_tbl"

    end_row = ws.max_row
    end_col = ws.max_column
    if end_row >= 2 and end_col >= 1:
        ref = f"A1:{ws.cell(end_row, end_col).coordinate}"
        tbl = Table(displayName=table_name, ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tbl)


# ---------------------------
# Core analysis
# ---------------------------
def build_op_task_dataframe(schedule: Dict[str, Any]) -> pd.DataFrame:
    wf_tasks = schedule.get("schedule", {}).get("workflow_task_list", []) or []

    rows: List[Dict[str, Any]] = []
    for wf in wf_tasks:
        if (wf.get("workflow") or "") != "wf_tool":
            continue

        module_id = wf.get("id")
        module_name = wf.get("name")
        fab = wf.get("fab")

        for ph in (wf.get("phase_task_list") or []):
            phase_id = ph.get("phase")
            phase_name = ph.get("name")

            for op_task in (ph.get("operation_task_list") or []):
                if is_dummy_op_task(op_task):
                    continue

                op = op_task.get("operation")
                op_task_id = op_task.get("id")
                op_task_name = op_task.get("name")

                start = parse_ymd(op_task.get("start_date") or ph.get("start_date"))
                end = parse_ymd(op_task.get("end_date") or ph.get("end_date"))

                wd = op_task.get("workload_days")
                try:
                    workload_days = float(wd) if wd is not None else np.nan
                except Exception:
                    workload_days = np.nan

                duration_days = (end - start).days + 1 if (start and end) else np.nan
                intensity = (
                    workload_days / duration_days
                    if (duration_days and duration_days > 0 and not np.isnan(workload_days))
                    else np.nan
                )

                rows.append(
                    dict(
                        module_id=module_id,
                        module_name=module_name,
                        fab=fab,
                        phase_id=phase_id,
                        phase_name=phase_name,
                        operation=op,
                        operation_task_id=op_task_id,
                        operation_task_name=op_task_name,
                        start_date=start,
                        end_date=end,
                        duration_days=duration_days,
                        workload_days=workload_days,
                        intensity_workload_per_day=intensity,
                    )
                )

    return pd.DataFrame(rows)


def build_workbook(env_path: Path, schedule_path: Path, out_path: Path) -> None:
    # env not required now, but keep it in inputs for future metrics/extensions
    _env = load_yaml(env_path)
    schedule = load_yaml(schedule_path)

    df = build_op_task_dataframe(schedule)

    plan_range = schedule.get("schedule", {}).get("plan_range", {}) or {}
    plan_start = parse_ymd(plan_range.get("start_date")) if plan_range.get("start_date") else None
    plan_end = parse_ymd(plan_range.get("end_date")) if plan_range.get("end_date") else None

    wb = Workbook()
    wb.remove(wb.active)

    if df.empty:
        ws = wb.create_sheet("Workload Analysis")
        ws["A1"] = "No wf_tool non-dummy operation_task rows found after filtering."
        wb.save(out_path)
        return

    # --- Overview
    overview = pd.DataFrame(
        [
            ["Plan start", plan_start],
            ["Plan end", plan_end],
            ["Modules (wf_tool)", int(df["module_id"].nunique())],
            ["Operation tasks (non-dummy)", int(len(df))],
            ["Total workload_days", float(df["workload_days"].sum())],
            ["Median workload_days (per op task)", float(df["workload_days"].median())],
            ["Avg workload_days (per op task)", float(df["workload_days"].mean())],
            ["Median duration_days (per op task)", float(df["duration_days"].median())],
            ["Avg intensity (workload/duration)", float(df["intensity_workload_per_day"].mean())],
        ],
        columns=["Metric", "Value"],
    )

    # --- Operation summary
    op_summary = (
        df.groupby("operation", dropna=False)
        .agg(
            task_count=("operation_task_id", "count"),
            avg_workload_days=("workload_days", "mean"),
            median_workload_days=("workload_days", "median"),
            min_workload_days=("workload_days", "min"),
            max_workload_days=("workload_days", "max"),
            avg_duration_days=("duration_days", "mean"),
            median_duration_days=("duration_days", "median"),
            avg_intensity=("intensity_workload_per_day", "mean"),
            median_intensity=("intensity_workload_per_day", "median"),
        )
        .reset_index()
        .sort_values(["operation"])
        .round(3)
    )

    # --- Phase summary
    phase_summary = (
        df.groupby(["phase_id", "phase_name"], dropna=False)
        .agg(
            task_count=("operation_task_id", "count"),
            avg_workload_days=("workload_days", "mean"),
            median_workload_days=("workload_days", "median"),
            avg_duration_days=("duration_days", "mean"),
            median_duration_days=("duration_days", "median"),
            avg_intensity=("intensity_workload_per_day", "mean"),
            median_intensity=("intensity_workload_per_day", "median"),
        )
        .reset_index()
        .sort_values(["phase_id"])
        .round(3)
    )

    # --- Module summary
    mod_key = ["module_id", "module_name", "fab"]
    mod_grp = df.groupby(mod_key, dropna=False)

    mod_min = mod_grp["start_date"].min()
    mod_max = mod_grp["end_date"].max()
    mod_total = mod_grp["workload_days"].sum(min_count=1)
    mod_ops = mod_grp["operation_task_id"].count()
    mod_phases = mod_grp["phase_id"].nunique()

    mod_duration_days = (pd.to_datetime(mod_max) - pd.to_datetime(mod_min)).dt.days + 1
    mod_avg_per_day = mod_total.values / mod_duration_days.values

    module_summary = (
        pd.DataFrame(
            {
                "module_id": mod_total.index.get_level_values(0),
                "module_name": mod_total.index.get_level_values(1),
                "fab": mod_total.index.get_level_values(2),
                "operation_task_count": mod_ops.values,
                "phase_count": mod_phases.values,
                "module_start": mod_min.values,
                "module_end": mod_max.values,
                "module_duration_days": mod_duration_days.values,
                "total_workload_days": mod_total.values,
                "avg_workload_days_per_calendar_day": mod_avg_per_day,
            }
        )
        .sort_values(["module_id"])
        .round(3)
    )

    # --- Distributions (histogram bins)
    bins = [0, 25, 50, 75, 100, 150, 200, 300, 500, 1000, float("inf")]
    labels = ["0-25", "25-50", "50-75", "75-100", "100-150", "150-200", "200-300", "300-500", "500-1000", "1000+"]

    dist_rows: List[Dict[str, Any]] = []
    for op, sub in df.groupby("operation"):
        s = sub["workload_days"].dropna()
        if s.empty:
            continue
        cats = pd.cut(s, bins=bins, labels=labels, right=False, include_lowest=True)
        counts = cats.value_counts().reindex(labels, fill_value=0)
        for rng, cnt in counts.items():
            dist_rows.append({"operation": op, "workload_range": rng, "count": int(cnt)})

    dist_df = pd.DataFrame(dist_rows)
    if not dist_df.empty:
        dist_df = dist_df.sort_values(["operation", "workload_range"])

    # --- Raw data (pretty dates)
    raw = df.copy()
    raw["start_date"] = raw["start_date"].apply(lambda d: d.strftime("%Y/%m/%d") if isinstance(d, date) else "")
    raw["end_date"] = raw["end_date"].apply(lambda d: d.strftime("%Y/%m/%d") if isinstance(d, date) else "")

    # write sheets
    add_df_sheet(wb, "Overview", overview, table_name="OverviewTbl")
    add_df_sheet(wb, "Operation Summary", op_summary, table_name="OpSummaryTbl")
    add_df_sheet(wb, "Phase Summary", phase_summary, table_name="PhaseSummaryTbl")
    add_df_sheet(wb, "Module Summary", module_summary, table_name="ModuleSummaryTbl")
    if not dist_df.empty:
        add_df_sheet(wb, "Distributions", dist_df, table_name="DistTbl")
    add_df_sheet(wb, "Data (op tasks)", raw, table_name="DataTbl")

    wb.save(out_path)


# ---------------------------
# CLI
# ---------------------------
def main(argv: List[str]) -> int:
    if len(argv) < 4:
        print(
            "Usage:\n"
            "  python workload_analysis_dashboard.py EnvConfig.yaml Schedule.yaml Output.xlsx\n",
            file=sys.stderr,
        )
        return 2

    env_path = Path(argv[1])
    schedule_path = Path(argv[2])
    out_path = Path(argv[3])

    if not env_path.exists():
        raise FileNotFoundError(env_path)
    if not schedule_path.exists():
        raise FileNotFoundError(schedule_path)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(env_path, schedule_path, out_path)

    print(f"OK: wrote {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))