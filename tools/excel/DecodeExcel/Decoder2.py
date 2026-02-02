# Decoder2.py
# ---------------------------------------------------------------------
# Generates EnvConfig.yaml + Schedule.yaml for the Timefold scheduler from 3 Excels:
#   1) 20260105 SU_Others.xlsm
#       - Worker list (name/company/manager flag)
#       - Worker unavailable dates (RED cells)
#       - Worker daily "what they did" text matrix (normal cells)
#       - Personal business (GREY cells) -> fixed assignments
#   2) SU_Others_予定表_2025_新規製番リスト_20260127.xlsx (sheet "CSV")
#       - Defines tool-install tasks (modules) + per-phase start dates (P1..P4) + P4 end
#       - Also provides Customer company / Country / Fab name for EnvConfig
#   3) スキル集計_20260127.xlsx (sheet "Sheet1")
#       - Main source of worker skill levels (p1,p2,p3,p4) using 合計/Level and 最小/Level
#
# Output format changes vs Decoder.py:
#   - wf_tool operations are now generic "p1,p2,p3,p4" (not f22p1...)
#   - Task phase IDs are "e{n}_p{phase}" and operation_task.operation is "p{phase}"
#   - Worker skill_map keys are "p1,p2,p3,p4" (plus other_op, personal_business_op)
#
# NOTE:
#   - If a worker appears in skill excel (same name+company), we use those skill levels.
#   - If not, we infer skill>=1 for phases they appear to work on in SU_Others (Flexible assignments),
#     otherwise skill defaults to 0.
#
# ---------------------------------------------------------------------

import re
import math
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import load_workbook


# ============================================================
# Helpers
# ============================================================

def _to_ymd(dt) -> str:
    """Convert pandas/datetime to 'YYYY/MM/DD'."""
    if isinstance(dt, pd.Timestamp):
        return dt.strftime("%Y/%m/%d")
    if isinstance(dt, datetime):
        return dt.strftime("%Y/%m/%d")
    return str(dt)


def _norm(s: str) -> str:
    """Normalize string for matching (trim, collapse spaces, lower)."""
    if not isinstance(s, str):
        return ""
    s = s.replace("　", " ")
    s = s.replace("（", "(").replace("）", ")")
    return " ".join(s.split()).lower()


def load_sheet_as_df(path: str, sheet_name: str) -> pd.DataFrame:
    """
    Load a sheet using openpyxl (read_only=False to avoid ReadOnlyWorksheet bug)
    and convert it to a pandas DataFrame.
    """
    wb = load_workbook(path, data_only=True, read_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    df = pd.DataFrame(rows)
    return df


def _as_timestamp(v):
    """Best-effort convert excel cell value into pd.Timestamp or None."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, pd.Timestamp):
        return v
    if isinstance(v, datetime):
        return pd.Timestamp(v)
    if isinstance(v, str) and v.strip():
        dt = pd.to_datetime(v.strip(), errors="coerce")
        if isinstance(dt, pd.Timestamp) and not pd.isna(dt):
            return dt
    return None


# ============================================================
# Code extraction from SU_Others cell text
# (新規製番 codes are usually like 530N02716A, 830300179A, 852Z00771A)
# ============================================================

TOOLCODE_RE = re.compile(r"\d{3}[A-Z0-9]\d{5}A")


def extract_tool_code(s: str):
    """
    Extract code like 530N02716A, 830300179A, 852Z00771A from a string.
    Returns the FIRST match if there are multiple (e.g. '530N01814A_852Z00771A').
    """
    if not isinstance(s, str):
        return None
    m = TOOLCODE_RE.search(s)
    if m:
        return m.group(0)
    return None


# ============================================================
# SU_Others: worker info + raw matrix for assignments
# ============================================================

IGNORE_WHITE_TEXT = {"FI", "FO"}  # exact tokens to ignore

GREY_RGB_LAST6 = {"A6A6A6", "BFBFBF", "D9D9D9", "808080"}
RED_RGB_LAST6 = {"FF0000"}  # last-6 of FFFF0000


def _cell_rgb_last6(cell):
    """Return last 6 hex chars of an rgb fill (upper), or None."""
    fill = getattr(cell, "fill", None)
    if fill is None:
        return None
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return None
    if getattr(fg, "type", None) != "rgb":
        return None
    rgb = getattr(fg, "rgb", None)
    if not rgb:
        return None
    return str(rgb).upper()[-6:]  # normalize FFA6A6A6 -> A6A6A6


def _find_date_header_ws(ws, max_scan_rows=12):
    """
    Find the header row containing datetime values.
    Returns:
      - date_row_idx (1-based)
      - date_cols (1-based col idx list)
      - dt_by_col {col_idx -> pd.Timestamp}
    """
    for r in range(1, max_scan_rows + 1):
        row_vals = [cell.value for cell in ws[r]]
        if any(isinstance(v, (pd.Timestamp, datetime)) for v in row_vals):
            date_cols = [
                c for c, v in enumerate(row_vals, start=1)
                if isinstance(v, (pd.Timestamp, datetime))
            ]
            if not date_cols:
                continue
            dt_by_col = {c: pd.Timestamp(row_vals[c - 1]) for c in date_cols}
            return r, date_cols, dt_by_col
    raise RuntimeError("Could not find date header row in SU_Others.")


def parse_su_others(path: str, sheet_names=("予定表_2024", "予定表_2025")):
    """
    Build:
      - worker_company_list
      - worker_list (name/company/is_manager/unavailable_dates)
      - plan_range
      - worker_date_map: (worker_id, Timestamp) -> raw cell string (normal work)
      - worker_personal_map: (worker_id, Timestamp) -> raw cell string (grey personal business)

    Rules (date intersect area only):
      - RED  => worker unavailable_dates (no assignments)
      - GREY => personal business assignments
      - White cell text 'FI'/'FO' => ignore completely
      - Others => normal assignment source
    """
    wb = load_workbook(path, data_only=True, read_only=False)

    used_sheets = [s for s in sheet_names if s in wb.sheetnames]
    if not used_sheets:
        raise ValueError(f"None of {sheet_names} found in {path}. Available: {wb.sheetnames}")

    worker_company_map = {}
    worker_company_list = []

    def get_worker_company_id(company_name: str) -> str:
        company_name = str(company_name).strip()
        if company_name not in worker_company_map:
            cid = f"wc{len(worker_company_map) + 1}"
            worker_company_map[company_name] = cid
            worker_company_list.append({
                "id": cid,
                "name": company_name,
                "annual_overtime_limit": 360,
                "monthly_overtime_limit": 40,
                "unavailable_dates": [],
            })
        return worker_company_map[company_name]

    # Merge by (company, name)
    worker_key_to_id = {}
    worker_acc = {}  # key -> {id,name,company,is_manager,unavailable_set}

    worker_date_map = {}      # normal text cells
    worker_personal_map = {}  # grey text cells

    plan_start = None
    plan_end = None

    for sheet_name in used_sheets:
        ws = wb[sheet_name]

        date_row_idx, date_cols, dt_by_col = _find_date_header_ws(ws)
        worker_start_row = date_row_idx + 2

        # plan_range update
        if dt_by_col:
            s = min(dt_by_col.values())
            e = max(dt_by_col.values())
            plan_start = s if plan_start is None else min(plan_start, s)
            plan_end = e if plan_end is None else max(plan_end, e)

        # stop early when reaching empty tail
        blank_streak = 0

        # Performance note:
        #   Avoid calling ws.max_row in a loop (it can be slow on large styled sheets).
        #   Iterate rows directly and stop when we hit a long blank tail.
        max_col = max(date_cols) if date_cols else 5
        for r, row_cells in enumerate(ws.iter_rows(min_row=worker_start_row, min_col=1, max_col=max_col), start=worker_start_row):
            company = row_cells[0].value
            name = row_cells[1].value

            if name is None or str(name).strip() == "":
                blank_streak += 1
                if blank_streak >= 30:
                    break
                continue
            blank_streak = 0

            company_str = "" if company is None else str(company).strip()
            name_str = str(name).strip()

            free_slot = row_cells[4].value if len(row_cells) >= 5 else None  # E column
            is_manager = bool(isinstance(free_slot, str) and "責" in free_slot)

            key = (company_str, name_str)
            if key not in worker_key_to_id:
                wid = f"w{len(worker_key_to_id) + 1:03d}"
                worker_key_to_id[key] = wid
                company_id = get_worker_company_id(company_str)
                worker_acc[key] = {
                    "id": wid,
                    "name": name_str,
                    "worker_company": company_id,
                    "is_manager": is_manager,
                    "unavailable_set": set(),
                }
            else:
                if is_manager:
                    worker_acc[key]["is_manager"] = True

            wid = worker_key_to_id[key]

            # date intersect only
            for c in date_cols:
                dt = dt_by_col[c]
                cell = row_cells[c - 1]
                rgb6 = _cell_rgb_last6(cell)

                # red => unavailable
                if rgb6 in RED_RGB_LAST6:
                    worker_acc[key]["unavailable_set"].add(_to_ymd(dt))
                    continue

                val = cell.value
                if not (isinstance(val, str) and val.strip()):
                    continue
                text = val.strip()

                # ignore white tokens
                if text.upper() in IGNORE_WHITE_TEXT:
                    continue

                # grey => personal business
                if rgb6 in GREY_RGB_LAST6:
                    worker_personal_map[(wid, dt)] = text
                else:
                    worker_date_map[(wid, dt)] = text


    worker_list = []
    for acc in worker_acc.values():
        worker_list.append({
            "id": acc["id"],
            "name": acc["name"],
            "worker_company": acc["worker_company"],
            "is_manager": acc["is_manager"],
            # skill_map filled later (after we read skill excel + infer from assignments)
            "skill_map": {},
            "fab_suitability_map": [],
            "unavailable_dates": [{"date": d} for d in sorted(acc["unavailable_set"])],
        })

    plan_range = {
        "start_date": _to_ymd(plan_start),
        "end_date": _to_ymd(plan_end),
    }

    return {
        "worker_company_list": worker_company_list,
        "worker_company_map": worker_company_map,
        "worker_list": worker_list,
        "plan_range": plan_range,
        "worker_date_map": worker_date_map,
        "worker_personal_map": worker_personal_map,
        "worker_key_to_id": worker_key_to_id,
    }


# ============================================================
# Task list from 新規製番リスト (sheet "CSV")
# ============================================================

def parse_tasks_from_csv(path: str, sheet_name: str = "CSV"):
    """
    From SU_Others_予定表_2025_新規製番リスト_*.xlsx, sheet "CSV":
      - Build tool tasks for schedule.workflow_task_list
      - Build code_to_phases: code -> per-phase meta (phase_id/start/end/op_id)
      - Gather customers/countries/fabs for EnvConfig
    """
    df = pd.read_excel(path, sheet_name=sheet_name)

    # Normalize column names (some have newlines)
    colmap = {c: str(c).replace("\n", "").strip() for c in df.columns}
    df = df.rename(columns=colmap)

    # Required columns (Japanese)
    # - 新規製番
    # - ユーザー名 (customer company)
    # - 国 (country)
    # - ファブ名 (fab/factory)
    # - 工程1..4開始可能日, 工程4終了予定日
    code_col = [c for c in df.columns if "新規製番" in c]
    if not code_col:
        raise ValueError("Could not find 新規製番 column in CSV sheet.")
    code_col = code_col[0]

    cust_col = "ユーザー名"
    country_col = "国"
    fab_col = "ファブ名"

    p1s_col = "工程１開始可能日"
    p2s_col = "工程２開始可能日"
    p3s_col = "工程３開始可能日"
    p4s_col = "工程４開始可能日"
    p4e_col = "工程４終了予定日"

    # Optional provided end-date columns (fallback only)
    p1e_col = "工程１終了予定日" if "工程１終了予定日" in df.columns else None
    p2e_col = "工程２終了予定日" if "工程２終了予定日" in df.columns else None
    p3e_col = "工程３終了予定日" if "工程３終了予定日" in df.columns else None

    tool_tasks = []
    code_to_phases = defaultdict(list)
    all_dates = []

    # Collect raw meta for envconfig
    task_meta_rows = []

    task_counter = 1
    for _, row in df.iterrows():
        code_raw = row.get(code_col)
        if not isinstance(code_raw, str) or not code_raw.strip():
            continue
        code = code_raw.strip()

        customer = row.get(cust_col)
        customer = str(customer).strip() if isinstance(customer, str) and str(customer).strip() else "OTHER"

        country = row.get(country_col)
        country = str(country).strip() if isinstance(country, str) and str(country).strip() else "Other"

        fab_name = row.get(fab_col)
        fab_name = str(fab_name).strip() if isinstance(fab_name, str) and str(fab_name).strip() else "Other"

        p1s = _as_timestamp(row.get(p1s_col))
        p2s = _as_timestamp(row.get(p2s_col))
        p3s = _as_timestamp(row.get(p3s_col))
        p4s = _as_timestamp(row.get(p4s_col))
        p4e = _as_timestamp(row.get(p4e_col))

        # fallbacks (if the sheet provides them)
        p1e_fallback = _as_timestamp(row.get(p1e_col)) if p1e_col else None
        p2e_fallback = _as_timestamp(row.get(p2e_col)) if p2e_col else None
        p3e_fallback = _as_timestamp(row.get(p3e_col)) if p3e_col else None

        # Skip rows that have literally no date info at all
        if not any([p1s, p2s, p3s, p4s, p4e, p1e_fallback, p2e_fallback, p3e_fallback]):
            continue

        task_id = f"e{task_counter}"
        task_counter += 1

        # Compute phase start/end using user rule:
        # - start: given StartBy
        # - end  : (next phase start - 1 day), p4 end = given
        # - fallback to provided ResolveBy columns if needed
        starts = {1: p1s, 2: p2s, 3: p3s, 4: p4s}
        ends_fallback = {1: p1e_fallback, 2: p2e_fallback, 3: p3e_fallback, 4: p4e}

        # If some start is missing, try to infer from fallback end or neighboring phases
        # (best-effort; keep it simple and safe)
        # If p1 start missing but p1 end exists -> start=end
        for ph in (1, 2, 3, 4):
            if starts[ph] is None and ends_fallback.get(ph) is not None:
                starts[ph] = ends_fallback[ph]

        # Determine end by next phase start - 1
        ends = {1: None, 2: None, 3: None, 4: ends_fallback.get(4)}
        for ph in (1, 2, 3):
            nxt = starts.get(ph + 1)
            if nxt is not None:
                ends[ph] = nxt - pd.Timedelta(days=1)

        # If computed end missing, use fallback end
        for ph in (1, 2, 3):
            if ends[ph] is None and ends_fallback.get(ph) is not None:
                ends[ph] = ends_fallback[ph]

        # Final safety: if start missing, try to use next known start/end
        for ph in (1, 2, 3, 4):
            if starts[ph] is None:
                # try next start
                for j in range(ph + 1, 5):
                    if starts.get(j) is not None:
                        starts[ph] = starts[j]
                        break
            if starts[ph] is None and ends.get(ph) is not None:
                starts[ph] = ends[ph]

        # Ensure end exists; if still missing, use start
        for ph in (1, 2, 3, 4):
            if ends.get(ph) is None:
                ends[ph] = starts.get(ph)

        # Normalize: end >= start
        for ph in (1, 2, 3, 4):
            if starts[ph] is None:
                # worst-case fallback (should be rare)
                starts[ph] = pd.Timestamp("2025-01-01")
            if ends[ph] is None:
                ends[ph] = starts[ph]
            if ends[ph] < starts[ph]:
                ends[ph] = starts[ph]

            all_dates.append(starts[ph])
            all_dates.append(ends[ph])

        # Build schedule.workflow_task_list for this task
        phase_task_list = []
        for ph in (1, 2, 3, 4):
            phase_id = f"{task_id}_p{ph}"
            start = starts[ph]
            end = ends[ph]
            workload_days = int((end - start).days) + 1

            phase_task_list.append({
                "id": phase_id,
                "name": {1: "Module Setup", 2: "Hardware Setup", 3: "Function Setup", 4: "Utility"}.get(ph, f"P{ph}"),
                "phase": f"tool_p{ph}",
                "start_date": _to_ymd(start),
                "end_date": _to_ymd(end),
                "operation_task_list": [{
                    "id": phase_id,
                    "name": {1: "Module Setup", 2: "Hardware Setup", 3: "Function Setup", 4: "Utility"}.get(ph, f"P{ph}"),
                    "operation": f"p{ph}",
                    "workload_days": workload_days,
                }],
            })

            code_to_phases[code].append({
                "phase_index": ph,
                "phase_id": phase_id,
                "start": start,
                "end": end,
                "operation": f"p{ph}",
                "task_id": task_id,
            })

        tool_tasks.append({
            "id": task_id,
            "name": code,            # make the visible name be the code
            "workflow": "wf_tool",
            "fab": None,             # filled later after we create fab ids
            "phase_task_list": phase_task_list,
            "module_code": code,     # internal only, stripped before YAML
            "customer": customer,    # internal only
            "country": country,      # internal only
            "fab_name": fab_name,    # internal only
        })

        task_meta_rows.append((code, customer, country, fab_name))

    return {
        "tool_tasks": tool_tasks,
        "code_to_phases": code_to_phases,
        "date_list": all_dates,
        "task_meta_rows": task_meta_rows,
    }


# ============================================================
# Skill aggregation (スキル集計_*.xlsx)
# ============================================================

def _skill_level(total_level, min_level):
    """
    Convert 合計/Level and 最小/Level into final skill level.
    Rules (as specified):
      - total<=20 and min=0 -> 0
      - total<=20 and min>0 -> min
      - total>20 and min=0 -> at least 1
      - in general: level = max(bucket(total), min)
        where bucket(total) = 0 if total<=20 else floor(total/20)
      - cap to 5
    """
    try:
        total = 0 if total_level is None or (isinstance(total_level, float) and pd.isna(total_level)) else int(total_level)
    except Exception:
        total = 0
    try:
        mn = 0 if min_level is None or (isinstance(min_level, float) and pd.isna(min_level)) else int(min_level)
    except Exception:
        mn = 0

    bucket = 0 if total <= 20 else (total // 20)
    lvl = max(bucket, mn)
    return max(0, min(5, lvl))


def parse_skill_excel(path: str, sheet_name: str = "Sheet1"):
    """
    Fast parser for スキル集計_*.xlsx without openpyxl (this file contains pivot caches and can be slow).

    Returns dict keyed by (norm_company, norm_name) -> {"p1":lvl, "p2":lvl, "p3":lvl, "p4":lvl}

    We only need a small subset of Sheet1:
      - Row 4: group labels (1:Module Setup / 2:Hardware Setup / 3:Function Setup)
      - Row 6: headers (氏名, 所属, 合計 / Level, 最小 / Level ...)
      - Rows 7+: data
    """
    import zipfile
    import xml.etree.ElementTree as ET

    # --- shared strings ---
    def load_shared_strings(z):
        try:
            data = z.read("xl/sharedStrings.xml")
        except KeyError:
            return []
        root = ET.fromstring(data)
        ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        strings = []
        for si in root.findall("s:si", ns):
            # collect all <t> inside (some are rich text)
            ts = si.findall(".//s:t", ns)
            strings.append("".join([t.text or "" for t in ts]))
        return strings

    # --- map sheet_name -> worksheet xml path ---
    def sheet_xml_path(z, target_name):
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        ns = {
            "w": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        rid = None
        for sh in wb.findall("w:sheets/w:sheet", ns):
            if sh.get("name") == target_name:
                rid = sh.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                break
        if not rid:
            raise ValueError(f"Sheet '{target_name}' not found in {path}.")

        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        nsr = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
        for rel in rels.findall("rel:Relationship", nsr):
            if rel.get("Id") == rid:
                target = rel.get("Target")
                if not target:
                    break
                # Target like "worksheets/sheet1.xml"
                return "xl/" + target.lstrip("/")
        raise RuntimeError(f"Could not resolve xml path for sheet '{target_name}'.")

    def col_letters_to_index(letters: str) -> int:
        letters = letters.upper()
        n = 0
        for ch in letters:
            if "A" <= ch <= "Z":
                n = n * 26 + (ord(ch) - ord("A") + 1)
        return n

    def cell_ref_to_rc(a1: str):
        # e.g. "G6" -> (row=6, col=7)
        m = re.match(r"^([A-Za-z]+)(\d+)$", a1)
        if not m:
            return None, None
        col = col_letters_to_index(m.group(1))
        row = int(m.group(2))
        return row, col

    # We only care these columns:
    needed_cols = {1, 2, 7, 8, 9, 10, 11, 12}  # A,B,G..L

    with zipfile.ZipFile(path) as z:
        shared = load_shared_strings(z)
        sheet_path = sheet_xml_path(z, sheet_name)
        root = ET.fromstring(z.read(sheet_path))

    ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    # Read rows into a dict: row_idx -> {col_idx -> value}
    rows = {}
    for row in root.findall(".//s:sheetData/s:row", ns):
        r_idx = int(row.get("r"))
        if r_idx < 4:
            continue
        if r_idx > 6 and r_idx > 20000:
            break
        # we still parse all rows >=4 because we need row4/6 and data rows
        row_map = {}
        for c in row.findall("s:c", ns):
            a1 = c.get("r")
            rr, cc = cell_ref_to_rc(a1)
            if cc not in needed_cols:
                continue
            t = c.get("t")  # 's' for shared string
            v = c.find("s:v", ns)
            if v is None or v.text is None:
                continue
            raw = v.text
            if t == "s":
                try:
                    row_map[cc] = shared[int(raw)]
                except Exception:
                    row_map[cc] = ""
            else:
                # number; keep as int if possible
                if raw.isdigit():
                    row_map[cc] = int(raw)
                else:
                    try:
                        row_map[cc] = float(raw)
                    except Exception:
                        row_map[cc] = raw
        if row_map:
            rows[r_idx] = row_map

    # group labels from row 4 (start columns 7,9,11 typically)
    row4 = rows.get(4, {})
    g1 = 7 if str(row4.get(7, "")).strip() == "1:Module Setup" else None
    g2 = 9 if str(row4.get(9, "")).strip() == "2:Hardware Setup" else None
    g3 = 11 if str(row4.get(11, "")).strip() == "3:Function Setup" else None

    if not (g1 and g2 and g3):
        # fallback: search within needed cols
        def find_label(label):
            for col in sorted(needed_cols):
                if str(row4.get(col, "")).strip() == label:
                    return col
            return None
        g1 = find_label("1:Module Setup")
        g2 = find_label("2:Hardware Setup")
        g3 = find_label("3:Function Setup")

    if not (g1 and g2 and g3):
        raise RuntimeError("Could not find group labels (1/2/3) on row 4 in skill sheet.")

    p1_total, p1_min = g1, g1 + 1
    p2_total, p2_min = g2, g2 + 1
    p3_total, p3_min = g3, g3 + 1

    # Headers row6 to find A/B meaning
    # (We assume A=氏名, B=所属 in this file; but still keep robust.)
    # We'll just use col 1 as name, col 2 as company.
    skill_map = {}
    for r_idx in sorted(rows.keys()):
        if r_idx <= 6:
            continue
        row = rows[r_idx]
        name = row.get(1)
        comp = row.get(2, "")
        if not (isinstance(name, str) and name.strip()):
            continue
        name_s = name.strip()
        comp_s = str(comp).strip() if isinstance(comp, str) and str(comp).strip() else ""

        p1_lvl = _skill_level(row.get(p1_total, 0), row.get(p1_min, 0))
        p2_lvl = _skill_level(row.get(p2_total, 0), row.get(p2_min, 0))
        p3_lvl = _skill_level(row.get(p3_total, 0), row.get(p3_min, 0))

        skill_map[(_norm(comp_s), _norm(name_s))] = {
            "p1": p1_lvl,
            "p2": p2_lvl,
            "p3": p3_lvl,
            "p4": p3_lvl,
        }

    return skill_map


# ============================================================
# Build assignments (SU_Others -> task list via code)
# ============================================================

def build_assignments(su_data: dict, task_data: dict):
    code_to_phases = task_data["code_to_phases"]
    worker_date_map = su_data["worker_date_map"]
    worker_personal_map = su_data["worker_personal_map"]

    known_assign_map = defaultdict(list)  # (wid, phase_id) -> [date,...]

    misc_label_dates = defaultdict(set)
    misc_worker_label_dates = defaultdict(list)

    # personal business (grey)
    personal_label_dates = defaultdict(set)
    personal_worker_label_dates = defaultdict(list)

    # Track which worker did which phase (for later skill inference)
    inferred_worker_phase = defaultdict(set)  # wid -> {"p1","p2",...}

    # ----- normal cells -----
    for (wid, dt), text in worker_date_map.items():
        code = extract_tool_code(text)
        if code and code in code_to_phases:
            # assign to the phase whose window includes this day
            matched = False
            for phase_meta in code_to_phases[code]:
                ps = phase_meta["start"]
                pe = phase_meta["end"]
                if ps <= dt <= pe:
                    known_assign_map[(wid, phase_meta["phase_id"])].append(dt)
                    inferred_worker_phase[wid].add(phase_meta["operation"])  # "p1".. "p4"
                    matched = True
                    break
            if not matched:
                # If we couldn't match into any phase window, treat as misc label to avoid losing data
                label = text.strip()
                misc_label_dates[label].add(dt)
                misc_worker_label_dates[(wid, label)].append(dt)
        else:
            label = text.strip()
            misc_label_dates[label].add(dt)
            misc_worker_label_dates[(wid, label)].append(dt)

    # ----- grey cells => personal business -----
    for (wid, dt), text in worker_personal_map.items():
        label = text.strip() if isinstance(text, str) and text.strip() else "personal_business"
        personal_label_dates[label].add(dt)
        personal_worker_label_dates[(wid, label)].append(dt)

    assignments = []

    # ---------- tool assignments (Flexible) ----------
    for (wid, phase_id), dates in known_assign_map.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        work_date_list = [{"hour": 12, "date": _to_ymd(d)} for d in uniq_dates]
        assignments.append({
            "worker": wid,
            "operation_task": phase_id,
            "start_date": _to_ymd(uniq_dates[0]),
            "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list,
            "plan_flexibility": "Flexible",
        })

    # ---------- dummy workflow tasks for misc work ----------
    misc_tasks = []
    misc_label_to_phase = {}  # label -> {"phase_id", "start", "end"}
    misc_counter = 1

    for label, dates in misc_label_dates.items():
        if not dates:
            continue
        start = min(dates)
        end = max(dates)
        task_id = f"misc_{misc_counter}"
        misc_counter += 1
        phase_id = f"{task_id}_p1"
        misc_label_to_phase[label] = {"phase_id": phase_id, "start": start, "end": end}

        workload_days = int((end - start).days) + 1

        misc_tasks.append({
            "id": task_id,
            "name": label,
            "workflow": "wf_other",
            "fab": "f_other",
            "phase_task_list": [{
                "id": phase_id,
                "name": "Other work",
                "phase": "other_p1",
                "start_date": _to_ymd(start),
                "end_date": _to_ymd(end),
                "operation_task_list": [{
                    "id": phase_id,
                    "name": "Other work",
                    "operation": "other_op",
                    "workload_days": workload_days,
                }],
            }],
        })

    # ---------- misc assignments (Fixed) ----------
    for (wid, label), dates in misc_worker_label_dates.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        phase_meta = misc_label_to_phase.get(label)
        if not phase_meta:
            continue
        phase_id = phase_meta["phase_id"]
        work_date_list = [{"hour": 8, "date": _to_ymd(d)} for d in uniq_dates]

        assignments.append({
            "worker": wid,
            "operation_task": phase_id,
            "start_date": _to_ymd(uniq_dates[0]),
            "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list,
            "plan_flexibility": "Fixed",
        })

    # ========== personal business tasks + assignments ==========
    personal_tasks = []
    personal_label_to_phase = {}
    pb_counter = 1

    for label, dates in personal_label_dates.items():
        if not dates:
            continue
        start = min(dates)
        end = max(dates)

        task_id = f"pb_{pb_counter}"
        pb_counter += 1
        phase_id = f"{task_id}_p1"
        personal_label_to_phase[label] = {"phase_id": phase_id, "start": start, "end": end}

        workload_days = int((end - start).days) + 1

        personal_tasks.append({
            "id": task_id,
            "name": label,
            "workflow": "wf_personal_business",
            "fab": "f_other",
            "phase_task_list": [{
                "id": phase_id,
                "name": "Personal Business",
                "phase": "pb_p1",
                "start_date": _to_ymd(start),
                "end_date": _to_ymd(end),
                "operation_task_list": [{
                    "id": phase_id,
                    "name": "Personal Business",
                    "operation": "personal_business_op",
                    "workload_days": workload_days,
                }],
            }],
        })

    for (wid, label), dates in personal_worker_label_dates.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        phase_meta = personal_label_to_phase.get(label)
        if not phase_meta:
            continue
        phase_id = phase_meta["phase_id"]
        work_date_list = [{"hour": 8, "date": _to_ymd(d)} for d in uniq_dates]

        assignments.append({
            "worker": wid,
            "operation_task": phase_id,
            "start_date": _to_ymd(uniq_dates[0]),
            "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list,
            "plan_flexibility": "Fixed",
        })

    return assignments, misc_tasks, personal_tasks, inferred_worker_phase


# ============================================================
# Build EnvConfig + Schedule and dump YAML
# ============================================================

def build_env_and_schedule_v2(
    su_others_path: str,
    task_csv_path: str,
    skill_excel_path: str,
    envconfig_out: str = "EnvConfig_from_excel_v2.yaml",
    schedule_out: str = "Schedule_from_excel_v2.yaml",
):
    # ---------- Parse inputs ----------
    su_data = parse_su_others(su_others_path)
    task_data = parse_tasks_from_csv(task_csv_path)
    skill_data = parse_skill_excel(skill_excel_path)

    # ---------- Build EnvConfig dynamic lists ----------
    # Customer companies, Countries (as regions), Fabs
    customer_name_to_id = {"OTHER": "c_other"}
    region_name_to_id = {"Other": "r_other"}
    fab_name_to_id = {"Other": "f_other"}

    customer_company_list = [{
        "id": "c_other",
        "name": "OTHER",
        "unavailable_dates": [],
    }]

    region_list = [{
        "id": "r_other",
        "name": "Other",
        "max_stay_on": 90,
        "max_annual_stay": 240,
        "stay_off_interval": 3,
        "unavailable_dates": [{"weekly": {"weekdays": ["sat", "sun"]}}],
    }]

    fab_list = [{
        "id": "f_other",
        "name": "Other",
        "region": "r_other",
        "customer_company": "c_other",
        "unavailable_dates": [],
    }]

    def get_customer_id(name: str) -> str:
        nm = name.strip() if isinstance(name, str) and name.strip() else "OTHER"
        if nm not in customer_name_to_id:
            cid = f"c{len(customer_name_to_id)}"  # c1, c2, ...
            customer_name_to_id[nm] = cid
            customer_company_list.append({
                "id": cid,
                "name": nm,
                "unavailable_dates": [],
            })
        return customer_name_to_id[nm]

    def get_region_id(country: str) -> str:
        nm = country.strip() if isinstance(country, str) and country.strip() else "Other"
        if nm not in region_name_to_id:
            rid = f"r{len(region_name_to_id)}"  # r1, r2, ...
            region_name_to_id[nm] = rid
            region_list.append({
                "id": rid,
                "name": nm,
                "max_stay_on": 90,
                "max_annual_stay": 240,
                "stay_off_interval": 3,
                "unavailable_dates": [{"weekly": {"weekdays": ["sat", "sun"]}}],
            })
        return region_name_to_id[nm]

    def get_fab_id(fab_name: str, country: str, customer: str) -> str:
        nm = fab_name.strip() if isinstance(fab_name, str) and fab_name.strip() else "Other"
        if nm not in fab_name_to_id:
            fid = f"f{len(fab_name_to_id)}"  # f1, f2, ...
            fab_name_to_id[nm] = fid
            fab_list.append({
                "id": fid,
                "name": nm,
                "region": get_region_id(country),
                "customer_company": get_customer_id(customer),
                "unavailable_dates": [],
            })
        return fab_name_to_id[nm]

    # Assign fab ids into tool_tasks based on CSV meta
    for t in task_data["tool_tasks"]:
        fid = get_fab_id(t.get("fab_name"), t.get("country"), t.get("customer"))
        t["fab"] = fid

    # ---------- ENVIRONMENT / workflows ----------
    environment = {
        "workflow_list": [
            {
                "id": "wf_tool",
                "name": "Tool Install",
                "phase_list": [
                    {
                        "id": "tool_p1",
                        "name": "Module Setup",
                        "operation_list": [{
                            "id": "p1",
                            "name": "Module Setup",
                            "work_hours": [8],
                            "min_worker_num": 1,
                            "max_worker_num": 3,
                        }],
                    },
                    {
                        "id": "tool_p2",
                        "name": "Hardware Setup",
                        "operation_list": [{
                            "id": "p2",
                            "name": "Hardware Setup",
                            "work_hours": [8],
                            "min_worker_num": 1,
                            "max_worker_num": 3,
                        }],
                    },
                    {
                        "id": "tool_p3",
                        "name": "Function Setup",
                        "operation_list": [{
                            "id": "p3",
                            "name": "Function Setup",
                            "work_hours": [8],
                            "min_worker_num": 1,
                            "max_worker_num": 3,
                        }],
                    },
                    {
                        "id": "tool_p4",
                        "name": "Utility",
                        "operation_list": [{
                            "id": "p4",
                            "name": "Utility",
                            "work_hours": [8],
                            "min_worker_num": 1,
                            "max_worker_num": 3,
                        }],
                    },
                ],
            },
            {
                # Dummy workflow for "other" work that only appears in SU_Others
                "id": "wf_other",
                "name": "Other work (from SU_Others)",
                "phase_list": [{
                    "id": "other_p1",
                    "name": "Other work",
                    "operation_list": [{
                        "id": "other_op",
                        "name": "Other work",
                        "work_hours": [8],
                        "min_worker_num": 1,
                        "max_worker_num": 3,
                    }],
                }],
            },
            {
                "id": "wf_personal_business",
                "name": "Personal Business (from SU_Others grey cells)",
                "phase_list": [{
                    "id": "pb_p1",
                    "name": "Personal Business",
                    "operation_list": [{
                        "id": "personal_business_op",
                        "name": "Personal Business",
                        "work_hours": [8],
                        "min_worker_num": 1,
                        "max_worker_num": 1,
                    }],
                }],
            },
        ],
        "fab_list": fab_list,
        "region_list": region_list,
        "customer_company_list": customer_company_list,
        "worker_company_list": su_data["worker_company_list"],
        "worker_list": su_data["worker_list"],
        "transite_day_map": [],
    }

    # ---------- Assignments + extra tasks ----------
    assignments, misc_tasks, personal_tasks, inferred_worker_phase = build_assignments(su_data, task_data)

    # ---------- Fill worker skill_map ----------
    # 1) initialize all workers with 0 for p1..p4
    # 2) if skill excel has entry -> use it
    # 3) else infer skill>=1 for phases they actually worked on (Flexible tool assignments)
    wid_to_worker = {w["id"]: w for w in environment["worker_list"]}

    # Build a (norm_company, norm_name) lookup from current workers
    # We also need the original strings to build the key.
    worker_company_id_to_name = {wc["id"]: wc["name"] for wc in su_data["worker_company_list"]}

    workers_with_skill_excel = set()

    for w in environment["worker_list"]:
        w["skill_map"] = {
            "p1": 0, "p2": 0, "p3": 0, "p4": 0,
            "other_op": 1,
            "personal_business_op": 1,
        }
        comp_name = worker_company_id_to_name.get(w["worker_company"], "")
        key = (_norm(comp_name), _norm(w["name"]))
        if key in skill_data:
            w["skill_map"].update(skill_data[key])
            workers_with_skill_excel.add(w["id"])

    for wid, ops in inferred_worker_phase.items():
        if wid in workers_with_skill_excel:
            continue
        w = wid_to_worker.get(wid)
        if not w:
            continue
        for op in ops:
            if op in ("p1", "p2", "p3", "p4"):
                w["skill_map"][op] = max(w["skill_map"].get(op, 0), 1)

    # ---------- SCHEDULE ----------
    # plan_range from combined dates
    all_dates = []
    all_dates.append(pd.to_datetime(su_data["plan_range"]["start_date"]))
    all_dates.append(pd.to_datetime(su_data["plan_range"]["end_date"]))
    all_dates.extend(task_data["date_list"])
    all_dates = [d for d in all_dates if isinstance(d, pd.Timestamp)]
    all_dates.sort()

    plan_range = {
        "start_date": _to_ymd(all_dates[0]) if all_dates else su_data["plan_range"]["start_date"],
        "end_date": _to_ymd(all_dates[-1]) if all_dates else su_data["plan_range"]["end_date"],
    }

    # Remove internal keys from tool tasks before writing YAML
    tool_tasks_for_yaml = []
    for t in task_data["tool_tasks"]:
        t_copy = dict(t)
        t_copy.pop("module_code", None)
        t_copy.pop("customer", None)
        t_copy.pop("country", None)
        t_copy.pop("fab_name", None)
        tool_tasks_for_yaml.append(t_copy)

    schedule = {
        "plan_range": plan_range,
        "workflow_task_list": tool_tasks_for_yaml + misc_tasks + personal_tasks,
        "assignment_list": assignments,
    }

    env_root = {"environment": environment}
    sch_root = {"schedule": schedule}

    # Prefer C-based dumper if available (much faster on big YAML)
    _BaseDumper = getattr(yaml, "CSafeDumper", yaml.SafeDumper)

    class NoAliasDumper(_BaseDumper):
        def ignore_aliases(self, data):
            return True

    with open(envconfig_out, "w", encoding="utf-8") as f:
        yaml.dump(
            env_root,
            f,
            Dumper=NoAliasDumper,
            sort_keys=False,
            allow_unicode=True,
            width=4096,
        )

    with open(schedule_out, "w", encoding="utf-8") as f:
        yaml.dump(
            sch_root,
            f,
            Dumper=NoAliasDumper,
            sort_keys=False,
            allow_unicode=True,
            width=4096,
        )

    return env_root, sch_root


# ============================================================
# Entrypoint
# ============================================================

if __name__ == "__main__":
    su_file = "20260105 SU_Others.xlsm"
    task_file = "SU_Others_予定表_2025_新規製番リスト_20260127.xlsx"
    skill_file = "スキル集計_20260127.xlsx"

    su_path = Path(su_file)
    task_path = Path(task_file)
    skill_path = Path(skill_file)

    if su_path.exists() and task_path.exists() and skill_path.exists():
        build_env_and_schedule_v2(
            str(su_path),
            str(task_path),
            str(skill_path),
            envconfig_out="EnvConfig_from_excel_decoder2.yaml",
            schedule_out="Schedule_from_excel_decoder2.yaml",
        )
        print("EnvConfig_from_excel_decoder2.yaml and Schedule_from_excel_decoder2.yaml have been written.")
    else:
        print("Please fix input file paths at the bottom of Decoder2.py.")
