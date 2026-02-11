
# Decoder3.py
# ---------------------------------------------------------------------
# Generates EnvConfig.yaml + Schedule.yaml for the Timefold scheduler from Excels.
#
# Decoder3 major change:
# - "新規製番リスト" provides planned phase dates; "SU_Others" provides actual execution span.
# - For each module code, we compute planned phase ratios (by planned phase lengths),
#   then "shift" (rescale) the plan onto the actual span found in SU_Others.
# - The shifted phase windows are used for:
#     (1) task phase start/end in Schedule.yaml
#     (2) mapping SU_Others daily assignments -> phase assignments
#     (3) inferring worker skills from SU_Others assignments
#     (4) worker-day workload counting per phase
# - Modules with invalid planned date fields are CUT from 新規製番リスト and treated as "not listed"
#   (so any SU_Others occurrences become dummy "other" tasks).
#
# Outputs:
#   - EnvConfig.yaml
#   - Schedule.yaml
#   - TransformationLog.txt
#
# ---------------------------------------------------------------------
# CONFIG (edit here)
# ---------------------------------------------------------------------
READ_ALL_DATA = True
DATE_RANGE = None   # e.g. "2026/01/01-2026/03/31" or None

# Workload calculation for wf_tool operation_task_list[].workload_days:
#   OptionA = window-days (end-start+1) from shifted phase window
#   OptionB = worker-days (1 worker x 1 day = 1), counted from shifted SU_Others assignments
# If True  => workload_days = max(OptionA, OptionB)
# If False => workload_days = OptionB only
WORKLOAD_USE_WINDOW_DAYS = False

# Transformation log filename
TRANSFORMATION_LOG = "TransformationLog.txt"
# ---------------------------------------------------------------------

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import unicodedata
import math

import pandas as pd
import yaml
from openpyxl import load_workbook

# ============================================================
# Helpers
# ============================================================

def _to_ymd(dt) -> str:
    if isinstance(dt, pd.Timestamp):
        return dt.strftime("%Y/%m/%d")
    if isinstance(dt, datetime):
        return pd.Timestamp(dt).strftime("%Y/%m/%d")
    return str(dt)

def _as_timestamp(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, pd.Timestamp):
        return v.normalize()
    if isinstance(v, datetime):
        return pd.Timestamp(v).normalize()
    if isinstance(v, str):
        s = v.strip()
        if not s or s.upper() == "N/A":
            return None
        dt = pd.to_datetime(s, errors="coerce")
        if isinstance(dt, pd.Timestamp) and not pd.isna(dt):
            return dt.normalize()
    return None

def _parse_date_range(date_range_str):
    if not date_range_str:
        return None, None
    s = str(date_range_str).strip()
    if not s:
        return None, None
    parts = [p.strip() for p in s.replace("〜", "-").split("-") if p.strip()]
    if len(parts) != 2:
        raise ValueError(f"DATE_RANGE must be like 'YYYY/MM/DD-YYYY/MM/DD'. Got: {date_range_str}")
    a = pd.to_datetime(parts[0], errors="coerce")
    b = pd.to_datetime(parts[1], errors="coerce")
    if not isinstance(a, pd.Timestamp) or pd.isna(a) or not isinstance(b, pd.Timestamp) or pd.isna(b):
        raise ValueError(f"DATE_RANGE could not be parsed. Got: {date_range_str}")
    if b < a:
        a, b = b, a
    return a.normalize(), b.normalize()

def _overlaps(a_start, a_end, b_start, b_end) -> bool:
    return (a_start <= b_end) and (a_end >= b_start)

_WS_RE = re.compile(r"\s+")
_ZERO_WIDTH = {"\u200b", "\u200c", "\u200d", "\ufeff"}

def _clean_text(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("　", " ")
    for z in _ZERO_WIDTH:
        s = s.replace(z, "")
    return s

def _norm_name(s: str) -> str:
    s = _clean_text(s)
    s = _WS_RE.sub("", s).strip()
    return s

def load_sheet_as_df(path: str, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    return pd.DataFrame(rows)

def load_sheet_as_df_with_header(path: str, sheet_name: str, header_row: int = 0) -> pd.DataFrame:
    df_raw = load_sheet_as_df(path, sheet_name)
    if df_raw.empty:
        return df_raw
    if header_row < 0 or header_row >= len(df_raw):
        raise ValueError(f"header_row={header_row} out of range for sheet {sheet_name} in {path}")
    headers = df_raw.iloc[header_row].tolist()
    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = headers
    df = df.reset_index(drop=True)
    return df

# ============================================================
# Code extraction from SU_Others cell text
# ============================================================

TOOLCODE_RE = re.compile(r"\d{3}[A-Z0-9]\d{5}A")

def extract_tool_code(s: str):
    if not isinstance(s, str):
        return None
    m = TOOLCODE_RE.search(s)
    return m.group(0) if m else None

# ============================================================
# SU_Others parsing
# ============================================================

IGNORE_WHITE_TEXT = {"FI", "FO"}
GREY_RGB_LAST6 = {"A6A6A6", "BFBFBF", "D9D9D9", "808080"}
RED_RGB_LAST6 = {"FF0000"}

def _cell_rgb_last6(cell):
    fill = getattr(cell, "fill", None)
    if fill is None:
        return None
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return None
    if getattr(fg, "type", None) != "rgb":
        return None
    rgb = getattr(fg, "rgb", None)
    if not rgb:
        return None
    return str(rgb).upper()[-6:]

def _find_date_header_ws(ws, max_scan_rows=12):
    for r in range(1, max_scan_rows + 1):
        row_vals = [cell.value for cell in ws[r]]
        if any(isinstance(v, (pd.Timestamp, datetime)) for v in row_vals):
            date_cols = [c for c, v in enumerate(row_vals, start=1)
                         if isinstance(v, (pd.Timestamp, datetime))]
            if not date_cols:
                continue
            dt_by_col = {c: pd.Timestamp(row_vals[c - 1]).normalize() for c in date_cols}
            return r, date_cols, dt_by_col
    raise RuntimeError("Could not find date header row in SU_Others.")

def parse_su_others(path: str, sheet_names=("予定表_2025",), date_filter=None):
    wb = load_workbook(path, data_only=True, read_only=False)
    used_sheets = [s for s in sheet_names if s in wb.sheetnames]
    if not used_sheets:
        raise ValueError(f"None of {sheet_names} found in {path}. Available: {wb.sheetnames}")

    f_start, f_end = date_filter if date_filter else (None, None)

    worker_company_map = {}
    worker_company_list = []

    def get_worker_company_id(company_name: str) -> str:
        company_name = str(company_name).strip()
        if company_name not in worker_company_map:
            cid = f"wc{len(worker_company_map) + 1}"
            worker_company_map[company_name] = cid
            worker_company_list.append({
                "id": cid,
                "name": company_name,
                "annual_overtime_limit": 10000,
                "monthly_overtime_limit": 10000,
                "unavailable_dates": [],
            })
        return worker_company_map[company_name]

    worker_key_to_id = {}
    worker_acc = {}
    worker_date_map = {}
    worker_personal_map = {}

    plan_start = None
    plan_end = None

    for sheet_name in used_sheets:
        ws = wb[sheet_name]
        date_row_idx, date_cols, dt_by_col = _find_date_header_ws(ws)

        if f_start is not None and f_end is not None:
            date_cols = [c for c in date_cols if f_start <= dt_by_col[c] <= f_end]
        if not date_cols:
            continue

        worker_start_row = date_row_idx + 2
        s = min(dt_by_col[c] for c in date_cols)
        e = max(dt_by_col[c] for c in date_cols)
        plan_start = s if plan_start is None else min(plan_start, s)
        plan_end = e if plan_end is None else max(plan_end, e)

        blank_streak = 0
        max_col = max(date_cols)

        for r, row_cells in enumerate(ws.iter_rows(min_row=worker_start_row, min_col=1, max_col=max_col), start=worker_start_row):
            company = row_cells[0].value
            name = row_cells[1].value

            if name is None or str(name).strip() == "":
                blank_streak += 1
                if blank_streak >= 30:
                    break
                continue
            blank_streak = 0

            company_str = "" if company is None else str(company).strip()
            name_str = str(name).strip()

            free_slot = row_cells[4].value if len(row_cells) >= 5 else None
            is_manager = bool(isinstance(free_slot, str) and "責" in free_slot)

            key = _norm_name(name_str)
            if key not in worker_key_to_id:
                wid = f"w{len(worker_key_to_id) + 1:03d}"
                worker_key_to_id[key] = wid
                company_id = get_worker_company_id(company_str)
                worker_acc[key] = {
                    "id": wid,
                    "name": name_str,
                    "worker_company": company_id,
                    "is_manager": is_manager,
                    "unavailable_set": set(),
                }
            else:
                if is_manager:
                    worker_acc[key]["is_manager"] = True

            wid = worker_key_to_id[key]

            for c in date_cols:
                dt = dt_by_col[c]
                cell = row_cells[c - 1]
                rgb6 = _cell_rgb_last6(cell)

                if rgb6 in RED_RGB_LAST6:
                    worker_acc[key]["unavailable_set"].add(_to_ymd(dt))
                    continue

                val = cell.value
                if not (isinstance(val, str) and val.strip()):
                    continue
                text = val.strip()

                if text.upper() in IGNORE_WHITE_TEXT:
                    continue

                if rgb6 in GREY_RGB_LAST6:
                    worker_personal_map[(wid, dt)] = text
                else:
                    worker_date_map[(wid, dt)] = text

    worker_list = []
    for acc in worker_acc.values():
        worker_list.append({
            "id": acc["id"],
            "name": acc["name"],
            "worker_company": acc["worker_company"],
            "is_manager": acc["is_manager"],
            "skill_map": {},
            "fab_suitability_map": [],
            "unavailable_dates": [{"date": d} for d in sorted(acc["unavailable_set"])],
        })

    plan_range = {
        "start_date": _to_ymd(plan_start) if plan_start is not None else "2025/01/01",
        "end_date": _to_ymd(plan_end) if plan_end is not None else "2025/01/01",
    }

    return {
        "worker_company_list": worker_company_list,
        "worker_company_map": worker_company_map,
        "worker_list": worker_list,
        "plan_range": plan_range,
        "worker_date_map": worker_date_map,
        "worker_personal_map": worker_personal_map,
    }

# ============================================================
# Task list from 新規製番リスト (sheet "CSV") + planned ratios
# ============================================================

def parse_tasks_from_csv_v3(path: str, sheet_name: str = "CSV", date_filter=None):
    df = load_sheet_as_df_with_header(path, sheet_name, header_row=0)
    colmap = {c: str(c).replace("\n", "").strip() for c in df.columns}
    df = df.rename(columns=colmap)

    code_col = [c for c in df.columns if "新規製番" in c]
    if not code_col:
        raise ValueError("Could not find 新規製番 column in CSV sheet.")
    code_col = code_col[0]

    cust_col = "ユーザー名"
    country_col = "国"
    fab_col = "ファブ名"

    p1s_col = "工程１開始可能日"
    p2s_col = "工程２開始可能日"
    p3s_col = "工程３開始可能日"
    p4s_col = "工程４開始可能日"
    p4e_col = "工程４終了予定日"

    f_start, f_end = date_filter if date_filter else (None, None)

    # outputs
    valid_rows = []
    cut_rows = []  # (code, reason)
    planned_meta = {}  # code -> dict with planned phase windows/lengths/percents + customer/country/fab_name
    all_dates = []

    for _, row in df.iterrows():
        code_raw = row.get(code_col)
        if not isinstance(code_raw, str) or not code_raw.strip():
            continue
        code = code_raw.strip()

        customer = row.get(cust_col)
        customer = str(customer).strip() if isinstance(customer, str) and str(customer).strip() else "OTHER"
        country = row.get(country_col)
        country = str(country).strip() if isinstance(country, str) and str(country).strip() else "Other"
        fab_name = row.get(fab_col)
        fab_name = str(fab_name).strip() if isinstance(fab_name, str) and str(fab_name).strip() else "Other"

        p1s = _as_timestamp(row.get(p1s_col))
        p2s = _as_timestamp(row.get(p2s_col))
        p3s = _as_timestamp(row.get(p3s_col))
        p4s = _as_timestamp(row.get(p4s_col))
        p4e = _as_timestamp(row.get(p4e_col))

        if any(x is None for x in [p1s, p2s, p3s, p4s, p4e]):
            cut_rows.append((code, "blank or N/A in phase date columns"))
            continue

        # order check
        if not (p1s <= p2s <= p3s <= p4s <= p4e):
            cut_rows.append((code, "wrong date order (p1<=p2<=p3<=p4<=p4end violated)"))
            continue

        starts = {1: p1s, 2: p2s, 3: p3s, 4: p4s}
        ends = {
            1: (p2s - pd.Timedelta(days=1)).normalize(),
            2: (p3s - pd.Timedelta(days=1)).normalize(),
            3: (p4s - pd.Timedelta(days=1)).normalize(),
            4: p4e,
        }

        # validate non-negative windows (should hold if order ok, but keep safe)
        bad = False
        for ph in (1,2,3,4):
            if ends[ph] < starts[ph]:
                cut_rows.append((code, f"phase {ph} end < start after applying end=next_start-1 rule"))
                bad = True
                break
        if bad:
            continue

        overall_start = min(starts.values())
        overall_end = max(ends.values())

        # filter by overlap
        if f_start is not None and f_end is not None:
            if not _overlaps(overall_start, overall_end, f_start, f_end):
                continue

        phase_len = {ph: int((ends[ph]-starts[ph]).days)+1 for ph in (1,2,3,4)}
        total_len = sum(phase_len.values())
        if total_len <= 0:
            cut_rows.append((code, "total planned duration <= 0"))
            continue

        phase_pct = {ph: (phase_len[ph] / total_len) for ph in (1,2,3,4)}

        planned_meta[code] = {
            "customer": customer,
            "country": country,
            "fab_name": fab_name,
            "starts": starts,
            "ends": ends,
            "phase_len": phase_len,
            "phase_pct": phase_pct,
            "overall_start": overall_start,
            "overall_end": overall_end,
            "total_len": total_len,
        }

        all_dates.extend([overall_start, overall_end])
        valid_rows.append(code)

    return {
        "valid_codes": valid_rows,
        "planned_meta": planned_meta,
        "cut_rows": cut_rows,
        "date_list": all_dates,
    }

# ============================================================
# Shifting / Rescaling plan onto SU_Others actual span
# ============================================================

def _allocate_phase_lengths(actual_total_days: int, planned_phase_len: dict):
    """
    Allocate integer lengths for phases 1..4 so that:
      - sum(lengths) == actual_total_days
      - ratios follow planned_phase_len proportions
    Uses floor + largest remainder; allows zeros if actual_total_days is very small.
    """
    phs = [1,2,3,4]
    total_planned = sum(max(0,int(planned_phase_len.get(ph,0))) for ph in phs)
    if total_planned <= 0:
        # fallback: equal split
        base = actual_total_days // 4
        rem = actual_total_days - base*4
        out = {1: base, 2: base, 3: base, 4: base}
        for i in range(rem):
            out[phs[i]] += 1
        return out

    raw = {ph: (actual_total_days * (planned_phase_len.get(ph,0)/total_planned)) for ph in phs}
    flo = {ph: int(math.floor(raw[ph])) for ph in phs}
    s = sum(flo.values())
    rem = actual_total_days - s

    # distribute remainder by fractional part
    frac = sorted(phs, key=lambda ph: (raw[ph]-flo[ph]), reverse=True)
    out = dict(flo)
    for i in range(rem):
        out[frac[i % len(frac)]] += 1

    # if rounding caused negative/overflow, fix defensively
    s2 = sum(out.values())
    if s2 != actual_total_days:
        # adjust last phase
        out[4] += (actual_total_days - s2)
    return out

def build_shifted_meta(planned_meta: dict, su_data: dict | None):
    """
    Returns:
      shifted_meta[code] = dict(shifted starts/ends, actual_first/last, allocation)
      code_to_shifted_phases[code] = list of phase dicts for assignment mapping
      extras for logging
    """
    shifted_meta = {}
    code_to_shifted_phases = defaultdict(list)

    # build quick index for SU_Others occurrences
    code_occ = defaultdict(list)  # code -> list of (dt, wid, text)
    if su_data is not None:
        for (wid, dt), text in su_data["worker_date_map"].items():
            code = extract_tool_code(text)
            if code:
                code_occ[code].append((dt, wid, text))

    for code, meta in planned_meta.items():
        # default (no shift)
        starts = meta["starts"]
        ends = meta["ends"]
        plan_overall_start = meta["overall_start"]
        plan_overall_end = meta["overall_end"]

        occ = code_occ.get(code, [])
        if occ:
            actual_first = min(x[0] for x in occ)
            actual_last = max(x[0] for x in occ)
            actual_total = int((actual_last-actual_first).days) + 1
            if actual_total <= 0:
                actual_total = 1

            alloc = _allocate_phase_lengths(actual_total, meta["phase_len"])

            shifted_starts = {}
            shifted_ends = {}
            cur = actual_first
            for ph in (1,2,3,4):
                ln = int(alloc.get(ph,0))
                if ln <= 0:
                    shifted_starts[ph] = cur
                    shifted_ends[ph] = (cur - pd.Timedelta(days=1)).normalize()  # empty window
                    continue
                shifted_starts[ph] = cur
                shifted_ends[ph] = (cur + pd.Timedelta(days=ln-1)).normalize()
                cur = (shifted_ends[ph] + pd.Timedelta(days=1)).normalize()

            # force last end to actual_last (in case of tiny drift)
            shifted_ends[4] = actual_last

            shifted_meta[code] = {
                "plan": meta,
                "had_su_match": True,
                "actual_first": actual_first,
                "actual_last": actual_last,
                "actual_total": actual_total,
                "alloc_days": alloc,
                "shifted_starts": shifted_starts,
                "shifted_ends": shifted_ends,
                "occ_sample": sorted(occ, key=lambda x: x[0])[:3],  # up to 3 earliest samples
                "occ_last_sample": sorted(occ, key=lambda x: x[0])[-3:],  # up to 3 latest samples
            }
        else:
            shifted_meta[code] = {
                "plan": meta,
                "had_su_match": False,
                "actual_first": None,
                "actual_last": None,
                "actual_total": None,
                "alloc_days": None,
                "shifted_starts": starts,
                "shifted_ends": ends,
                "occ_sample": [],
                "occ_last_sample": [],
            }

        for ph in (1,2,3,4):
            code_to_shifted_phases[code].append({
                "phase_index": ph,
                "start": shifted_meta[code]["shifted_starts"][ph],
                "end": shifted_meta[code]["shifted_ends"][ph],
                "operation": f"p{ph}",
            })

    return shifted_meta, code_to_shifted_phases, code_occ

# ============================================================
# Skill aggregation (スキル集計_*.xlsx) (same as decoder2)
# ============================================================

def _skill_level(total_level, min_level):
    try:
        total = 0 if total_level is None or (isinstance(total_level, float) and pd.isna(total_level)) else int(total_level)
    except Exception:
        total = 0
    try:
        mn = 0 if min_level is None or (isinstance(min_level, float) and pd.isna(min_level)) else int(min_level)
    except Exception:
        mn = 0

    bucket = 0 if total <= 20 else (total // 20)
    lvl = max(bucket, mn)
    return max(0, min(5, lvl))

def parse_skill_excel(path: str, sheet_name: str = "Sheet1"):
    df = load_sheet_as_df(path, sheet_name)

    header_row = None
    for r in range(0, min(20, len(df))):
        row = df.iloc[r].astype(str).tolist()
        if any("氏名" in c for c in row) and any("所属" in c for c in row):
            header_row = r
            break
    if header_row is None:
        header_row = 6

    headers = df.iloc[header_row].tolist()
    col_index = {str(h).strip(): i for i, h in enumerate(headers) if str(h).strip() != "nan"}

    def find_col_contains(keyword):
        for k, i in col_index.items():
            if keyword in k:
                return i
        return None

    name_c = find_col_contains("氏名")
    comp_c = find_col_contains("所属")

    def find_first_cell(text):
        for r in range(len(df)):
            for c in range(df.shape[1]):
                v = df.iat[r, c]
                if isinstance(v, str) and v.strip() == text:
                    return r, c
        return None, None

    g1_r, g1_c = find_first_cell("1:Module Setup")
    g2_r, g2_c = find_first_cell("2:Hardware Setup")
    g3_r, g3_c = find_first_cell("3:Function Setup")
    if g1_c is None or g2_c is None or g3_c is None:
        raise RuntimeError("Could not find group labels (1/2/3) in skill sheet.")

    p1_total, p1_min = g1_c, g1_c + 1
    p2_total, p2_min = g2_c, g2_c + 1
    p3_total, p3_min = g3_c, g3_c + 1

    start_row = max(g1_r, g2_r, g3_r) + 2

    skill_levels = {}
    people_meta = {}

    for r in range(start_row, len(df)):
        name = df.iat[r, name_c] if name_c is not None else None
        if not (isinstance(name, str) and name.strip()):
            continue
        comp = df.iat[r, comp_c] if comp_c is not None else ""
        name_s = name.strip()
        comp_s = str(comp).strip() if isinstance(comp, str) and str(comp).strip() else ""

        p1_lvl = _skill_level(df.iat[r, p1_total], df.iat[r, p1_min])
        p2_lvl = _skill_level(df.iat[r, p2_total], df.iat[r, p2_min])
        p3_lvl = _skill_level(df.iat[r, p3_total], df.iat[r, p3_min])

        key = _norm_name(name_s)
        skill_levels[key] = {"p1": p1_lvl, "p2": p2_lvl, "p3": p3_lvl, "p4": p3_lvl}
        if key not in people_meta:
            people_meta[key] = {"company": comp_s, "name": name_s}

    return skill_levels, people_meta

# ============================================================
# Build assignments using SHIFTED phase windows
# ============================================================

def build_assignments_v3(
    su_data: dict,
    shifted_code_to_phases: dict,
    valid_code_set: set,
    date_filter=None
):
    worker_date_map = su_data["worker_date_map"]
    worker_personal_map = su_data["worker_personal_map"]
    f_start, f_end = date_filter if date_filter else (None, None)

    known_assign_map = defaultdict(list)

    misc_label_dates = defaultdict(set)
    misc_worker_label_dates = defaultdict(list)

    personal_label_dates = defaultdict(set)
    personal_worker_label_dates = defaultdict(list)

    inferred_worker_phase = defaultdict(set)

    # for log: dummy tool-code-ish labels (in SU_Others but not in 新規製番リスト valid set)
    dummy_tool_labels = defaultdict(set)  # code -> set(full_text)

    # 1) NORMAL CELLS
    for (wid, dt), text in worker_date_map.items():
        if f_start is not None and f_end is not None and (dt < f_start or dt > f_end):
            continue

        code = extract_tool_code(text)

        if code and (code in valid_code_set) and (code in shifted_code_to_phases):
            matched = False
            for phase_meta in shifted_code_to_phases[code]:
                ps = phase_meta["start"]
                pe = phase_meta["end"]
                if ps is None or pe is None:
                    continue
                if ps <= dt <= pe:
                    phase_id = phase_meta["phase_id"]
                    known_assign_map[(wid, phase_id)].append(dt)
                    inferred_worker_phase[wid].add(phase_meta["operation"])
                    matched = True
                    break
            if matched:
                continue

        # otherwise: dummy "other"
        label = text.strip() if isinstance(text, str) else ""
        if not label:
            label = "other"
        misc_label_dates[label].add(dt)
        misc_worker_label_dates[(wid, label)].append(dt)

        if code and code not in valid_code_set:
            dummy_tool_labels[code].add(label)

    # 2) GREY CELLS = Personal Business
    for (wid, dt), text in worker_personal_map.items():
        if f_start is not None and f_end is not None and (dt < f_start or dt > f_end):
            continue
        label = text.strip() if isinstance(text, str) and text.strip() else "personal_business"
        personal_label_dates[label].add(dt)
        personal_worker_label_dates[(wid, label)].append(dt)

    # 3) TOOL ASSIGNMENTS (Flexible) + OptionB worker-day counting
    assignments = []
    phase_workerday_count = defaultdict(int)

    tmp_phase_to_worker_dates = defaultdict(set)  # (phase_id, wid) -> set(dt)
    for (wid, phase_id), dates in known_assign_map.items():
        for d in dates:
            tmp_phase_to_worker_dates[(phase_id, wid)].add(d)
    for (phase_id, wid), dset in tmp_phase_to_worker_dates.items():
        phase_workerday_count[phase_id] += len(dset)

    for (wid, phase_id), dates in known_assign_map.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        work_date_list = [{"hour": 10, "date": _to_ymd(d)} for d in uniq_dates]
        assignments.append({
            "worker": wid,
            "operation_task": phase_id,
            "start_date": _to_ymd(uniq_dates[0]),
            "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list,
            "plan_flexibility": "Flexible",
        })

    # 4) MISC ("OTHER") TASKS + FIXED ASSIGNMENTS
    misc_tasks = []
    misc_label_to_phase = {}
    misc_counter = 1

    for label, dates in misc_label_dates.items():
        if not dates:
            continue
        start = min(dates)
        end = max(dates)
        task_id = f"misc_{misc_counter}"
        misc_counter += 1
        phase_id = f"{task_id}_p1"
        misc_label_to_phase[label] = {"phase_id": phase_id, "start": start, "end": end}

        workload_days = int((end - start).days) + 1

        misc_tasks.append({
            "id": task_id,
            "name": label,
            "workflow": "wf_other",
            "fab": "f_other",   # Decoder3 change: never null
            "phase_task_list": [{
                "id": phase_id,
                "name": "Other work",
                "phase": "other_p1",
                "start_date": _to_ymd(start),
                "end_date": _to_ymd(end),
                "operation_task_list": [{
                    "id": phase_id,
                    "name": "Other work",
                    "operation": "other_op",
                    "workload_days": workload_days,
                }],
            }],
        })

    for (wid, label), dates in misc_worker_label_dates.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        meta = misc_label_to_phase.get(label)
        if not meta:
            continue
        phase_id = meta["phase_id"]
        work_date_list = [{"hour": 10, "date": _to_ymd(d)} for d in uniq_dates]
        assignments.append({
            "worker": wid,
            "operation_task": phase_id,
            "start_date": _to_ymd(uniq_dates[0]),
            "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list,
            "plan_flexibility": "Fixed",
        })

    # 5) PERSONAL BUSINESS TASKS + FIXED ASSIGNMENTS
    personal_tasks = []
    pb_label_to_phase = {}
    pb_counter = 1

    for label, dates in personal_label_dates.items():
        if not dates:
            continue
        start = min(dates)
        end = max(dates)
        task_id = f"pb_{pb_counter}"
        pb_counter += 1
        phase_id = f"{task_id}_p1"
        pb_label_to_phase[label] = {"phase_id": phase_id, "start": start, "end": end}

        workload_days = int((end - start).days) + 1

        personal_tasks.append({
            "id": task_id,
            "name": label,
            "workflow": "wf_personal_business",
            "fab": None,
            "phase_task_list": [{
                "id": phase_id,
                "name": "Personal Business",
                "phase": "pb_p1",
                "start_date": _to_ymd(start),
                "end_date": _to_ymd(end),
                "operation_task_list": [{
                    "id": phase_id,
                    "name": "Personal Business",
                    "operation": "personal_business_op",
                    "workload_days": workload_days,
                }],
            }],
        })

    for (wid, label), dates in personal_worker_label_dates.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        meta = pb_label_to_phase.get(label)
        if not meta:
            continue
        phase_id = meta["phase_id"]
        work_date_list = [{"hour": 10, "date": _to_ymd(d)} for d in uniq_dates]
        assignments.append({
            "worker": wid,
            "operation_task": phase_id,
            "start_date": _to_ymd(uniq_dates[0]),
            "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list,
            "plan_flexibility": "Fixed",
        })

    return assignments, misc_tasks, personal_tasks, inferred_worker_phase, phase_workerday_count, dummy_tool_labels

# ============================================================
# Transformation log building
# ============================================================

def _format_phase_line(ph, start, end, extra=""):
    s = _to_ymd(start) if start is not None else "N/A"
    e = _to_ymd(end) if end is not None else "N/A"
    if extra:
        return f"  - P{ph}: {s} - {e} {extra}"
    return f"  - P{ph}: {s} - {e}"

def write_transformation_log(
    out_path: str,
    cut_rows: list,
    shifted_meta: dict,
    worker_id_to_name: dict,
    workload_zero_phase_ids: list,
    dummy_tool_labels: dict,
):
    lines = []
    lines.append("Decoder3 Transformation Log")
    lines.append("")

    # CUT OUT
    lines.append("---------------------- CUT OUT (ignored from 新規製番リスト) ----------------------")
    if not cut_rows:
        lines.append("(none)")
    else:
        for code, reason in cut_rows:
            lines.append(f"- {code}: {reason}")
    lines.append("")

    # SHIFTING DATE
    lines.append("---------------------- SHIFTING DATE (plan -> actual) ----------------------")
    for code in sorted(shifted_meta.keys()):
        m = shifted_meta[code]
        plan = m["plan"]
        lines.append(f"module: {code}")

        # planned
        lines.append("planned (新規製番リスト):")
        total_len = plan["total_len"]
        for ph in (1,2,3,4):
            pl = plan["phase_len"][ph]
            pct = plan["phase_pct"][ph] * 100.0
            lines.append(_format_phase_line(ph, plan["starts"][ph], plan["ends"][ph], extra=f"(len={pl}d, {pct:.1f}%)"))
        lines.append(f"  planned overall: {_to_ymd(plan['overall_start'])} - {_to_ymd(plan['overall_end'])} (total={total_len}d)")

        # actual
        if m["had_su_match"]:
            lines.append("actual (SU_Others):")
            lines.append(f"  first date: {_to_ymd(m['actual_first'])}")
            for dt, wid, text in m["occ_sample"]:
                nm = worker_id_to_name.get(wid, wid)
                lines.append(f"    - {_to_ymd(dt)} / {wid}({nm}) / {text}")
            lines.append(f"  last date:  {_to_ymd(m['actual_last'])}")
            for dt, wid, text in m["occ_last_sample"]:
                nm = worker_id_to_name.get(wid, wid)
                lines.append(f"    - {_to_ymd(dt)} / {wid}({nm}) / {text}")

            # shifted
            lines.append("shifted result (used in Schedule.yaml):")
            for ph in (1,2,3,4):
                ln = m["alloc_days"].get(ph, 0) if m["alloc_days"] else 0
                lines.append(_format_phase_line(ph, m["shifted_starts"][ph], m["shifted_ends"][ph], extra=f"(alloc={ln}d)"))
            lines.append(f"  shifted overall: {_to_ymd(m['actual_first'])} - {_to_ymd(m['actual_last'])} (total={m['actual_total']}d)")
        else:
            lines.append("actual (SU_Others): NOT FOUND -> no shift (kept planned dates)")
            lines.append("shifted result (used in Schedule.yaml): (same as planned)")
        lines.append("")

    # workerday == 0
    lines.append("---------------------- WORKLOAD WARNING (worker-days == 0) ----------------------")
    if not workload_zero_phase_ids:
        lines.append("(none)")
    else:
        for pid in workload_zero_phase_ids:
            lines.append(f"- phase_id: {pid} (no assigned worker-days in SU_Others after shifting)")
    lines.append("")

    # dummy tool codes
    lines.append("---------------------- DUMMY MODULES (SU_Others tool-code not in 新規製番リスト) ----------------------")
    if not dummy_tool_labels:
        lines.append("(none)")
    else:
        for code in sorted(dummy_tool_labels.keys()):
            labels = sorted(dummy_tool_labels[code])
            lines.append(f"- {code}:")
            for lb in labels[:50]:
                lines.append(f"    - {lb}")
            if len(labels) > 50:
                lines.append(f"    ... ({len(labels)-50} more)")
    lines.append("")

    Path(out_path).write_text("\n".join(lines), encoding="utf-8")

# ============================================================
# Build EnvConfig + Schedule and dump YAML
# ============================================================

def build_env_and_schedule_decoder3(
    su_others_path: str,
    task_csv_path: str,
    skill_excel_path: str,
    envconfig_out: str = "EnvConfig.yaml",
    schedule_out: str = "Schedule.yaml",
    log_out: str = TRANSFORMATION_LOG,
    read_all_data: bool = True,
    date_range: str | None = None,
    workload_use_window_days: bool = True,
):
    date_filter = _parse_date_range(date_range) if date_range else (None, None)
    f_start, f_end = date_filter

    # 1) Planned meta from 新規製番リスト (strict validate)
    task_meta = parse_tasks_from_csv_v3(task_csv_path, date_filter=date_filter if f_start is not None else None)
    planned_meta = task_meta["planned_meta"]
    cut_rows = task_meta["cut_rows"]
    valid_code_set = set(planned_meta.keys())

    # 2) Skills from スキル集計
    skill_levels, skill_people = parse_skill_excel(skill_excel_path)

    # 3) SU_Others
    if read_all_data:
        su_data = parse_su_others(su_others_path, date_filter=date_filter if f_start is not None else None)
        worker_company_list = su_data["worker_company_list"]
        worker_list = su_data["worker_list"]
    else:
        su_data = None
        comp_to_id = {}
        worker_company_list = []

        def get_wc_id(cn: str):
            cn = str(cn).strip() if isinstance(cn, str) else ""
            if cn not in comp_to_id:
                cid = f"wc{len(comp_to_id) + 1}"
                comp_to_id[cn] = cid
                worker_company_list.append({
                    "id": cid,
                    "name": cn,
                    "annual_overtime_limit": 10000,
                    "monthly_overtime_limit": 10000,
                    "unavailable_dates": [],
                })
            return comp_to_id[cn]

        keys = sorted(skill_people.keys())
        worker_list = []
        for i, k in enumerate(keys, start=1):
            meta = skill_people[k]
            worker_list.append({
                "id": f"w{i:03d}",
                "name": meta["name"],
                "worker_company": get_wc_id(meta["company"]),
                "is_manager": False,
                "skill_map": {},
                "fab_suitability_map": [],
                "unavailable_dates": [],
            })

    # 4) Merge workers from skill excel (name-only key)
    wc_id_to_name = {wc["id"]: wc["name"] for wc in worker_company_list}
    wc_name_to_id = {wc["name"]: wc["id"] for wc in worker_company_list}

    def get_or_create_wc_id(company_name: str) -> str:
        company_name = str(company_name).strip() if isinstance(company_name, str) else ""
        if company_name not in wc_name_to_id:
            cid = f"wc{len(wc_name_to_id) + 1}"
            wc_name_to_id[company_name] = cid
            worker_company_list.append({
                "id": cid,
                "name": company_name,
                "annual_overtime_limit": 10000,
                "monthly_overtime_limit": 10000,
                "unavailable_dates": [],
            })
            wc_id_to_name[cid] = company_name
        return wc_name_to_id[company_name]

    worker_by_key = {_norm_name(w["name"]): w for w in worker_list}

    next_worker_num = 1
    if worker_list:
        try:
            next_worker_num = max(int(w["id"][1:]) for w in worker_list) + 1
        except Exception:
            next_worker_num = len(worker_list) + 1

    for key, meta in skill_people.items():
        if key in worker_by_key:
            continue
        cid = get_or_create_wc_id(meta["company"])
        new_w = {
            "id": f"w{next_worker_num:03d}",
            "name": meta["name"],
            "worker_company": cid,
            "is_manager": False,
            "skill_map": {},
            "fab_suitability_map": [],
            "unavailable_dates": [],
        }
        next_worker_num += 1
        worker_list.append(new_w)
        worker_by_key[key] = new_w

    worker_id_to_name = {w["id"]: w["name"] for w in worker_list}

    # 5) Build shifted meta using SU_Others
    shifted_meta, code_to_shifted_phases_simple, _code_occ = build_shifted_meta(planned_meta, su_data if read_all_data else None)

    # 6) Build tool tasks + code_to_phases (for assignments) with shifted windows
    tool_tasks = []
    code_to_phases = defaultdict(list)
    all_dates = []

    task_counter = 1
    for code in task_meta["valid_codes"]:
        meta = shifted_meta.get(code)
        if not meta:
            continue

        plan = meta["plan"]
        customer, country, fab_name = plan["customer"], plan["country"], plan["fab_name"]

        task_id = f"e{task_counter}"
        task_counter += 1

        phase_task_list = []
        for ph in (1,2,3,4):
            phase_id = f"{task_id}_p{ph}"
            start = meta["shifted_starts"][ph]
            end = meta["shifted_ends"][ph]
            if end < start:
                # empty window; collapse to start
                end = start
            nm = {1: "Module Setup", 2: "Hardware Setup", 3: "Function Setup", 4: "Utility"}.get(ph, f"P{ph}")
            workload_days_a = int((end - start).days) + 1

            phase_task_list.append({
                "id": phase_id,
                "name": nm,
                "phase": f"tool_p{ph}",
                "start_date": _to_ymd(start),
                "end_date": _to_ymd(end),
                "operation_task_list": [{
                    "id": phase_id,
                    "name": nm,
                    "operation": f"p{ph}",
                    "workload_days": workload_days_a,  # temp, overwritten later by option logic
                }],
            })

            code_to_phases[code].append({
                "phase_index": ph,
                "phase_id": phase_id,
                "start": start,
                "end": end,
                "operation": f"p{ph}",
            })

            all_dates.extend([start, end])

        tool_tasks.append({
            "id": task_id,
            "name": code,
            "workflow": "wf_tool",
            "fab": None,     # filled later
            "phase_task_list": phase_task_list,
            "module_code": code,   # internal
            "customer": customer,  # internal
            "country": country,    # internal
            "fab_name": fab_name,  # internal
        })

    # 7) customer/region/fab lists from tasks (same ids policy)
    customer_name_to_id = {"OTHER": "c_other"}
    region_name_to_id = {"Other": "r_other"}
    fab_name_to_id = {"Other": "f_other"}

    customer_company_list = [{"id": "c_other", "name": "OTHER", "unavailable_dates": []}]
    region_list = [{
        "id": "r_other",
        "name": "Other",
        "max_stay_on": 90,
        "max_annual_stay": 240,
        "stay_off_interval": 3,
        "unavailable_dates": [],
    }]
    fab_list = [{"id": "f_other", "name": "Other", "region": "r_other", "customer_company": "c_other", "unavailable_dates": []}]

    def get_customer_id(name: str) -> str:
        nm = name.strip() if isinstance(name, str) and name.strip() else "OTHER"
        if nm not in customer_name_to_id:
            cid = f"c{len(customer_name_to_id)}"
            customer_name_to_id[nm] = cid
            customer_company_list.append({"id": cid, "name": nm, "unavailable_dates": []})
        return customer_name_to_id[nm]

    def get_region_id(country: str) -> str:
        nm = country.strip() if isinstance(country, str) and country.strip() else "Other"
        if nm not in region_name_to_id:
            rid = f"r{len(region_name_to_id)}"
            region_name_to_id[nm] = rid
            region_list.append({
                "id": rid,
                "name": nm,
                "max_stay_on": 90,
                "max_annual_stay": 240,
                "stay_off_interval": 3,
                "unavailable_dates": [],
            })
        return region_name_to_id[nm]

    def get_fab_id(fab_name: str, country: str, customer: str) -> str:
        nm = fab_name.strip() if isinstance(fab_name, str) and fab_name.strip() else "Other"
        if nm not in fab_name_to_id:
            fid = f"f{len(fab_name_to_id)}"
            fab_name_to_id[nm] = fid
            fab_list.append({
                "id": fid,
                "name": nm,
                "region": get_region_id(country),
                "customer_company": get_customer_id(customer),
                "unavailable_dates": [],
            })
        return fab_name_to_id[nm]

    for t in tool_tasks:
        fid = get_fab_id(t.get("fab_name"), t.get("country"), t.get("customer"))
        t["fab"] = fid

    # Decoder3 change: transit gap involving r_other is 0, otherwise default 1
    def build_transite_day_map(region_list, days_default: int = 1):
        region_ids = [r["id"] for r in region_list if r.get("id")]
        out = []
        for fr in region_ids:
            for to in region_ids:
                if fr == to:
                    continue
                d = 0 if (fr == "r_other" or to == "r_other") else days_default
                out.append({"from": fr, "to": to, "days": d})
        return out

    transite_day_map = build_transite_day_map(region_list, days_default=1)

    # 8) environment
    environment = {
        "workflow_list": [
            {
                "id": "wf_tool",
                "name": "Tool Install",
                "phase_list": [
                    {"id": "tool_p1", "name": "Module Setup", "operation_list": [{"id": "p1", "name": "Module Setup", "work_hours": [8, 10, 12], "min_worker_num": 1, "max_worker_num": 3}]},
                    {"id": "tool_p2", "name": "Hardware Setup", "operation_list": [{"id": "p2", "name": "Hardware Setup", "work_hours": [8, 10, 12], "min_worker_num": 1, "max_worker_num": 3}]},
                    {"id": "tool_p3", "name": "Function Setup", "operation_list": [{"id": "p3", "name": "Function Setup", "work_hours": [8, 10, 12], "min_worker_num": 1, "max_worker_num": 3}]},
                    {"id": "tool_p4", "name": "Utility", "operation_list": [{"id": "p4", "name": "Utility", "work_hours": [8, 10, 12], "min_worker_num": 1, "max_worker_num": 3}]},
                ],
            },
            {
                "id": "wf_other",
                "name": "Other work (from SU_Others)",
                "phase_list": [{
                    "id": "other_p1",
                    "name": "Other work",
                    "operation_list": [{"id": "other_op", "name": "Other work", "work_hours": [8, 10, 12], "min_worker_num": 1, "max_worker_num": 8}],
                }],
            },
            {
                "id": "wf_personal_business",
                "name": "Personal Business",
                "phase_list": [{
                    "id": "pb_p1",
                    "name": "Personal Business",
                    "operation_list": [{"id": "personal_business_op", "name": "Personal Business", "work_hours": [8, 10, 12], "min_worker_num": 1, "max_worker_num": 8}],
                }],
            },
        ],
        "fab_list": fab_list,
        "region_list": region_list,
        "customer_company_list": customer_company_list,
        "worker_company_list": worker_company_list,
        "transite_day_map": transite_day_map,
        "worker_list": worker_list,
    }

    # 9) assignments
    if read_all_data and su_data is not None:
        assignments, misc_tasks, personal_tasks, inferred_worker_phase, phase_workerday_count, dummy_tool_labels = build_assignments_v3(
            su_data, code_to_phases, valid_code_set, date_filter=date_filter if f_start is not None else None
        )
    else:
        assignments, misc_tasks, personal_tasks = [], [], []
        inferred_worker_phase = defaultdict(set)
        phase_workerday_count = defaultdict(int)
        dummy_tool_labels = {}

    # 10) fill skills (base + inferred + excel max)
    for w in environment["worker_list"]:
        w["skill_map"] = {
            "p1": 0, "p2": 0, "p3": 0, "p4": 0,
            "other_op": 1, "personal_business_op": 1
        }

    for wid, ops in inferred_worker_phase.items():
        w = next((x for x in environment["worker_list"] if x["id"] == wid), None)
        if not w:
            continue
        for op in ops:
            if op in ("p1", "p2", "p3", "p4"):
                w["skill_map"][op] = max(w["skill_map"].get(op, 0), 1)

    for w in environment["worker_list"]:
        key = _norm_name(w["name"])
        excel_map = skill_levels.get(key)
        if not excel_map:
            continue
        for op in ("p1", "p2", "p3", "p4"):
            w["skill_map"][op] = max(int(w["skill_map"].get(op, 0)), int(excel_map.get(op, 0)))

    # 11) plan_range
    if read_all_data and su_data is not None:
        try:
            all_dates.append(pd.to_datetime(su_data["plan_range"]["start_date"]))
            all_dates.append(pd.to_datetime(su_data["plan_range"]["end_date"]))
        except Exception:
            pass
    all_dates = [d for d in all_dates if isinstance(d, pd.Timestamp)]
    all_dates.sort()
    plan_range = {"start_date": _to_ymd(all_dates[0]), "end_date": _to_ymd(all_dates[-1])} if all_dates else {"start_date": "2025/01/01", "end_date": "2025/01/01"}

    # strip internal keys
    tool_tasks_for_yaml = []
    for t in tool_tasks:
        t2 = dict(t)
        for k in ("module_code", "customer", "country", "fab_name"):
            t2.pop(k, None)
        tool_tasks_for_yaml.append(t2)

    # 12) workload_days overwrite + collect workerday==0 warnings
    workload_zero_phase_ids = []
    for t in tool_tasks_for_yaml:
        if t.get("workflow") != "wf_tool":
            continue
        for pt in t.get("phase_task_list", []):
            phase_id = pt.get("id")  # eN_pX
            b = int(phase_workerday_count.get(phase_id, 0))
            if b == 0:
                workload_zero_phase_ids.append(phase_id)
            for ot in pt.get("operation_task_list", []):
                a = int(ot.get("workload_days", 0))
                if workload_use_window_days:
                    ot["workload_days"] = max(a, b)
                else:
                    ot["workload_days"] = b

    schedule = {
        "plan_range": plan_range,
        "workflow_task_list": tool_tasks_for_yaml + misc_tasks + personal_tasks,
        "assignment_list": assignments,
    }

    env_root = {"environment": environment}
    sch_root = {"schedule": schedule}

    _BaseDumper = getattr(yaml, "CSafeDumper", yaml.SafeDumper)

    class NoAliasDumper(_BaseDumper):
        def ignore_aliases(self, data):
            return True

    with open(envconfig_out, "w", encoding="utf-8") as f:
        yaml.dump(env_root, f, Dumper=NoAliasDumper, sort_keys=False, allow_unicode=True, width=4096)

    with open(schedule_out, "w", encoding="utf-8") as f:
        yaml.dump(sch_root, f, Dumper=NoAliasDumper, sort_keys=False, allow_unicode=True, width=4096)

    # transformation log
    write_transformation_log(
        log_out,
        cut_rows=cut_rows,
        shifted_meta=shifted_meta,
        worker_id_to_name=worker_id_to_name,
        workload_zero_phase_ids=sorted(set(workload_zero_phase_ids)),
        dummy_tool_labels=dummy_tool_labels,
    )

    return env_root, sch_root, shifted_meta


if __name__ == "__main__":
    su_file = "20260105 SU_Others.xlsm"
    task_file = "SU_Others_予定表_2025_新規製番リスト_20260127.xlsx"
    skill_file = "スキル集計_20260127.xlsx"

    su_path = Path(su_file)
    task_path = Path(task_file)
    skill_path = Path(skill_file)

    if (not READ_ALL_DATA or su_path.exists()) and task_path.exists() and skill_path.exists():
        build_env_and_schedule_decoder3(
            str(su_path),
            str(task_path),
            str(skill_path),
            envconfig_out="EnvConfig.yaml",
            schedule_out="Schedule.yaml",
            log_out=TRANSFORMATION_LOG,
            read_all_data=READ_ALL_DATA,
            date_range=DATE_RANGE,
            workload_use_window_days=WORKLOAD_USE_WINDOW_DAYS,
        )
        print("EnvConfig.yaml, Schedule.yaml, and TransformationLog.txt have been written.")
    else:
        print("Please fix input file paths at the bottom of Decoder3.py.")
