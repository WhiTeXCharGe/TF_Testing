
# Decoder3.py
# ---------------------------------------------------------------------
# Generates EnvConfig.yaml + Schedule.yaml for the Timefold scheduler from Excels.
#
# Decoder3 major change:
# - "新規製番リスト" provides planned phase dates; "SU_Others" provides actual execution span.
# - For each module code, we compute planned phase ratios (by planned phase lengths),
#   then "shift" (rescale) the plan onto the actual span found in SU_Others.
# - The shifted phase windows are used for:
#     (1) task phase start/end in Schedule.yaml
#     (2) mapping SU_Others daily assignments -> phase assignments
#     (3) inferring worker skills from SU_Others assignments
#     (4) worker-day workload counting per phase
# - Modules with invalid planned date fields are CUT from 新規製番リスト and treated as "not listed"
#   (so any SU_Others occurrences become dummy "other" tasks).
#
# Outputs:
#   - EnvConfig.yaml
#   - Schedule.yaml
#   - TransformationLog.txt
#
# ---------------------------------------------------------------------
# CONFIG (edit here)
# ---------------------------------------------------------------------
READ_ALL_DATA = True
DATE_RANGE = None   # e.g. "2026/01/01-2026/03/31" or None

#Default maximum worker for normal task
DEFAULT_MAX_WORKER = 8
# Workload calculation for wf_tool operation_task_list[].workload_days:
#   OptionA = window-days (end-start+1) from shifted phase window
#   OptionB = worker-days (1 worker x 1 day = 1), counted from shifted SU_Others assignments
# If True  => workload_days = max(OptionA, OptionB)
# If False => workload_days = OptionB only
WORKLOAD_USE_WINDOW_DAYS = False

# Transformation log filename
TRANSFORMATION_LOG = "TransformationLog.txt"

# If SU_Others actual span is completely outside the planned window
# by more than this many days, cut the module from 新規製番リスト
CUT_DISTANCE_DAYS = 365

# Shift policy:
# True  => shifting timeline = unique worked days (compressed)
# False => shifting timeline = continuous span (actual_start..actual_end)
SHIFT_USE_WORKED_DAYS = False

# Pre-cut policy:
# If True => before shifting, run "phase-zero workload" sanity check using span-based split,
#            and if any phase would get 0 worked days => cut whole module to dummy.
CUT_MODULE_IF_PHASE_ZERO_WORKLOAD = True

# If planned module has NO SU_Others match after cutting:
# A) False => keep module in Schedule.yaml (no shift), workload will be 0 and show in WORKLOAD WARNING
# B) True  => skip module entirely (do not write to Schedule.yaml, and do not show in WORKLOAD WARNING)
SKIP_MODULE_IF_NO_SU_MATCH = True  # Option A default

# If SU_Others has fewer than this many UNIQUE worked days for a module code,
# treat it as dummy "other" (do not shift, do not create wf_tool tasks).
MIN_WORKED_DAYS_FOR_TOOL = 4
# ---------------------------------------------------------------------

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import unicodedata
import math

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
# ============================================================
# Helpers
# ============================================================

def _to_ymd(dt) -> str:
    if isinstance(dt, pd.Timestamp):
        return dt.strftime("%Y/%m/%d")
    if isinstance(dt, datetime):
        return pd.Timestamp(dt).strftime("%Y/%m/%d")
    return str(dt)

def _as_timestamp(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, pd.Timestamp):
        return v.normalize()
    if isinstance(v, datetime):
        return pd.Timestamp(v).normalize()
    if isinstance(v, str):
        s = v.strip()
        if not s or s.upper() == "N/A":
            return None
        dt = pd.to_datetime(s, errors="coerce")
        if isinstance(dt, pd.Timestamp) and not pd.isna(dt):
            return dt.normalize()
    return None

def _parse_date_range(date_range_str):
    if not date_range_str:
        return None, None
    s = str(date_range_str).strip()
    if not s:
        return None, None
    parts = [p.strip() for p in s.replace("〜", "-").split("-") if p.strip()]
    if len(parts) != 2:
        raise ValueError(f"DATE_RANGE must be like 'YYYY/MM/DD-YYYY/MM/DD'. Got: {date_range_str}")
    a = pd.to_datetime(parts[0], errors="coerce")
    b = pd.to_datetime(parts[1], errors="coerce")
    if not isinstance(a, pd.Timestamp) or pd.isna(a) or not isinstance(b, pd.Timestamp) or pd.isna(b):
        raise ValueError(f"DATE_RANGE could not be parsed. Got: {date_range_str}")
    if b < a:
        a, b = b, a
    return a.normalize(), b.normalize()

def _overlaps(a_start, a_end, b_start, b_end) -> bool:
    return (a_start <= b_end) and (a_end >= b_start)

_WS_RE = re.compile(r"\s+")
_ZERO_WIDTH = {"\u200b", "\u200c", "\u200d", "\ufeff"}

def _clean_text(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("　", " ")
    for z in _ZERO_WIDTH:
        s = s.replace(z, "")
    return s

def _norm_name(s: str) -> str:
    s = _clean_text(s)
    s = _WS_RE.sub("", s).strip()
    return s

def load_sheet_as_df(path: str, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    return pd.DataFrame(rows)

def load_sheet_as_df_with_header(path: str, sheet_name: str, header_row: int = 0) -> pd.DataFrame:
    df_raw = load_sheet_as_df(path, sheet_name)
    if df_raw.empty:
        return df_raw
    if header_row < 0 or header_row >= len(df_raw):
        raise ValueError(f"header_row={header_row} out of range for sheet {sheet_name} in {path}")
    headers = df_raw.iloc[header_row].tolist()
    df = df_raw.iloc[header_row + 1:].copy()
    df.columns = headers
    df = df.reset_index(drop=True)
    return df

def _planned_actual_gap_days(planned_start, planned_end, actual_start, actual_end) -> int:
    """
    If actual overlaps planned -> gap=0.
    If actual is fully before planned -> gap = planned_start - actual_end
    If actual is fully after planned -> gap = actual_start - planned_end
    """
    if planned_start is None or planned_end is None or actual_start is None or actual_end is None:
        return 0
    if actual_end < planned_start:
        return int((planned_start - actual_end).days)
    if actual_start > planned_end:
        return int((actual_start - planned_end).days)
    return 0

def _remember_original_text(su_data: dict, wid: str, dt: pd.Timestamp, old_text: str, new_text: str):
    """
    Remember original SU cell text so we can restore it in outputs (Schedule.yaml / log),
    while keeping the internal modified text for matching logic.
    """
    m = su_data.setdefault("su_outlier_original_text", {})
    # key includes wid+date (enough because 1 cell = 1 text)
    k = (wid, _to_ymd(dt))
    # store only once (keep earliest/original)
    if k not in m:
        m[k] = {"old": old_text, "new": new_text}

def cut_su_short_span_modules_to_dummy(
    su_data: dict,
    min_unique_worked_days: int = 4,
    planned_meta: dict | None = None,
):
    """
    If a tool-code appears in SU_Others but its UNIQUE worked days are too few (< min_unique_worked_days),
    convert ALL its occurrences into dummy by breaking the tool-code pattern.

    Returns correction records for TransformationLog, same shape as cut_su_outlier_cells().
    """
    if not su_data:
        return []
    worker_date_map = su_data.get("worker_date_map", {})
    if not worker_date_map:
        return []

    # Gather occurrences: code -> list of (dt, wid, raw_text)
    code_to_occ = defaultdict(list)
    for (wid, dt), text in list(worker_date_map.items()):
        code = extract_tool_code(text)
        if not code:
            continue
        # optional: only enforce for codes that exist in planned_meta (real planned modules)
        if planned_meta is not None and code not in planned_meta:
            continue
        code_to_occ[code].append((dt, wid, text))

    corrections = []

    for code, occ in code_to_occ.items():
        uniq_days = sorted(set(dt for dt, _, _ in occ))
        if len(uniq_days) >= int(min_unique_worked_days):
            continue

        broken = _break_tool_code(code)
        for dt, wid, text in occ:
            if not isinstance(text, str):
                continue
            old = text
            new = old.replace(code, broken, 1)
            if new == old:
                continue

            _remember_original_text(su_data, wid, dt, old, new)
            worker_date_map[(wid, dt)] = new
            corrections.append({
                "wid": wid,
                "date": _to_ymd(dt),
                "code": code,
                "text": old,
                "reason": f"short-span cut: unique worked days {len(uniq_days)} < {min_unique_worked_days} (treated as dummy other)",
            })

    su_data["su_short_span_corrections"] = corrections
    return corrections


def cut_module_if_large_cluster_gap(
    su_data: dict,
    gap_days_threshold: int,
    planned_meta: dict | None = None,
):
    """
    If a module has multiple clusters separated by > gap_days_threshold,
    cut ALL its occurrences to dummy (no shifting).
    """
    if not su_data:
        return []

    worker_date_map = su_data.get("worker_date_map", {})
    if not worker_date_map:
        return []

    code_to_dates = defaultdict(list)

    for (wid, dt), text in worker_date_map.items():
        code = extract_tool_code(text)
        if not code:
            continue
        if planned_meta is not None and code not in planned_meta:
            continue
        code_to_dates[code].append(dt)

    corrections = []

    for code, dts in code_to_dates.items():
        uniq = sorted(set(dts))
        if len(uniq) <= 1:
            continue

        clusters = _cluster_by_date_gap(uniq, gap_days=1)

        if len(clusters) <= 1:
            continue

        # check gap between clusters
        cut_flag = False
        for i in range(len(clusters) - 1):
            gap = (clusters[i+1][0] - clusters[i][-1]).days
            if gap > gap_days_threshold:
                cut_flag = True
                break

        if not cut_flag:
            continue

        broken = _break_tool_code(code)

        for (wid, dt), text in list(worker_date_map.items()):
            if extract_tool_code(text) != code:
                continue
            old = text
            new = old.replace(code, broken, 1)
            if new == old:
                continue

            _remember_original_text(su_data, wid, dt, old, new)
            worker_date_map[(wid, dt)] = new

            corrections.append({
                "wid": wid,
                "date": _to_ymd(dt),
                "code": code,
                "text": old,
                "reason": f"module cut: cluster gap > {gap_days_threshold}",
            })

    su_data["su_large_gap_corrections"] = corrections
    return corrections

def cut_module_if_phase_zero_workload(
    su_data: dict,
    planned_meta: dict,
):
    """
    Detect modules where, if we split the ACTUAL SPAN (actual_start..actual_end) by planned ratios,
    some phases receive 0 worked days from SU_Others evidence.
    If so, cut the entire module to dummy (break tool code everywhere).
    """
    if not su_data:
        return []
    worker_date_map = su_data.get("worker_date_map", {})
    if not worker_date_map:
        return []

    # code -> list of (dt, wid, text)
    code_to_occ = defaultdict(list)
    for (wid, dt), text in worker_date_map.items():
        code = extract_tool_code(text)
        if not code:
            continue
        if code not in planned_meta:
            continue
        code_to_occ[code].append((dt, wid, text))

    corrections = []

    for code, occ in code_to_occ.items():
        worked_days = sorted(set(dt for dt, _, _ in occ))
        if not worked_days:
            continue

        actual_start = worked_days[0]
        actual_end = worked_days[-1]
        actual_total_span = int((actual_end - actual_start).days) + 1
        if actual_total_span <= 0:
            continue

        # Build continuous span days
        span_days = [actual_start + pd.Timedelta(days=i) for i in range(actual_total_span)]

        # Allocate span lengths by planned ratio
        alloc = _allocate_phase_lengths(actual_total_span, planned_meta[code]["phase_len"])

        # Split span days into phase buckets
        phase_span_days = {}
        idx = 0
        for ph in (1, 2, 3, 4):
            ln = int(alloc.get(ph, 0))
            phase_span_days[ph] = span_days[idx: idx + ln] if ln > 0 else []
            idx += ln

        worked_set = set(worked_days)

        # Count worked days inside each phase bucket
        phase_worked_counts = {
            ph: sum(1 for d in phase_span_days[ph] if d in worked_set)
            for ph in (1, 2, 3, 4)
        }

        # If any phase has 0 workload -> cut whole module
        if any(phase_worked_counts[ph] == 0 for ph in (1, 2, 3, 4)):
            broken = _break_tool_code(code)

            for dt, wid, text in occ:
                if not isinstance(text, str):
                    continue
                old = text
                new = old.replace(code, broken, 1)
                if new == old:
                    continue
                _remember_original_text(su_data, wid, dt, old, new)
                worker_date_map[(wid, dt)] = new
                corrections.append({
                    "wid": wid,
                    "date": _to_ymd(dt),
                    "code": code,
                    "text": old,
                    "reason": f"module cut: phase zero workload under span-split check (counts={phase_worked_counts})",
                })

    su_data["su_phase_zero_corrections"] = corrections
    return corrections
# ============================================================
# Code extraction from SU_Others cell text
# ============================================================

TOOLCODE_RE = re.compile(r"\d{3}[A-Z0-9]\d{5}A")

def extract_tool_code(s: str):
    if not isinstance(s, str):
        return None
    m = TOOLCODE_RE.search(s)
    return m.group(0) if m else None

# ============================================================
# SU_Others parsing
# ============================================================

IGNORE_WHITE_TEXT = {"FI", "FO"}
GREY_RGB_LAST6 = {"A6A6A6", "BFBFBF", "D9D9D9", "808080"}
RED_RGB_LAST6 = {"FF0000"}
# Common Excel indexed greys (these vary by file/theme; expand if needed)
GREY_INDEXED = {15, 22, 23, 24, 25, 26, 27, 28, 29}
RED_INDEXED  = {10}  # sometimes "red"

def _color_to_rgb6(color):
    """
    Convert openpyxl color object to RGB last6 if possible.
    Supports rgb and indexed. Theme colors are not directly resolvable
    without theme parsing, so we return None for theme.
    """
    if color is None:
        return None
    ctype = getattr(color, "type", None)
    if ctype == "rgb":
        rgb = getattr(color, "rgb", None)
        return str(rgb).upper()[-6:] if rgb else None
    if ctype == "indexed":
        idx = getattr(color, "indexed", None)
        if idx is None:
            return None
        try:
            rgb = COLOR_INDEX[idx]
            return str(rgb).upper()[-6:] if rgb else None
        except Exception:
            return None
    # theme / auto / etc.
    return None

def _cell_fill_rgb6(cell):
    """
    Try multiple places: fgColor/start_color/bgColor/end_color.
    """
    fill = getattr(cell, "fill", None)
    if fill is None:
        return None

    # Prefer fgColor if pattern fill, but some files store in start_color
    candidates = []
    fg = getattr(fill, "fgColor", None)
    bg = getattr(fill, "bgColor", None)
    st = getattr(fill, "start_color", None)
    en = getattr(fill, "end_color", None)
    candidates.extend([fg, st, bg, en])

    for col in candidates:
        rgb6 = _color_to_rgb6(col)
        if rgb6:
            return rgb6
    return None

def _cell_fill_indexed(cell):
    fill = getattr(cell, "fill", None)
    if fill is None:
        return None
    for col in [getattr(fill, "fgColor", None), getattr(fill, "start_color", None),
                getattr(fill, "bgColor", None), getattr(fill, "end_color", None)]:
        if col is None:
            continue
        if getattr(col, "type", None) == "indexed":
            return getattr(col, "indexed", None)
    return None

def _is_red_cell(cell) -> bool:
    rgb6 = _cell_fill_rgb6(cell)
    if rgb6 and rgb6 in RED_RGB_LAST6:
        return True
    idx = _cell_fill_indexed(cell)
    if idx is not None and idx in RED_INDEXED:
        return True
    return False

def _is_grey_cell(cell) -> bool:
    rgb6 = _cell_fill_rgb6(cell)
    if rgb6 and rgb6 in GREY_RGB_LAST6:
        return True

    idx = _cell_fill_indexed(cell)
    if idx is not None and idx in GREY_INDEXED:
        return True

    # NEW: theme-based greys
    fill = getattr(cell, "fill", None)
    if fill is not None:
        for col in [getattr(fill, "fgColor", None), getattr(fill, "start_color", None),
                    getattr(fill, "bgColor", None), getattr(fill, "end_color", None)]:
            if _is_theme_grey(col):
                return True

    return False


def _cell_rgb_last6(cell):
    fill = getattr(cell, "fill", None)
    if fill is None:
        return None
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return None
    if getattr(fg, "type", None) != "rgb":
        return None
    rgb = getattr(fg, "rgb", None)
    if not rgb:
        return None
    return str(rgb).upper()[-6:]

def _find_date_header_ws(ws, max_scan_rows=12):
    for r in range(1, max_scan_rows + 1):
        row_vals = [cell.value for cell in ws[r]]
        if any(isinstance(v, (pd.Timestamp, datetime)) for v in row_vals):
            date_cols = [c for c, v in enumerate(row_vals, start=1)
                         if isinstance(v, (pd.Timestamp, datetime))]
            if not date_cols:
                continue
            dt_by_col = {c: pd.Timestamp(row_vals[c - 1]).normalize() for c in date_cols}
            return r, date_cols, dt_by_col
    raise RuntimeError("Could not find date header row in SU_Others.")

def _is_theme_grey(color) -> bool:
    """
    Heuristic: many 'grey' fills in corporate Excels are theme-based.
    We can't resolve to RGB without parsing the theme, so we detect by (theme + tint).
    """
    if color is None:
        return False
    if getattr(color, "type", None) != "theme":
        return False

    tint = getattr(color, "tint", None)
    theme = getattr(color, "theme", None)

    # If tint exists, it's often used to make greys (light/dark).
    # These ranges catch most grey shades in practice.
    if tint is not None:
        try:
            t = float(tint)
            # common greys: around -0.25, -0.15, 0.25, 0.35, 0.5 etc
            if -0.6 <= t <= 0.6:
                return True
        except Exception:
            pass

    # fallback: if theme is set at all and pattern fill is used,
    # treat it as grey-ish rather than missing PB.
    if theme is not None:
        return True

    return False

def parse_su_others(path: str, sheet_names=("予定表_2025",), date_filter=None):
    wb = load_workbook(path, data_only=True, read_only=False)
    used_sheets = [s for s in sheet_names if s in wb.sheetnames]
    if not used_sheets:
        raise ValueError(f"None of {sheet_names} found in {path}. Available: {wb.sheetnames}")

    f_start, f_end = date_filter if date_filter else (None, None)

    worker_company_map = {}
    worker_company_list = []

    def get_worker_company_id(company_name: str) -> str:
        company_name = str(company_name).strip()
        if company_name not in worker_company_map:
            cid = f"wc{len(worker_company_map) + 1}"
            worker_company_map[company_name] = cid
            worker_company_list.append({
                "id": cid,
                "name": company_name,
                "annual_overtime_limit": 10000,
                "monthly_overtime_limit": 10000,
                "unavailable_dates": [],
            })
        return worker_company_map[company_name]

    worker_key_to_id = {}
    worker_acc = {}
    worker_date_map = {}
    worker_personal_map = {}

    plan_start = None
    plan_end = None

    for sheet_name in used_sheets:
        ws = wb[sheet_name]
        date_row_idx, date_cols, dt_by_col = _find_date_header_ws(ws)

        if f_start is not None and f_end is not None:
            date_cols = [c for c in date_cols if f_start <= dt_by_col[c] <= f_end]
        if not date_cols:
            continue

        worker_start_row = date_row_idx + 2
        s = min(dt_by_col[c] for c in date_cols)
        e = max(dt_by_col[c] for c in date_cols)
        plan_start = s if plan_start is None else min(plan_start, s)
        plan_end = e if plan_end is None else max(plan_end, e)

        blank_streak = 0
        max_col = max(date_cols)

        for r, row_cells in enumerate(ws.iter_rows(min_row=worker_start_row, min_col=1, max_col=max_col), start=worker_start_row):
            company = row_cells[0].value
            name = row_cells[1].value

            if name is None or str(name).strip() == "":
                blank_streak += 1
                if blank_streak >= 30:
                    break
                continue
            blank_streak = 0

            company_str = "" if company is None else str(company).strip()
            name_str = str(name).strip()

            free_slot = row_cells[4].value if len(row_cells) >= 5 else None
            is_manager = bool(isinstance(free_slot, str) and "責" in free_slot)

            key = _norm_name(name_str)
            if key not in worker_key_to_id:
                wid = f"w{len(worker_key_to_id) + 1:03d}"
                worker_key_to_id[key] = wid
                company_id = get_worker_company_id(company_str)
                worker_acc[key] = {
                    "id": wid,
                    "name": name_str,
                    "worker_company": company_id,
                    "is_manager": is_manager,
                    "unavailable_set": set(),
                }
            else:
                if is_manager:
                    worker_acc[key]["is_manager"] = True

            wid = worker_key_to_id[key]

            for c in date_cols:
                dt = dt_by_col[c]
                cell = row_cells[c - 1]
                rgb6 = _cell_rgb_last6(cell)

                # Red = unavailable
                if _is_red_cell(cell):
                    worker_acc[key]["unavailable_set"].add(_to_ymd(dt))
                    continue

                val = cell.value
                text = val.strip() if isinstance(val, str) else ""

                # Ignore explicit white markers
                if text.upper() in IGNORE_WHITE_TEXT:
                    continue
                fill = cell.fill
                fg = getattr(fill, "fgColor", None)
                # Grey rule (NEW):
                #   - Grey cell => Personal Business
                #   - EXCEPT: if text contains a tool-code pattern, treat as normal (NOT PB)
                if _is_grey_cell(cell):
                    if text and extract_tool_code(text):
                        # Grey but contains module/tool-code => NOT PB
                        worker_date_map[(wid, dt)] = text
                    else:
                        # Grey => PB (label can be "" or Japanese text)
                        worker_personal_map[(wid, dt)] = text  # store label
                    continue

                # Non-grey normal cells:
                if text == "":
                    continue

                worker_date_map[(wid, dt)] = text

    worker_list = []
    for acc in worker_acc.values():
        worker_list.append({
            "id": acc["id"],
            "name": acc["name"],
            "worker_company": acc["worker_company"],
            "is_manager": acc["is_manager"],
            "skill_map": {},
            "fab_suitability_map": [],
            "unavailable_dates": [{"date": d} for d in sorted(acc["unavailable_set"])],
        })

    plan_range = {
        "start_date": _to_ymd(plan_start) if plan_start is not None else "2025/01/01",
        "end_date": _to_ymd(plan_end) if plan_end is not None else "2025/01/01",
    }

    return {
        "worker_company_list": worker_company_list,
        "worker_company_map": worker_company_map,
        "worker_list": worker_list,
        "plan_range": plan_range,
        "worker_date_map": worker_date_map,
        "worker_personal_map": worker_personal_map,
    }


# ============================================================
# SU_Others outlier-cell cut (prevent horizontal drag-fill drift)
# ============================================================

def _break_tool_code(code: str) -> str:
    """Change a valid tool code string so extract_tool_code() won't match it."""
    if not isinstance(code, str) or len(code) < 2:
        return "OUTLIER"
    # Our TOOLCODE_RE ends with 'A'. Flip last char to 'X' to break the pattern.
    return code[:-1] + "X"


def _cluster_by_date_gap(sorted_dates, gap_days: int):
    """Cluster sorted pd.Timestamp dates by allowing up to gap_days between adjacent dates."""
    clusters = []
    cur = []
    for d in sorted_dates:
        if not cur:
            cur = [d]
            continue
        if (d - cur[-1]).days <= gap_days:
            cur.append(d)
        else:
            clusters.append(cur)
            cur = [d]
    if cur:
        clusters.append(cur)
    return clusters


def cut_su_outlier_cells(
    su_data: dict,
    cluster_gap_days: int = 7,
    keep_top_clusters: int = 1,
    far_gap_days: int = 60,
    small_cluster_max_unique_days: int = 7,
    cut_module_if_total_cells_lt: int = 4,
    cut_module_if_unique_days_lt: int = 4,
    planned_meta: dict | None = None,
    cut_if_far_from_planned_days: int = 90,
):
    """
    Some modules appear in SU_Others as one long continuous period, but also have
    a few isolated cells months later (often caused by Excel horizontal drag-fill
    mistakes). Those isolated cells can make shifting explode (first..last spans months).

    Strategy:
      - (Optional) If planned window exists: cut SU cells too far from planned window
      - If module has too few UNIQUE work days OR too few total cells: cut everything to dummy
      - Otherwise: cluster occurrence dates by gap<=cluster_gap_days
          * Keep top cluster(s) (largest by worker-days, then uniq days, then earliest)
          * Consider keeping additional clusters unless they look like drag-fill noise
      - NEW: After deciding kept_dates, if remaining evidence is too small (days/cells),
             cut the whole module to dummy (prevents "100 cells -> after cut only 1 day -> still shifted")

    Returns a list of correction records for TransformationLog.
    """
    if not su_data:
        return []

    worker_date_map = su_data.get("worker_date_map", {})
    if not worker_date_map:
        return []

    # Gather occurrences
    code_to_occ = defaultdict(list)  # code -> list of (dt, wid, text)
    for (wid, dt), text in list(worker_date_map.items()):
        code = extract_tool_code(text)
        if not code:
            continue
        code_to_occ[code].append((dt, wid, text))

    def _longest_consecutive_run(dts_sorted):
        """Longest run where each next day is previous + 1 day."""
        if not dts_sorted:
            return 0
        best = 1
        cur = 1
        for i in range(1, len(dts_sorted)):
            if dts_sorted[i] == dts_sorted[i - 1] + pd.Timedelta(days=1):
                cur += 1
                if cur > best:
                    best = cur
            else:
                cur = 1
        return best

    corrections = []

    def _cut_all_occurrences(code: str, occ_list: list, reason: str):
        """Break tool-code for ALL occurrences => they become dummy 'other'."""
        broken = _break_tool_code(code)
        for dt, wid, text in occ_list:
            if not isinstance(text, str):
                continue
            old = text
            new = old.replace(code, broken, 1)
            if new == old:
                continue

            _remember_original_text(su_data, wid, dt, old, new)
            worker_date_map[(wid, dt)] = new
            corrections.append({
                "wid": wid,
                "date": _to_ymd(dt),
                "code": code,
                "text": old,
                "reason": reason,
            })

    for code, occ in code_to_occ.items():
        # ------------------------------------------------------------
        # (Example: planned ends Mar, but SU has tail in Oct -> cut tail)
        # ------------------------------------------------------------
        if planned_meta is not None and code in planned_meta:
            pstart = planned_meta[code].get("overall_start")
            pend = planned_meta[code].get("overall_end")

            if pstart is not None and pend is not None:
                broken = _break_tool_code(code)

                for dt, wid, text in occ:
                    dist = _planned_actual_gap_days(pstart, pend, dt, dt)
                    if dist <= int(cut_if_far_from_planned_days):
                        continue
                    if not isinstance(text, str):
                        continue

                    old = text
                    new = old.replace(code, broken, 1)
                    if new == old:
                        continue

                    _remember_original_text(su_data, wid, dt, old, new)
                    worker_date_map[(wid, dt)] = new
                    corrections.append({
                        "wid": wid,
                        "date": _to_ymd(dt),
                        "code": code,
                        "text": old,
                        "reason": f"planned-window far cut: {dist}d away (> {cut_if_far_from_planned_days})",
                    })

        # ------------------------------------------------------------
        # Pre-check A: too few UNIQUE work days (before clustering)
        # ------------------------------------------------------------
        uniq_days_pre = sorted(set([x[0] for x in occ]))
        if len(uniq_days_pre) < int(cut_module_if_unique_days_lt):
            _cut_all_occurrences(
                code,
                occ,
                f"module cut: unique SU_Others work days < {cut_module_if_unique_days_lt} "
                f"(unique_days={len(uniq_days_pre)})",
            )
            continue

        # ------------------------------------------------------------
        # Pre-check B: too few total worker-day cells (before clustering)
        # ------------------------------------------------------------
        if len(occ) < int(cut_module_if_total_cells_lt):
            _cut_all_occurrences(
                code,
                occ,
                f"module cut: total SU_Others cells < {cut_module_if_total_cells_lt} "
                f"(cells={len(occ)})",
            )
            continue

        if len(occ) <= 1:
            continue

        # ------------------------------------------------------------
        # PER-WORKER OUTLIER CUT (your requested behavior)
        # For each (code, wid), cluster that worker's dates ONLY.
        # If the worker's best cluster is not "continuous enough" (run < 2),
        # cut ALL that worker's occurrences for this code into dummy.
        # Otherwise keep only the best cluster for that worker and cut the rest.
        # ------------------------------------------------------------

        occ_by_wid = defaultdict(list)  # wid -> list of (dt, text)
        for dt, wid, text in occ:
            occ_by_wid[wid].append((dt, text))

        broken = _break_tool_code(code)

        for wid, wid_occ in occ_by_wid.items():
            # unique sorted dates for THIS worker only
            wid_dates = sorted(set(dt for dt, _ in wid_occ))
            if len(wid_dates) <= 1:
                # single-day involvement => treat as outlier (cut it)
                for dt, text in wid_occ:
                    if not isinstance(text, str):
                        continue
                    old = text
                    new = old.replace(code, broken, 1)
                    if new == old:
                        continue
                    _remember_original_text(su_data, wid, dt, old, new)
                    worker_date_map[(wid, dt)] = new
                    corrections.append({
                        "wid": wid,
                        "date": _to_ymd(dt),
                        "code": code,
                        "text": old,
                        "reason": "per-worker cut: only 1 day for this worker on this module (treated as outlier)",
                    })
                continue

            # cluster THIS worker's dates
            wid_clusters = _cluster_by_date_gap(wid_dates, gap_days=cluster_gap_days)

            # pick "best cluster" for this worker:
            # prefer largest uniq-days, then earliest start
            wid_clusters.sort(key=lambda cl: (len(cl), -int(cl[0].timestamp())), reverse=True)
            best = wid_clusters[0]

            # longest consecutive run INSIDE best cluster
            best_run = _longest_consecutive_run(best)

            # If this worker does not have a consecutive run of >=2 days,
            # cut ALL occurrences for this worker for this code.
            if best_run < 2:
                for dt, text in wid_occ:
                    if not isinstance(text, str):
                        continue
                    old = text
                    new = old.replace(code, broken, 1)
                    if new == old:
                        continue
                    _remember_original_text(su_data, wid, dt, old, new)
                    worker_date_map[(wid, dt)] = new
                    corrections.append({
                        "wid": wid,
                        "date": _to_ymd(dt),
                        "code": code,
                        "text": old,
                        "reason": f"per-worker cut: no consecutive run (best_run={best_run}) for this worker on this module",
                    })
                continue

            # Keep ALL "good" clusters; cut only "bad" clusters.
            # A "good" cluster = has at least 2 consecutive days inside it.

            def _cluster_has_run2(cluster_dates):
                return _longest_consecutive_run(cluster_dates) >= 2

            # Decide which dates to keep for this worker+code
            keep_dates = set()
            for cl in wid_clusters:
                if _cluster_has_run2(cl):
                    keep_dates.update(cl)

            # If nothing qualifies, fall back to old behavior: cut all
            if not keep_dates:
                for dt, text in wid_occ:
                    if not isinstance(text, str):
                        continue
                    old = text
                    new = old.replace(code, broken, 1)
                    if new == old:
                        continue
                    _remember_original_text(su_data, wid, dt, old, new)
                    worker_date_map[(wid, dt)] = new
                    corrections.append({
                        "wid": wid,
                        "date": _to_ymd(dt),
                        "code": code,
                        "text": old,
                        "reason": "per-worker cut: no cluster has >=2 consecutive days (treated as outliers)",
                    })
                continue

            # Cut only dates NOT in keep_dates
            for dt, text in wid_occ:
                if dt in keep_dates:
                    continue
                if not isinstance(text, str):
                    continue
                old = text
                new = old.replace(code, broken, 1)
                if new == old:
                    continue
                _remember_original_text(su_data, wid, dt, old, new)
                worker_date_map[(wid, dt)] = new
                corrections.append({
                    "wid": wid,
                    "date": _to_ymd(dt),
                    "code": code,
                    "text": old,
                    "reason": "per-worker cut: cluster is too small / not consecutive (outlier)",
                })

        # Done for this code
        continue

    # attach for optional downstream use
    su_data["su_outlier_corrections"] = corrections
    return corrections


# ============================================================
# Task list from 新規製番リスト (sheet "CSV") + planned ratios
# ============================================================

def parse_tasks_from_csv_v3(path: str, sheet_name: str = "CSV", date_filter=None):
    df = load_sheet_as_df_with_header(path, sheet_name, header_row=0)
    colmap = {c: str(c).replace("\n", "").strip() for c in df.columns}
    df = df.rename(columns=colmap)

    code_col = [c for c in df.columns if "新規製番" in c]
    if not code_col:
        raise ValueError("Could not find 新規製番 column in CSV sheet.")
    code_col = code_col[0]

    cust_col = "ユーザー名"
    country_col = "国"
    fab_col = "ファブ名"

    p1s_col = "工程１開始可能日"
    p2s_col = "工程２開始可能日"
    p3s_col = "工程３開始可能日"
    p4s_col = "工程４開始可能日"
    p4e_col = "工程４終了予定日"

    f_start, f_end = date_filter if date_filter else (None, None)

    # outputs
    valid_rows = []
    cut_rows = []  # (code, reason)
    planned_meta = {}  # code -> dict with planned phase windows/lengths/percents + customer/country/fab_name
    all_dates = []

    for _, row in df.iterrows():
        code_raw = row.get(code_col)
        if not isinstance(code_raw, str) or not code_raw.strip():
            continue
        code = code_raw.strip()

        customer = row.get(cust_col)
        customer = str(customer).strip() if isinstance(customer, str) and str(customer).strip() else "OTHER"
        country = row.get(country_col)
        country = str(country).strip() if isinstance(country, str) and str(country).strip() else "Other"
        fab_name = row.get(fab_col)
        fab_name = str(fab_name).strip() if isinstance(fab_name, str) and str(fab_name).strip() else "Other"

        p1s = _as_timestamp(row.get(p1s_col))
        p2s = _as_timestamp(row.get(p2s_col))
        p3s = _as_timestamp(row.get(p3s_col))
        p4s = _as_timestamp(row.get(p4s_col))
        p4e = _as_timestamp(row.get(p4e_col))

        if any(x is None for x in [p1s, p2s, p3s, p4s, p4e]):
            cut_rows.append((code, "blank or N/A in phase date columns"))
            continue

        # order check
        if not (p1s <= p2s <= p3s <= p4s <= p4e):
            cut_rows.append((code, "wrong date order (p1<=p2<=p3<=p4<=p4end violated)"))
            continue

        starts = {1: p1s, 2: p2s, 3: p3s, 4: p4s}
        ends = {
            1: (p2s - pd.Timedelta(days=1)).normalize(),
            2: (p3s - pd.Timedelta(days=1)).normalize(),
            3: (p4s - pd.Timedelta(days=1)).normalize(),
            4: p4e,
        }

        # validate non-negative windows (should hold if order ok, but keep safe)
        bad = False
        for ph in (1,2,3,4):
            if ends[ph] < starts[ph]:
                cut_rows.append((code, f"phase {ph} end < start after applying end=next_start-1 rule"))
                bad = True
                break
        if bad:
            continue

        overall_start = min(starts.values())
        overall_end = max(ends.values())

        # filter by overlap
        if f_start is not None and f_end is not None:
            if not _overlaps(overall_start, overall_end, f_start, f_end):
                continue

        phase_len = {ph: int((ends[ph]-starts[ph]).days)+1 for ph in (1,2,3,4)}
        total_len = sum(phase_len.values())
        if total_len <= 0:
            cut_rows.append((code, "total planned duration <= 0"))
            continue

        phase_pct = {ph: (phase_len[ph] / total_len) for ph in (1,2,3,4)}

        planned_meta[code] = {
            "customer": customer,
            "country": country,
            "fab_name": fab_name,
            "starts": starts,
            "ends": ends,
            "phase_len": phase_len,
            "phase_pct": phase_pct,
            "overall_start": overall_start,
            "overall_end": overall_end,
            "total_len": total_len,
        }

        all_dates.extend([overall_start, overall_end])
        valid_rows.append(code)

    return {
        "valid_codes": valid_rows,
        "planned_meta": planned_meta,
        "cut_rows": cut_rows,
        "date_list": all_dates,
    }

# ============================================================
# Shifting / Rescaling plan onto SU_Others actual span
# ============================================================

def _allocate_phase_lengths(actual_total_days: int, planned_phase_len: dict):
    """
    Allocate integer lengths for phases 1..4 so that:
      - sum(lengths) == actual_total_days
      - ratios follow planned_phase_len proportions
      - if actual_total_days >= 4: each phase gets at least 1 day
    """
    phs = [1, 2, 3, 4]

    # Defensive
    if actual_total_days <= 0:
        return {1: 1, 2: 1, 3: 1, 4: 1}

    total_planned = sum(max(0, int(planned_phase_len.get(ph, 0))) for ph in phs)
    if total_planned <= 0:
        # equal-ish split fallback
        base = actual_total_days // 4
        rem = actual_total_days - base * 4
        out = {1: base, 2: base, 3: base, 4: base}
        for i in range(rem):
            out[phs[i]] += 1
        # enforce min-1 if possible
        if actual_total_days >= 4:
            for ph in phs:
                if out[ph] == 0:
                    out[ph] = 1
        # re-balance sum (just in case)
        s = sum(out.values())
        out[4] += (actual_total_days - s)
        return out

    # ---- base proportional allocation (floor + remainder) ----
    raw = {ph: (actual_total_days * (planned_phase_len.get(ph, 0) / total_planned)) for ph in phs}
    flo = {ph: int(math.floor(raw[ph])) for ph in phs}
    s = sum(flo.values())
    rem = actual_total_days - s

    frac = sorted(phs, key=lambda ph: (raw[ph] - flo[ph]), reverse=True)
    out = dict(flo)
    for i in range(rem):
        out[frac[i % len(frac)]] += 1

    # Adjust last phase if drift
    s2 = sum(out.values())
    if s2 != actual_total_days:
        out[4] += (actual_total_days - s2)

    # ---- NEW: enforce min 1 day per phase if possible ----
    if actual_total_days >= 4:
        zeros = [ph for ph in phs if out.get(ph, 0) <= 0]
        if zeros:
            # donors: phases that can spare days (must stay >= 1)
            def pick_donor():
                donors = [p for p in phs if out.get(p, 0) > 1]
                if not donors:
                    return None
                # steal from the largest allocation first
                donors.sort(key=lambda p: out[p], reverse=True)
                return donors[0]

            for z in zeros:
                donor = pick_donor()
                if donor is None:
                    break  # should not happen when actual_total_days>=4, but stay safe
                out[donor] -= 1
                out[z] = 1

            # re-check sum (keep invariant)
            s3 = sum(out.values())
            if s3 != actual_total_days:
                out[4] += (actual_total_days - s3)

    return out

def build_shifted_meta(planned_meta: dict, su_data: dict | None):
    orig_map = su_data.get("su_outlier_original_text", {}) if su_data else {}

    shifted_meta = {}
    code_to_shifted_phases = defaultdict(list)

    # build quick index for SU_Others occurrences
    code_occ = defaultdict(list)  # code -> list of (dt, wid, text)
    if su_data is not None:
        for (wid, dt), text in su_data["worker_date_map"].items():
            code = extract_tool_code(text)
            if code:
                disp = text
                k = (wid, _to_ymd(dt))
                if k in orig_map:
                    disp = orig_map[k]["old"]
                code_occ[code].append((dt, wid, disp))

    for code, meta in planned_meta.items():
        starts = meta["starts"]
        ends = meta["ends"]

        occ = code_occ.get(code, [])
        if occ:
            worked_days = sorted(set(x[0] for x in occ))
            worked_set = set(worked_days)

            actual_start = worked_days[0]
            actual_end = worked_days[-1]

            # timeline for shifting split
            if SHIFT_USE_WORKED_DAYS:
                timeline_days = worked_days  # compressed (worked days only)
            else:
                total_span = int((actual_end - actual_start).days) + 1
                timeline_days = [actual_start + pd.Timedelta(days=i) for i in range(total_span)]

            timeline_total = len(timeline_days)

            # Allocate by planned ratio on the shifting timeline
            alloc_span = _allocate_phase_lengths(timeline_total, meta["phase_len"])

            # Build per-phase segments on timeline_days
            # - seg = raw segment days (worked-days or span-days depending on policy)
            # - phase_days = only worked days inside that segment (always used for mapping)
            phase_days = {}
            seg_edges = {}  # ph -> (seg_start, seg_end) from timeline_days slicing

            idx = 0
            for ph in (1, 2, 3, 4):
                ln = int(alloc_span.get(ph, 0))
                seg = timeline_days[idx: idx + ln] if ln > 0 else []
                if seg:
                    seg_edges[ph] = (seg[0], seg[-1])
                else:
                    # empty bucket fallback
                    seg_edges[ph] = (actual_start, actual_start)

                if SHIFT_USE_WORKED_DAYS:
                    # seg already equals worked days
                    phase_days[ph] = seg
                else:
                    # seg is continuous span; keep only actual worked days for mapping
                    phase_days[ph] = [d for d in seg if d in worked_set]

                idx += ln

            # shifted window edges MUST come from segment edges (span edges),
            # otherwise SHIFT_USE_WORKED_DAYS=False collapses windows incorrectly.
            shifted_starts = {ph: seg_edges[ph][0] for ph in (1, 2, 3, 4)}
            shifted_ends   = {ph: seg_edges[ph][1] for ph in (1, 2, 3, 4)}

            alloc_worked = {ph: len(phase_days.get(ph, [])) for ph in (1, 2, 3, 4)}

            shifted_meta[code] = {
                "plan": meta,
                "had_su_match": True,
                "actual_first": actual_start,
                "actual_last": actual_end,

                # actual worked days count (what you *meant* by "actual_total")
                "actual_total": len(worked_days),

                # split lengths on timeline (span-days when SHIFT_USE_WORKED_DAYS=False)
                "alloc_span_days": alloc_span,

                # worked-day counts per phase (always meaningful)
                "alloc_worked_days": alloc_worked,

                "phase_days": phase_days,  # IMPORTANT for dt-in-date_set mapping
                "shifted_starts": shifted_starts,
                "shifted_ends": shifted_ends,

                "occ_sample": sorted(occ, key=lambda x: x[0])[:3],
                "occ_last_sample": sorted(occ, key=lambda x: x[0])[-3:],
            }
        else:
            shifted_meta[code] = {
                "plan": meta,
                "had_su_match": False,
                "actual_first": None,
                "actual_last": None,
                "actual_total": None,
                "alloc_span_days": None,
                "alloc_worked_days": None,
                "phase_days": None,
                "shifted_starts": starts,
                "shifted_ends": ends,
                "occ_sample": [],
                "occ_last_sample": [],
            }

        # Build mapping metadata used later in assignment building
        for ph in (1, 2, 3, 4):
            ds = None
            if shifted_meta[code].get("had_su_match") and shifted_meta[code].get("phase_days"):
                ds = set(shifted_meta[code]["phase_days"][ph])

            code_to_shifted_phases[code].append({
                "phase_index": ph,
                "start": shifted_meta[code]["shifted_starts"][ph],
                "end": shifted_meta[code]["shifted_ends"][ph],
                "operation": f"p{ph}",
                "date_set": ds,
            })

    return shifted_meta, code_to_shifted_phases, code_occ


# ============================================================
# Skill aggregation (スキル集計_*.xlsx) (same as decoder2)
# ============================================================

def _skill_level(total_level, min_level):
    try:
        total = 0 if total_level is None or (isinstance(total_level, float) and pd.isna(total_level)) else int(total_level)
    except Exception:
        total = 0
    try:
        mn = 0 if min_level is None or (isinstance(min_level, float) and pd.isna(min_level)) else int(min_level)
    except Exception:
        mn = 0

    bucket = 0 if total <= 20 else (total // 20)
    lvl = max(bucket, mn)
    return max(0, min(5, lvl))

def parse_skill_excel(path: str, sheet_name: str = "Sheet1"):
    df = load_sheet_as_df(path, sheet_name)

    header_row = None
    for r in range(0, min(20, len(df))):
        row = df.iloc[r].astype(str).tolist()
        if any("氏名" in c for c in row) and any("所属" in c for c in row):
            header_row = r
            break
    if header_row is None:
        header_row = 6

    headers = df.iloc[header_row].tolist()
    col_index = {str(h).strip(): i for i, h in enumerate(headers) if str(h).strip() != "nan"}

    def find_col_contains(keyword):
        for k, i in col_index.items():
            if keyword in k:
                return i
        return None

    name_c = find_col_contains("氏名")
    comp_c = find_col_contains("所属")

    def find_first_cell(text):
        for r in range(len(df)):
            for c in range(df.shape[1]):
                v = df.iat[r, c]
                if isinstance(v, str) and v.strip() == text:
                    return r, c
        return None, None

    g1_r, g1_c = find_first_cell("1:Module Setup")
    g2_r, g2_c = find_first_cell("2:Hardware Setup")
    g3_r, g3_c = find_first_cell("3:Function Setup")
    if g1_c is None or g2_c is None or g3_c is None:
        raise RuntimeError("Could not find group labels (1/2/3) in skill sheet.")

    p1_total, p1_min = g1_c, g1_c + 1
    p2_total, p2_min = g2_c, g2_c + 1
    p3_total, p3_min = g3_c, g3_c + 1

    start_row = max(g1_r, g2_r, g3_r) + 2

    skill_levels = {}
    people_meta = {}

    for r in range(start_row, len(df)):
        name = df.iat[r, name_c] if name_c is not None else None
        if not (isinstance(name, str) and name.strip()):
            continue
        comp = df.iat[r, comp_c] if comp_c is not None else ""
        name_s = name.strip()
        comp_s = str(comp).strip() if isinstance(comp, str) and str(comp).strip() else ""

        p1_lvl = _skill_level(df.iat[r, p1_total], df.iat[r, p1_min])
        p2_lvl = _skill_level(df.iat[r, p2_total], df.iat[r, p2_min])
        p3_lvl = _skill_level(df.iat[r, p3_total], df.iat[r, p3_min])

        key = _norm_name(name_s)
        skill_levels[key] = {"p1": p1_lvl, "p2": p2_lvl, "p3": p3_lvl, "p4": p3_lvl}
        if key not in people_meta:
            people_meta[key] = {"company": comp_s, "name": name_s}

    return skill_levels, people_meta

# ============================================================
# Build assignments using SHIFTED phase windows
# ============================================================

def build_assignments_v3(
    su_data: dict,
    shifted_code_to_phases: dict,
    valid_code_set: set,
    date_filter=None
):
    orig_map = su_data.get("su_outlier_original_text", {})  # (wid, ymd) -> {old,new}
    worker_date_map = su_data["worker_date_map"]
    worker_personal_map = su_data["worker_personal_map"]
    f_start, f_end = date_filter if date_filter else (None, None)

    known_assign_map = defaultdict(list)

    misc_label_dates = defaultdict(set)
    misc_worker_label_dates = defaultdict(list)

    personal_label_dates = defaultdict(set)
    personal_worker_label_dates = defaultdict(list)

    inferred_worker_phase = defaultdict(set)

    # for log: dummy tool-code-ish labels (in SU_Others but not in 新規製番リスト valid set)
    dummy_tool_labels = defaultdict(set)  # code -> set(full_text)

    # 1) NORMAL CELLS
    for (wid, dt), raw_text in worker_date_map.items():
        if f_start is not None and f_end is not None and (dt < f_start or dt > f_end):
            continue

        # IMPORTANT:
        # - use internal_text (possibly broken by outlier-cut) for code extraction
        # - use display_text (original) for label and outputs
        internal_text = raw_text
        display_text = raw_text

        k = (wid, _to_ymd(dt))
        if k in orig_map:
            display_text = orig_map[k]["old"]     # original text for label/log/output
            internal_text = orig_map[k]["new"]    # broken text for matching (dummy behavior)

        code = extract_tool_code(internal_text)

        if code and (code in valid_code_set) and (code in shifted_code_to_phases):
            matched = False
            for phase_meta in shifted_code_to_phases[code]:
                ds = phase_meta.get("date_set")
                if ds is not None:
                    if dt in ds:
                        phase_id = phase_meta["phase_id"]
                        known_assign_map[(wid, phase_id)].append(dt)
                        inferred_worker_phase[wid].add(phase_meta["operation"])
                        matched = True
                        break
                else:
                    ps = phase_meta["start"]
                    pe = phase_meta["end"]
                    if ps is None or pe is None:
                        continue
                    if ps <= dt <= pe:
                        phase_id = phase_meta["phase_id"]
                        known_assign_map[(wid, phase_id)].append(dt)
                        inferred_worker_phase[wid].add(phase_meta["operation"])
                        matched = True
                        break
            if matched:
                continue

        # otherwise: dummy "other" (Fixed)
        label = display_text.strip() if isinstance(display_text, str) else ""
        if not label:
            label = "other"
        misc_label_dates[label].add(dt)
        misc_worker_label_dates[(wid, label)].append(dt)

        # for log: tool-code-ish labels that are not in 新規製番リスト valid set
        # NOTE: use code extracted from DISPLAY (original) only for reporting, not matching
        code_disp = extract_tool_code(display_text)
        if code_disp and code_disp not in valid_code_set:
            dummy_tool_labels[code_disp].add(label)

    # 2) GREY CELLS (PB) = Personal Business (Fixed), grouped by label text
    for (wid, dt), text in worker_personal_map.items():
        if f_start is not None and f_end is not None and (dt < f_start or dt > f_end):
            continue
        label = text if isinstance(text, str) else ""
        # keep EXACT label including empty string
        personal_label_dates[label].add(dt)
        personal_worker_label_dates[(wid, label)].append(dt)

    # --- PB worker-day counts (unique wid+date), computed once ---
    personal_label_workerday = defaultdict(int)
    tmp_pb_label_wid_dates = defaultdict(set)  # label -> set((wid, dt))

    for (wid, label), dates in personal_worker_label_dates.items():
        for d in set(dates):
            tmp_pb_label_wid_dates[label].add((wid, d))

    for label, s in tmp_pb_label_wid_dates.items():
        personal_label_workerday[label] = len(s)
    # 3) TOOL ASSIGNMENTS (Flexible) + OptionB worker-day counting
    assignments = []
    phase_workerday_count = defaultdict(int)
    phase_worker_set = defaultdict(set)  # phase_id -> set(wid)

    tmp_phase_to_worker_dates = defaultdict(set)  # (phase_id, wid) -> set(dt)
    for (wid, phase_id), dates in known_assign_map.items():
        for d in dates:
            tmp_phase_to_worker_dates[(phase_id, wid)].add(d)

    for (phase_id, wid), dset in tmp_phase_to_worker_dates.items():
        phase_workerday_count[phase_id] += len(dset)
        phase_worker_set[phase_id].add(wid)

    phase_worker_count = {pid: len(wset) for pid, wset in phase_worker_set.items()}

    for (wid, phase_id), dates in known_assign_map.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        work_date_list = [{"hour": 10, "date": _to_ymd(d)} for d in uniq_dates]
        assignments.append({
            "worker": wid,
            "operation_task": phase_id,
            "start_date": _to_ymd(uniq_dates[0]),
            "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list,
            "plan_flexibility": "Flexible",
        })

    # worker-day counts for misc labels (unique wid+date)
    misc_label_workerday = defaultdict(int)
    tmp_label_wid_dates = defaultdict(set)  # label -> set((wid, dt))

    for (wid, label), dates in misc_worker_label_dates.items():
        for d in set(dates):
            tmp_label_wid_dates[label].add((wid, d))

    for label, s in tmp_label_wid_dates.items():
        misc_label_workerday[label] = len(s)

    # 4) MISC ("OTHER") TASKS + FIXED ASSIGNMENTS
    misc_tasks = []
    misc_label_to_phase = {}
    misc_counter = 1

    for label, dates in misc_label_dates.items():
        if not dates:
            continue
        start = min(dates)
        end = max(dates)
        task_id = f"misc_{misc_counter}"
        misc_counter += 1
        phase_id = f"{task_id}_p1"
        misc_label_to_phase[label] = {"phase_id": phase_id, "start": start, "end": end}

        # Workload should match assignment_list: 1 worker x 1 day = 1
        workload_days = int(misc_label_workerday.get(label, 0))
        if workload_days <= 0:
            # fallback safety (should rarely happen)
            workload_days = int((end - start).days) + 1

        misc_tasks.append({
            "id": task_id,
            "name": label,
            "workflow": "wf_other",
            "fab": "f_other",   
            "phase_task_list": [{
                "id": phase_id,
                "name": "Other work",
                "phase": "other_p1",
                "start_date": _to_ymd(start),
                "end_date": _to_ymd(end),
                "operation_task_list": [{
                    "id": phase_id,
                    "name": "Other work",
                    "operation": "other_op",
                    "workload_days": workload_days,
                }],
            }],
        })

    for (wid, label), dates in misc_worker_label_dates.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        meta = misc_label_to_phase.get(label)
        if not meta:
            continue
        phase_id = meta["phase_id"]
        work_date_list = [{"hour": 10, "date": _to_ymd(d)} for d in uniq_dates]
        assignments.append({
            "worker": wid,
            "operation_task": phase_id,
            "start_date": _to_ymd(uniq_dates[0]),
            "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list,
            "plan_flexibility": "Fixed",
        })

    for (wid, lb), dates in personal_worker_label_dates.items():
        for d in set(dates):
            tmp_pb_label_wid_dates[lb].add((wid, d))

    for lb, s in tmp_pb_label_wid_dates.items():
        personal_label_workerday[lb] = len(s)

    # 5) PERSONAL BUSINESS TASKS + FIXED ASSIGNMENTS
    personal_tasks = []
    pb_label_to_phase = {}
    pb_counter = 1

    for label, dates in personal_label_dates.items():
        if not dates:
            continue
        start = min(dates)
        end = max(dates)
        task_id = f"pb_{pb_counter}"
        pb_counter += 1
        phase_id = f"{task_id}_p1"
        pb_label_to_phase[label] = {"phase_id": phase_id, "start": start, "end": end}
            
        workload_days = int(personal_label_workerday.get(label, 0))
        if workload_days <= 0:
            workload_days = int((end - start).days) + 1

        personal_tasks.append({
            "id": task_id,
            "name": label,
            "workflow": "wf_personal_business",
            "fab": "f_other",
            "phase_task_list": [{
                "id": phase_id,
                "name": "Personal Business",
                "phase": "pb_p1",
                "start_date": _to_ymd(start),
                "end_date": _to_ymd(end),
                "operation_task_list": [{
                    "id": phase_id,
                    "name": "Personal Business",
                    "operation": "personal_business_op",
                    "workload_days": workload_days,
                }],
            }],
        })

    pb_worker_dates = defaultdict(list)
    for (wid, dt), _ in worker_personal_map.items():
        if f_start is not None and f_end is not None and (dt < f_start or dt > f_end):
            continue
        pb_worker_dates[wid].append(dt)

    # normalize
    for wid in list(pb_worker_dates.keys()):
        pb_worker_dates[wid] = sorted(set(pb_worker_dates[wid]))

    for (wid, label), dates in personal_worker_label_dates.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        meta = pb_label_to_phase.get(label)
        if not meta:
            continue
        phase_id = meta["phase_id"]
        work_date_list = [{"hour": 10, "date": _to_ymd(d)} for d in uniq_dates]
        assignments.append({
            "worker": wid,
            "operation_task": phase_id,
            "start_date": _to_ymd(uniq_dates[0]),
            "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list,
            "plan_flexibility": "Fixed",
        })

    return assignments, misc_tasks, personal_tasks, inferred_worker_phase, phase_workerday_count, phase_worker_count, pb_worker_dates, dummy_tool_labels
# ============================================================
# Transformation log building
# ============================================================

def _format_phase_line(ph, start, end, extra=""):
    s = _to_ymd(start) if start is not None else "N/A"
    e = _to_ymd(end) if end is not None else "N/A"
    if extra:
        return f"  - P{ph}: {s} - {e} {extra}"
    return f"  - P{ph}: {s} - {e}"

def write_transformation_log(
    out_path: str,
    cut_rows: list,
    shifted_meta: dict,
    worker_id_to_name: dict,
    workload_zero_phase_info: list,
    dummy_tool_labels: dict,
    su_outlier_corrections: list | None = None,
    outlier_cut_summary: dict | None = None,
    pb_worker_dates: dict | None = None,
):
    lines = []
    lines.append("Decoder3 Transformation Log")
    lines.append("")

    # CUT OUT
    lines.append("---------------------- CUT OUT (ignored from 新規製番リスト) ----------------------")
    if not cut_rows:
        lines.append("(none)")
    else:
        for code, reason in cut_rows:
            lines.append(f"- {code}: {reason}")
    lines.append("")

    # SU_Others outlier cells cut (pre-shift)
    lines.append("---------------------- SU_OTHERS OUTLIER CELLS CUT ----------------------")
    if not su_outlier_corrections:
        lines.append("(none)")
    else:
        by_code = defaultdict(list)
        for rec in su_outlier_corrections:
            by_code[rec.get("code", "?")].append(rec)
        for code in sorted(by_code.keys()):
            lines.append(f"module: {code}")
            recs = sorted(by_code[code], key=lambda r: (r.get("date", ""), r.get("wid", "")))
            for rec in recs[:2000]:
                wid = rec.get("wid", "?")
                nm = worker_id_to_name.get(wid, wid)
                lines.append(f"  - {rec.get('date')} / {wid}({nm})")
                lines.append(f"      {rec.get('text')}")
                lines.append(f"      reason: {rec.get('reason')}")
            if len(recs) > 2000:
                lines.append(f"  ... ({len(recs)-2000} more)")
            lines.append("")
    lines.append("")

    # SHIFTING DATE
    lines.append("---------------------- SHIFTING DATE (plan -> actual) ----------------------")
    for code in sorted(shifted_meta.keys()):
        m = shifted_meta[code]
        plan = m["plan"]
        lines.append(f"module: {code}")

        # planned
        lines.append("planned (新規製番リスト):")
        total_len = plan["total_len"]
        for ph in (1,2,3,4):
            pl = plan["phase_len"][ph]
            pct = plan["phase_pct"][ph] * 100.0
            lines.append(_format_phase_line(ph, plan["starts"][ph], plan["ends"][ph], extra=f"(len={pl}d, {pct:.1f}%)"))
        lines.append(f"  planned overall: {_to_ymd(plan['overall_start'])} - {_to_ymd(plan['overall_end'])} (total={total_len}d)")

        # actual
        if m["had_su_match"]:
            lines.append("shifted result (used in Schedule.yaml):")
            alloc_span = m.get("alloc_span_days") or m.get("alloc_days")  # support older logs
            alloc_worked = m.get("alloc_worked_days")

            for ph in (1, 2, 3, 4):
                ln_span = alloc_span.get(ph, 0) if alloc_span else 0
                extra = f"(alloc_span={ln_span}d)"
                if alloc_worked:
                    extra += f", worked_in_phase={alloc_worked.get(ph, 0)}d"
                lines.append(_format_phase_line(ph, m["shifted_starts"][ph], m["shifted_ends"][ph], extra=extra))

            lines.append(
                f"  shifted overall: {_to_ymd(m['actual_first'])} - {_to_ymd(m['actual_last'])} "
                f"(worked_total={m.get('actual_total')}d)"
            )
        else:
            lines.append("actual (SU_Others): NOT FOUND -> no shift (kept planned dates)")
            lines.append("shifted result (used in Schedule.yaml): (same as planned)")
        lines.append("")

    # workerday == 0
    lines.append("---------------------- WORKLOAD WARNING (worker-days == 0) ----------------------")
    if not workload_zero_phase_info:
        lines.append("(none)")
    else:
        for pid, mod in workload_zero_phase_info:
            if mod:
                lines.append(f"- phase_id: {pid} / module: {mod} (no assigned worker-days in SU_Others after shifting)")
            else:
                lines.append(f"- phase_id: {pid} (no assigned worker-days in SU_Others after shifting)")
    lines.append("")

    lines.append("---------------------- DUMMY MODULES ----------------------")
    lines.append("(1) SU_Others tool-code not in 新規製番リスト:")
    if not dummy_tool_labels:
        lines.append("  (none)")
    else:
        for code in sorted(dummy_tool_labels.keys()):
            labels = sorted(dummy_tool_labels[code])
            lines.append(f"- {code}:")
            for lb in labels[:50]:
                lines.append(f"    - {lb}")
            if len(labels) > 50:
                lines.append(f"    ... ({len(labels)-50} more)")

    lines.append("")
    lines.append("(2) SU_Others outlier-cut modules (tool-code intentionally broken -> treated as dummy 'other'):")
    if not outlier_cut_summary:
        lines.append("  (none)")
    else:
        for code in sorted(outlier_cut_summary.keys()):
            lines.append(f"- {code}:")
            for txt in outlier_cut_summary[code][:20]:
                lines.append(f"    - {txt}")
    lines.append("")

    lines.append("---------------------- PERSONAL BUSINESS (grey empty cells) ----------------------")
    if not pb_worker_dates:
        lines.append("(none)")
    else:
        for wid in sorted(pb_worker_dates.keys()):
            nm = worker_id_to_name.get(wid, wid)
            dts = pb_worker_dates[wid]
            lines.append(f"- {wid}({nm}) : {len(dts)} days")
            # print compact ranges
            # (simple range formatter)
            ranges = []
            if dts:
                start = dts[0]
                prev = dts[0]
                for cur in dts[1:]:
                    if cur == prev + pd.Timedelta(days=1):
                        prev = cur
                    else:
                        ranges.append((start, prev))
                        start = cur
                        prev = cur
                ranges.append((start, prev))
            for a, b in ranges[:200]:
                if a == b:
                    lines.append(f"    - {_to_ymd(a)}")
                else:
                    lines.append(f"    - {_to_ymd(a)} ~ {_to_ymd(b)}")
            if len(ranges) > 200:
                lines.append(f"    ... ({len(ranges)-200} more ranges)")
    lines.append("")

    Path(out_path).write_text("\n".join(lines), encoding="utf-8")

# ============================================================
# Build EnvConfig + Schedule and dump YAML
# ============================================================

def build_env_and_schedule_decoder3(
    su_others_path: str,
    task_csv_path: str,
    skill_excel_path: str,
    envconfig_out: str = "EnvConfig.yaml",
    schedule_out: str = "Schedule.yaml",
    log_out: str = TRANSFORMATION_LOG,
    read_all_data: bool = True,
    date_range: str | None = None,
    workload_use_window_days: bool = True,
):
    date_filter = _parse_date_range(date_range) if date_range else (None, None)
    f_start, f_end = date_filter

    # 1) Planned meta from 新規製番リスト (strict validate)
    task_meta = parse_tasks_from_csv_v3(task_csv_path, date_filter=date_filter if f_start is not None else None)
    planned_meta = task_meta["planned_meta"]
    cut_rows = task_meta["cut_rows"]
    valid_code_set = set(planned_meta.keys())

    # 2) Skills from スキル集計
    skill_levels, skill_people = parse_skill_excel(skill_excel_path)

    # 3) SU_Others
    su_outlier_corrections = []
    su_short_span_corrections = []
    if read_all_data:
        su_data = parse_su_others(su_others_path, date_filter=date_filter if f_start is not None else None)

        # Human error defense: horizontal drag-fill sometimes leaves a valid tool-code
        # in an isolated/non-contiguous cluster far away from the main work period.
        # Those cells should become dummy "other" so shifting uses the real span.
        su_outlier_corrections = cut_su_outlier_cells(
            su_data,
            cluster_gap_days=7,
            far_gap_days=60,
            small_cluster_max_unique_days=7,
            cut_module_if_unique_days_lt=4,
            cut_module_if_total_cells_lt=4,
            planned_meta=planned_meta,
            cut_if_far_from_planned_days=90
        )
        # short-span modules (< 4 unique worked days) must become dummy "other"
        su_short_span_corrections = cut_su_short_span_modules_to_dummy(
            su_data,
            min_unique_worked_days=MIN_WORKED_DAYS_FOR_TOOL,
            planned_meta=planned_meta
        )
        su_phase_zero_corrections = []
        if CUT_MODULE_IF_PHASE_ZERO_WORKLOAD:
            su_phase_zero_corrections = cut_module_if_phase_zero_workload(
                su_data,
                planned_meta=planned_meta
            )

        all_su_corrections = (
            su_outlier_corrections
            + su_short_span_corrections
            + su_phase_zero_corrections
        )
        # Summarize outlier-cut modules so they appear in CUT OUT + DUMMY MODULES sections
        outlier_cut_summary = defaultdict(list)  # code -> list of sample original texts
        for rec in all_su_corrections:
            code = rec.get("code")
            txt = rec.get("text")  # original text (A version)
            if code and txt:
                if len(outlier_cut_summary[code]) < 20:   # keep small sample
                    outlier_cut_summary[code].append(txt)

        # Also show these in CUT OUT section (as requested)
        for code in sorted(outlier_cut_summary.keys()):
            cut_rows.append((code, "SU_Others OUTLIER CUT: occurrences were converted to dummy 'other' (Fixed)"))

        worker_company_list = su_data["worker_company_list"]
        worker_list = su_data["worker_list"]

        # any code that got short-span-cut must not be output as wf_tool
        short_span_codes = set()
        for rec in su_short_span_corrections:
            code = rec.get("code")
            if code:
                short_span_codes.add(code)

        if short_span_codes:
            for code in sorted(short_span_codes):
                cut_rows.append((code, f"DUMMY: SU_Others unique worked days < {MIN_WORKED_DAYS_FOR_TOOL} (converted to wf_other)"))
                planned_meta.pop(code, None)   # remove from planned modules
            # keep task_meta["valid_codes"] consistent
            task_meta["valid_codes"] = [c for c in task_meta["valid_codes"] if c in planned_meta]
    else:
        su_data = None
        comp_to_id = {}
        worker_company_list = []

        def get_wc_id(cn: str):
            cn = str(cn).strip() if isinstance(cn, str) else ""
            if cn not in comp_to_id:
                cid = f"wc{len(comp_to_id) + 1}"
                comp_to_id[cn] = cid
                worker_company_list.append({
                    "id": cid,
                    "name": cn,
                    "annual_overtime_limit": 10000,
                    "monthly_overtime_limit": 10000,
                    "unavailable_dates": [],
                })
            return comp_to_id[cn]

        keys = sorted(skill_people.keys())
        worker_list = []
        for i, k in enumerate(keys, start=1):
            meta = skill_people[k]
            worker_list.append({
                "id": f"w{i:03d}",
                "name": meta["name"],
                "worker_company": get_wc_id(meta["company"]),
                "is_manager": False,
                "skill_map": {},
                "fab_suitability_map": [],
                "unavailable_dates": [],
            })

    # 4) Merge workers from skill excel (name-only key)
    wc_id_to_name = {wc["id"]: wc["name"] for wc in worker_company_list}
    wc_name_to_id = {wc["name"]: wc["id"] for wc in worker_company_list}

    def get_or_create_wc_id(company_name: str) -> str:
        company_name = str(company_name).strip() if isinstance(company_name, str) else ""
        if company_name not in wc_name_to_id:
            cid = f"wc{len(wc_name_to_id) + 1}"
            wc_name_to_id[company_name] = cid
            worker_company_list.append({
                "id": cid,
                "name": company_name,
                "annual_overtime_limit": 10000,
                "monthly_overtime_limit": 10000,
                "unavailable_dates": [],
            })
            wc_id_to_name[cid] = company_name
        return wc_name_to_id[company_name]

    worker_by_key = {_norm_name(w["name"]): w for w in worker_list}

    next_worker_num = 1
    if worker_list:
        try:
            next_worker_num = max(int(w["id"][1:]) for w in worker_list) + 1
        except Exception:
            next_worker_num = len(worker_list) + 1

    for key, meta in skill_people.items():
        if key in worker_by_key:
            continue
        cid = get_or_create_wc_id(meta["company"])
        new_w = {
            "id": f"w{next_worker_num:03d}",
            "name": meta["name"],
            "worker_company": cid,
            "is_manager": False,
            "skill_map": {},
            "fab_suitability_map": [],
            "unavailable_dates": [],
        }
        next_worker_num += 1
        worker_list.append(new_w)
        worker_by_key[key] = new_w

    worker_id_to_name = {w["id"]: w["name"] for w in worker_list}

    # 5) Build shifted meta using SU_Others
    shifted_meta, code_to_shifted_phases_simple, _code_occ = build_shifted_meta(planned_meta, su_data if read_all_data else None)
    
    # 5.25) OPTION B: if SU_Others NOT FOUND -> skip module entirely (do not output in Schedule.yaml)
    # This also prevents it from appearing in WORKLOAD WARNING because tool tasks won't exist.
    if SKIP_MODULE_IF_NO_SU_MATCH:
        no_su_codes = [code for code, m in shifted_meta.items() if not m.get("had_su_match")]
        if no_su_codes:
            for code in no_su_codes:
                # log it like "CUT OUT" so everyone sees why it's missing
                cut_rows.append((code, "SKIPPED: SU_Others NOT FOUND; module omitted from Schedule.yaml"))

                # remove from planned_meta + shifted_meta so it never becomes a tool task
                planned_meta.pop(code, None)
                shifted_meta.pop(code, None)

            # recompute valid_code_set and valid_codes list
            valid_code_set = set(planned_meta.keys())
            task_meta["valid_codes"] = [c for c in task_meta["valid_codes"] if c in valid_code_set]

    # 5.5) EXTRA CUT: actual span too far from planned window
    cut_due_to_distance = []
    if read_all_data and su_data is not None:
        for code, m in list(shifted_meta.items()):
            if not m.get("had_su_match"):
                continue

            plan = m["plan"]
            gap = _planned_actual_gap_days(
                plan["overall_start"], plan["overall_end"],
                m["actual_first"], m["actual_last"]
            )

            if gap > CUT_DISTANCE_DAYS:
                cut_due_to_distance.append((code, gap))

        if cut_due_to_distance:
            for code, gap in cut_due_to_distance:
                cut_rows.append((code, f"actual span is outside planned window by {gap} days (> {CUT_DISTANCE_DAYS})"))

                # remove it so it behaves like "not listed"
                planned_meta.pop(code, None)
                shifted_meta.pop(code, None)

            # recompute valid_code_set and valid_codes list
            valid_code_set = set(planned_meta.keys())
            task_meta["valid_codes"] = [c for c in task_meta["valid_codes"] if c in valid_code_set]

    # 6) Build tool tasks + code_to_phases (for assignments) with shifted windows
    tool_tasks = []
    code_to_phases = defaultdict(list)
    all_dates = []

    task_counter = 1
    for code in task_meta["valid_codes"]:
        meta = shifted_meta.get(code)
        if not meta:
            continue

        plan = meta["plan"]
        customer, country, fab_name = plan["customer"], plan["country"], plan["fab_name"]

        task_id = f"e{task_counter}"
        task_counter += 1

        phase_task_list = []
        for ph in (1,2,3,4):
            phase_id = f"{task_id}_p{ph}"
            start = meta["shifted_starts"][ph]
            end = meta["shifted_ends"][ph]
            if end < start:
                # empty window; collapse to start
                end = start
            nm = {1: "Module Setup", 2: "Hardware Setup", 3: "Function Setup", 4: "Utility"}.get(ph, f"P{ph}")
            workload_days_a = len(meta["phase_days"][ph]) if (meta.get("had_su_match") and meta.get("phase_days")) else int((end-start).days)+1


            phase_task_list.append({
                "id": phase_id,
                "name": nm,
                "phase": f"tool_p{ph}",
                "start_date": _to_ymd(start),
                "end_date": _to_ymd(end),
                "operation_task_list": [{
                    "id": phase_id,
                    "name": nm,
                    "operation": f"p{ph}",
                    "workload_days": workload_days_a,  # temp, overwritten later by option logic
                }],
            })

            ds = None
            if meta.get("had_su_match") and meta.get("phase_days"):
                ds = set(meta["phase_days"].get(ph, []))

            code_to_phases[code].append({
                "phase_index": ph,
                "phase_id": phase_id,
                "start": start,
                "end": end,
                "operation": f"p{ph}",
                "date_set": ds,   # IMPORTANT
            })

            all_dates.extend([start, end])

        tool_tasks.append({
            "id": task_id,
            "name": code,
            "workflow": "wf_tool",
            "fab": None,     # filled later
            "phase_task_list": phase_task_list,
            "module_code": code,   # internal
            "customer": customer,  # internal
            "country": country,    # internal
            "fab_name": fab_name,  # internal
        })

    # 7) customer/region/fab lists from tasks (same ids policy)
    customer_name_to_id = {"OTHER": "c_other"}
    region_name_to_id = {"Other": "r_other"}
    fab_name_to_id = {"Other": "f_other"}

    customer_company_list = [{"id": "c_other", "name": "OTHER", "unavailable_dates": []}]
    region_list = [{
        "id": "r_other",
        "name": "Other",
        "max_stay_on": 10000,
        "max_annual_stay": 10000,
        "stay_off_interval": 3,
        "unavailable_dates": [],
    }]
    fab_list = [{"id": "f_other", "name": "Other", "region": "r_other", "customer_company": "c_other", "unavailable_dates": []}]

    def get_customer_id(name: str) -> str:
        nm = name.strip() if isinstance(name, str) and name.strip() else "OTHER"
        if nm not in customer_name_to_id:
            cid = f"c{len(customer_name_to_id)}"
            customer_name_to_id[nm] = cid
            customer_company_list.append({"id": cid, "name": nm, "unavailable_dates": []})
        return customer_name_to_id[nm]

    def get_region_id(country: str) -> str:
        nm = country.strip() if isinstance(country, str) and country.strip() else "Other"
        if nm not in region_name_to_id:
            rid = f"r{len(region_name_to_id)}"
            region_name_to_id[nm] = rid
            region_list.append({
                "id": rid,
                "name": nm,
                "max_stay_on": 10000,
                "max_annual_stay": 10000,
                "stay_off_interval": 3,
                "unavailable_dates": [],
            })
        return region_name_to_id[nm]

    def get_fab_id(fab_name: str, country: str, customer: str) -> str:
        nm = fab_name.strip() if isinstance(fab_name, str) and fab_name.strip() else "Other"
        if nm not in fab_name_to_id:
            fid = f"f{len(fab_name_to_id)}"
            fab_name_to_id[nm] = fid
            fab_list.append({
                "id": fid,
                "name": nm,
                "region": get_region_id(country),
                "customer_company": get_customer_id(customer),
                "unavailable_dates": [],
            })
        return fab_name_to_id[nm]

    for t in tool_tasks:
        fid = get_fab_id(t.get("fab_name"), t.get("country"), t.get("customer"))
        t["fab"] = fid

    # Decoder3 change: transit gap involving r_other is 0, otherwise default 1
    def build_transite_day_map(region_list, days_default: int = 1):
        region_ids = [r["id"] for r in region_list if r.get("id")]
        out = []
        for fr in region_ids:
            for to in region_ids:
                if fr == to:
                    continue
                d = 0 if (fr == "r_other" or to == "r_other") else days_default
                out.append({"from": fr, "to": to, "days": d})
        return out

    transite_day_map = build_transite_day_map(region_list, days_default=1)

    # 8) environment
    environment = {
        "workflow_list": [
            {
                "id": "wf_tool",
                "name": "Tool Install",
                "phase_list": [
                    {"id": "tool_p1", "name": "Module Setup", "operation_list": [{"id": "p1", "name": "Module Setup", "work_hours": [10], "min_worker_num": 1, "max_worker_num": DEFAULT_MAX_WORKER}]},
                    {"id": "tool_p2", "name": "Hardware Setup", "operation_list": [{"id": "p2", "name": "Hardware Setup", "work_hours": [10], "min_worker_num": 1, "max_worker_num": DEFAULT_MAX_WORKER}]},
                    {"id": "tool_p3", "name": "Function Setup", "operation_list": [{"id": "p3", "name": "Function Setup", "work_hours": [10], "min_worker_num": 1, "max_worker_num": DEFAULT_MAX_WORKER}]},
                    {"id": "tool_p4", "name": "Utility", "operation_list": [{"id": "p4", "name": "Utility", "work_hours": [10], "min_worker_num": 1, "max_worker_num": DEFAULT_MAX_WORKER}]},
                ],
            },
            {
                "id": "wf_other",
                "name": "Other work (from SU_Others)",
                "phase_list": [{
                    "id": "other_p1",
                    "name": "Other work",
                    "operation_list": [{"id": "other_op", "name": "Other work", "work_hours": [10], "min_worker_num": 1, "max_worker_num": 28}],
                }],
            },
            {
                "id": "wf_personal_business",
                "name": "Personal Business",
                "phase_list": [{
                    "id": "pb_p1",
                    "name": "Personal Business",
                    "operation_list": [{"id": "personal_business_op", "name": "Personal Business", "work_hours": [10], "min_worker_num": 1, "max_worker_num": 42}],
                }],
            },
        ],
        "fab_list": fab_list,
        "region_list": region_list,
        "customer_company_list": customer_company_list,
        "worker_company_list": worker_company_list,
        "transite_day_map": transite_day_map,
        "worker_list": worker_list,
    }

    # 9) assignments
    if read_all_data and su_data is not None:
        assignments, misc_tasks, personal_tasks, inferred_worker_phase, phase_workerday_count, phase_worker_count, pb_worker_dates, dummy_tool_labels = build_assignments_v3(
            su_data, code_to_phases, valid_code_set, date_filter=date_filter if f_start is not None else None
        )
    else:
        assignments, misc_tasks, personal_tasks = [], [], []
        inferred_worker_phase = defaultdict(set)
        phase_workerday_count = defaultdict(int)
        dummy_tool_labels = {}
        phase_worker_count = {}

    # 10) fill skills (base + inferred + excel max)
    for w in environment["worker_list"]:
        w["skill_map"] = {
            "p1": 0, "p2": 0, "p3": 0, "p4": 0,
            "other_op": 1, "personal_business_op": 1
        }

    for wid, ops in inferred_worker_phase.items():
        w = next((x for x in environment["worker_list"] if x["id"] == wid), None)
        if not w:
            continue
        for op in ops:
            if op in ("p1", "p2", "p3", "p4"):
                w["skill_map"][op] = max(w["skill_map"].get(op, 0), 1)

    for w in environment["worker_list"]:
        key = _norm_name(w["name"])
        excel_map = skill_levels.get(key)
        if not excel_map:
            continue
        for op in ("p1", "p2", "p3", "p4"):
            w["skill_map"][op] = max(int(w["skill_map"].get(op, 0)), int(excel_map.get(op, 0)))

    # 11) plan_range
    if read_all_data and su_data is not None:
        try:
            all_dates.append(pd.to_datetime(su_data["plan_range"]["start_date"]))
            all_dates.append(pd.to_datetime(su_data["plan_range"]["end_date"]))
        except Exception:
            pass
    all_dates = [d for d in all_dates if isinstance(d, pd.Timestamp)]
    all_dates.sort()
    plan_range = {"start_date": _to_ymd(all_dates[0]), "end_date": _to_ymd(all_dates[-1])} if all_dates else {"start_date": "2025/01/01", "end_date": "2025/01/01"}

    # strip internal keys
    tool_tasks_for_yaml = []
    for t in tool_tasks:
        t2 = dict(t)
        for k in ("module_code", "customer", "country", "fab_name"):
            t2.pop(k, None)
        tool_tasks_for_yaml.append(t2)

    # Map phase_id (eN_pX) -> module code (task name)
    phase_id_to_module_code = {}
    for t in tool_tasks_for_yaml:
        mod = t.get("name", "")
        for pt in t.get("phase_task_list", []):
            pid = pt.get("id")
            if pid:
                phase_id_to_module_code[pid] = mod

    # 12) workload_days overwrite + collect workerday==0 warnings
    workload_zero_phase_info = []
    for t in tool_tasks_for_yaml:
        if t.get("workflow") != "wf_tool":
            continue
        for pt in t.get("phase_task_list", []):
            phase_id = pt.get("id")  # eN_pX

            b = int(phase_workerday_count.get(phase_id, 0))
            if b == 0:
                workload_zero_phase_info.append((phase_id, phase_id_to_module_code.get(phase_id, "")))

            # NEW: unique worker count => max_worker_num
            uniq_w = int(phase_worker_count.get(phase_id, 0))

            for ot in pt.get("operation_task_list", []):
                a = int(ot.get("workload_days", 0))
                if workload_use_window_days:
                    ot["workload_days"] = max(a, b)
                else:
                    ot["workload_days"] = b
                # If the real worker assigned was exceed default the min/max of the worker is have special worker maimum count
                if uniq_w > DEFAULT_MAX_WORKER:
                    # Put min/max into Schedule.yaml operation_task_list
                    ot["min_worker_num"] = 1

                    # if nobody worked in SU_Others, keep it safe:
                    # - either 1
                    # - or keep workflow default (3) if you prefer
                    ot["max_worker_num"] = (uniq_w if uniq_w > 0 else DEFAULT_MAX_WORKER)
                #The recommend worker number is depend on the worker real assigned 
                ot["recommends_worker_min"] = (uniq_w if uniq_w > 0 else DEFAULT_MAX_WORKER)
                ot["recommends_worker_max"] = (uniq_w if uniq_w > 0 else DEFAULT_MAX_WORKER)
                

    schedule = {
        "plan_range": plan_range,
        "workflow_task_list": tool_tasks_for_yaml + misc_tasks + personal_tasks,
        "assignment_list": assignments,
    }

    env_root = {"environment": environment}
    sch_root = {"schedule": schedule}

    _BaseDumper = getattr(yaml, "CSafeDumper", yaml.SafeDumper)

    class NoAliasDumper(_BaseDumper):
        def ignore_aliases(self, data):
            return True

    with open(envconfig_out, "w", encoding="utf-8") as f:
        yaml.dump(env_root, f, Dumper=NoAliasDumper, sort_keys=False, allow_unicode=True, width=4096)

    with open(schedule_out, "w", encoding="utf-8") as f:
        yaml.dump(sch_root, f, Dumper=NoAliasDumper, sort_keys=False, allow_unicode=True, width=4096)

    # transformation log
    write_transformation_log(
        log_out,
        cut_rows=cut_rows,
        shifted_meta=shifted_meta,
        worker_id_to_name=worker_id_to_name,
        workload_zero_phase_info=sorted(set(workload_zero_phase_info)),
        dummy_tool_labels=dummy_tool_labels,
        su_outlier_corrections=all_su_corrections,
        outlier_cut_summary=outlier_cut_summary,
        pb_worker_dates=pb_worker_dates,
    )

    return env_root, sch_root, shifted_meta


if __name__ == "__main__":
    su_file = "20260105 SU_Others.xlsm"
    task_file = "SU_Others_予定表_2025_新規製番リスト_20260127.xlsx"
    skill_file = "スキル集計_20260127.xlsx"

    su_path = Path(su_file)
    task_path = Path(task_file)
    skill_path = Path(skill_file)

    if (not READ_ALL_DATA or su_path.exists()) and task_path.exists() and skill_path.exists():
        build_env_and_schedule_decoder3(
            str(su_path),
            str(task_path),
            str(skill_path),
            envconfig_out="EnvConfig.yaml",
            schedule_out="Schedule.yaml",
            log_out=TRANSFORMATION_LOG,
            read_all_data=READ_ALL_DATA,
            date_range=DATE_RANGE,
            workload_use_window_days=WORKLOAD_USE_WINDOW_DAYS,
        )
        print("EnvConfig.yaml, Schedule.yaml, and TransformationLog.txt have been written.")
    else:
        print("Please fix input file paths at the bottom of Decoder3.py.")