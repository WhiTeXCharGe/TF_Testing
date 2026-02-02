# compare_workers_names.py
# ------------------------------------------------------------
# Compare worker NAME lists between:
#   1) 20260105 SU_Others.xlsm  (予定表_2024 / 予定表_2025)
#   2) スキル集計_20260127.xlsx (table with columns: 氏名, 所属)
#
# Output YAML: worker_compare_names.yaml
#   - su_others_names_not_in_skill
#   - skill_names_not_in_su_others
#   - same_name_but_company_diff (name exists in both, but company set differs)
#
# Normalization:
# - NFKC (handles full-width/half-width)
# - trim + collapse spaces
# ------------------------------------------------------------
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import yaml
from openpyxl import load_workbook


def norm_text(s) -> str:
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("　", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def read_su_pairs(path: str, sheet_names=("予定表_2024", "予定表_2025")):
    wb = load_workbook(path, data_only=True, read_only=False)
    pairs = set()

    used = [s for s in sheet_names if s in wb.sheetnames]
    if not used:
        used = [s for s in wb.sheetnames if str(s).startswith("予定表")]

    for sh in used:
        ws = wb[sh]

        # detect date header row like decoder
        date_row = None
        for r in range(1, 16):
            vals = [c.value for c in ws[r]]
            if any(hasattr(v, "year") and hasattr(v, "month") for v in vals):
                date_row = r
                break

        start_r = (date_row + 2) if date_row else 1

        blank = 0
        for r in range(start_r, ws.max_row + 1):
            company = ws.cell(row=r, column=1).value
            name = ws.cell(row=r, column=2).value

            name_n = norm_text(name)
            if not name_n:
                blank += 1
                if blank >= 30:
                    break
                continue
            blank = 0

            comp_n = norm_text(company)
            pairs.add((name_n, comp_n))

    return pairs


def read_skill_pairs(path: str):
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[wb.sheetnames[0]]

    # header row might not be row 1, so scan
    header_row = None
    max_scan = min(60, ws.max_row)
    for r in range(1, max_scan + 1):
        row_vals = [norm_text(c.value) for c in ws[r]]
        if "氏名" in row_vals and "所属" in row_vals:
            header_row = r
            break

    if header_row is None:
        raise RuntimeError("Could not find header row containing 氏名 and 所属.")

    header_vals = [norm_text(c.value) for c in ws[header_row]]
    name_col_idx = header_vals.index("氏名") + 1
    comp_col_idx = header_vals.index("所属") + 1

    pairs = set()
    blank = 0
    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=name_col_idx).value
        comp = ws.cell(row=r, column=comp_col_idx).value

        name_n = norm_text(name)
        if not name_n:
            blank += 1
            if blank >= 30:
                break
            continue
        blank = 0

        comp_n = norm_text(comp)
        pairs.add((name_n, comp_n))

    return pairs


def main():
    su_path = Path("20260105 SU_Others.xlsm")
    skill_path = Path("スキル集計_20260127.xlsx")

    su_pairs = read_su_pairs(str(su_path))
    skill_pairs = read_skill_pairs(str(skill_path))

    su_names = {n for n, _ in su_pairs}
    skill_names = {n for n, _ in skill_pairs}

    su_names_not_in_skill = sorted(list(su_names - skill_names))
    skill_names_not_in_su = sorted(list(skill_names - su_names))

    su_by_name = defaultdict(set)
    sk_by_name = defaultdict(set)
    for n, c in su_pairs:
        su_by_name[n].add(c)
    for n, c in skill_pairs:
        sk_by_name[n].add(c)

    same_name_diff_company = []
    for n in sorted(su_names & skill_names):
        if su_by_name[n] != sk_by_name[n]:
            same_name_diff_company.append({
                "name": n,
                "su_others_company_list": sorted(list(su_by_name[n])),
                "skill_company_list": sorted(list(sk_by_name[n])),
            })

    out = {
        "su_others_names_not_in_skill": su_names_not_in_skill,
        "skill_names_not_in_su_others": skill_names_not_in_su,
        "same_name_but_company_diff": same_name_diff_company,
        "stats": {
            "su_others_unique_names": len(su_names),
            "skill_unique_names": len(skill_names),
            "su_names_not_in_skill": len(su_names_not_in_skill),
            "skill_names_not_in_su": len(skill_names_not_in_su),
            "same_name_company_diff": len(same_name_diff_company),
        }
    }

    with open("worker_compare_names.yaml", "w", encoding="utf-8") as f:
        yaml.safe_dump(out, f, sort_keys=False, allow_unicode=True, width=4096)

    print("Wrote worker_compare_names.yaml")


if __name__ == "__main__":
    main()
