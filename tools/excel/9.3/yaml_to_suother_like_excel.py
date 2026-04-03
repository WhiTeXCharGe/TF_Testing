# yaml_to_suother_like_schedule.py
# ------------------------------------------------------------
# Convert Timefold output (EnvConfig.yaml + Schedule.yaml)
# into a SU_Others-like "Employees x Dates" Excel.
#
# Differences from export_schedule_excel.py Sheet2:
# - Intersect cells show MODULE "name" (module code) instead of task id.
# - Same base code (before first "_") uses same color.
# - Red is reserved ONLY for unavailable dates.
# - Left columns: company(worker_company) | name | manager | today_task
# - Sheet name: "Schedule"
# - Date headers include YEAR (YYYY/MM/DD)
#
# Usage:
#   python yaml_to_suother_like_schedule.py EnvConfig.yaml Schedule.yaml out.xlsx
#
# Optional env var:
#   EXPORT_TODAY_YMD=2025/10/14   (controls "today_task" column)
# ------------------------------------------------------------

import os
import re
import sys
import yaml
from datetime import datetime, date, timedelta
from collections import defaultdict
import hashlib

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter


# ---------------- CONFIG (easy to adjust) ----------------
DATE_COL_WIDTH = 3.5   # width for the employee x date cells

META_COL_WIDTHS = {
    "company": 20,
    "name": 20,
    "role": 16,
    "manager": 10,
    "today_task": 30,
}

HEADER_ROW_HEIGHT = 18
DATA_ROW_HEIGHT = 18
HEADER_ROWS = 3
MONTH_ROW = 1
DAY_ROW = 2
DOW_ROW = 3
DATA_START_ROW = 4

# Filter huge worker_list:
# True  -> include everyone in EnvConfig.worker_list (can be 800+)
# False -> include only workers who have at least one assignment OR one unavailable day within plan range
INCLUDE_ALL_WORKERS = False

# If multiple module assignments in same day, join by newline.
JOIN_MULTIPLE_WITH_NEWLINE = True

# Unavailable date fill (must be RED-ish)
UNAV_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

BLACK = Font(color="000000")
BOLD_BLACK = Font(bold=True, color="000000")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False, )
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)

THIN = Side(style="thin", color="999999")
BORDER_THIN = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

PB_WORKFLOW_ID = "wf_personal_business"
PB_COLOR = "898989"
PB_FILL = PatternFill(start_color=PB_COLOR, end_color=PB_COLOR, fill_type="solid")
PB_FONT = Font(color="FFFFFF")   # optional; use white so it’s readable on grey
# Flexible highlight mode:
# False -> current behavior (black text + PALETTE_FLEX)
# True  -> flexible cells use WHITE text + DARK palette (strong highlight)
FLEX_WHITE_TEXT_MODE = False

# Optional: only apply to flexible cells (fixed stays black)
FLEX_FONT_WHITE = Font(color="FFFFFF", bold=False)
FLEX_FONT_BLACK = Font(color="000000", bold=False)

# Light palette (avoid red + avoid too-dark; text must remain black)
PALETTE_FIXED = [
    "FFF2CC", "D9EAD3", "CFE2F3", "D9D2E9", "FCE5CD", "EAD1DC", "D0E0E3", "E2EFDA",
    "DEEAF6", "E7E6E6", "C9DAF8", "D9E1F2", "D0CECE", "E2F0D9", "DDEBF7", "F8CBAD",
]

# One palette for ALL module background colors (stable by workflow_task_list order)
# Excludes reds and greys; suitable with BLACK text (and WHITE when needed)
MODULE_PALETTE = [
    "FFE599", "FFD966", "F9CB9C", "F6B26B", "FCE5CD",
    "B6D7A8", "93C47D", "6AA84F", "D9EAD3", "E2EFDA",
    "9FC5E8", "6FA8DC", "3D85C6", "A4C2F4", "CFE2F3",
    "B4A7D6", "8E7CC3", "674EA7", "D9D2E9", "EAD1DC",
    "A2C4C9", "76A5AF", "45818E", "D0E0E3", "DDEBF7",
    "C27BA0", "D5A6BD", "E6B8AF", "F8CBAD", "FFF2CC",
]

#  search/output sheet
SEARCH_KEYWORD = ""         # example: "1001A" ; empty = disable
SEARCH_SHEET_PREFIX = "Search_"

# role-only filter for main sheet
ROLE_FILTER_EXACT = ""      # example: "QC(OJT)" ; empty = disable

# Red outline for keyword hit bars
HIT_RED = "FF0000"
HIT_SIDE_MEDIUM = Side(style="medium", color=HIT_RED)
# ---------------- Utilities ----------------
def _d(s) -> date:
    """Parse 'YYYY/MM/DD' (or 'YYYY-MM-DD') to date."""
    if isinstance(s, date):
        return s
    s = str(s).strip().replace("-", "/")
    return datetime.strptime(s, "%Y/%m/%d").date()


def daterange(start: date, end: date):
    d = start
    while d <= end:
        yield d
        d += timedelta(days=1)


def load_yaml(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def normalize_module_code(name: str) -> str:
    """
    Same module code may appear with different suffix name:
      '530N02621A_JP_Rapidus_新規' and '530N02621A'
    -> base is text before first '_'
    """
    if not name:
        return ""
    s = str(name).strip()
    return s.split("_", 1)[0]

def stable_palette_color(key: str, palette: list[str], namespace: str) -> str:
    """
    Stable across runs/machines.
    namespace lets you separate mappings for company vs module if you want.
    """
    s = (namespace + "::" + (key or "")).encode("utf-8")
    h = hashlib.md5(s).hexdigest()
    idx = int(h[:8], 16) % len(palette)
    return palette[idx]

def build_module_fill_by_order(modules, module_name_by_id):
    """
    Stable module color assignment by the order in Schedule.yaml workflow_task_list.
    Uses normalized module code to keep suffix variants consistent:
      530N02621A_JP_Rapidus_新規  and  530N02621A  -> same base -> same color
    """
    base_to_color = {}
    module_id_to_fill = {}

    color_idx = 0
    for m in modules or []:
        mid = m.get("id")
        name = (m.get("name") or "")
        base = normalize_module_code(name)

        # Keep same base code same color (first occurrence decides)
        if base not in base_to_color:
            base_to_color[base] = MODULE_PALETTE[color_idx % len(MODULE_PALETTE)]
            color_idx += 1

        col = base_to_color[base]
        module_id_to_fill[mid] = PatternFill(start_color=col, end_color=col, fill_type="solid")

    # fallback fill for unknowns (not red/grey)
    unknown_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    return module_id_to_fill, unknown_fill

def parse_unavailable_dates(raw, plan_start: date, plan_end: date) -> set:
    """
    Supports these shapes:
      A) unavailable_dates:
           - date: 2025/01/25
           - date: 2025/01/26

      B) unavailable_dates:
           - 2025/01/25
           - 2025/01/26

      C) unavailable_dates:
           - single:
               days: [2025/01/25, ...]
           - weekly:
               weekdays: [sat, sun, ...]

    (C is kept for compatibility with older configs.)
    """
    off = set()
    if raw is None:
        return off

    def add_single(obj):
        try:
            dd = _d(obj)
        except Exception:
            return
        if plan_start <= dd <= plan_end:
            off.add(dd)

    def weekday_to_int(w):
        m = {
            "mon": 0, "monday": 0,
            "tue": 1, "tues": 1, "tuesday": 1,
            "wed": 2, "weds": 2, "wednesday": 2,
            "thu": 3, "thur": 3, "thurs": 3, "thursday": 3,
            "fri": 4, "friday": 4,
            "sat": 5, "saturday": 5,
            "sun": 6, "sunday": 6,
        }
        return m.get(str(w).strip().lower())

    items = raw if isinstance(raw, list) else [raw]
    weekly = set()

    for item in items:
        if item is None:
            continue

        #  Case A: dict with {"date": "..."}
        if isinstance(item, dict) and "date" in item:
            add_single(item["date"])
            continue

        # Case C: single/weekly schema
        if isinstance(item, dict):
            single = item.get("single")
            if isinstance(single, dict):
                days = single.get("days")
                if isinstance(days, list):
                    for x in days:
                        add_single(x)

            weekly_obj = item.get("weekly")
            if isinstance(weekly_obj, dict):
                wds = weekly_obj.get("weekdays")
                if isinstance(wds, list):
                    for w in wds:
                        wd = weekday_to_int(w)
                        if wd is not None:
                            weekly.add(wd)
            continue

        # Case B: plain scalar date
        add_single(item)

    if weekly:
        for dd in daterange(plan_start, plan_end):
            if dd.weekday() in weekly:
                off.add(dd)

    return off



def build_op_task_index(workflow_task_list):
    """
    Build mapping: operation_task_list.id -> module_id
    Uses Schedule.yaml task structure (no guessing).
    """
    idx = {}
    for m in workflow_task_list or []:
        mid = m["id"]
        for ph in (m.get("phase_task_list") or []):
            for ot in (ph.get("operation_task_list") or []):
                idx[ot["id"]] = mid
    return idx

def company_fill(comp_key: str):
    if not comp_key:
        return None
    col = stable_palette_color(comp_key, PALETTE_FIXED, namespace="company")  # stable
    return PatternFill(start_color=col, end_color=col, fill_type="solid")

def copy_cell_style(src, dst):
    dst.font = src.font.copy()
    dst.fill = src.fill.copy()
    dst.border = src.border.copy()
    dst.alignment = src.alignment.copy()
    dst.number_format = src.number_format

def find_keyword_runs_for_row(ws, row_idx, start_col_dates, end_col_dates, keyword):
    """
    Return continuous runs [(start_col, end_col), ...] where the cell value contains keyword.
    """
    runs = []
    in_run = False
    run_start = None

    for col in range(start_col_dates, end_col_dates + 1):
        val = ws.cell(row=row_idx, column=col).value
        hit = keyword in str(val) if val is not None else False

        if hit and not in_run:
            in_run = True
            run_start = col
        elif not hit and in_run:
            runs.append((run_start, col - 1))
            in_run = False
            run_start = None

    if in_run:
        runs.append((run_start, end_col_dates))

    return runs

def apply_red_outline_run(ws, row_idx, start_col, end_col):
    """
    Draw a red outline around one continuous bar on a single row.
    """
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col)

        left = HIT_SIDE_MEDIUM if col == start_col else cell.border.left
        right = HIT_SIDE_MEDIUM if col == end_col else cell.border.right
        top = HIT_SIDE_MEDIUM
        bottom = HIT_SIDE_MEDIUM

        cell.border = Border(
            left=left,
            right=right,
            top=top,
            bottom=bottom
        )

def build_search_sheet(
    wb,
    source_ws,
    keyword,
    headers_count,
    start_col_dates,
    dates,
    meta_col_widths,
    header_rows=3
):
    if not keyword:
        return

    sheet_name = f"{SEARCH_SHEET_PREFIX}{keyword}"
    # Excel sheet name max 31 chars
    sheet_name = sheet_name[:31]

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws2 = wb.create_sheet(sheet_name)

    last_row = source_ws.max_row
    last_col = start_col_dates + len(dates) - 1

    # copy column widths
    for col in range(1, last_col + 1):
        col_letter = get_column_letter(col)
        if source_ws.column_dimensions[col_letter].width is not None:
            ws2.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width

    # copy header rows 1..3
    for r in range(1, header_rows + 1):
        ws2.row_dimensions[r].height = source_ws.row_dimensions[r].height
        for c in range(1, last_col + 1):
            src = source_ws.cell(row=r, column=c)
            dst = ws2.cell(row=r, column=c, value=src.value)
            copy_cell_style(src, dst)

    # copy merged month cells
    for merged in source_ws.merged_cells.ranges:
        if merged.min_row <= header_rows:
            ws2.merge_cells(str(merged))

    out_row = DATA_START_ROW
    for src_row in range(DATA_START_ROW, last_row + 1):
        runs = find_keyword_runs_for_row(source_ws, src_row, start_col_dates, last_col, keyword)
        if not runs:
            continue

        # copy whole row
        ws2.row_dimensions[out_row].height = source_ws.row_dimensions[src_row].height
        for c in range(1, last_col + 1):
            src = source_ws.cell(row=src_row, column=c)
            dst = ws2.cell(row=out_row, column=c, value=src.value)
            copy_cell_style(src, dst)

        # apply red outline only to matching bars
        for run_start, run_end in runs:
            apply_red_outline_run(ws2, out_row, run_start, run_end)

        out_row += 1

    ws2.freeze_panes = source_ws.freeze_panes
    ws2.auto_filter.ref = f"A{DOW_ROW}:{get_column_letter(last_col)}{ws2.max_row}"
# ---------------- Main exporter ----------------
def main(env_path: str, sched_path: str, out_path: str):
    env = load_yaml(env_path)
    sched = load_yaml(sched_path)

    env_root = env.get("environment", env)
    sched_root = sched.get("schedule", sched)

    # plan range (FULL YEAR RANGE)
    plan_start = _d(sched_root["plan_range"]["start_date"])
    plan_end = _d(sched_root["plan_range"]["end_date"])
    dates = list(daterange(plan_start, plan_end))

    # env lookups
    worker_companies = {c["id"]: c for c in env_root.get("worker_company_list", [])}
    workers = {w["id"]: w for w in env_root.get("worker_list", [])}

    # schedule lookups
    modules = sched_root.get("workflow_task_list", []) or []
    module_name_by_id = {m["id"]: (m.get("name") or "") for m in modules}
    module_workflow_by_id = {m["id"]: (m.get("workflow") or "") for m in modules}
    module_fill_by_id, unknown_fill = build_module_fill_by_order(modules, module_name_by_id)

    op_task_to_module = build_op_task_index(modules)

    # assignments expanded to per-day:
    # (wid, date) -> [module_id...]
    cell_modules = defaultdict(list)
    assigned_workers = set()
    for a in (sched_root.get("assignment_list") or []):
        wid = a.get("worker")
        op_task = a.get("operation_task")
        mid = op_task_to_module.get(op_task)

        # IMPORTANT: If not found, keep it visible but DON'T break.
        # However, in correct data this should be found.
        if not mid:
            # fallback: do NOT invent "OTHER"; just keep op_task as key
            mid = f"UNKNOWN::{op_task}"

        wd_key = "work_date_lsit" if "work_date_lsit" in a else "work_date_list"
        for ditem in (a.get(wd_key) or []):
            dd = _d(ditem.get("date"))
            if plan_start <= dd <= plan_end:
                flex = str(a.get("plan_flexibility", "")).strip()   # "Flexible" / "Fixed" / maybe blank
                is_flexible = (flex.lower() == "flexible")

                cell_modules[(wid, dd)].append((mid, is_flexible))
                assigned_workers.add(wid)

    # worker unavailable map (and filter set)
    worker_off = {}
    workers_with_off = set()
    for wid, w in workers.items():
        off = parse_unavailable_dates(w.get("unavailable_dates"), plan_start, plan_end)
        worker_off[wid] = off
        if off:
            workers_with_off.add(wid)

    # decide which workers to include
    if INCLUDE_ALL_WORKERS:
        worker_ids = list(workers.keys())
    else:
        worker_ids = sorted(assigned_workers.union(workers_with_off))

    # stable sort: company then name
    def company_name_of(wid):
        w = workers.get(wid, {})
        wc = w.get("worker_company")
        return (worker_companies.get(wc, {}) or {}).get("name", wc or "")

    def worker_name_of(wid):
        return (workers.get(wid, {}) or {}).get("name", wid)

    worker_ids.sort(key=lambda wid: (company_name_of(wid), worker_name_of(wid)))

    # today for today_task column
    today = date.today()
    if os.environ.get("EXPORT_TODAY_YMD"):
        try:
            today = _d(os.environ["EXPORT_TODAY_YMD"])
        except Exception:
            pass

    # workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    headers = ["company", "name", "role", "manager", "today_task"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=DOW_ROW, column=j, value=h)   # row 3
        c.font = BOLD_BLACK
        c.alignment = CENTER

        c.border = BORDER_THIN
    # apply meta column widths
    for j, h in enumerate(headers, start=1):
        col_letter = get_column_letter(j)
        if h in META_COL_WIDTHS:
            ws.column_dimensions[col_letter].width = META_COL_WIDTHS[h]
    start_col_dates = len(headers) + 1

    # --- 3-row date header: Month / Day / Weekday(kanji) ---
    dow_kanji = ["月", "火", "水", "木", "金", "土", "日"]

    # row 2 + 3: day + weekday
    for k, dd in enumerate(dates):
        col = start_col_dates + k
        dt = datetime(dd.year, dd.month, dd.day)

        # Day row (value is real date; display only day)
        c_day = ws.cell(row=DAY_ROW, column=col, value=dt)
        c_day.number_format = "dd"
        c_day.font = BOLD_BLACK
        c_day.alignment = CENTER
        c_day.border = BORDER_THIN

        # Weekday row (kanji)
        c_dow = ws.cell(row=DOW_ROW, column=col, value=dow_kanji[dd.weekday()])
        c_dow.font = BOLD_BLACK
        c_dow.alignment = CENTER
        c_dow.border = BORDER_THIN

        ws.column_dimensions[get_column_letter(col)].width = DATE_COL_WIDTH

    # row 1: Month labels, merged across same month
    # Example: "1月" spanning all columns of that month
    month_start_k = 0
    while month_start_k < len(dates):
        m = dates[month_start_k].month
        y = dates[month_start_k].year
        month_end_k = month_start_k
        while month_end_k + 1 < len(dates) and dates[month_end_k + 1].month == m and dates[month_end_k + 1].year == y:
            month_end_k += 1

        c1 = start_col_dates + month_start_k
        c2 = start_col_dates + month_end_k

        label = f"{m}月"
        c_month = ws.cell(row=MONTH_ROW, column=c1, value=label)
        c_month.font = BOLD_BLACK
        c_month.alignment = CENTER
        c_month.border = BORDER_THIN

        # fill borders for merged area
        for cc in range(c1, c2 + 1):
            ws.cell(row=MONTH_ROW, column=cc).border = BORDER_THIN

        if c2 > c1:
            ws.merge_cells(start_row=MONTH_ROW, start_column=c1, end_row=MONTH_ROW, end_column=c2)

        month_start_k = month_end_k + 1

    # header row heights
    ws.row_dimensions[MONTH_ROW].height = HEADER_ROW_HEIGHT
    ws.row_dimensions[DAY_ROW].height = HEADER_ROW_HEIGHT
    ws.row_dimensions[DOW_ROW].height = HEADER_ROW_HEIGHT


    ws.freeze_panes = get_column_letter(start_col_dates) + str(DATA_START_ROW)

    # color maps
    company_color = {}
    module_color = {}

    for r, wid in enumerate(worker_ids, start=DATA_START_ROW):
        wcfg = workers.get(wid, {}) or {}
        wname = wcfg.get("name", wid)
        is_mgr = bool(wcfg.get("is_manager", False))
        wrole = str(wcfg.get("role", "") or "")
        wc_id = wcfg.get("worker_company")
        wc_name = (worker_companies.get(wc_id, {}) or {}).get("name", wc_id or "")

        # company label color (consistent per company)
        # company label color (consistent per company)
        comp_key = wc_name or wc_id or ""
        comp_fill = company_fill(comp_key)

        # left columns (create cells first)
        cA = ws.cell(row=r, column=1, value=wc_name)
        cB = ws.cell(row=r, column=2, value=wname)
        cRole = ws.cell(row=r, column=3, value=wrole)
        cC = ws.cell(row=r, column=4, value="Yes" if is_mgr else "")
        cD = ws.cell(row=r, column=5, value="")

        # style left columns
        for c in (cA, cB, cRole, cC, cD):
            c.border = BORDER_THIN
            c.font = BOLD_BLACK if c in (cA, cB) else BLACK
            c.alignment = CENTER if c in (cA, cC) else LEFT

        # apply company color ONLY to company + name columns (A,B)
        if comp_fill:
            for c in (cA, cB):
                c.fill = comp_fill
                c.font = Font(bold=c.font.bold, color="000000")

        # manager column should be blank fill (no company color)
        # today_task column will be colored later by today's task fill (not company color)

        # grid cells
        todays_text = ""
        todays_fill = None
        todays_font = None
        for k, dd in enumerate(dates):
            col = start_col_dates + k
            cell = ws.cell(row=r, column=col, value="")
            cell.border = BORDER_THIN
            cell.alignment = LEFT
            cell.font = BLACK

            # unavailable -> RED
            if dd in worker_off.get(wid, set()):
                cell.fill = UNAV_FILL
                continue

            entries = cell_modules.get((wid, dd), [])
            if not entries:
                continue

            mid, is_flexible = entries[0]  # first only

            # module name
            if str(mid).startswith("UNKNOWN::"):
                module_name = str(mid).replace("UNKNOWN::", "")
            else:
                module_name = module_name_by_id.get(mid, mid)

            cell.value = module_name

            # ---- PB override: one fixed grey color ----
            wf_id = module_workflow_by_id.get(mid, "")
            if wf_id == PB_WORKFLOW_ID:
                cell.fill = PB_FILL
                cell.font = PB_FONT
                if dd == today:
                    todays_text = module_name
                    todays_fill = PB_FILL
                    todays_font = PB_FONT
                continue
            # ------------------------------------------

            # background fill: stable by workflow_task_list order (NOT affected by flexibility)
            task_fill = module_fill_by_id.get(mid, unknown_fill)
            cell.fill = task_fill

            # flexible => only text becomes white if enabled
            cell_font = (FLEX_FONT_WHITE if (is_flexible and FLEX_WHITE_TEXT_MODE) else BLACK)
            cell.font = cell_font

            if dd == today:
                todays_text = module_name
                todays_fill = task_fill 
                todays_font = cell_font 

        if todays_text:
            cD.value = todays_text
            if todays_fill:
                cD.fill = todays_fill
                cD.font = todays_font or BLACK


        ws.row_dimensions[r].height = DATA_ROW_HEIGHT
    last_row = ws.max_row
    last_col = start_col_dates + len(dates) - 1
    ws.auto_filter.ref = f"A{DOW_ROW}:{get_column_letter(last_col)}{last_row}"

    build_search_sheet(
        wb=wb,
        source_ws=ws,
        keyword=SEARCH_KEYWORD,
        headers_count=len(headers),
        start_col_dates=start_col_dates,
        dates=dates,
        meta_col_widths=META_COL_WIDTHS,
        header_rows=HEADER_ROWS
    )

    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python yaml_to_suother_like_schedule.py EnvConfig.yaml Schedule.yaml out.xlsx")
        sys.exit(2)

    main(sys.argv[1], sys.argv[2], sys.argv[3])