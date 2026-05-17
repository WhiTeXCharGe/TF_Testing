# export_schedule_excel.py
# ---------------------------------------------------------------------
# Builds an Excel with four sheets from Schedule.yaml + EnvConfig.yaml
# SHEET 1 -> "Tasks x Dates" (meta cols + required/assigned per task,
#             per-day cells highlight: start/lightblue, deadline/red,
#             window breach/purple, ordering breach/blue, staffing minmax/pink.
#             assigned_hours cell yellow if under-assigned; shows per-day manager tag ★[AB,CD])
#             NEW: grey for weekends & fab/region/customer off; brown for "assigned on closed day" breach.
# SHEET 2 -> "Employees x Dates" (company | employee | skills,
#             per-day cells highlight: skill mismatch/orange)
#             NEW: grey for personal & worker-company off; brown for "assigned on personal/company off" breach.
# SHEET 3 -> "Dashboard" (KPIs + tables + CHARTS, including Team Quality by Block + Scatter)
# SHEET 4 -> "Breaches" (window / ordering / skill mismatch / minmax)
#             NEW: Unavailable-date breaches for Tasks and Employees.
# ---------------------------------------------------------------------
# # OUT marker (day after segment end)
# out_mark_day = seg_end + timedelta(days=1)
# if (plan_start <= out_mark_day <= plan_end) and (not _is_weekend(out_mark_day)):
#     move_markers[wname].append({"type": "out", "date": out_mark_day, "region": cur_r})

# # ...

# # IN marker (day before next segment start)
# in_mark_day = next_start - timedelta(days=1)
# if (plan_start <= in_mark_day <= plan_end) and (not _is_weekend(in_mark_day)):
#     move_markers[wname].append({"type": "in", "date": in_mark_day, "region": next_r})
import os
import math
import yaml
import re
try:
    from yaml import CSafeLoader as _YamlLoader
except Exception:
    _YamlLoader = None
from argparse import ArgumentParser
from datetime import date, datetime, timedelta
from collections import defaultdict, Counter
from itertools import combinations

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series  # NEW: ScatterChart
from openpyxl.worksheet.table import Table, TableStyleInfo
# ------------------------------- COLORS --------------------------------
LIGHT_BLUE   = "ADD8E6"   # start marker (not a violation)
RED          = "FF9999"   # deadline marker (not a violation)

PURPLE_WIN   = "E6B8F7"   # phase window breach (violation)
BLUE_ORDER   = "9DC3E6"   # phase ordering breach (violation)
PINK_MINMAX  = "F6B5C9"   # staffing min/max breach (violation)
YELLOW_UNDER = "FFF2CC"   # under-assigned task (violation) -> assigned_hours cell
ORANGE_SKILL = "F8CBAD"   # skill mismatch (violation) -> sheet2 per-day cells
MANAGER_MISSING = "FFE699" # highlight 'manager' cell when a task has no manager

# NEW (calendar visuals)
GREY_CLOSED   = "DDDDDD"  # closed day shading (weekend/unavailable)
BROWN_UNAV_BREACH = "C79D6B"  # assignment placed on a closed day (breach)

# ------------------------------- CONFIG --------------------------------
# Required hours from Schedule.yaml only:
#   if workload_hours present -> use directly
#   if workload_days present  -> multiply by UNIT_HOUR
REQUIRED_HOURS_MODE = True

OT_THRESHOLD = 8
CAP_HOURS    = 12
UNIT_HOUR    = 10  # hours per workload_day (used when workload_days is given)

# Team quality scoring constants
BALANCE_K = 25.0  # higher = stricter penalty per dev from op_avg_skill
# 0..PREF_MAX_LEVEL levels, 0 = NG, higher = more preferred
PREF_MAX_LEVEL = 3


# ------------------------------ UTILITIES ------------------------------
def _d(s):
    """Parse 'YYYY/MM/DD' or 'YYYY-MM-DD' to date."""
    if isinstance(s, date):
        return s
    s = str(s).replace("-", "/")
    return datetime.strptime(s, "%Y/%m/%d").date()

def _hfmt(h):
    return f"{int(h)}H"

def _safe(dct, key):
    try:
        return dct[key]
    except Exception:
        return None
def _to_int_or_default(v, default=0):
    """Convert to int, with safe default."""
    if v is None:
        return default
    try:
        return int(v)
    except Exception:
        return default

def has_skill_for_assignment(skills: dict, module_id: str, op_id: str) -> bool:
    """
    Support multiple skill_map key styles.
    Returns True if the worker has a positive skill level for this op.

    Supported keys (checked in this order):
      1) op_id only                    -> e.g. 'p1', 'p1o1', 'other_op'
      2) module_id + op_id             -> e.g. 'f22_3p1o1'
      3) module base + op_id           -> e.g. 'f22p1o1'  (when module_id like 'f22_3')
    """
    if not skills:
        return False

    # 1) Plain op id
    lvl = skills.get(op_id)
    if _to_int_or_default(lvl, 0) > 0:
        return True

    # 2) Module + op id (exact)
    if module_id:
        key2 = f"{module_id}{op_id}"
        lvl2 = skills.get(key2)
        if _to_int_or_default(lvl2, 0) > 0:
            return True

        # 3) Module base (before first underscore) + op id
        base = str(module_id).split("_")[0]
        if base and base != module_id:
            key3 = f"{base}{op_id}"
            lvl3 = skills.get(key3)
            if _to_int_or_default(lvl3, 0) > 0:
                return True

    return False

    # 1) Plain op id, e.g. 'p1o1'
    lvl = skills.get(op_id)
    if _to_int_or_default(lvl, 0) > 0:
        return True

    # 2) Module + op id, e.g. 'm1p1o1'
    if module_id:
        key2 = f"{module_id}{op_id}"
        lvl2 = skills.get(key2)
        if _to_int_or_default(lvl2, 0) > 0:
            return True

    return False
 
def _initials(name: str) -> str:
    if not name:
        return ""
    parts = [p for p in str(name).strip().split() if p]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][0] + parts[-1][0]).upper()

def _is_weekend(d: date) -> bool:
    return d.weekday() >= 5  # 5=Sat, 6=Sun


def _normalize_pref_level(level, max_level=PREF_MAX_LEVEL):
    x = _to_int_or_default(level, default=0)
    if x < 0:
        x = 0
    if x > max_level:
        x = max_level
    return (x / max_level) if max_level > 0 else 0.0

def compute_preference_score(region_level, company_level, max_level=PREF_MAX_LEVEL):
    """
    Return a 0..100 preference score combining region(country) and company preferences.
    - If both are high -> close to 100
    - If one is 0 -> middle/low
    - If both 0 -> 0
    """
    r = _normalize_pref_level(region_level, max_level)
    c = _normalize_pref_level(company_level, max_level)
    return round(100.0 * (r + c) / 2.0, 1)


def _norm_flex(x):
    s = str(x or "").strip().lower()
    if s == "fixed":
        return "fix"
    if s == "flexible":
        return "flex"
    return ""


def _merge_scope(values):
    vals = {v for v in values if v in ("fix", "flex")}
    if not vals:
        return ""
    if len(vals) >= 2:
        return "both"
    return next(iter(vals))


def _scope_matches(page_scope, row_scope):
    return row_scope == page_scope or row_scope == "both"

# ------------------------------- LOADERS -------------------------------
def load_env(env_path):
    with open(env_path, "r", encoding="utf-8") as f:
        env = (yaml.load(f, Loader=_YamlLoader) if _YamlLoader else yaml.safe_load(f))
    env = env.get("environment", {})

    workflows = {w["id"]: w for w in env.get("workflow_list", [])}
    fabs      = {f["id"]: f for f in env.get("fab_list", [])}
    regions   = {r["id"]: r for r in env.get("region_list", [])}
    customers = {c["id"]: c for c in env.get("customer_company_list", [])}
    wcompanies= {c["id"]: c for c in env.get("worker_company_list", [])}
    workers   = {w["id"]: w for w in env.get("worker_list", [])}
    

    # operation meta: workflow_id -> phase_id -> op_id -> meta
    op_meta = defaultdict(lambda: defaultdict(dict))
    for wf in env.get("workflow_list", []):
        for ph in wf.get("phase_list", []):
            for op in ph.get("operation_list", []):
                op_meta[wf["id"]][ph["id"]][op["id"]] = {
                    "name": op.get("name", op["id"]),
                    "work_hours": op.get("work_hours", [8,10,12]),
                    "min_worker_num": op.get("min_worker_num", 1),
                    "max_worker_num": op.get("max_worker_num", 99),
                }
    affinity_tags = {t["id"]: t.get("weight", 0) for t in env.get("affinity_tag", [])}

    return {
        "workflows": workflows,
        "fabs": fabs,
        "regions": regions,
        "customers": customers,
        "workers": workers,
        "worker_companies": wcompanies,
        "op_meta": op_meta,
        "transite_day_map": env.get("transite_day_map", []),
        "affinity_tags": affinity_tags,
    }

def load_schedule(path):
    with open(path, "r", encoding="utf-8") as f:
        s = (yaml.load(f, Loader=_YamlLoader) if _YamlLoader else yaml.safe_load(f))
    s = s.get("schedule", s)

    plan_start = _d(s["plan_range"]["start_date"])
    plan_end   = _d(s["plan_range"]["end_date"])
    modules    = s.get("workflow_task_list", [])
    asg_raw    = s.get("assignment_list", [])

    # --------- expand to per-day assignments ----------
    assignments = []
    # temp holder to merge blocks
    per_key_dates = defaultdict(list)   # (worker, op_task, flex) -> [dates]
    for a in asg_raw:
        wd_key = "work_date_lsit" if "work_date_lsit" in a else "work_date_list"
        op_task = a["operation_task"]
        wid     = a["worker"]
        flex    = str(a.get("plan_flexibility", "") or "").strip()
        wdates = []
        for ditem in a.get(wd_key, []) or []:
            d = _d(ditem["date"])
            h = int(ditem["hour"])
            assignments.append({
                "worker": wid,
                "operation_task": op_task,
                "date": d,
                "hours": h,
                "plan_flexibility": flex,
            })
            wdates.append(d)
        per_key_dates[(wid, op_task, flex)].extend(wdates)

    # --------- COALESCE assignment blocks by (worker, op, flex) ----------
    assignment_blocks = []
    for (wid, op_task, flex), dates in per_key_dates.items():
        if not dates:
            continue
        dates = sorted(set(dates))
        sd, ed = dates[0], dates[-1]
        assignment_blocks.append({
            "worker": wid,
            "operation_task": op_task,
            "start_date": sd,
            "end_date": ed,
            "work_dates": dates,
            "plan_flexibility": flex,
        })

    return plan_start, plan_end, modules, assignments, assignment_blocks


# -------------------------  UNAVAILABILITY CAL ---------------------
def parse_unavailable_dates(raw, plan_start: date, plan_end: date):
    """
    Python port of EmployeeSchedule.parseUnavailableDates():
    Supports shapes like:
      - unavailable_dates: 2025/09/10
      - unavailable_dates: [2025/09/10, 2025/09/11]
      - unavailable_dates:
          single:
            days: [2025/09/10, 2025/09/11]
      - unavailable_dates:
          weekly:
            weekdays: [sat, sun]
      - unavailable_dates:
        - { single: { days: [...] } }
        - { weekly: { weekdays: [...] } }

    Returns a set[date] within [plan_start, plan_end].
    """
    off = set()
    if raw is None:
        return off

    # horizon dates (inclusive)
    horizon = (plan_end - plan_start).days + 1
    all_dates = [plan_start + timedelta(days=i) for i in range(horizon)]

    def add_single(obj):
        """Add one scalar date if valid."""
        try:
            d = _d(obj)
        except Exception:
            return
        if plan_start <= d <= plan_end:
            off.add(d)

    # weekday parser (mon/tue/... or full name)
    def parse_weekday(name: str):
        t = str(name).strip().lower()
        if t in ("mon", "monday"):
            return 0
        if t in ("tue", "tues", "tuesday"):
            return 1
        if t in ("wed", "weds", "wednesday"):
            return 2
        if t in ("thu", "thur", "thurs", "thursday"):
            return 3
        if t in ("fri", "friday"):
            return 4
        if t in ("sat", "saturday"):
            return 5
        if t in ("sun", "sunday"):
            return 6
        return None

    # Normalise raw into a list
    if isinstance(raw, list):
        items = raw
    elif isinstance(raw, dict):
        # top-level {single:{...}} or {weekly:{...}}
        items = [raw]
    else:
        # scalar -> one date string
        add_single(raw)
        return off

    weekly_off = set()  # set of weekday ints (0=Mon .. 6=Sun)

    for item in items:
        if item is None:
            continue

        if isinstance(item, dict):
            # --- single: { days: [...] } ---
            single_obj = item.get("single")
            if isinstance(single_obj, dict):
                days_obj = single_obj.get("days")
                if isinstance(days_obj, list):
                    for d_obj in days_obj:
                        add_single(d_obj)

            # --- weekly: { weekdays: [sat, sun] } ---
            weekly_obj = item.get("weekly")
            if isinstance(weekly_obj, dict):
                wdays_obj = weekly_obj.get("weekdays")
                if isinstance(wdays_obj, list):
                    for wd in wdays_obj:
                        dow = parse_weekday(wd)
                        if dow is not None:
                            weekly_off.add(dow)
        else:
            # plain scalar inside list -> treat as date string
            add_single(item)

    # Expand weekly patterns into actual dates
    if weekly_off:
        for d in all_dates:
            if d.weekday() in weekly_off:
                off.add(d)

    return off


def build_unavailability(plan_start: date, plan_end: date, env):
    """
    Build date sets for weekends and unavailability by fab/region/customer/worker-company + personal worker off.

    NOTE:
    - 'weekends' here is still the simple global Sat/Sun visual shading.
    - fab_off / region_off / customer_off / worker_company_off / worker_off
      use the same richer 'unavailable_dates' schema as EmployeeSchedule.buildCalendars().
    """
    def daterange(start: date, end: date):
        d = start
        while d <= end:
            yield d
            d += timedelta(days=1)

    # Global visual weekend (Sat/Sun) – used for generic grey background & "weekend" reason.
    weekends = {d for d in daterange(plan_start, plan_end) if d.weekday() >= 5}

    fabs      = env["fabs"]
    regions   = env["regions"]
    customers = env["customers"]
    workers   = env["workers"]
    wcompanies= env["worker_companies"]

    fab_off            = defaultdict(set)
    region_off         = defaultdict(set)
    customer_off       = defaultdict(set)
    worker_company_off = defaultdict(set)
    worker_off         = defaultdict(set)

    fab_region   = {}
    fab_customer = {}

    # ---- fab -> off + region/customer maps ----
    for fid, f in fabs.items():
        fab_region[fid]   = f.get("region")
        fab_customer[fid] = f.get("customer_company")
        raw = f.get("unavailable_dates")
        fab_off[fid] = parse_unavailable_dates(raw, plan_start, plan_end)

    # ---- region off ----
    for rid, r in regions.items():
        raw = r.get("unavailable_dates")
        region_off[rid] = parse_unavailable_dates(raw, plan_start, plan_end)

    # ---- customer off ----
    for cid, c in customers.items():
        raw = c.get("unavailable_dates")
        customer_off[cid] = parse_unavailable_dates(raw, plan_start, plan_end)

    # ---- worker company off + OT limits (limits only used on Java side) ----
    for wcid, c in wcompanies.items():
        raw = c.get("unavailable_dates")
        worker_company_off[wcid] = parse_unavailable_dates(raw, plan_start, plan_end)

    # ---- personal worker off ----
    for wid, w in workers.items():
        raw = w.get("unavailable_dates")
        worker_off[wid] = parse_unavailable_dates(raw, plan_start, plan_end)

    return {
        "weekends": weekends,
        "fab_off": dict(fab_off),
        "region_off": dict(region_off),
        "customer_off": dict(customer_off),
        "worker_company_off": dict(worker_company_off),
        "worker_off": dict(worker_off),
        "fab_region": fab_region,
        "fab_customer": fab_customer,
    }

def build_op_task_index(modules):
    """
    Map Schedule.yaml operation_task_list.id  ->  meta we need for excel
    Example: 'f22_3_p3' -> {m_id:'f22_3', phase:'tool_p3', op_id:'f22p3', op_name:'Phase 3'}
    """
    idx = {}
    for m in modules:
        m_id = m["id"]
        for ph in m.get("phase_task_list", []):
            ph_id = ph.get("phase")
            for ot in ph.get("operation_task_list", []):
                idx[ot["id"]] = {
                    "m_id": m_id,
                    "phase": ph_id,
                    "op_id": ot.get("operation"),
                    "op_name": ot.get("name", "") or ot.get("operation", ""),
                    # Optional per-task staffing override (Schedule.yaml)
                    "min_worker_num": ot.get("min_worker_num"),
                    "max_worker_num": ot.get("max_worker_num"),
                    # NEW: recommended staffing range (soft)
                    "recommends_worker_min": ot.get("recommends_worker_min"),
                    "recommends_worker_max": ot.get("recommends_worker_max"),
                }
    return idx


def parse_op_task_ids(op_task: str, op_task_index: dict):
    """Return (module_id, op_id) for an operation_task string, robust across dataset formats."""
    if not op_task:
        return ("", "")
    meta = op_task_index.get(op_task) if op_task_index else None
    if meta:
        return (meta.get("m_id", "") or "", meta.get("op_id", "") or "")
    # Fallbacks:
    # - dataset9 style: 'f22_3_p3' -> module 'f22_3', op 'f22p3' or 'p3'
    m = re.match(r"^(.*)_p(\d+)$", str(op_task))
    if m:
        mid = m.group(1)
        op = f"{mid.split('_')[0]}p{m.group(2)}"  # best-effort
        return (mid, op)
    # - legacy style without underscores: 'm1p1o1' -> split at first 'p'
    s = str(op_task)
    i = s.find("p")
    if i > 0:
        return (s[:i], s[i:])
    return (s, "")
# ---------------------------- AGGREGATIONS -----------------------------
def build_maps(env, modules, assignments):
    workers   = env["workers"]
    fabs      = env["fabs"]
    regions   = env["regions"]
    customers = env["customers"]
    wcompanies= env["worker_companies"]
    workers   = env["workers"]

    mod_map = {m["id"]: m for m in modules}

    # module start & phase end
    module_start = {}
    phase_end = {}  # (module_id, phase_id) -> date
    op_phase_of_module = {}  # (module_id, op_id) -> phase_id
    
    op_task_index = build_op_task_index(modules)

    # Per-(module, op) staffing overrides from Schedule.yaml (min/max can be None)
    schedule_minmax = {}  # (m_id, op_id) -> (min_worker_num, max_worker_num)
    for _ot_id, meta in op_task_index.items():
        m_id0 = meta.get("m_id")
        op_id0 = meta.get("op_id")
        if not m_id0 or not op_id0:
            continue
        mn = meta.get("min_worker_num")
        mx = meta.get("max_worker_num")
        if mn is not None or mx is not None:
            schedule_minmax[(m_id0, op_id0)] = (mn, mx)

    module_workflow = {m["id"]: (m.get("workflow") or m.get("workflow_id")) for m in modules}
    for m in modules:
        starts = []
        for ph in m.get("phase_task_list", []):
            p_start = _d(ph["start_date"])
            p_end   = _d(ph["end_date"])
            starts.append(p_start)
            phase_end[(m["id"], ph["phase"])] = p_end
            for ot in ph.get("operation_task_list", []):
                op_phase_of_module[(m["id"], ot["operation"])] = ph["phase"]
        module_start[m["id"]] = min(starts) if starts else None

    # worker info
    worker_name = {wid: workers[wid].get("name", wid) for wid in workers}
    worker_company_name = {
        wid: wcompanies.get(workers[wid].get("worker_company"), {}).get(
            "name", workers[wid].get("worker_company")
        )
        for wid in workers
    }
    worker_company_id = {wid: workers[wid].get("worker_company") for wid in workers}
    worker_skills = {wid: workers[wid].get("skill_map", {}) for wid in workers}
    worker_is_manager = {wid: bool(workers[wid].get("is_manager", False)) for wid in workers}
    name_is_manager   = {worker_name[wid]: worker_is_manager[wid] for wid in workers}
    worker_region_pref = {}
    worker_company_pref = {}
    for wid, cfg in workers.items():
        rmap = {}
        cmap = {}
        for item in cfg.get("fab_suitability_map") or []:
            kind = item.get("kind")
            smap = item.get("suitability") or {}
            if kind == "region":
                rmap.update(smap)
            elif kind == "customer_company":
                cmap.update(smap)
        worker_region_pref[wid] = rmap
        worker_company_pref[wid] = cmap

    # module metadata for sheet 1
    mod_meta_cols = {}
    for m in modules:
        fab_id = m.get("fab", "")
        fab    = fabs.get(fab_id, {})
        region = regions.get(fab.get("region"), {})
        cust   = customers.get(fab.get("customer_company"), {})
        mod_meta_cols[m["id"]] = {
            "module": m["id"],
            "module_name": m.get("name",""),
            "fab_id": fab_id,
            "fab_name": fab.get("name",""),
            "region": region.get("name",""),
            "customer": cust.get("name",""),
        }

    # catalog rows: (module, op)
    module_ops = []
    for m in modules:
        for ph in m.get("phase_task_list", []):
            for ot in ph.get("operation_task_list", []):
                module_ops.append((m["id"], ot["operation"], ot.get("name","")))

    # map: per-day strings for Tasks x Dates
    tde = defaultdict(list)  # (module_id, op_id, date) -> ["AA(12H)" ...]
    # map: per-day for Employees x Dates
    edt = defaultdict(list)  # (company, worker_name, date) -> ["e1p1o2 (12H)"]
    # stats
    emp_total_hours = Counter()
    emp_workdays    = Counter()
    per_day_total   = Counter()
    day_op_heads    = defaultdict(int)  # (module_id, op_id, date) -> heads

    # also keep a per-cell list of assignments to evaluate violations later
    per_cell_assigns_task = defaultdict(list)     # key: (m_id, op_id, date) -> list of dicts
    per_cell_assigns_emp  = defaultdict(list)     # key: (comp, wname, date) -> list of dicts

    for a in assignments:
        # NOTE: no more global weekend skip here. All work days go into maps.
        wid   = a["worker"]
        wname = worker_name.get(wid, wid)
        comp  = worker_company_name.get(wid, "")
        ot_id = a["operation_task"]              # e.g. f22_3_p3
        meta  = op_task_index.get(ot_id)

        if meta:
            m_id  = meta["m_id"]                 # f22_3
            op_id = meta["op_id"]                # f22p3
        else:
            # fallback for old datasets (keep your old split if you want)
            idx   = ot_id.find("p")
            m_id  = ot_id[:idx] if idx > 0 else ot_id
            op_id = ot_id[idx:] if idx > 0 else ""

        # sheet1 strings
        tde[(m_id, op_id, a["date"])].append(
            f'{wname}({_hfmt(a["hours"])})'
        )
        per_cell_assigns_task[(m_id, op_id, a["date"])].append({
            "worker": wid, "wname": wname, "hours": a["hours"]
        })

        # sheet2 strings
        edt[(comp, wname, a["date"])].append(f'{m_id}{op_id} ({_hfmt(a["hours"])})')
        per_cell_assigns_emp[(comp, wname, a["date"])].append({
            "worker": wid, "m_id": m_id, "op_id": op_id, "hours": a["hours"]
        })

        emp_total_hours[wname] += a["hours"]
        per_day_total[a["date"]] += a["hours"]
        day_op_heads[(m_id, op_id, a["date"])] += 1

    managers_by_task = defaultdict(set)  # (m_id, op_id) -> {manager_names}
    for a in assignments:
        wid = a["worker"]
        if worker_is_manager.get(wid, False):
            ot_id = a["operation_task"]
            meta = op_task_index.get(ot_id)
            if not meta:
                continue
            m_id  = meta["m_id"]
            op_id = meta["op_id"]
            managers_by_task[(m_id, op_id)].add(worker_name.get(wid, wid))

    # employee workdays
    for (comp, wname, d), items in edt.items():
        if items:
            emp_workdays[wname] += 1

    return {
        "mod_map": mod_map,
        "module_start": module_start,
        "phase_end": phase_end,
        "op_phase_of_module": op_phase_of_module,
        "op_task_index": op_task_index,
        "schedule_minmax": schedule_minmax,
        "module_workflow": module_workflow,
        "mod_meta_cols": mod_meta_cols,
        "module_ops": module_ops,
        "tde": tde,
        "edt": edt,
        "worker_name": worker_name,
        "worker_company_name": worker_company_name,
        "worker_company_id": worker_company_id,            # NEW
        "worker_skills": worker_skills,
        "worker_is_manager": worker_is_manager,
        "name_is_manager": name_is_manager,
        "emp_total_hours": emp_total_hours,
        "emp_workdays": emp_workdays,
        "per_day_total": per_day_total,
        "day_op_heads": day_op_heads,
        "per_cell_assigns_task": per_cell_assigns_task,
        "per_cell_assigns_emp": per_cell_assigns_emp,
        "managers_by_task": managers_by_task,
        "worker_region_pref": worker_region_pref,
        "worker_company_pref": worker_company_pref,
    }

def build_preference_map_rows(env, maps):
    """
    Build rows to show each employee's preference map.

    Returns:
      rows_region:  list of dicts with keys
        employee, company, region_id, region_name, level
      rows_company: list of dicts with keys
        employee, company, customer_id, customer_name, level
    """
    workers   = env.get("workers", {})
    regions   = env.get("regions", {})
    customers = env.get("customers", {})

    worker_region_pref  = maps.get("worker_region_pref", {})
    worker_company_pref = maps.get("worker_company_pref", {})
    worker_name         = maps["worker_name"]
    worker_company_name = maps["worker_company_name"]

    region_ids   = sorted(regions.keys())
    customer_ids = sorted(customers.keys())

    rows_region = []
    for wid, cfg in workers.items():
        wname     = worker_name.get(wid, cfg.get("name", wid))
        comp_name = worker_company_name.get(wid, "")
        r_map     = worker_region_pref.get(wid, {}) or {}
        for rid in region_ids:
            region_cfg = regions.get(rid, {}) or {}
            lvl = _to_int_or_default(r_map.get(rid), default=0)
            rows_region.append({
                "employee":    wname,
                "company":     comp_name,
                "region_id":   rid,
                "region_name": region_cfg.get("name", rid),
                "level":       lvl,
            })

    rows_region.sort(key=lambda r: (r["employee"], r["region_id"]))

    rows_company = []
    for wid, cfg in workers.items():
        wname     = worker_name.get(wid, cfg.get("name", wid))
        comp_name = worker_company_name.get(wid, "")
        c_map     = worker_company_pref.get(wid, {}) or {}
        for cid in customer_ids:
            cust_cfg = customers.get(cid, {}) or {}
            lvl = _to_int_or_default(c_map.get(cid), default=0)
            rows_company.append({
                "employee":       wname,
                "company":        comp_name,
                "customer_id":    cid,
                "customer_name":  cust_cfg.get("name", cid),
                "level":          lvl,
            })

    rows_company.sort(key=lambda r: (r["employee"], r["customer_id"]))
    return rows_region, rows_company

def build_preference_match_rows(env, modules, assignments, maps):
    """
    Build rows for 'Preference match (per employee x fab)'.

    One row per (worker, fab) with:
      - employee, company
      - fab_id, fab_name, region, customer_company
      - region_pref, company_pref (0..3)
      - pref_score (0..100)
      - assign_days, assign_hours, modules list
    """
    fabs      = env.get("fabs", {})
    regions   = env.get("regions", {})
    customers = env.get("customers", {})

    worker_region_pref  = maps.get("worker_region_pref", {})
    worker_company_pref = maps.get("worker_company_pref", {})
    worker_name         = maps["worker_name"]
    worker_company_name = maps["worker_company_name"]
    mod_meta_cols       = maps["mod_meta_cols"]

    # aggregate per (worker, fab)
    per_w_fab = {}   # (wid, fab_id) -> agg
    for a in assignments:
        wid     = a["worker"]
        op_task = a["operation_task"]
        m_id, _op_id = parse_op_task_ids(op_task, maps.get("op_task_index", {}))

        meta   = mod_meta_cols.get(m_id, {}) or {}
        fab_id = meta.get("fab_id")
        if not fab_id:
            continue

        key = (wid, fab_id)
        agg = per_w_fab.get(key)
        if not agg:
            agg = {
                "worker": wid,
                "fab_id": fab_id,
                "dates": set(),
                "hours": 0,
                "modules": set(),
            }
            per_w_fab[key] = agg

        agg["dates"].add(a["date"])
        agg["hours"] += a["hours"]
        agg["modules"].add(m_id)

    rows = []
    for (wid, fab_id), agg in per_w_fab.items():
        wname     = worker_name.get(wid, wid)
        comp_name = worker_company_name.get(wid, "")

        fab_cfg = fabs.get(fab_id, {}) or {}
        rid     = fab_cfg.get("region")
        cid     = fab_cfg.get("customer_company")

        region_obj = regions.get(rid, {}) or {}
        cust_obj   = customers.get(cid, {}) or {}

        region_name   = region_obj.get("name", rid)
        customer_name = cust_obj.get("name", cid)

        r_map = worker_region_pref.get(wid, {}) or {}
        c_map = worker_company_pref.get(wid, {}) or {}

        r_lvl = _to_int_or_default(r_map.get(rid), default=0)
        c_lvl = _to_int_or_default(c_map.get(cid), default=0)

        score = compute_preference_score(r_lvl, c_lvl)

        rows.append({
            "employee":       wname,
            "company":        comp_name,
            "fab_id":         fab_id,
            "fab_name":       fab_cfg.get("name", fab_id),
            "region":         region_name,
            "customer":       customer_name,
            "region_pref":    r_lvl,
            "company_pref":   c_lvl,
            "pref_score":     score,
            "assign_days":    len(agg["dates"]),
            "assign_hours":   agg["hours"],
            "modules":        ", ".join(sorted(agg["modules"])),
        })

    # sort by employee, then fab, high score first
    rows.sort(key=lambda r: (r["employee"], r["fab_id"] or "", -r["pref_score"]))
    return rows

# ----------------------- REQUIRED HOURS (task/module) ------------------
def compute_required_hours_task_module(modules):
    """Return:
       - req_task[(m_id, op_id)] = required hours (workload_hours directly, or workload_days * UNIT_HOUR)
       - req_module[m_id]        = sum of its tasks in hours
    """
    req_task   = defaultdict(int)
    req_module = defaultdict(int)
    for m in modules:
        m_id = m["id"]
        total_hours = 0
        for ph in m.get("phase_task_list", []):
            for ot in ph.get("operation_task_list", []):
                if "workload_hours" in ot:
                    hours = int(ot["workload_hours"])
                else:
                    hours = int(ot.get("workload_days", 0)) * UNIT_HOUR
                req_task[(m_id, ot["operation"])] += hours
                total_hours += hours
        req_module[m_id] = total_hours
    return req_task, req_module

# --------------------------- VIOLATION DETECTION -----------------------
def build_effective_phase_window(modules, maps):
    """
    Effective phase window rules:
    - For wf_tool:
        start_date = module phase1 start (earliest phase start in that module)
        end_date   = that phase own end_date
    - For other workflows:
        start_date = phase own start_date
        end_date   = phase own end_date
    Returns:
        dict[(m_id, phase_id)] = (effective_start, effective_end)
    """
    phase_window = {}

    module_workflow = maps.get("module_workflow", {}) or {}

    for m in modules:
        m_id = m.get("id")
        if not m_id:
            continue

        phase_list = list(m.get("phase_task_list", []) or [])
        if not phase_list:
            continue

        wf = module_workflow.get(m_id) or m.get("workflow") or m.get("workflow_id")

        # module earliest phase start
        module_start = None
        for ph in phase_list:
            try:
                ps = _d(ph.get("start_date"))
            except Exception:
                continue
            if module_start is None or ps < module_start:
                module_start = ps

        for ph in phase_list:
            ph_id = ph.get("phase")
            if not ph_id:
                continue
            try:
                own_start = _d(ph.get("start_date"))
                own_end   = _d(ph.get("end_date"))
            except Exception:
                continue

            if wf == "wf_tool":
                eff_start = module_start if module_start else own_start
                eff_end   = own_end
            else:
                eff_start = own_start
                eff_end   = own_end

            phase_window[(m_id, ph_id)] = (eff_start, eff_end)

    return phase_window

def detect_violations(env, modules, assignments, maps, cal, plan_start=None, plan_end=None):
    """Build lookups for coloring and Sheet breach tables (Sheet4/5/6)."""

    # ---------------- Effective Phase windows ----------------
    phase_window = build_effective_phase_window(modules, maps)

    # ---------------- Phase order predecessor map ----------------
    # Works for phases like "tool_p1", "tool_p2", ... (no numeric parsing).
    phase_prev = {}
    for m in modules:
        seq = [ph["phase"] for ph in m.get("phase_task_list", [])]
        for i in range(1, len(seq)):
            phase_prev[(m["id"], seq[i])] = seq[i - 1]

    # ---------------- Derived from assignments ----------------
    last_assigned_in_phase = defaultdict(lambda: defaultdict(lambda: None))  # m_id -> phase -> last_date
    assigned_task = defaultdict(int)  # (m_id, op_id) -> hours
    assigned_mod  = defaultdict(int)  # m_id -> hours

    # ---------------- Calendar breach storage ----------------
    unavail_task_cells = set()  # (m_id, op_id, date)
    unavail_emp_cells  = set()  # (company_name, worker_name, date)
    tbl_unavail_task   = []     # [date, module, op_id, reason]
    tbl_unavail_emp    = []     # [date, worker, company, reason, module, op_id]

    # ---------------- Preference breach storage ----------------
    tbl_pref_breach  = []   # [worker, company, module, fab_id, region, customer_company, reason]
    pref_breach_seen = set()

    op_task_index   = maps["op_task_index"]     # operation_task_list.id -> meta (m_id, phase, op_id, name)
    module_workflow = maps["module_workflow"]   # m_id -> workflow_id

    # ---------------- Fixed/Flexible scope maps ----------------
    task_day_scopes = defaultdict(set)    # (m_id, op_id, iso_date) -> {'fix','flex'}
    task_scopes = defaultdict(set)        # (m_id, op_id) -> {'fix','flex'}
    phase_scopes = defaultdict(set)       # (m_id, phase_id) -> {'fix','flex'}
    emp_day_scopes = defaultdict(set)     # (worker, company, iso_date) -> {'fix','flex'}
    emp_scopes = defaultdict(set)         # (worker, company) -> {'fix','flex'}
    worker_scopes = defaultdict(set)      # worker -> {'fix','flex'}

    # ---------------- Scan assignments once ----------------
    for a in assignments:
        ot_id = a["operation_task"]
        meta_ot = op_task_index.get(ot_id)
        if not meta_ot:
            # unknown op_task id -> skip safely
            continue

        m_id  = meta_ot["m_id"]
        op_id = meta_ot["op_id"]
        d     = a["date"]
        wid   = a["worker"]
        scope = _norm_flex(a.get("plan_flexibility"))

        assigned_task[(m_id, op_id)] += a["hours"]
        assigned_mod[m_id]           += a["hours"]

        # last assigned in its phase (from maps)
        ph = maps["op_phase_of_module"].get((m_id, op_id))
        date_key = d.isoformat()
        task_day_scopes[(m_id, op_id, date_key)].add(scope)
        task_scopes[(m_id, op_id)].add(scope)
        if ph:
            phase_scopes[(m_id, ph)].add(scope)
        prev = last_assigned_in_phase[m_id].get(ph)
        if prev is None or d > prev:
            last_assigned_in_phase[m_id][ph] = d

        # ---------- Unavailable breaches (task-side: fab/region/customer) ----------
        mod_meta = maps["mod_meta_cols"].get(m_id, {}) or {}
        fab_id = mod_meta.get("fab_id")
        rid    = cal["fab_region"].get(fab_id)
        cid    = cal["fab_customer"].get(fab_id)

        reasons = []
        if fab_id and d in cal["fab_off"].get(fab_id, set()):
            reasons.append(f"fab_off({fab_id})")
        if rid and d in cal["region_off"].get(rid, set()):
            reasons.append(f"region_off({rid})")
        if cid and d in cal["customer_off"].get(cid, set()):
            reasons.append(f"customer_off({cid})")

        if reasons:
            unavail_task_cells.add((m_id, op_id, d))
            tbl_unavail_task.append([d.isoformat(), m_id, op_id, ", ".join(reasons)])

        # ---------- Unavailable breaches (employee-side: personal + worker-company) ----------
        wname = maps["worker_name"].get(wid, wid)
        comp_name = maps["worker_company_name"].get(wid, "")
        emp_day_scopes[(wname, comp_name, date_key)].add(scope)
        emp_scopes[(wname, comp_name)].add(scope)
        worker_scopes[wname].add(scope)
        wco = maps.get("worker_company_id", {}).get(wid)

        reasons_e = []
        if d in cal["worker_off"].get(wid, set()):
            reasons_e.append("personal_off")
        if wco and d in cal["worker_company_off"].get(wco, set()):
            reasons_e.append(f"worker_company_off({wco})")

        if reasons_e:
            unavail_emp_cells.add((comp_name, wname, d))
            tbl_unavail_emp.append([d.isoformat(), wname, comp_name, ", ".join(reasons_e), m_id, op_id])

        # ---------- Preference breach (suitability level 0) ----------
        region_pref_map  = maps.get("worker_region_pref", {}).get(wid, {}) or {}
        company_pref_map = maps.get("worker_company_pref", {}).get(wid, {}) or {}

        r_lvl = region_pref_map.get(rid, None)
        c_lvl = company_pref_map.get(cid, None)

        reason_tags = []
        if r_lvl is not None and _to_int_or_default(r_lvl, default=0) == 0:
            reason_tags.append("country/region=0")
        if c_lvl is not None and _to_int_or_default(c_lvl, default=0) == 0:
            reason_tags.append("company=0")

        if reason_tags:
            reason_str = " & ".join(reason_tags)
            key = (wid, m_id, reason_str)
            if key not in pref_breach_seen:
                tbl_pref_breach.append([
                    wname,
                    comp_name,
                    m_id,
                    fab_id or "",
                    rid or "",
                    cid or "",
                    reason_str,
                ])
                pref_breach_seen.add(key)

    # ---------------- Staffing min/max, window, ordering ----------------
    op_meta = env["op_meta"]

    win_breach_cells    = set()
    order_breach_cells  = set()
    minmax_breach_cells = set()
    skill_mismatch_cells = set()

    tbl_window = []  # [date, module, phase, op, worker, reason, phase_start, phase_end]
    tbl_order  = []  # [date, module, later_phase, op, worker, required_prev_phase_last_date]
    tbl_skill  = []  # [date, worker, company, module, op_id]
    tbl_minmax = []  # [date, module, op, heads, min, max, status]

    # --- window & ordering & minmax (sheet1) ---
    for (m_id, op_id, dt), heads in maps["day_op_heads"].items():
        ph = maps["op_phase_of_module"].get((m_id, op_id))

        # window breach
        if ph:
            start, end = phase_window.get((m_id, ph), (None, None))
            if (start and dt < start) or (end and dt > end):
                win_breach_cells.add((m_id, op_id, dt))
                for A in maps["per_cell_assigns_task"].get((m_id, op_id, dt), []):
                    tbl_window.append([
                        dt.isoformat(),
                        m_id,
                        ph,
                        op_id,
                        A["wname"],
                        "early" if (start and dt < start) else "late",
                        start.isoformat() if start else "",
                        end.isoformat() if end else "",
                    ])

            # ordering breach using phase_prev map
            prev_phase = phase_prev.get((m_id, ph))
            if prev_phase:
                prev_last = last_assigned_in_phase[m_id].get(prev_phase)
                if prev_last and dt <= prev_last:
                    order_breach_cells.add((m_id, op_id, dt))
                    for A in maps["per_cell_assigns_task"].get((m_id, op_id, dt), []):
                        tbl_order.append([
                            dt.isoformat(),
                            m_id,
                            ph,
                            op_id,
                            A["wname"],
                            prev_last.isoformat(),
                        ])

        # min/max staffing (workflow differs per module!)
        wf_id = module_workflow.get(m_id)
        meta = (op_meta.get(wf_id, {}) if wf_id else {}).get(ph, {}).get(op_id, {}) or {}

        # Prefer per-task min/max from Schedule.yaml when provided; otherwise fall back to EnvConfig workflow defaults
        sched_mnmx = maps.get("schedule_minmax", {}).get((m_id, op_id))
        if sched_mnmx is not None:
            mn_raw, mx_raw = sched_mnmx
        else:
            mn_raw, mx_raw = (None, None)

        # Resolve with safe defaults
        if mn_raw is None:
            mn_raw = meta.get("min_worker_num", 1)
        if mx_raw is None:
            mx_raw = meta.get("max_worker_num", 999)

        min_w = int(mn_raw) if mn_raw is not None else 1
        max_w = int(mx_raw) if mx_raw is not None else 999

        status = ""
        if heads < min_w:
            status = f"below min ({heads}<{min_w})"
        elif heads > max_w:
            status = f"above max ({heads}>{max_w})"

        if status:
            minmax_breach_cells.add((m_id, op_id, dt))
            tbl_minmax.append([dt.isoformat(), m_id, op_id, heads, min_w, max_w, status])

    # --- skill mismatch (sheet2 only) ---
    for (comp, wname, dt), lst in maps["per_cell_assigns_emp"].items():
        mismatch = False
        for info in lst:
            wid = info["worker"]
            skills = maps["worker_skills"].get(wid, {}) or {}
            if not has_skill_for_assignment(skills, info["m_id"], info["op_id"]):
                mismatch = True
                tbl_skill.append([dt.isoformat(), wname, comp, info["m_id"], info["op_id"]])
        if mismatch:
            skill_mismatch_cells.add((comp, wname, dt))

    # --- tasks with NO manager at all (across the whole horizon) ---
    tbl_no_manager = []
    for (m_id, op_id, _op_name) in maps.get("module_ops", []):
        mgrs = maps["managers_by_task"].get((m_id, op_id), set())
        if not mgrs:
            tbl_no_manager.append([m_id, op_id])

    # ===== Region presence & move-gap checks =====
    presence_map, segments, tbl_overstay, tbl_move_gap, move_markers = build_region_presence(
        plan_start=plan_start, plan_end=plan_end,
        env=env, modules=modules, assignments=assignments, maps=maps, cal=cal
    )

    # ===== Overtime limit breaches (monthly / annual) =====
    workers       = env["workers"]
    wcompanies    = env["worker_companies"]
    worker_name   = maps["worker_name"]
    worker_compid = maps["worker_company_id"]

    # Per worker-day total hours (by worker ID)
    emp_day_hours = defaultdict(int)
    for a in assignments:
        emp_day_hours[(a["worker"], a["date"])] += a["hours"]

    emp_year_month_ot = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))  # wid -> year -> month -> ot
    emp_year_total_ot = defaultdict(lambda: defaultdict(int))                      # wid -> year -> ot

    for (wid, dt), h in emp_day_hours.items():
        ot = max(0, h - OT_THRESHOLD)
        if ot <= 0:
            continue
        y = dt.year
        m = dt.month
        emp_year_month_ot[wid][y][m] += ot
        emp_year_total_ot[wid][y]    += ot

    tbl_ot_month  = []  # [worker_name, company_name, year, month, ot_hours, monthly_limit]
    tbl_ot_annual = []  # [worker_name, company_name, year, ot_hours, annual_limit]

    for wid, year_map in emp_year_month_ot.items():
        wname = worker_name.get(wid, wid)
        wco   = worker_compid.get(wid)
        comp_cfg = wcompanies.get(wco, {}) if wco else {}
        comp_name     = comp_cfg.get("name", wco) if wco else ""
        monthly_limit = comp_cfg.get("monthly_overtime_limit", None)
        annual_limit  = comp_cfg.get("annual_overtime_limit", None)

        # Monthly breaches
        for y, month_map in year_map.items():
            for m, ot_val in month_map.items():
                if isinstance(monthly_limit, (int, float)) and ot_val > monthly_limit:
                    tbl_ot_month.append([wname, comp_name, y, m, ot_val, monthly_limit])

        # Annual breaches
        if isinstance(annual_limit, (int, float)):
            for y, total_ot in emp_year_total_ot.get(wid, {}).items():
                if total_ot > annual_limit:
                    tbl_ot_annual.append([wname, comp_name, y, total_ot, annual_limit])

    return {
        "assigned_task": assigned_task,
        "assigned_mod": assigned_mod,
        "last_assigned_in_phase": last_assigned_in_phase,

        "win_breach_cells": win_breach_cells,
        "order_breach_cells": order_breach_cells,
        "minmax_breach_cells": minmax_breach_cells,
        "skill_mismatch_cells": skill_mismatch_cells,

        "tbl_window": tbl_window,
        "tbl_order": tbl_order,
        "tbl_skill": tbl_skill,
        "tbl_minmax": tbl_minmax,
        "tbl_no_manager": tbl_no_manager,

        # Unavailable / calendar
        "unavail_task_cells": unavail_task_cells,
        "unavail_emp_cells": unavail_emp_cells,
        "tbl_unavail_task": tbl_unavail_task,
        "tbl_unavail_emp": tbl_unavail_emp,

        # Region presence & moves
        "presence_map": presence_map,
        "presence_segments": segments,
        "tbl_overstay": tbl_overstay,
        "tbl_move_gap": tbl_move_gap,
        "move_markers": move_markers,

        # Overtime limit breaches
        "tbl_ot_month": tbl_ot_month,
        "tbl_ot_annual": tbl_ot_annual,

        # Preference breaches
        "tbl_pref_breach": tbl_pref_breach,

        # Fixed/Flexible scope maps
        "task_day_scopes": {k: _merge_scope(v) for k, v in task_day_scopes.items()},
        "task_scopes": {k: _merge_scope(v) for k, v in task_scopes.items()},
        "phase_scopes": {k: _merge_scope(v) for k, v in phase_scopes.items()},
        "emp_day_scopes": {k: _merge_scope(v) for k, v in emp_day_scopes.items()},
        "emp_scopes": {k: _merge_scope(v) for k, v in emp_scopes.items()},
        "worker_scopes": {k: _merge_scope(v) for k, v in worker_scopes.items()},
    }



# ----------------------- Skill distribution (unchanged) ----------------
def build_skill_distribution(env):
    workers = env["workers"]
    op_meta = env["op_meta"]

    # Collect all op_ids from env (across workflows and phases)
    all_ops = set()
    for wf_id, phases in op_meta.items():
        for ph_id, ops in phases.items():
            for op_id in ops.keys():
                all_ops.add(op_id)

    # Discover levels present
    level_set = set()
    for w in workers.values():
        smap = w.get("skill_map", {}) or {}
        for lvl in smap.values():
            try:
                iv = int(lvl)
                if iv > 0:
                    level_set.add(iv)
            except Exception:
                pass
    if not level_set:
        level_set = {1, 2, 3, 4, 5}
    levels = sorted(level_set)

    # Count employees per op_id per exact level (>0)
    from collections import Counter, defaultdict
    counts = defaultdict(Counter)  # op_id -> Counter(level -> count)

    for w in workers.values():
        smap = w.get("skill_map", {}) or {}
        for op_id, lvl in smap.items():
            if op_id not in all_ops:
                continue
            try:
                iv = int(lvl)
            except Exception:
                continue
            if iv > 0:
                counts[op_id][iv] += 1

    def op_sort_key(op_id: str):
        try:
            p_part, o_part = op_id.split("o", 1)
            p_num = int(p_part.replace("p", "")) if p_part.startswith("p") else 999
            o_num = int(o_part)
            return (p_num, o_num, op_id)
        except Exception:
            return (999, 999, op_id)

    rows = []
    for op_id in sorted(all_ops, key=op_sort_key):
        lvl_counts = [counts[op_id].get(l, 0) for l in levels]
        total = sum(lvl_counts)
        rows.append((op_id, total, lvl_counts))

    return rows, levels

# -------------------- NEW: op average skill per op_id ------------------
def compute_op_avg_skill(env):
    """
    Mean skill per op_id among all workers who have level > 0 for that op.
    Returns: dict op_id -> float
    """
    workers = env["workers"]
    sums = defaultdict(int)
    cnts = defaultdict(int)
    for w in workers.values():
        smap = w.get("skill_map", {}) or {}
        for op_id, lvl in smap.items():
            try:
                iv = int(lvl)
            except Exception:
                continue
            if iv > 0:
                sums[op_id] += iv
                cnts[op_id] += 1
    return {op: (sums[op] / cnts[op]) if cnts[op] else 0.0 for op in set(list(sums.keys()) + list(cnts.keys()))}

# --------- build "Team Quality by Block" aggregated rows ----------
def build_blocks_quality(modules, assignments, assignment_blocks, env, maps):
    """
    Returns list of dict rows, one per block:
    key fields:
      module, op_id, phase, start_date, end_date, heads, hours, days,
      managers, company_mix, max_pairs, same_company_pairs, cohesion_pct,
      op_avg_skill, team_avg_skill, balance_dev, balance_score, variety_score
    Grouping heuristic: (module, op_id, start_date, end_date, hours)
    """
    worker_name = maps["worker_name"]
    worker_is_manager = maps["worker_is_manager"]
    worker_company_id = maps["worker_company_id"]
    worker_company_name = maps["worker_company_name"]
    worker_skills = maps["worker_skills"]
    op_phase_of_module = maps["op_phase_of_module"]

    # Build per-employee blocks keyed by (module, op_id, start, end, hours)
    # First, we need hours per worker per assignment block: infer from assignments on first work day.
    # Build quick lookup: (worker, module, op_id, date) -> hours
    per_w_mop_day_hours = {}
    for a in assignments:
        wid = a["worker"]
        op_task = a["operation_task"]
        m_id, op_id = parse_op_task_ids(op_task, maps.get("op_task_index", {}))
        per_w_mop_day_hours[(wid, m_id, op_id, a["date"])] = a["hours"]

    # Collate blocks from assignment_blocks
    # group key
    def blk_key(wid, op_task, start_date, end_date):
        m_id, op_id = parse_op_task_ids(op_task, maps.get("op_task_index", {}))
        # infer hours: take hours on the first work date if present; else 8
        hours = 8
        wdates = []
        for ab in assignment_blocks:
            if ab["worker"] == wid and ab["operation_task"] == op_task and ab["start_date"] == start_date and ab["end_date"] == end_date:
                wdates = ab.get("work_dates", []) or []
                break
        if wdates:
            h = per_w_mop_day_hours.get((wid, m_id, op_id, wdates[0]), None)
            if isinstance(h, int):
                hours = h
        days = (end_date - start_date).days + 1 if (start_date and end_date) else len(wdates)
        return (m_id, op_id, start_date, end_date, hours, days)

    # group: block_key -> set of workers
    block_workers = defaultdict(set)
    # also union managers across all days within the block
    block_managers = defaultdict(set)

    for ab in assignment_blocks:
        wid = ab["worker"]
        op_task = ab["operation_task"]
        sd = ab["start_date"]; ed = ab["end_date"]
        if not sd or not ed:
            continue
        key = blk_key(wid, op_task, sd, ed)
        block_workers[key].add(wid)
        if maps["worker_is_manager"].get(wid, False):
            block_managers[key].add(worker_name.get(wid, wid))

    op_avg = compute_op_avg_skill(env)

    rows = []
    for (m_id, op_id, sd, ed, hours, days), wset in block_workers.items():
        heads = len(wset)
        # company mix
        comp_counts = Counter(worker_company_name.get(wid, "") for wid in wset)
        comp_text = " | ".join(f"{c or '':s}×{n}" for c, n in sorted(comp_counts.items(), key=lambda kv: (-kv[1], kv[0] or "")))

        # cohesion
        max_pairs = heads * (heads - 1) // 2
        same_company_pairs = sum(n * (n - 1) // 2 for n in comp_counts.values())
        cohesion_pct = (same_company_pairs / max_pairs * 100.0) if max_pairs > 0 else None

        # skills for this op
        levels = []
        for wid in wset:
            lvl = 0
            try:
                lvl = int((worker_skills.get(wid) or {}).get(op_id, 0))
            except Exception:
                lvl = 0
            levels.append(max(0, lvl))
        team_avg_skill = (sum(levels) / len(levels)) if levels else 0.0

        # variety via Shannon index over levels present (ignore zeros if you want)
        lvl_counts = Counter([lv for lv in levels if lv > 0])
        if lvl_counts:
            total = sum(lvl_counts.values())
            H = 0.0
            for n in lvl_counts.values():
                p = n / total
                if p > 0:
                    H -= p * math.log(p)
            Hmax = math.log(len(lvl_counts)) if len(lvl_counts) > 0 else 1.0
            variety_score = (H / Hmax * 100.0) if Hmax > 0 else 0.0
        else:
            variety_score = 0.0

        op_mean = op_avg.get(op_id, 0.0)
        balance_dev = abs(team_avg_skill - op_mean)
        balance_score = max(0.0, 100.0 - BALANCE_K * balance_dev)

        phase = maps["op_phase_of_module"].get((m_id, op_id), "")
        mgrs = sorted(block_managers.get((m_id, op_id, sd, ed, hours, days), set()))
        mgr_text = ", ".join(mgrs)

        rows.append({
            "module": m_id,
            "op_id": op_id,
            "phase": phase,
            "start_date": sd.isoformat() if sd else "",
            "end_date": ed.isoformat() if ed else "",
            "heads": heads,
            "hours": hours,
            "days": days,
            "managers": mgr_text,
            "company_mix": comp_text,
            "max_pairs": max_pairs,
            "same_company_pairs": same_company_pairs,
            "cohesion_pct": round(cohesion_pct, 1) if cohesion_pct is not None else "",
            "op_avg_skill": round(op_mean, 2),
            "team_avg_skill": round(team_avg_skill, 2),
            "balance_dev": round(balance_dev, 3),
            "balance_score": round(balance_score, 1),
            "variety_score": round(variety_score, 1),
        })

    # stable sort
    rows.sort(key=lambda r: (r["module"], r["phase"], r["op_id"], r["start_date"]))
    return rows
# ------------------------------ REGION POLICY HELPERS ------------------------------
def _region_policy(regions, rid):
    """Return (stay_limit_days, stay_gap_days) for region id, robust to different key names."""
    if not rid or rid not in regions:
        return (None, 0)
    r = regions[rid] or {}

    # --- your schema first ---
    stay_limit = r.get("max_stay_on")         # e.g., 60
    stay_gap   = r.get("stay_off_interval")   # e.g., 5

    # --- legacy / alternate names as fallback ---
    if stay_limit is None:
        stay_limit = (
            r.get("stay_limit") or r.get("max_stay") or r.get("max_stay_days")
            or r.get("visa_stay_limit")
        )
    if stay_gap is None:
        stay_gap = (
            r.get("stay_gap") or r.get("min_gap") or r.get("stay_break_gap")
            or r.get("visa_stay_gap") or r.get("interval_off")
        )

    try:
        stay_limit = int(stay_limit) if stay_limit is not None else None
    except Exception:
        stay_limit = None
    try:
        stay_gap = int(stay_gap) if stay_gap is not None else 0
    except Exception:
        stay_gap = 0

    return (stay_limit, stay_gap)


def _daterange_inclusive(a: date, b: date):
    d = a
    while d <= b:
        yield d
        d += timedelta(days=1)

# -------------------- BUILD PER-WORKER REGION PRESENCE & MOVES --------------------
def build_region_presence(plan_start, plan_end, env, modules, assignments, maps, cal):
    """
    Derives, for each worker:
      - continuous 'presence segments' in a region (including non-working days),
        where continuity is broken only if the gap between working days exceeds the region stay_gap.
      - move events between regions and whether the off-gap before entering the new region was sufficient.
    Returns:
      presence_map: dict[(worker_name, date)] = region_id or None
      segments: dict[worker_name] = list of {region, start, end, days}
      tbl_overstay: rows [worker, region, start, end, days, stay_limit]
      tbl_move_gap: rows [worker, from_region, to_region, out_date, in_date, gap_days, required_gap]
      move_markers: dict[worker_name] = list of {"type": "out"/"in", "date": <date>, "region": id}
                    (used only by Sheet 5 to write the "out"/"in" strings)
    """
    workers = env["workers"]
    regions = env["regions"]

    # Helper: get region for an assignment (via module -> fab -> region)
    def region_for_assignment(a):
        op_task = a["operation_task"]
        m_id, _op = parse_op_task_ids(op_task, maps.get("op_task_index", {}))
        meta = maps["mod_meta_cols"].get(m_id, {})
        fab_id = meta.get("fab_id")
        rid = cal["fab_region"].get(fab_id)
        return rid, m_id

    # transit time between regions (days)
    transit_days = {}
    for tr in (env.get("transite_day_map") or []):
        try:
            k = (tr.get("from"), tr.get("to"))
            transit_days[k] = int(tr.get("days", 0))
        except Exception:
            pass
    
    # Collect per-worker working dates with their regions AND Fixed/Flexible scope.
    # Important for breach split sheets:
    #   - Fixed -> Fixed transit gap belongs to fix page.
    #   - Any pair that touches Flexible (Fixed->Flexible, Flexible->Fixed, Flexible->Flexible) belongs to flex page.
    perw_date_region_scopes = defaultdict(set)  # (wname, date, region_id) -> {'fix','flex'}
    worker_name = maps["worker_name"]

    for a in assignments:
        wname = worker_name.get(a["worker"], a["worker"])
        rid, _ = region_for_assignment(a)
        if rid:
            perw_date_region_scopes[(wname, a["date"], rid)].add(_norm_flex(a.get("plan_flexibility")))

    perw_dates = defaultdict(list)  # wname -> list of (date, region_id, scope)
    for (wname, dt, rid), scopes in perw_date_region_scopes.items():
        perw_dates[wname].append((dt, rid, _merge_scope(scopes)))

    def _transition_scope(left_scope, right_scope):
        vals = {left_scope, right_scope}
        if "flex" in vals or "both" in vals:
            return "flex"
        if vals == {"fix"}:
            return "fix"
        return _merge_scope(vals)

    # Sort and build continuous presence segments
    presence_map = {}           # (wname, date) -> region_id
    segments = defaultdict(list)
    move_markers = defaultdict(list)
    tbl_overstay = []
    tbl_move_gap = []

    # For dates outside any segment, presence_map stays empty (no color).
    for wname, lst in perw_dates.items():
        lst.sort(key=lambda t: t[0])  # by date
        if not lst:
            continue

        # Build raw "stops" grouped by region, with allowed gaps (stay_gap) for continuity
        i = 0
        while i < len(lst):
            start_date, cur_r, start_scope = lst[i]
            segment_scope_values = {start_scope}
            (stay_limit, stay_gap) = _region_policy(regions, cur_r)
            if stay_gap is None: stay_gap = 0
            # If this is the first segment for this worker, also place an initial "in"
            if i == 0:
                pre_in_day = start_date - timedelta(days=1)
                if plan_start <= pre_in_day <= plan_end:
                    # avoid duplicate if somehow already present
                    already = any(m.get("type") == "in" and m.get("date") == pre_in_day
                                for m in move_markers[wname])
                    if not already:
                        move_markers[wname].append({"type": "in", "date": pre_in_day, "region": cur_r})

            end_date = start_date
            j = i + 1
            while j < len(lst):
                d, r, sc = lst[j]
                if r != cur_r:
                    break
                gap_days = (d - end_date).days - 1  # in-between non-working days count
                if gap_days <= stay_gap:
                    end_date = d
                    segment_scope_values.add(sc)
                    j += 1
                else:
                    break
            seg_end = end_date
            seg_scope = _merge_scope(segment_scope_values)
            segments[wname].append({
                "region": cur_r,
                "start": start_date,
                "end": seg_end,
                "days": (seg_end - start_date).days + 1
            })

            if stay_limit is not None:
                days = (seg_end - start_date).days + 1
                if days > stay_limit:
                    tbl_overstay.append([
                        wname, cur_r,
                        start_date.isoformat(), seg_end.isoformat(),
                        days, stay_limit
                    ])

            for d in _daterange_inclusive(start_date, seg_end):
                presence_map[(wname, d)] = cur_r

            out_mark_day = seg_end + timedelta(days=1)
            if plan_start <= out_mark_day <= plan_end:
                move_markers[wname].append({"type": "out", "date": out_mark_day, "region": cur_r})

            i = j
            if i < len(lst):
                next_start, next_r, next_scope0 = lst[i]
                in_mark_day = next_start - timedelta(days=1)
                if plan_start <= in_mark_day <= plan_end:
                    move_markers[wname].append({"type": "in", "date": in_mark_day, "region": next_r})

                (_, next_gap) = _region_policy(regions, next_r)
                if next_gap is None:
                    (_, old_gap) = _region_policy(regions, cur_r)
                    next_gap = old_gap if old_gap is not None else 0

                # Transit requirement strictly governs inter-region moves.
                # If there is a mapping for (cur_r -> next_r), use it.
                # If not, fall back to region stay_gap (next_gap).
                if cur_r != next_r:
                    transit_req = transit_days.get((cur_r, next_r), None)
                    required_gap = transit_req if transit_req is not None else next_gap
                else:
                    required_gap = 0  # same region → no “move” gap needed

                actual_gap = (next_start - seg_end).days - 1

                if actual_gap < required_gap:
                    pair_scope = _transition_scope(seg_scope, next_scope0)
                    tbl_move_gap.append([
                        wname, cur_r, next_r,
                        seg_end.isoformat(), next_start.isoformat(),
                        actual_gap, required_gap, pair_scope
                    ])
    return presence_map, segments, tbl_overstay, tbl_move_gap, move_markers

# ------------------------------- WRITERS -------------------------------
def write_sheet_tasks_dates(wb, plan_start, plan_end, env, maps, modules, vios, req_task, cal):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="999999")
    border_thin = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws = wb.create_sheet("Tasks x Dates")

    # headers incl. required/assigned per task
    headers = ["module", "module_name", "fab_id", "fab_name", "region", "customer", "task","manager",
               "required_hours", "assigned_hours"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h); c.font = bold

    # date columns
    dates = []
    d = plan_start
    while d <= plan_end:
        dates.append(d); d += timedelta(days=1)

    for j, dt in enumerate(dates, start=len(headers)+1):
        c = ws.cell(row=1, column=j, value=dt.isoformat())
        c.font = bold
        ws.column_dimensions[get_column_letter(j)].width = 28  # a bit wider for ★ tags

    widths = [8, 18, 8, 14, 12, 10, 18, 18, 16, 16]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = get_column_letter(len(headers)+1) + "2"

    # coloring fills
    fill_start   = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    fill_deadln  = PatternFill(start_color=RED,        end_color=RED,        fill_type="solid")
    fill_win     = PatternFill(start_color=PURPLE_WIN, end_color=PURPLE_WIN, fill_type="solid")
    fill_order   = PatternFill(start_color=BLUE_ORDER, end_color=BLUE_ORDER, fill_type="solid")
    fill_minmax  = PatternFill(start_color=PINK_MINMAX,end_color=PINK_MINMAX,fill_type="solid")
    fill_under   = PatternFill(start_color=YELLOW_UNDER,end_color=YELLOW_UNDER,fill_type="solid")
    # NEW
    fill_grey    = PatternFill(start_color=GREY_CLOSED, end_color=GREY_CLOSED, fill_type="solid")
    fill_unavail = PatternFill(start_color=BROWN_UNAV_BREACH, end_color=BROWN_UNAV_BREACH, fill_type="solid")

    # sort rows
    def mod_sort_key(m_id):
        try: return int(m_id[1:])
        except: return 999
    rows = sorted(maps["module_ops"], key=lambda x: (mod_sort_key(x[0]), x[1]))

    # helper for module start & phase end
    module_start = maps["module_start"]
    phase_end    = maps["phase_end"]
    op_phase     = maps["op_phase_of_module"]

    # write rows
    for r_idx, (m_id, op_id, op_name) in enumerate(rows, start=2):
        meta = maps["mod_meta_cols"][m_id]
        ws.cell(row=r_idx, column=1, value=meta["module"]).font = bold
        ws.cell(row=r_idx, column=2, value=meta["module_name"]).font = bold
        ws.cell(row=r_idx, column=3, value=meta["fab_id"]).font = bold
        ws.cell(row=r_idx, column=4, value=meta["fab_name"]).font = bold
        ws.cell(row=r_idx, column=5, value=meta["region"]).font = bold
        ws.cell(row=r_idx, column=6, value=meta["customer"]).font = bold
        ws.cell(row=r_idx, column=7, value=f"{op_id} {op_name}").font = bold

        # managers — list unique manager names for this task (all managers)
        mgr_names = sorted(maps["managers_by_task"].get((m_id, op_id), []))
        mgr_text  = " | ".join(mgr_names) if mgr_names else ""
        c_mgr = ws.cell(row=r_idx, column=8, value=mgr_text); c_mgr.alignment = center
        if not mgr_names:
            c_mgr.fill = PatternFill(start_color=MANAGER_MISSING, end_color=MANAGER_MISSING, fill_type="solid")

        # required/assigned per task
        req = req_task.get((m_id, op_id), 0)
        asg = vios["assigned_task"].get((m_id, op_id), 0)
        ws.cell(row=r_idx, column=9,  value=req).alignment = center
        c_asg = ws.cell(row=r_idx, column=10, value=asg); c_asg.alignment = center
        if asg < req:
            c_asg.fill = fill_under  # highlight assigned_hours cell only

        # pre-compute fab/region/customer ids (for grey shading)
        fab_id = meta["fab_id"]
        rid = cal["fab_region"].get(fab_id)
        cid = cal["fab_customer"].get(fab_id)

        # date cells
        m_start = module_start.get(m_id)
        ph_id   = op_phase.get((m_id, op_id))
        p_end   = phase_end.get((m_id, ph_id))

        for j, dt in enumerate(dates, start=len(headers)+1):
            txt = " | ".join(sorted(maps["tde"].get((m_id, op_id, dt), [])))

            c = ws.cell(row=r_idx, column=j, value=txt)

            # ----- GREY OUT fab/region/customer unavailable (no global weekend) -----
            is_closed = (
                (fab_id and dt in cal["fab_off"].get(fab_id, set())) or
                (rid and dt in cal["region_off"].get(rid, set())) or
                (cid and dt in cal["customer_off"].get(cid, set()))
            )
            if is_closed:
                c.fill = fill_grey
            # -------------------------------------------------------------------

            # decor (non-violation)
            if m_start and dt == m_start:
                c.fill = fill_start
            if p_end and dt == p_end:
                c.fill = fill_deadln

            # violations overlays (priority: minmax > window > ordering > unavailable-breach)
            key = (m_id, op_id, dt)
            if key in vios["minmax_breach_cells"]:
                c.fill = fill_minmax
            elif key in vios["win_breach_cells"]:
                c.fill = fill_win
            elif key in vios["order_breach_cells"]:
                c.fill = fill_order
            # NEW: assignment on unavailable day (brown)
            elif key in vios["unavail_task_cells"]:
                c.fill = fill_unavail

    # row heights
    ws.row_dimensions[1].height = 22
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 36

    # legend at bottom
    legend_row = ws.max_row + 2
    ws.cell(row=legend_row,   column=1, value="Legend").font = bold
    def legend(label, fill, col):
        cell = ws.cell(row=legend_row+1, column=col, value=label)
        cell.fill = fill; cell.alignment = center
    legend("Start",   fill_start,   1)
    legend("Deadline",fill_deadln,  2)
    legend("Window breach", fill_win, 3)
    legend("Ordering breach", fill_order, 4)
    legend("Staffing min/max breach", fill_minmax, 5)
    legend("Under-assigned (assigned_hours)", fill_under, 6)
    # NEW legend items
    legend("Closed (fab/region/customer off)", PatternFill(start_color=GREY_CLOSED, end_color=GREY_CLOSED, fill_type="solid"), 7)
    legend("Assigned on closed day", PatternFill(start_color=BROWN_UNAV_BREACH, end_color=BROWN_UNAV_BREACH, fill_type="solid"), 8)


def write_sheet_employees_dates(wb, plan_start, plan_end, env, maps, vios, cal):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws = wb.create_sheet("Employees x Dates")

    base_headers = ["company", "employee", "skills", "Manager"]
    for i, h in enumerate(base_headers, start=1):
        c = ws.cell(row=1, column=i, value=h); c.font = bold

    # dates
    dates = []
    d = plan_start
    while d <= plan_end:
        dates.append(d); d += timedelta(days=1)

    for j, dt in enumerate(dates, start=len(base_headers)+1):
        c = ws.cell(row=1, column=j, value=dt.isoformat()); c.font = bold
        ws.column_dimensions[get_column_letter(j)].width = 30

    workdays_col = len(base_headers) + len(dates) + 1
    workhours_col= len(base_headers) + len(dates) + 2
    ws.cell(row=1, column=workdays_col, value="Workdays").font = bold
    ws.cell(row=1, column=workhours_col, value="WorkHours").font = bold

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 54
    ws.column_dimensions["D"].width = 10  # Manager
    ws.column_dimensions[get_column_letter(workdays_col)].width = 11
    ws.column_dimensions[get_column_letter(workhours_col)].width = 11
    ws.freeze_panes = "E2"

    # skills text
    def skills_text(wname):
        wid = None
        for k, v in env["workers"].items():
            if v.get("name") == wname:
                wid = k; break
        if wid is None:
            return ""
        items = list(maps["worker_skills"].get(wid, {}).items())
        def keyer(k):
            try:
                p, o = k.split("o")
                return (int(p[1:]), int(o))
            except Exception:
                return (999, 999)
        items.sort(key=lambda kv: keyer(kv[0]))
        return ", ".join(f"{k}:{v}" for k, v in items)

    # roster
    roster = []
    seen = set()
    for (comp, wname, d), _ in maps["edt"].items():
        if (comp, wname) not in seen:
            roster.append((comp, wname)); seen.add((comp, wname))
    roster.sort(key=lambda t: (t[0] or "", t[1] or ""))

    fill_skill  = PatternFill(start_color=ORANGE_SKILL, end_color=ORANGE_SKILL, fill_type="solid")
    fill_grey   = PatternFill(start_color=GREY_CLOSED, end_color=GREY_CLOSED, fill_type="solid")  # personal off only
    fill_unavail= PatternFill(start_color=BROWN_UNAV_BREACH, end_color=BROWN_UNAV_BREACH, fill_type="solid")  # breaches

    for i, (comp, wname) in enumerate(roster, start=2):
        ws.row_dimensions[i].height = 34
        ws.cell(row=i, column=1, value=comp or "").alignment = left
        ws.cell(row=i, column=2, value=wname or "").alignment = left
        ws.cell(row=i, column=3, value=skills_text(wname)).alignment = left
        ws.cell(row=i, column=4, value="True" if maps["name_is_manager"].get(wname, False) else "").alignment = center

        # resolve worker id once per row
        wid = None
        for k, v in env["workers"].items():
            if v.get("name") == wname:
                wid = k; break
        wco = env["workers"].get(wid, {}).get("worker_company") if wid else None

        for j, dt in enumerate(dates, start=len(base_headers)+1):
            txt = " | ".join(sorted(maps["edt"].get((comp, wname, dt), [])))
            c = ws.cell(row=i, column=j, value=txt)

            # ----- GREY OUT only personal unavailable dates (no weekend / no company-off) -----
            personal_off = bool(wid) and (dt in cal["worker_off"].get(wid, set()))
            if personal_off:
                c.fill = fill_grey
            # ---------------------------------------------------------------------------

            # skill mismatch (orange)
            if (comp, wname, dt) in vios["skill_mismatch_cells"]:
                c.fill = fill_skill

            # assigned on personal/company off day (breach color)
            if (comp, wname, dt) in vios["unavail_emp_cells"]:
                c.fill = fill_unavail

        ws.cell(row=i, column=workdays_col, value=maps["emp_workdays"].get(wname, 0)).alignment = center
        ws.cell(row=i, column=workhours_col, value=maps["emp_total_hours"].get(wname, 0)).alignment = center

    # legend
    legend_row = ws.max_row + 2
    ws.cell(row=legend_row, column=1, value="Legend").font = bold
    cell = ws.cell(row=legend_row+1, column=1, value="Skill mismatch")
    cell.fill = fill_skill; cell.alignment = center
    cell = ws.cell(row=legend_row+1, column=2, value="Personal unavailable day")
    cell.fill = fill_grey; cell.alignment = center
    cell = ws.cell(row=legend_row+1, column=3, value="Assigned on closed day (personal/company)")
    cell.fill = fill_unavail; cell.alignment = center


# ----------------------- SHEET 3 : DASHBOARD (plan) ------------------
def write_sheet_dashboard_plan(wb, plan_start, plan_end, env, maps, modules, assignments, assignment_blocks, req_module):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws = wb.create_sheet("Dashboard (Plan)")
    for col in range(1, 120):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # ================= KPIs (overall) =================
    unique_workers = len({a["worker"] for a in assignments})

    avg_hours = 0.0
    if maps["emp_workdays"]:
        total_hours = sum(maps["emp_total_hours"].values())
        total_days  = sum(maps["emp_workdays"].values())
        if total_days:
            avg_hours = total_hours / total_days

    # build per-employee-day hours to compute OT and cap
    per_emp_day = defaultdict(int)
    for (comp, wname, d), items in maps["edt"].items():
        day_hours = 0
        for s in items:
            try:
                h = int(s.split("(")[-1].rstrip("H)"))
            except Exception:
                h = 0
            day_hours += h
        per_emp_day[(wname, d)] += day_hours

    ot_hours = 0
    for (_, _dte), h in per_emp_day.items():
        if h > OT_THRESHOLD:
            ot_hours += (h - OT_THRESHOLD)

    cap_breach_cnt = sum(1 for h in per_emp_day.values() if h > CAP_HOURS)
    cap_breach_pct = (100.0 * cap_breach_cnt / max(1, len(per_emp_day))) if per_emp_day else 0.0

    # Completion % by module
    assigned_by_module = defaultdict(int)
    for a in assignments:
        op_task = a["operation_task"]
        m_id, _op = parse_op_task_ids(op_task, maps.get("op_task_index", {}))
        assigned_by_module[m_id] += a["hours"]

    total_req = sum(req_module.values())
    completion_pct = (100.0 * sum(assigned_by_module.values()) / total_req) if total_req else 100.0

    ws["A1"].value = "KPIs"
    ws["A1"].font = bold
    kpi_rows = [
        ("Unique workers", unique_workers),
        ("Avg hours/worker-day", round(avg_hours, 2)),
        ("Overtime hours(>8)", ot_hours),
        ("Cap breach(>12h) days", cap_breach_cnt),
        ("Cap breach(%)", f"{cap_breach_pct:.1f}%"),
        ("Completion", f"{completion_pct:.1f}%"),
    ]
    for i, (k, v) in enumerate(kpi_rows, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

    # ================= Progress by module =================
    start_row = 10
    ws.cell(row=start_row, column=1, value="Progress by module").font = bold
    hdr = ["module", "required_hours", "assigned_hours", "%complete",
           "planned_end", "last_assigned", "delay(vs plan)"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=start_row+1, column=j, value=h).font = bold

    # last_assigned per module
    last_assigned_by_module = {}
    for a in assignments:
        op_task = a["operation_task"]
        m_id_a, _op = parse_op_task_ids(op_task, maps.get("op_task_index", {}))
        prev = last_assigned_by_module.get(m_id_a)
        if prev is None or a["date"] > prev:
            last_assigned_by_module[m_id_a] = a["date"]

    fill_under = PatternFill(start_color=YELLOW_UNDER, end_color=YELLOW_UNDER, fill_type="solid")

    prog_rows = []
    for m in modules:
        m_id = m["id"]
        req = req_module.get(m_id, 0)
        asg = assigned_by_module.get(m_id, 0)
        pct = (100.0 * asg / req) if req else 100.0
        planned_end = max(_d(ph["end_date"]) for ph in m.get("phase_task_list", [])) if m.get("phase_task_list") else None
        last_asg = last_assigned_by_module.get(m_id)
        delay = None
        if planned_end and last_asg:
            delay = (last_asg - planned_end).days
        prog_rows.append((m_id, req, asg, pct,
                          planned_end.isoformat() if planned_end else "",
                          last_asg.isoformat() if last_asg else "",
                          delay if delay is not None else ""))

    for i, row in enumerate(prog_rows, start=start_row+2):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            if j in (3, 4):  # assigned_hours, %complete
                req = row[1]; asg = row[2]
                if isinstance(req, (int, float)) and isinstance(asg, (int, float)) and asg < req:
                    c.fill = fill_under

    # ================= Total hours per day =================
    th_start = start_row + 2 + len(prog_rows) + 2
    ws.cell(row=th_start, column=1, value="Total hours per day").font = bold
    ws.cell(row=th_start+1, column=1, value="date").font = bold
    ws.cell(row=th_start+1, column=2, value="total_hours").font = bold

    day_rows = []
    d = plan_start
    while d <= plan_end:
        day_rows.append((d.strftime("%m/%d"), maps["per_day_total"].get(d, 0)))
        d += timedelta(days=1)

    for i, (dd, hrs) in enumerate(day_rows, start=th_start+2):
        ws.cell(row=i, column=1, value=dd)
        ws.cell(row=i, column=2, value=int(hrs))

    # Line chart for total hours per day
    anchor_line = "N2"
    if day_rows:
        chart2 = LineChart()
        chart2.title = "Total hours per day"
        chart2.y_axis.title = "hours"
        chart2.x_axis.title = "date"
        chart2.x_axis.tickLblSkip = 9
        dr0 = th_start + 1
        mrows = len(day_rows)
        cats = Reference(ws, min_col=1, min_row=dr0+1, max_row=dr0+mrows)
        vals = Reference(ws, min_col=2, min_row=dr0,   max_row=dr0+mrows)
        chart2.add_data(vals, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.height = 12
        chart2.width = 24
        ws.add_chart(chart2, anchor_line)

    # ===== Skill distribution table (by op_id x level) =====
    skill_rows, levels = build_skill_distribution(env)

    if skill_rows:
        insert_row = ws.max_row + 3
        ws.cell(row=insert_row, column=1, value="Employee skill distribution").font = bold
        insert_row += 1

        headers = ["skill", "total"] + [str(l) for l in levels]
        for j, h in enumerate(headers, start=1):
            ws.cell(row=insert_row, column=j, value=h).font = bold
        insert_row += 1

        for (op_id, total, lvl_counts) in skill_rows:
            ws.cell(row=insert_row, column=1, value=op_id)
            ws.cell(row=insert_row, column=2, value=int(total))
            for idx, c in enumerate(lvl_counts, start=3):
                ws.cell(row=insert_row, column=idx, value=int(c))
            insert_row += 1

        ws.cell(row=insert_row, column=1, value="TOTAL").font = bold
        ws.cell(row=insert_row, column=2, value=sum(r[1] for r in skill_rows)).font = bold
        for idx, l in enumerate(levels, start=3):
            ws.cell(
                row=insert_row,
                column=idx,
                value=sum(r[2][idx-3] for r in skill_rows)
            ).font = bold

    # ===== Team Quality by Block =====
    q0 = ws.max_row + 3
    ws.cell(row=q0, column=1, value="Team Quality by Block").font = bold
    headers = [
        "module", "op_id", "phase", "start_date", "end_date", "heads", "hours", "days",
        "managers", "company_mix", "max_pairs", "same_company_pairs", "cohesion_%",
        "op_avg_skill", "team_avg_skill", "balance_dev", "balance_score", "variety_score"
    ]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=q0+1, column=j, value=h).font = bold

    block_rows = build_blocks_quality(modules, assignments, assignment_blocks, env, maps)
    for i, r in enumerate(block_rows, start=q0+2):
        ws.cell(row=i, column=1,  value=r["module"])
        ws.cell(row=i, column=2,  value=r["op_id"])
        ws.cell(row=i, column=3,  value=r["phase"])
        ws.cell(row=i, column=4,  value=r["start_date"])
        ws.cell(row=i, column=5,  value=r["end_date"])
        ws.cell(row=i, column=6,  value=r["heads"])
        ws.cell(row=i, column=7,  value=r["hours"])
        ws.cell(row=i, column=8,  value=r["days"])
        ws.cell(row=i, column=9,  value=r["managers"])
        ws.cell(row=i, column=10, value=r["company_mix"])
        ws.cell(row=i, column=11, value=r["max_pairs"])
        ws.cell(row=i, column=12, value=r["same_company_pairs"])
        ws.cell(row=i, column=13, value=r["cohesion_pct"] if r["cohesion_pct"] != "" else None)
        ws.cell(row=i, column=14, value=r["op_avg_skill"])
        ws.cell(row=i, column=15, value=r["team_avg_skill"])
        ws.cell(row=i, column=16, value=r["balance_dev"])
        ws.cell(row=i, column=17, value=r["balance_score"])
        ws.cell(row=i, column=18, value=r["variety_score"])

    if block_rows:
        chart = ScatterChart()
        chart.title = "Balance vs Variety (by block)"
        chart.x_axis.title = "balance_dev (|team_avg - op_avg|)"
        chart.y_axis.title = "variety_score"
        first_row = q0 + 2
        last_row = q0 + 1 + len(block_rows)
        xref = Reference(ws, min_col=16, min_row=first_row, max_row=last_row)
        yref = Reference(ws, min_col=18, min_row=first_row, max_row=last_row)
        s = Series(yref, xref, title_from_data=False)
        chart.series.append(s)
        chart.height = 15
        chart.width = 28
        ws.add_chart(chart, "T{}".format(q0))

# =========================== SHEET 4: Recommend staffing deviation (soft) ==========================
def write_sheet_recommend_staffing(wb, plan_start, plan_end, modules, maps):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws = wb.create_sheet("Recommend (Soft)")

    ws.cell(row=1, column=1, value="Recommend staffing deviation (soft)").font = bold

    rec_hdr = [
        "module", "op_id", "phase",
        "phase_start", "phase_end",
        "rec_min", "rec_max",
        "avg_heads",
        "max_worker",
        "diff_avg_vs_rec"
    ]
    for j, h in enumerate(rec_hdr, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = bold
        c.alignment = center

    for col in range(1, len(rec_hdr) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16

    op_task_index = maps.get("op_task_index", {}) or {}

    # (m_id, op_id) -> (rec_min, rec_max)
    rec_range = {}
    for _ot_id, meta in op_task_index.items():
        m_id = meta.get("m_id")
        op_id = meta.get("op_id")
        if not m_id or not op_id:
            continue

        rmin = meta.get("recommends_worker_min")
        rmax = meta.get("recommends_worker_max")

        if rmin is None and rmax is None:
            continue
        if rmin is None:
            rmin = rmax
        if rmax is None:
            rmax = rmin

        try:
            rmin = int(rmin)
            rmax = int(rmax)
        except Exception:
            continue

        if rmin > rmax:
            rmin, rmax = rmax, rmin

        rec_range[(m_id, op_id)] = (rmin, rmax)

    # use same effective phase window rule as breaches
    phase_window = build_effective_phase_window(modules, maps)

    rows = []
    for (m_id, op_id, _op_name_unused) in (maps.get("module_ops") or []):
        if (m_id, op_id) not in rec_range:
            continue

        ph_id = maps.get("op_phase_of_module", {}).get((m_id, op_id))
        w = phase_window.get((m_id, ph_id))
        if not w:
            continue

        w_start, w_end = w
        if not w_start or not w_end:
            continue

        start = max(plan_start, w_start)
        end   = min(plan_end, w_end)
        if start > end:
            continue

        rec_min, rec_max = rec_range[(m_id, op_id)]

        total_heads_assigned = 0
        assigned_days = 0
        max_heads = 0

        d = start
        while d <= end:
            heads = int(maps.get("day_op_heads", {}).get((m_id, op_id, d), 0))

            if heads > 0:
                assigned_days += 1
                total_heads_assigned += heads

            if heads > max_heads:
                max_heads = heads

            d += timedelta(days=1)

        avg_heads = (total_heads_assigned / assigned_days) if assigned_days > 0 else 0.0

        # distance from recommended range
        if avg_heads < rec_min:
            diff_avg_vs_rec = rec_min - avg_heads
        elif avg_heads > rec_max:
            diff_avg_vs_rec = avg_heads - rec_max
        else:
            diff_avg_vs_rec = 0.0

        rows.append((
            m_id,
            op_id,
            ph_id or "",
            start.isoformat(),
            end.isoformat(),
            rec_min,
            rec_max,
            round(avg_heads, 2),
            max_heads,
            round(diff_avg_vs_rec, 2),
        ))

    # sort: biggest gap from recommend first, then biggest worker peak
    rows.sort(key=lambda r: (-r[9], -r[8], r[0], r[1]))

    out_row = 3
    if rows:
        for row in rows:
            for j, v in enumerate(row, start=1):
                c = ws.cell(row=out_row, column=j, value=v)
                c.alignment = center if j >= 4 else left
            out_row += 1
    else:
        ws.cell(row=3, column=1, value="No rows: recommends_worker_min/max not found, or no matching tasks in module_ops.")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(rec_hdr))

    ws.freeze_panes = "A3"

    last_row = ws.max_row
    last_col = len(rec_hdr)
    if last_row >= 2 and last_col >= 1:
        ws.auto_filter.ref = f"A2:{get_column_letter(last_col)}{last_row}"

    if rows:
        ref = f"A2:{get_column_letter(last_col)}{ws.max_row}"
        tbl = Table(displayName="RecommendSoftTbl", ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tbl)
    

 #=========================== SHEET 5: write sheet dashboard employee =============================

def write_sheet_dashboard_employees(
    wb, plan_start, plan_end, env, maps,
    modules, assignments, assignment_blocks, req_module, vios
):
    """
    Dashboard (Employees) now includes:
      - Assignment utilization
      - Overtime & capacity
      - Workload balance
      - Top 20 hours bar chart
      - OT limit rules & OT usage
      - Assignment distribution by task (per employee)
      - Region move summary (per employee, using same logic as Moving plan)
      - Preference match (per employee x fab)
    """
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws = wb.create_sheet("Dashboard (Employees)")
    for col in range(1, 120):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # Build per-employee-per-day hours table once for this sheet (by employee name)
    from collections import defaultdict, Counter
    per_emp_day = defaultdict(int)
    for (comp, wname, d), items in maps["edt"].items():
        day_hours = 0
        for s in items:
            try:
                h = int(s.split("(")[-1].rstrip("H)"))
            except Exception:
                h = 0
            day_hours += h
        per_emp_day[(wname, d)] += day_hours

    # ================= Assignment Utilization =================
    util_start = 1
    ws.cell(row=util_start, column=1, value="Assignment Utilization").font = bold
    util_hdr = ["employee", "start_date", "end_date", "utilization%"]
    for j, h in enumerate(util_hdr, start=1):
        ws.cell(row=util_start+1, column=j, value=h).font = bold

    worker_name = maps["worker_name"]
    agg = {}  # wid -> {"name":..., "start":date, "end":date, "worked": set(date)}
    for blk in assignment_blocks:
        wid = blk["worker"]
        nm = worker_name.get(wid, wid)
        sd = blk.get("start_date")
        ed = blk.get("end_date")
        wds = set(blk.get("work_dates", []))
        if wid not in agg:
            agg[wid] = {"name": nm, "start": sd, "end": ed, "worked": set(wds)}
        else:
            if sd and (agg[wid]["start"] is None or sd < agg[wid]["start"]):
                agg[wid]["start"] = sd
            if ed and (agg[wid]["end"] is None or ed > agg[wid]["end"]):
                agg[wid]["end"] = ed
            agg[wid]["worked"].update(wds)

    util_rows = []
    for wid, info in agg.items():
        nm = info["name"]
        sd = info["start"]
        ed = info["end"]
        worked_days = len(info["worked"])
        if sd and ed:
            planned_days = (ed - sd).days + 1
            pct = (worked_days / planned_days * 100.0) if planned_days > 0 else 0.0
        else:
            planned_days = 0
            pct = 0.0
        util_rows.append((
            nm,
            sd.isoformat() if sd else "",
            ed.isoformat() if ed else "",
            f"{pct:.1f}%"
        ))

    util_rows.sort(key=lambda r: (r[0] or ""))

    for i, row in enumerate(util_rows, start=util_start+2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    # ================= Overtime & capacity (per-employee) =================
    b0 = util_start + 2 + len(util_rows) + 2
    ws.cell(row=b0, column=1, value="Overtime & capacity").font = bold
    ws.cell(row=b0+1, column=1, value="employee").font = bold
    ws.cell(row=b0+1, column=2, value="total_ot_hours").font = bold

    ot_by_emp = Counter()
    for (wname, d), h in per_emp_day.items():
        if h > OT_THRESHOLD:
            ot_by_emp[wname] += (h - OT_THRESHOLD)
    ot_table = sorted(ot_by_emp.items(), key=lambda kv: kv[1], reverse=True)

    for i, (name, hrs) in enumerate(ot_table, start=b0+2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=int(hrs))

    # ================= Workload balance by employee =================
    d0 = b0 + 2 + len(ot_table) + 2
    ws.cell(row=d0, column=1, value="Workload balance (per employee)").font = bold
    hdr3 = ["employee", "workdays", "total_hours", "avg/day", "stdev/day", "CoV"]
    for j, h in enumerate(hdr3, start=1):
        ws.cell(row=d0+1, column=j, value=h).font = bold

    per_emp_day_list = defaultdict(list)
    for (w, d), h in per_emp_day.items():
        per_emp_day_list[w].append(h)

    balance_rows = []
    for wname in sorted(maps["emp_total_hours"].keys()):
        days = maps["emp_workdays"].get(wname, 0)
        tot  = maps["emp_total_hours"].get(wname, 0)
        avg  = (tot / days) if days else 0.0
        lst  = per_emp_day_list.get(wname, [])
        stdev = 0.0
        if len(lst) >= 2:
            m = sum(lst) / len(lst)
            stdev = math.sqrt(sum((x - m) ** 2 for x in lst) / (len(lst) - 1))
        cov = (stdev / avg * 100.0) if avg else 0.0
        balance_rows.append((wname, days, int(tot), round(avg, 2), round(stdev, 2), f"{cov:.1f}%"))

    for i, row in enumerate(balance_rows, start=d0+2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    # ================= Top 20 total hours + bar chart =================
    top_col = 10
    ws.cell(row=d0, column=top_col, value="Top 20 total hours").font = bold
    ws.cell(row=d0+1, column=top_col,   value="employee").font = bold
    ws.cell(row=d0+1, column=top_col+1, value="total_hours").font = bold

    balance_rows_top20 = sorted(balance_rows, key=lambda r: r[2], reverse=True)[:20]
    for i, row in enumerate(balance_rows_top20, start=d0+2):
        ws.cell(row=i, column=top_col,   value=row[0])
        ws.cell(row=i, column=top_col+1, value=row[2])

    anchor_tot20 = "H2"
    if balance_rows_top20:
        chart4 = BarChart()
        chart4.type = "bar"
        chart4.title = "Total hours by employee (Top 20)"
        chart4.x_axis.title = "hours"
        chart4.y_axis.title = "employee"
        bal0 = d0 + 1
        mrows = len(balance_rows_top20)
        cats = Reference(ws, min_col=top_col,   min_row=bal0+1, max_row=bal0+mrows)
        vals = Reference(ws, min_col=top_col+1, min_row=bal0,   max_row=bal0+mrows)
        chart4.add_data(vals, titles_from_data=True)
        chart4.set_categories(cats)
        chart4.height = 12
        chart4.width = 24
        ws.add_chart(chart4, anchor_tot20)

    # ================= Overtime limit rules by company =================
    ot_rules_row = ws.max_row + 3
    ws.cell(row=ot_rules_row, column=1, value="Overtime limit rules by company").font = bold
    ot_rules_row += 1

    ws.cell(row=ot_rules_row, column=1, value="Company").font = bold
    ws.cell(row=ot_rules_row, column=2, value="Monthly OT limit (h)").font = bold
    ws.cell(row=ot_rules_row, column=3, value="Annual OT limit (h)").font = bold

    wcompanies = env.get("worker_companies", {})
    comp_limits = {}
    for cid, comp in wcompanies.items():
        comp_limits[cid] = {
            "name": comp.get("name", cid),
            "monthly_limit": comp.get("monthly_overtime_limit", None),
            "annual_limit": comp.get("annual_overtime_limit", None),
        }

    for cid, cfg in sorted(comp_limits.items(), key=lambda kv: (kv[1]["name"] or "", kv[0])):
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=cfg["name"])
        ws.cell(row=r, column=2, value=cfg["monthly_limit"] if cfg["monthly_limit"] is not None else "")
        ws.cell(row=r, column=3, value=cfg["annual_limit"]  if cfg["annual_limit"]  is not None else "")

    note_row = ws.max_row + 1
    ws.cell(
        row=note_row, column=1,
        value=f"Note: 'Overtime hours' here = max(0, daily_hours - {OT_THRESHOLD})."
    ).alignment = center

    # ================= Overtime usage by employee (monthly & annual) =================
    table_row = ws.max_row + 3
    ws.cell(row=table_row, column=1, value="Overtime usage by employee (monthly & annual)").font = bold
    table_row += 1

    emp_year_month_ot = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))  # wname -> year -> month -> ot
    years_set = set()
    for (wname, d), h in per_emp_day.items():
        ot = max(0, h - OT_THRESHOLD)
        if ot <= 0:
            continue
        y = d.year
        m = d.month
        emp_year_month_ot[wname][y][m] += ot
        years_set.add(y)

    if years_set:
        years = sorted(years_set)
    else:
        years = list(range(plan_start.year, plan_end.year + 1))

    name_to_company = {}
    for wid, w in env.get("workers", {}).items():
        nm = w.get("name", wid)
        cid = w.get("worker_company")
        cname = wcompanies.get(cid, {}).get("name", cid) if cid else ""
        if nm not in name_to_company:
            name_to_company[nm] = (cid, cname)

    ws.cell(row=table_row, column=1, value="employee").font = bold
    ws.cell(row=table_row, column=2, value="company").font = bold

    col = 3
    for y in years:
        for m in range(1, 12 + 1):
            ws.cell(row=table_row, column=col, value=f"{y}-{m:02d}").font = bold
            col += 1
        ws.cell(row=table_row, column=col, value=f"{y}-total").font = bold
        col += 1

    fill_ot_breach = PatternFill(start_color=PINK_MINMAX, end_color=PINK_MINMAX, fill_type="solid")

    employee_names = sorted({w for (w, _d) in per_emp_day.keys()})

    for idx, wname in enumerate(employee_names, start=1):
        r = table_row + idx
        ws.cell(row=r, column=1, value=wname)

        cid, cname = name_to_company.get(wname, (None, ""))
        ws.cell(row=r, column=2, value=cname)

        cfg = comp_limits.get(cid, {}) if cid else {}
        monthly_limit = cfg.get("monthly_limit", None)
        annual_limit  = cfg.get("annual_limit", None)

        col = 3
        for y in years:
            year_total = 0
            month_map = emp_year_month_ot.get(wname, {}).get(y, {})

            for m in range(1, 12 + 1):
                ot_val = int(month_map.get(m, 0))
                year_total += ot_val
                c = ws.cell(row=r, column=col, value=ot_val)
                if isinstance(monthly_limit, (int, float)) and ot_val > monthly_limit:
                    c.fill = fill_ot_breach
                col += 1

            c_tot = ws.cell(row=r, column=col, value=int(year_total))
            if isinstance(annual_limit, (int, float)) and year_total > annual_limit:
                c_tot.fill = fill_ot_breach
            col += 1

    # ================= Assignment distribution by task (per employee) =================
    task_row0 = ws.max_row + 3
    ws.cell(row=task_row0, column=1, value="Assignment distribution by task (per employee)").font = bold
    task_row0 += 1

    task_headers = [
        "employee", "company",
        "op_id", "op_name",
        "modules",
        "assign_days", "assign_hours",
    ]
    for j, h in enumerate(task_headers, start=1):
        ws.cell(row=task_row0, column=j, value=h).font = bold
    task_row0 += 1

    op_name_map = {}
    for m_id, op_id, op_name in maps["module_ops"]:
        if op_id not in op_name_map:
            op_name_map[op_id] = op_name or op_id

    per_emp_task = defaultdict(lambda: {"dates": set(), "hours": 0, "modules": set(), "wid": None})
    for a in assignments:
        wid = a["worker"]
        op_task = a["operation_task"]
        m_id, op_id = parse_op_task_ids(op_task, maps.get("op_task_index", {}))

        key = (wid, op_id)
        agg = per_emp_task[key]
        agg["wid"] = wid
        agg["dates"].add(a["date"])
        agg["hours"] += a["hours"]
        agg["modules"].add(m_id)

    task_rows = []
    for (wid, op_id), agg in per_emp_task.items():
        wname = maps["worker_name"].get(wid, wid)
        cid, cname = name_to_company.get(wname, (None, ""))
        op_name = op_name_map.get(op_id, op_id)
        assign_days = len(agg["dates"])
        assign_hours = agg["hours"]
        modules_str = ", ".join(sorted(agg["modules"]))

        task_rows.append((
            wname,
            cname,
            op_id, op_name,
            modules_str,
            assign_days, assign_hours,
        ))

    task_rows.sort(key=lambda r: (r[0] or "", r[2] or ""))

    for i, row in enumerate(task_rows, start=task_row0):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    # ================= Region move summary (per employee) =================
    move_row0 = ws.max_row + 3
    ws.cell(row=move_row0, column=1, value="Region move summary (per employee)").font = bold
    move_row0 += 1

    move_headers = [
        "employee", "company",
        "total_regions_with_assignments",
        "total_stay_days",
        "moves_between_regions",
        "moves_per_100_days",
        "routes_summary",
    ]
    for j, h in enumerate(move_headers, start=1):
        ws.cell(row=move_row0, column=j, value=h).font = bold
    move_row0 += 1

    from collections import Counter as Ctr
    regions = env.get("regions", {})
    segments = vios.get("presence_segments", {}) or {}

    move_rows = []
    for wname, seg_list in segments.items():
        if not seg_list:
            continue
        segs = sorted(seg_list, key=lambda s: s["start"])
        unique_regions = set()
        total_days = 0
        moves = 0
        route_counts = Ctr()

        prev_r = None
        for seg in segs:
            r = seg.get("region")
            days = seg.get("days")
            if days is None:
                sd = seg.get("start")
                ed = seg.get("end")
                if sd and ed:
                    days = (ed - sd).days + 1
                else:
                    days = 0
            total_days += days
            if r:
                unique_regions.add(r)
            if prev_r is not None and r is not None and r != prev_r:
                moves += 1
                route_counts[(prev_r, r)] += 1
            prev_r = r

        if total_days <= 0 and not route_counts:
            continue

        moves_per_100 = (moves / total_days * 100.0) if total_days > 0 else 0.0

        pieces = []
        for (r_from, r_to), n in sorted(route_counts.items(), key=lambda x: (x[0][0] or "", x[0][1] or "")):
            from_lbl = regions.get(r_from, {}).get("name", r_from) if r_from else str(r_from)
            to_lbl   = regions.get(r_to, {}).get("name", r_to) if r_to else str(r_to)
            pieces.append(f"{from_lbl}->{to_lbl}×{n}")
        routes_str = " | ".join(pieces) if pieces else ""

        cid, cname = name_to_company.get(wname, (None, ""))
        move_rows.append((
            wname,
            cname,
            len(unique_regions),
            total_days,
            moves,
            f"{moves_per_100:.1f}",
            routes_str,
        ))

    move_rows.sort(key=lambda r: (-r[4], r[0] or ""))

    for i, row in enumerate(move_rows, start=move_row0):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    # ================= Preference match (per employee x fab) =================
    pref_row = ws.max_row + 3
    ws.cell(row=pref_row, column=1, value="Preference match (per employee x fab)").font = bold
    pref_row += 1

    headers = [
        "employee", "company",
        "fab_id", "fab_name",
        "region", "customer_company",
        "region_pref(level)", "company_pref(level)",
        "pref_score(0-100)",
        "assign_days", "assign_hours",
        "modules",
    ]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=pref_row, column=j, value=h).font = bold
    pref_row += 1

    pref_rows = build_preference_match_rows(env, modules, assignments, maps)
    for i, r in enumerate(pref_rows, start=pref_row):
        ws.cell(row=i, column=1,  value=r["employee"])
        ws.cell(row=i, column=2,  value=r["company"])
        ws.cell(row=i, column=3,  value=r["fab_id"])
        ws.cell(row=i, column=4,  value=r["fab_name"])
        ws.cell(row=i, column=5,  value=r["region"])
        ws.cell(row=i, column=6,  value=r["customer"])
        ws.cell(row=i, column=7,  value=r["region_pref"])
        ws.cell(row=i, column=8,  value=r["company_pref"])
        ws.cell(row=i, column=9,  value=r["pref_score"])
        ws.cell(row=i, column=10, value=r["assign_days"])
        ws.cell(row=i, column=11, value=r["assign_hours"])
        ws.cell(row=i, column=12, value=r["modules"])



 #=========================== SHEET 5 :write sheet breashes plan ==================================

def write_sheet_breaches_plan(wb, vios, page_scope="fix", sheet_name="Bre Plan fix"):
    bold = Font(bold=True)
    ws = wb.create_sheet(sheet_name)

    task_day_scopes = vios.get("task_day_scopes", {}) or {}
    task_scopes = vios.get("task_scopes", {}) or {}
    phase_scopes = vios.get("phase_scopes", {}) or {}

    def scope_window(row):
        # row: [date, module, phase, op_id, worker, reason, phase_start, phase_end]
        return _merge_scope({
            task_day_scopes.get((row[1], row[3], row[0]), ""),
            phase_scopes.get((row[1], row[2]), ""),
        })

    def scope_order(row):
        return _merge_scope({
            task_day_scopes.get((row[1], row[3], row[0]), ""),
            phase_scopes.get((row[1], row[2]), ""),
        })

    def scope_unavail_task(row):
        return _merge_scope({task_day_scopes.get((row[1], row[2], row[0]), "")})

    tbl_window = [r for r in (vios.get("tbl_window", []) or []) if _scope_matches(page_scope, scope_window(r))]
    tbl_order = [r for r in (vios.get("tbl_order", []) or []) if _scope_matches(page_scope, scope_order(r))]
    tbl_unavail_task = [r for r in (vios.get("tbl_unavail_task", []) or []) if _scope_matches(page_scope, scope_unavail_task(r))]

    summary = [
        ("Phase window breaches", len(tbl_window)),
        ("Phase ordering breaches", len(tbl_order)),
        ("Unavailable breaches (Tasks)", len(tbl_unavail_task)),
    ]

    ws.cell(row=1, column=1, value="Breach counts (rows)").font = bold
    ws.cell(row=2, column=1, value="category").font = bold
    ws.cell(row=2, column=2, value="count").font = bold
    for i, (k, v) in enumerate(summary, start=3):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=int(v))

    row = 3 + len(summary) + 2
    def write_table(title, headers, rows):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = bold
        row += 1
        for j, h in enumerate(headers, start=1):
            ws.cell(row=row, column=j, value=h).font = bold
        row += 1
        for r in rows:
            for j, v in enumerate(r, start=1):
                ws.cell(row=row, column=j, value=v)
            row += 1
        row += 2

    write_table(
        "Phase window breaches",
        ["date", "module", "phase", "op_id", "worker", "reason", "phase_start", "phase_end"],
        sorted(tbl_window, key=lambda x: (x[1], x[2], x[0]))
    )

    write_table(
        "Phase ordering breaches",
        ["date", "module", "phase(later)", "op_id", "worker", "required_prev_phase_last_date"],
        sorted(tbl_order, key=lambda x: (x[1], x[2], x[0]))
    )

    write_table(
        "Unavailable breaches (Tasks)",
        ["date", "module", "op_id", "reason"],
        sorted(tbl_unavail_task, key=lambda x: (x[1], x[2], x[0]))
    )

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18

# ====================== SHEET 6 :Write sheet breaches employee =========================

def write_sheet_breaches_employee(wb, vios, page_scope="fix", sheet_name="Bre Emp fix"):
    bold = Font(bold=True)
    ws = wb.create_sheet(sheet_name)

    task_day_scopes = vios.get("task_day_scopes", {}) or {}
    task_scopes = vios.get("task_scopes", {}) or {}
    emp_day_scopes = vios.get("emp_day_scopes", {}) or {}
    emp_scopes = vios.get("emp_scopes", {}) or {}
    worker_scopes = vios.get("worker_scopes", {}) or {}

    def scope_skill(row):
        return _merge_scope({
            emp_day_scopes.get((row[1], row[2], row[0]), ""),
            task_scopes.get((row[3], row[4]), ""),
        })

    def scope_minmax(row):
        return _merge_scope({task_day_scopes.get((row[1], row[2], row[0]), "")})

    def scope_no_manager(row):
        return _merge_scope({task_scopes.get((row[0], row[1]), "")})

    def scope_unavail_emp(row):
        return _merge_scope({
            emp_day_scopes.get((row[1], row[2], row[0]), ""),
            task_scopes.get((row[4], row[5]), ""),
        })

    def scope_worker_row(row):
        return _merge_scope({worker_scopes.get(row[0], "")})

    def scope_move_gap(row):
        # Pair-level scope: fix only for Fixed->Fixed; flex for Fixed->Flexible / Flexible->Fixed / Flexible->Flexible.
        if len(row) >= 8:
            return row[7]
        return scope_worker_row(row)

    def scope_pref(row):
        return _merge_scope({worker_scopes.get(row[0], "")})

    tbl_skill = [r for r in (vios.get("tbl_skill", []) or []) if _scope_matches(page_scope, scope_skill(r))]
    tbl_minmax = [r for r in (vios.get("tbl_minmax", []) or []) if _scope_matches(page_scope, scope_minmax(r))]
    tbl_no_manager = [r for r in (vios.get("tbl_no_manager", []) or []) if _scope_matches(page_scope, scope_no_manager(r))]
    tbl_unavail_emp = [r for r in (vios.get("tbl_unavail_emp", []) or []) if _scope_matches(page_scope, scope_unavail_emp(r))]
    tbl_overstay = [r for r in (vios.get("tbl_overstay", []) or []) if _scope_matches(page_scope, scope_worker_row(r))]
    tbl_move_gap = [r for r in (vios.get("tbl_move_gap", []) or []) if _scope_matches(page_scope, scope_move_gap(r))]
    tbl_ot_month = [r for r in (vios.get("tbl_ot_month", []) or []) if _scope_matches(page_scope, scope_worker_row(r))]
    tbl_ot_annual = [r for r in (vios.get("tbl_ot_annual", []) or []) if _scope_matches(page_scope, scope_worker_row(r))]
    tbl_pref_breach = [r for r in (vios.get("tbl_pref_breach", []) or []) if _scope_matches(page_scope, scope_pref(r))]

    summary = [
        ("Skill mismatches", len(tbl_skill)),
        ("Staffing min/max breaches", len(tbl_minmax)),
        ("Tasks with no manager", len(tbl_no_manager)),
        ("Unavailable breaches (Employees)", len(tbl_unavail_emp)),
        ("Region overstay", len(tbl_overstay)),
        ("Region change gap", len(tbl_move_gap)),
        ("Overtime limit breaches (monthly)", len(tbl_ot_month)),
        ("Overtime limit breaches (annual)", len(tbl_ot_annual)),
        ("Preference breaches (suitability level 0)", len(tbl_pref_breach)),
    ]

    ws.cell(row=1, column=1, value="Breach counts (rows)").font = bold
    ws.cell(row=2, column=1, value="category").font = bold
    ws.cell(row=2, column=2, value="count").font = bold
    for i, (k, v) in enumerate(summary, start=3):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=int(v))

    row = 3 + len(summary) + 2
    def write_table(title, headers, rows):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = bold
        row += 1
        for j, h in enumerate(headers, start=1):
            ws.cell(row=row, column=j, value=h).font = bold
        row += 1
        for r in rows:
            for j, v in enumerate(r, start=1):
                ws.cell(row=row, column=j, value=v)
            row += 1
        row += 2

    write_table(
        "Skill mismatches",
        ["date", "worker", "company", "module", "op_id"],
        sorted(tbl_skill, key=lambda x: (x[1], x[2], x[0]))
    )

    write_table(
        "Staffing min/max breaches",
        ["date", "module", "op_id", "heads", "min", "max", "status"],
        sorted(tbl_minmax, key=lambda x: (x[0], x[1], x[2]))
    )

    write_table(
        "Tasks with no manager",
        ["module", "op_id"],
        sorted(tbl_no_manager, key=lambda x: (x[0], x[1]))
    )

    write_table(
        "Unavailable breaches (Employees)",
        ["date", "worker", "company", "reason", "module", "op_id"],
        sorted(tbl_unavail_emp, key=lambda x: (x[1], x[0]))
    )

    write_table(
        "Region overstay (exceeded max stay in a country/region)",
        ["worker", "region", "start_date", "end_date", "days", "stay_limit"],
        sorted(tbl_overstay, key=lambda x: (x[0], x[1], x[2]))
    )

    write_table(
        "Region change without sufficient gap",
        ["worker", "from_region", "to_region", "out_date", "in_date", "gap_days", "required_gap"],
        [r[:7] for r in sorted(tbl_move_gap, key=lambda x: (x[0], x[3], x[4]))]
    )

    write_table(
        "Overtime limit breaches (monthly)",
        ["worker", "company", "year", "month", "ot_hours", "monthly_limit"],
        sorted(tbl_ot_month, key=lambda x: (x[0], x[2], x[3]))
    )

    write_table(
        "Overtime limit breaches (annual)",
        ["worker", "company", "year", "ot_hours", "annual_limit"],
        sorted(tbl_ot_annual, key=lambda x: (x[0], x[2]))
    )

    write_table(
        "Preference breaches (assigned with suitability level 0)",
        ["worker", "company", "module", "fab_id", "region", "customer_company", "reason"],
        sorted(tbl_pref_breach, key=lambda x: (x[0], x[2]))
    )

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18

# ===================== SHEET X: wf_tool one-day isolated work =====================
def write_sheet_wftool_isolated_workdays(wb, plan_start, plan_end, env, maps, modules, assignments, vios, cal):
    bold   = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws = wb.create_sheet("Isolated Days")
    ws.cell(row=1, column=1, value="isolated one-day work").font = bold

    hdr = [
        "employee",
        "company",
        "date",
        "module_id",
        "module_name",
    ]
    for j, h in enumerate(hdr, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = bold
        c.alignment = center

    for col in range(1, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 40

    module_workflow = maps.get("module_workflow", {}) or {}
    worker_name = maps.get("worker_name", {}) or {}

    # module_id -> module_name
    module_name_map = {}
    if modules:
        if isinstance(modules, list):
            for m in modules:
                mid = (m or {}).get("id")
                mname = (m or {}).get("name")
                if mid:
                    module_name_map[str(mid)] = str(mname or "")
        elif isinstance(modules, dict):
            for mid, m in modules.items():
                mname = (m or {}).get("name")
                module_name_map[str(mid)] = str(mname or "")

    # name -> wid
    name_to_wid = {}
    for wid, w in (env.get("workers", {}) or {}).items():
        nm = worker_name.get(wid, w.get("name", wid))
        if nm and nm not in name_to_wid:
            name_to_wid[nm] = wid

    roster = []
    seen = set()
    for (comp, wname, _dte), _ in maps.get("edt", {}).items():
        if (comp, wname) not in seen:
            roster.append((comp, wname))
            seen.add((comp, wname))
    roster.sort(key=lambda t: (t[0] or "", t[1] or ""))

    def daterange(a: date, b: date):
        d = a
        while d <= b:
            yield d
            d += timedelta(days=1)

    def is_white_neighbor_cell(comp: str, wname: str, dt: date) -> bool:
        wid = name_to_wid.get(wname)
        personal_off = bool(wid) and (dt in cal["worker_off"].get(wid, set()))
        if personal_off:
            return False
        if (comp, wname, dt) in (vios.get("skill_mismatch_cells") or set()):
            return False
        if (comp, wname, dt) in (vios.get("unavail_emp_cells") or set()):
            return False
        return True

    def parse_module_id_from_item_text(item_text: str):
        s0 = str(item_text)
        # longest match first, safer when one id is prefix of another
        mids = sorted(module_workflow.keys(), key=lambda x: len(str(x)), reverse=True)
        for mid in mids:
            if s0.startswith(str(mid)):
                return str(mid)
        return None

    def wf_tool_module_ids_for_day(comp: str, wname: str, dt: date):
        items = list(maps.get("edt", {}).get((comp, wname, dt), []) or [])
        if not items:
            return []

        module_ids = []
        for s in items:
            mid = parse_module_id_from_item_text(s)
            if not mid:
                continue
            if module_workflow.get(mid) == "wf_tool":
                module_ids.append(mid)

        return sorted(set(module_ids))

    rows = []
    start_dt = plan_start + timedelta(days=1)
    end_dt   = plan_end - timedelta(days=1)

    if start_dt <= end_dt:
        for (comp, wname) in roster:
            for dt in daterange(start_dt, end_dt):
                module_ids_today = wf_tool_module_ids_for_day(comp, wname, dt)
                if not module_ids_today:
                    continue

                ldt = dt - timedelta(days=1)
                rdt = dt + timedelta(days=1)

                left_raw  = maps.get("edt", {}).get((comp, wname, ldt), []) or []
                right_raw = maps.get("edt", {}).get((comp, wname, rdt), []) or []
                if left_raw or right_raw:
                    continue

                if not is_white_neighbor_cell(comp, wname, ldt):
                    continue
                if not is_white_neighbor_cell(comp, wname, rdt):
                    continue

                module_names_today = [
                    module_name_map.get(mid, "")
                    for mid in module_ids_today
                ]
                module_names_today = [x for x in module_names_today if x]

                rows.append((
                    wname,
                    comp,
                    dt.isoformat(),
                    " | ".join(module_ids_today),
                    " | ".join(sorted(set(module_names_today))),
                ))

    if not rows:
        ws.cell(row=3, column=1, value="No isolated one-day wf_tool work found")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(hdr))
        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(hdr))}2"
        return

    rows.sort(key=lambda r: (r[1] or "", r[0] or "", r[2]))
    out_r = 3
    for row in rows:
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=out_r, column=j, value=v)
            c.alignment = left if j in (1, 2, 4, 5) else center
        out_r += 1

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(hdr))}{ws.max_row}"

# ===================== SHEET Y: wf_tool module phase-gap (Schedule windows) =====================
# Track module-level gaps between consecutive phases:
#   gap_days = next_phase_start - prev_phase_end - 1
# Only for modules where module_workflow == "wf_tool"
#
def write_sheet_wftool_module_phase_gaps(wb, plan_start, plan_end, modules, maps, assignments):
    """
    Track gaps between consecutive phases for wf_tool modules, using *actual* assignments only.
    Columns:
      - module
      - phase_prev
      - last_assign   : last real work day in previous phase
      - phase_next
      - next_assign  : first real work day in next phase
      - gap_days    : (next_first - prev_last) - 1   [kept only if > 0]
    """
    bold   = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws = wb.create_sheet("Phase Gaps")
    ws.cell(row=1, column=1, value="Module phase gaps").font = bold

    hdr = [
        "module",
        "phase_prev",
        "last_assign",
        "phase_next",
        "next_assign",
        "gap_days",
    ]
    for j, h in enumerate(hdr, start=1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = bold
        c.alignment = center

    for col in range(1, len(hdr) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.freeze_panes = "A3"

    # Indices prepared earlier in your file
    module_workflow = maps.get("module_workflow", {}) or {}
    op_task_index   = maps.get("op_task_index", {}) or {} 

    # ---------- Build per-(module, phase) first/last assignment dates ----------
    from collections import defaultdict
    first_in_phase = defaultdict(lambda: defaultdict(lambda: None))  # m -> phase -> date
    last_in_phase  = defaultdict(lambda: defaultdict(lambda: None))  # m -> phase -> date

    for a in assignments:
        ot_id = a.get("operation_task")
        dt    = a.get("date")
        meta  = op_task_index.get(ot_id)
        if not meta:
            # Unknown op_task id -> skip safely (older dataset variants are handled by your other helpers)
            continue
        m_id  = meta.get("m_id")
        ph_id = meta.get("phase")
        if not m_id or not ph_id or not isinstance(dt, date):
            continue

        f_prev = first_in_phase[m_id].get(ph_id)
        if f_prev is None or dt < f_prev:
            first_in_phase[m_id][ph_id] = dt

        l_prev = last_in_phase[m_id].get(ph_id)
        if l_prev is None or dt > l_prev:
            last_in_phase[m_id][ph_id] = dt
    # --------------------------------------------------------------------------

    rows = []
    for m in modules:
        m_id = m.get("id")
        if not m_id:
            continue
        # robust 'wf_tool' detection whether YAML uses workflow or workflow_id
        wf = module_workflow.get(m_id) or m.get("workflow") or m.get("workflow_id")
        if wf != "wf_tool":
            continue

        phase_list = list(m.get("phase_task_list", []) or [])

        # sort by start_date to be safe (uses the same _d() helper you already have)
        def _ph_key(ph):
            try:
                return _d(ph.get("start_date"))
            except Exception:
                return date.min
        phase_list.sort(key=_ph_key)  # keeps phases in chronological order

        for i in range(len(phase_list) - 1):
            p1 = phase_list[i]
            p2 = phase_list[i + 1]
            try:
                p1_id = p1.get("phase", "")
                p2_id = p2.get("phase", "")
            except Exception:
                continue

            prev_last  = last_in_phase.get(m_id, {}).get(p1_id)
            next_first = first_in_phase.get(m_id, {}).get(p2_id)

            # Only compute a gap if both sides have assignments
            assign_gap = None
            if isinstance(prev_last, date) and isinstance(next_first, date):
                assign_gap = (next_first - prev_last).days - 1

            # Keep only positive gaps
            if assign_gap is not None and assign_gap > 0:
                rows.append((
                    m_id,
                    p1_id,
                    prev_last.isoformat(),
                    p2_id,
                    next_first.isoformat(),
                    int(assign_gap),
                ))

    if not rows:
        ws.cell(row=3, column=1, value="No positive assignment gaps found for wf_tool modules.")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(hdr))
        ws.auto_filter.ref = f"A2:{get_column_letter(len(hdr))}{ws.max_row}"
        return

    # Sort: biggest gaps first, then module/phase
    rows.sort(key=lambda r: (-r[5], r[0] or "", r[1] or ""))

    out_r = 3
    for row in rows:
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=out_r, column=j, value=v)
            c.alignment = left if j in (1, 2, 4) else center
        out_r += 1

    ws.auto_filter.ref = f"A2:{get_column_letter(len(hdr))}{ws.max_row}"

# ----------------------- AFFINITY ANALYSIS -----------------------
def build_affinity_analysis(env, modules, assignments, maps):
    """
    Analyse pair-level affinity for workers assigned to the same (module, operation).

    Affinity score for a pair = sum of weights of tags shared by both workers.
    Positive weight tags contribute positively; negative weight tags lower the score.

    Returns:
        group_rows  – one dict per (m_id, op_id) group
        pair_rows   – one dict per unique (worker_a, worker_b, m_id, op_id) pair
    """
    tag_weights = env.get("affinity_tags", {})          # tag_id -> weight
    workers     = env.get("workers", {})
    worker_name = maps["worker_name"]

    # worker_id -> frozenset of tag ids
    worker_tags = {}
    for wid, cfg in workers.items():
        worker_tags[wid] = frozenset(cfg.get("affinity") or [])

    # Collect unique workers per (m_id, op_id) group
    op_task_index = maps["op_task_index"]
    group_workers = defaultdict(set)   # (m_id, op_id) -> {wid}
    for a in assignments:
        ot_id = a["operation_task"]
        meta  = op_task_index.get(ot_id)
        if not meta:
            continue
        group_workers[(meta["m_id"], meta["op_id"])].add(a["worker"])

    # Build module/op name lookup
    op_name_map = {}  # (m_id, op_id) -> op_name
    mod_name_map = {}
    for m in modules:
        mod_name_map[m["id"]] = m.get("name", m["id"])
        for ph in m.get("phase_task_list", []):
            for ot in ph.get("operation_task_list", []):
                op_name_map[(m["id"], ot["operation"])] = ot.get("name", ot["operation"])

    pair_rows  = []
    group_rows = []

    for (m_id, op_id), wids in sorted(group_workers.items()):
        wid_list = sorted(wids)
        pairs = list(combinations(wid_list, 2))

        group_total = 0
        group_pos   = 0
        group_neg   = 0

        for w1, w2 in pairs:
            tags1  = worker_tags.get(w1, frozenset())
            tags2  = worker_tags.get(w2, frozenset())
            shared = tags1 & tags2
            score  = sum(tag_weights.get(t, 0) for t in shared)
            pos    = sum(tag_weights.get(t, 0) for t in shared if tag_weights.get(t, 0) > 0)
            neg    = sum(tag_weights.get(t, 0) for t in shared if tag_weights.get(t, 0) < 0)

            pair_rows.append({
                "module":      m_id,
                "module_name": mod_name_map.get(m_id, m_id),
                "op_id":       op_id,
                "op_name":     op_name_map.get((m_id, op_id), op_id),
                "worker_a":    worker_name.get(w1, w1),
                "worker_b":    worker_name.get(w2, w2),
                "shared_tags": ", ".join(sorted(shared)) if shared else "-",
                "score":       score,
                "pos_score":   pos,
                "neg_score":   neg,
            })
            group_total += score
            group_pos   += pos
            group_neg   += neg

        n_workers = len(wid_list)
        n_pairs   = len(pairs)
        avg_score = round(group_total / n_pairs, 2) if n_pairs else 0

        group_rows.append({
            "module":      m_id,
            "module_name": mod_name_map.get(m_id, m_id),
            "op_id":       op_id,
            "op_name":     op_name_map.get((m_id, op_id), op_id),
            "n_workers":   n_workers,
            "n_pairs":     n_pairs,
            "total_score": group_total,
            "avg_score":   avg_score,
            "pos_score":   group_pos,
            "neg_score":   group_neg,
        })

    # Sort: lowest (worst) avg_score first so problematic groups are visible at top
    group_rows.sort(key=lambda r: (r["avg_score"], r["module"], r["op_id"]))
    pair_rows.sort(key=lambda r: (r["score"], r["module"], r["op_id"], r["worker_a"], r["worker_b"]))

    return group_rows, pair_rows


def write_sheet_affinity(wb, env, modules, assignments, maps):
    """Affinity analysis sheet – inserted before Moving plan."""
    ws = wb.create_sheet("Affinity")

    bold   = Font(bold=True, name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right",  vertical="center")

    # ---- Colors ----
    HDR_FILL    = PatternFill("solid", start_color="2F5496", end_color="2F5496")  # dark blue
    HDR_FONT    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    TITLE_FONT  = Font(bold=True, name="Arial", size=13)
    SUB_FONT    = Font(bold=True, name="Arial", size=11)
    POS_FILL    = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")  # green
    NEG_FILL    = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")  # red
    ZERO_FILL   = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")  # yellow
    STRIPE_FILL = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")  # light grey
    SECT_FILL   = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")  # light blue

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    tag_weights = env.get("affinity_tags", {})
    group_rows, pair_rows = build_affinity_analysis(env, modules, assignments, maps)

    # ---- helper ----
    def score_fill(score):
        if score > 0:  return POS_FILL
        if score < 0:  return NEG_FILL
        return ZERO_FILL

    cur_row = 1

    # ===== TITLE =====
    ws.cell(row=cur_row, column=1, value="Affinity Analysis").font = TITLE_FONT
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=10)
    cur_row += 1

    # ===== Tag legend =====
    ws.cell(row=cur_row, column=1, value="Tag Legend").font = SUB_FONT
    cur_row += 1
    tag_hdr = ["Tag ID", "Weight", "Direction"]
    for ci, h in enumerate(tag_hdr, 1):
        c = ws.cell(row=cur_row, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = center; c.border = border
    cur_row += 1
    for tid, w in sorted(tag_weights.items()):
        direction = "Positive" if w > 0 else ("Negative" if w < 0 else "Neutral")
        vals = [tid, w, direction]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=cur_row, column=ci, value=v)
            c.alignment = center; c.border = border
            if ci == 2:
                c.fill = score_fill(w)
        cur_row += 1
    cur_row += 1  # blank separator
    tag_legend_freeze_row = cur_row  # everything above this row will be frozen

    # ===== SECTION 1 : Group Summary =====
    ws.cell(row=cur_row, column=1, value="Section 1 – Group Affinity Summary").font = SUB_FONT
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=10)
    ws.cell(row=cur_row, column=1).fill = SECT_FILL
    cur_row += 1

    grp_hdrs = [
        "Module", "Module Name", "Op ID", "Op Name",
        "Workers", "Pairs", "Total Score", "Avg Pair Score",
        "Positive", "Negative",
    ]
    for ci, h in enumerate(grp_hdrs, 1):
        c = ws.cell(row=cur_row, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = center; c.border = border
    grp_hdr_row = cur_row
    cur_row += 1

    for ri, r in enumerate(group_rows):
        fill = STRIPE_FILL if ri % 2 == 1 else None
        vals = [
            r["module"], r["module_name"], r["op_id"], r["op_name"],
            r["n_workers"], r["n_pairs"], r["total_score"], r["avg_score"],
            r["pos_score"], r["neg_score"],
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=cur_row, column=ci, value=v)
            c.border = border
            c.alignment = left if ci in (1, 2, 3, 4) else center
            if fill:
                c.fill = fill
        # Score coloring on total and avg columns
        ws.cell(row=cur_row, column=7).fill = score_fill(r["total_score"])
        ws.cell(row=cur_row, column=8).fill = score_fill(r["avg_score"])
        cur_row += 1

    grp_data_end = cur_row - 1
    cur_row += 2  # blank separator

    # ===== SECTION 2 : Pair Detail =====
    ws.cell(row=cur_row, column=1, value="Section 2 – Pair Detail").font = SUB_FONT
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=10)
    ws.cell(row=cur_row, column=1).fill = SECT_FILL
    cur_row += 1

    pair_hdrs = [
        "Module", "Module Name", "Op ID", "Op Name",
        "Worker A", "Worker B", "Shared Tags", "Score",
        "Positive", "Negative",
    ]
    for ci, h in enumerate(pair_hdrs, 1):
        c = ws.cell(row=cur_row, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = center; c.border = border
    pair_hdr_row = cur_row
    cur_row += 1

    for ri, r in enumerate(pair_rows):
        fill = STRIPE_FILL if ri % 2 == 1 else None
        vals = [
            r["module"], r["module_name"], r["op_id"], r["op_name"],
            r["worker_a"], r["worker_b"], r["shared_tags"], r["score"],
            r["pos_score"], r["neg_score"],
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=cur_row, column=ci, value=v)
            c.border = border
            c.alignment = left if ci in (1, 2, 3, 4, 5, 6, 7) else center
            if fill:
                c.fill = fill
        ws.cell(row=cur_row, column=8).fill = score_fill(r["score"])
        cur_row += 1

    pair_data_end = cur_row - 1

    # ===== Auto-filter =====
    ws.auto_filter.ref = (
        f"A{grp_hdr_row}:{get_column_letter(len(grp_hdrs))}{grp_data_end}"
    )

    # ===== Column widths =====
    col_widths = [12, 20, 8, 16, 10, 7, 13, 14, 10, 10]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Freeze tag legend at top; Section 1 and Section 2 scroll freely
    ws.freeze_panes = f"A{tag_legend_freeze_row}"


# ----------------------- SHEET : MOVING PLAN CALENDAR -----------------------
def write_sheet_moving_plan(wb, plan_start, plan_end, env, maps, vios):
    """
    Calendar per employee:
      - Cell is colored by region presence (continuous stay includes off days up to stay_gap).
      - Writes 'out' on the day AFTER leaving a region, 'in' on the day BEFORE entering the next region.
      - If an employee has no presence that day, cell is blank.
    """
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws = wb.create_sheet("Moving plan")

    # Region color palette (extend as needed; we cycle when exceeding)
    REGION_PALETTE = [
        ("FF9999"),  # red
        ("9DC3E6"),  # blue
        ("A9D18E"),  # green
        ("F6B5C9"),  # pink
        ("FFD966"),  # yellow
        ("C9C9C9"),  # grey
        ("B4A7D6"),  # purple
        ("F4B183"),  # orange
    ]

    regions = env["regions"] or {}
    # Create stable mapping region_id -> color
    region_ids = sorted(regions.keys())
    region_to_fill = {}
    for idx, rid in enumerate(region_ids):
        color = REGION_PALETTE[idx % len(REGION_PALETTE)]
        region_to_fill[rid] = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Header
    base_headers = ["company", "employee"]
    for i, h in enumerate(base_headers, start=1):
        c = ws.cell(row=1, column=i, value=h); c.font = bold

    # Dates row
    dates = []
    d = plan_start
    while d <= plan_end:
        dates.append(d); d += timedelta(days=1)

    for j, dt in enumerate(dates, start=len(base_headers)+1):
        c = ws.cell(row=1, column=j, value=dt.strftime("%Y-%m-%d"))
        c.font = bold
        ws.column_dimensions[get_column_letter(j)].width = 10

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.freeze_panes = "C2"

    # Roster from maps["edt"] to keep ordering consistent with other sheets
    roster = []
    seen = set()
    for (comp, wname, _dte), _ in maps["edt"].items():
        if (comp, wname) not in seen:
            roster.append((comp, wname)); seen.add((comp, wname))
    roster.sort(key=lambda t: (t[0] or "", t[1] or ""))

    # Quick lookups
    presence_map = vios.get("presence_map", {})          # (wname, date) -> region_id
    move_markers = vios.get("move_markers", {})          # wname -> list of {type, date, region}

    # Render rows
    for i, (comp, wname) in enumerate(roster, start=2):
        ws.row_dimensions[i].height = 20
        ws.cell(row=i, column=1, value=comp or "").alignment = left
        ws.cell(row=i, column=2, value=wname or "").alignment = left

        # For rapid 'in/out' lookup
        markers = {(m["date"], m["type"]): m.get("region") for m in move_markers.get(wname, [])}

        for j, dt in enumerate(dates, start=len(base_headers)+1):
            rid = presence_map.get((wname, dt))
            cell = ws.cell(row=i, column=j, value="")
            cell.alignment = center

            # Color by presence
            if rid in region_to_fill:
                cell.fill = region_to_fill[rid]

            # Write 'out' on the day AFTER leaving (provided by markers)
            if (dt, "out") in markers:
                cell.value = "out"
            # Write 'in' on the day BEFORE entering
            if (dt, "in") in markers:
                cell.value = "in"

    # Legend block
    legend_row = ws.max_row + 2
    ws.cell(row=legend_row, column=1, value="Legend (Region Colors)").font = bold
    rr = legend_row + 1
    c1 = 1
    for rid in region_ids:
        lab = f"{rid} : {regions[rid].get('name', rid)}"
        ws.cell(row=rr, column=c1, value=lab)
        box = ws.cell(row=rr, column=c1+1, value="")
        box.fill = region_to_fill[rid]
        rr += 1

# -------------------------------- MAIN ---------------------------------
def main():
    ap = ArgumentParser(description="Read Schedule.yaml + EnvConfig.yaml and export Excel.")
    ap.add_argument("--schedule", default="Schedule.yaml")
    ap.add_argument("--env",      default="EnvConfig.yaml")
    ap.add_argument("--out",      default="schedule_export.xlsx")
    args = ap.parse_args()

    plan_start, plan_end, modules, assignments, assignment_blocks = load_schedule(args.schedule)
    env  = load_env(args.env)

    # build calendar closures once
    cal = build_unavailability(plan_start, plan_end, env)

    maps = build_maps(env, modules, assignments)

    req_task, req_module = compute_required_hours_task_module(modules)
    vios = detect_violations(env, modules, assignments, maps, cal, plan_start, plan_end)

    wb = Workbook()
    # remove default
    del wb[wb.sheetnames[0]]

    # ===== SHEET 1 =====
    write_sheet_tasks_dates(wb, plan_start, plan_end, env, maps, modules, vios, req_task, cal)
    # ===== SHEET 2 =====
    write_sheet_employees_dates(wb, plan_start, plan_end, env, maps, vios, cal)

    # ===== SHEET 3: Dashboard (Plan) =====
    write_sheet_dashboard_plan(
        wb, plan_start, plan_end, env, maps,
        modules, assignments, assignment_blocks, req_module
    )

    # ===== SHEET 4: Recommend staffing deviation (soft) =====
    write_sheet_recommend_staffing(wb, plan_start, plan_end, modules, maps)

    # ===== SHEET 5: Dashboard (Employees) =====
    write_sheet_dashboard_employees(
        wb, plan_start, plan_end, env, maps,
        modules, assignments, assignment_blocks, req_module, vios
    )
    # ===== SHEET 6-9: Breaches split by Fixed/Flexible =====
    write_sheet_breaches_plan(wb, vios, page_scope="fix", sheet_name="Bre Plan fix")
    write_sheet_breaches_plan(wb, vios, page_scope="flex", sheet_name="Bre Plan flex")
    write_sheet_breaches_employee(wb, vios, page_scope="fix", sheet_name="Bre Emp fix")
    write_sheet_breaches_employee(wb, vios, page_scope="flex", sheet_name="Bre Emp flex")
    # ===== isolated 1-day work =====
    write_sheet_wftool_isolated_workdays(
        wb, plan_start, plan_end, env, maps, modules, assignments, vios, cal
    )
    # ===== module phase gaps =====
    write_sheet_wftool_module_phase_gaps(
        wb, plan_start, plan_end, modules, maps, assignments
    )
    # ===== Affinity analysis =====
    write_sheet_affinity(wb, env, modules, assignments, maps)
    # ===== SHEET 8: Moving plan =====
    write_sheet_moving_plan(wb, plan_start, plan_end, env, maps, vios)

    wb.save(args.out)
    print(f"Wrote Excel: {os.path.abspath(args.out)}")


if __name__ == "__main__":
    main()