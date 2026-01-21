"""Excelファイルの読み込み機能"""
from typing import List, Tuple, Optional
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


class ExcelReader:
    """Excelファイルを読み込むクラス"""
    
    def __init__(self, file_path: str, sheet_name: str):
        """
        Args:
            file_path: Excelファイルのパス
            sheet_name: シート名
        """
        self.workbook = load_workbook(file_path, data_only=False)
        self.sheet = self.workbook[sheet_name]
    
    def read_data1(self, columns: List[str], start_row: int, end_row: Optional[int] = None) -> dict[int, List[str]]:
        """
        データ1を読み込む: 指定された複数の列の全ての行のデータの文字列を取得
        
        Args:
            columns: 列名のリスト (例: ["A", "B", "C"])
            start_row: 開始行番号
            end_row: 終了行番号 (Noneの場合はシートの最終行まで)
        
        Returns:
            {行番号: [列1の値, 列2の値, ...], ...}
        """
        if end_row is None:
            end_row = self.sheet.max_row
        
        result = {}
        for row_num in range(start_row, end_row + 1):
            row_values = []
            for col in columns:
                cell = self.sheet[f"{col}{row_num}"]
                value = cell.value if cell.value is not None else ""
                row_values.append(str(value))
            result[row_num] = row_values
        
        return result
    
    def read_data2(self, top_left: str, bottom_right: str) -> List[dict]:
        """
        データ2を読み込む: 指定範囲のセルデータを書式付きで取得
        
        Args:
            top_left: 左上のセル (例: "J46")
            bottom_right: 右下のセル (例: "AC112")
        
        Returns:
            セル情報のリスト [
                {
                    'row': 行番号,
                    'col': 列名,
                    'value': セルの値,
                    'fill_color': 塗りつぶし色 (RGB形式、なければNone)
                },
                ...
            ]
        """
        # 範囲の解析
        start_col, start_row = self._parse_cell_address(top_left)
        end_col, end_row = self._parse_cell_address(bottom_right)
        
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        result = []
        for row_num in range(start_row, end_row + 1):
            for col_idx in range(start_col_idx, end_col_idx + 1):
                col_name = get_column_letter(col_idx)
                cell = self.sheet[f"{col_name}{row_num}"]
                
                # セルの値
                value = cell.value if cell.value is not None else ""
                
                # 塗りつぶし色の取得
                fill_color = self._get_fill_color(cell)
                
                result.append({
                    'row': row_num,
                    'col': col_name,
                    'value': str(value),
                    'fill_color': fill_color
                })
        
        return result
    
    def _parse_cell_address(self, address: str) -> Tuple[str, int]:
        """
        セルアドレスを列名と行番号に分割
        
        Args:
            address: セルアドレス (例: "J46")
        
        Returns:
            (列名, 行番号)
        """
        col = ""
        row = ""
        for char in address:
            if char.isalpha():
                col += char
            else:
                row += char
        return col, int(row)
    
    def read_data3(self, start_cell: str, end_cell: str) -> dict[str, str]:
        """
        データ3を読み込む: 月の情報を持つ1行のデータ
        
        Args:
            start_cell: 開始セル (例: "J3")
            end_cell: 終了セル (例: "RR3")
        
        Returns:
            {列名: セルの値, ...}
        
        Raises:
            ValueError: 開始と終了が異なる行の場合
        """
        start_col, start_row = self._parse_cell_address(start_cell)
        end_col, end_row = self._parse_cell_address(end_cell)
        
        if start_row != end_row:
            raise ValueError(
                f"データ3は1行分のデータである必要があります。"
                f"開始: {start_cell} (行{start_row}), 終了: {end_cell} (行{end_row})"
            )
        
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        result = {}
        for col_idx in range(start_col_idx, end_col_idx + 1):
            col_name = get_column_letter(col_idx)
            cell = self.sheet[f"{col_name}{start_row}"]
            value = cell.value if cell.value is not None else ""
            result[col_name] = str(value)
        
        return result
    
    def read_data4(self, start_cell: str, end_cell: str) -> dict[str, str]:
        """
        データ4を読み込む: 日の情報を持つ1行のデータ
        
        Args:
            start_cell: 開始セル (例: "J4")
            end_cell: 終了セル (例: "RR4")
        
        Returns:
            {列名: セルの値, ...}
        
        Raises:
            ValueError: 開始と終了が異なる行の場合
        """
        start_col, start_row = self._parse_cell_address(start_cell)
        end_col, end_row = self._parse_cell_address(end_cell)
        
        if start_row != end_row:
            raise ValueError(
                f"データ4は1行分のデータである必要があります。"
                f"開始: {start_cell} (行{start_row}), 終了: {end_cell} (行{end_row})"
            )
        
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        result = {}
        for col_idx in range(start_col_idx, end_col_idx + 1):
            col_name = get_column_letter(col_idx)
            cell = self.sheet[f"{col_name}{start_row}"]
            value = cell.value if cell.value is not None else ""
            result[col_name] = str(value)
        
        return result
    
    def _get_fill_color(self, cell) -> Optional[str]:
        """
        セルの塗りつぶし色を取得
        
        Args:
            cell: openpyxlのCellオブジェクト
        
        Returns:
            RGB形式の色 (例: "FFFF0000") またはNone
        """
        if cell.fill and cell.fill.start_color:
            # RGBカラーの取得
            color = cell.fill.start_color
            
            # rgbプロパティの取得（エラーハンドリング付き）
            if hasattr(color, 'rgb'):
                try:
                    rgb_value = color.rgb
                    if rgb_value and isinstance(rgb_value, str):
                        # "00000000"形式の場合は塗りつぶしなしとみなす
                        if rgb_value == "00000000" or rgb_value == "FFFFFFFF":
                            return None
                        # 16進数文字列として有効かチェック
                        if len(rgb_value) in [6, 8] and all(c in '0123456789ABCDEFabcdef' for c in rgb_value):
                            return rgb_value
                except (AttributeError, TypeError, ValueError):
                    # rgbプロパティへのアクセスでエラーが発生した場合
                    pass
            
            # インデックスカラーの場合
            if hasattr(color, 'index') and color.index:
                return f"INDEX_{color.index}"
        
        return None
    
    def close(self):
        """ワークブックを閉じる"""
        self.workbook.close()
