"""
diff_semantic.yamlを読み込んで、人間が読みやすいExcel形式のレポートを生成
"""
import yaml
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


class SemanticDiffReporter:
    """意味的な差分をExcelレポートに変換するクラス"""
    
    def __init__(self, diff_yaml_path: str):
        """
        Args:
            diff_yaml_path: diff_semantic.yamlのパス
        """
        self.diff_yaml_path = diff_yaml_path
        self.diff_data = None
    
    def load_diff(self):
        """差分YAMLファイルを読み込む"""
        with open(self.diff_yaml_path, 'r', encoding='utf-8') as f:
            self.diff_data = yaml.safe_load(f)
    
    def format_date(self, date_str: Optional[str]) -> str:
        """
        日付文字列を「YYYY/MM/DD」形式に変換
        
        Args:
            date_str: 日付文字列（例: "2026-01-05 00:00:00" または "2026/1/5"）
        
        Returns:
            フォーマットされた日付文字列（例: "2026/01/05"）
        """
        if not date_str:
            return ""
        
        try:
            # "2026-01-05 00:00:00" 形式の場合
            if ' ' in str(date_str):
                dt = datetime.strptime(str(date_str), "%Y-%m-%d %H:%M:%S")
                return dt.strftime("%Y/%m/%d")
            # "2026/1/5" 形式の場合
            elif '/' in str(date_str):
                parts = str(date_str).split('/')
                if len(parts) == 3:
                    year, month, day = parts
                    return f"{year}/{int(month):02d}/{int(day):02d}"
            # "2026-01-05" 形式の場合
            elif '-' in str(date_str):
                dt = datetime.strptime(str(date_str), "%Y-%m-%d")
                return dt.strftime("%Y/%m/%d")
        except Exception:
            pass
        
        return str(date_str)
    
    def parse_position(self, position_str: str) -> tuple:
        """
        位置文字列をパース
        
        Args:
            position_str: "(J, 46)" のような文字列
        
        Returns:
            (列名, 行番号) のタプル
        """
        # "(J, 46)" -> "J, 46"
        position_str = position_str.strip('()')
        parts = position_str.split(',')
        col = parts[0].strip()
        row = int(parts[1].strip())
        return col, row
    
    def format_texts(self, texts: List[str]) -> str:
        """
        テキストリストをフォーマット
        
        Args:
            texts: テキストのリスト
        
        Returns:
            フォーマットされた文字列
        """
        if not texts:
            return "（テキストなし）"
        return "、".join(f"「{text}」" for text in texts)
    
    def format_color(self, color: Optional[str]) -> str:
        """
        色をフォーマット
        
        Args:
            color: 色の文字列（Noneの場合は色なし）
        
        Returns:
            フォーマットされた文字列
        """
        if color is None:
            return "色なし"
        elif color.startswith("INDEX_"):
            return f"パレット色{color[6:]}"
        else:
            return f"色{color}"
    
    def generate_change_text(self, change: Dict[str, Any]) -> str:
        """
        1つの変更をテキスト化
        
        Args:
            change: 変更情報の辞書
        
        Returns:
            テキスト形式の変更内容
        """
        # data1の情報を取得
        data1 = change.get('data1', [])
        data1_str = "、".join(f"「{d}」" for d in data1 if d)
        
        # texts と color を取得
        texts = change.get('texts', [])
        texts_str = self.format_texts(texts)
        color = change.get('color')
        color_str = self.format_color(color)
        
        # before と after を取得
        before = change.get('before')
        after = change.get('after')
        
        # 行番号を取得（before または after から）
        row = None
        if isinstance(before, dict) and 'start' in before:
            _, row = self.parse_position(before['start'])
        elif isinstance(after, dict) and 'start' in after:
            _, row = self.parse_position(after['start'])
        
        row_str = f"{row}行" if row else "不明な行"
        
        # 変更の種類を判定してテキスト化
        if isinstance(before, dict) and isinstance(after, dict):
            # 移動または変更
            before_start_col, _ = self.parse_position(before['start'])
            before_end_col, _ = self.parse_position(before['end'])
            after_start_col, _ = self.parse_position(after['start'])
            after_end_col, _ = self.parse_position(after['end'])
            
            # 日付情報を取得してフォーマット
            before_start_date = self.format_date(before.get('start_date'))
            before_end_date = self.format_date(before.get('end_date'))
            after_start_date = self.format_date(after.get('start_date'))
            after_end_date = self.format_date(after.get('end_date'))
            
            # 位置変更のチェック
            position_changed = (before_start_col != after_start_col or before_end_col != after_end_col)
            # 日付変更のチェック
            date_changed = (before_start_date != after_start_date or before_end_date != after_end_date)
            
            result = f"・{data1_str}、{row_str}の、"
            
            if position_changed and date_changed:
                # 位置と日付の両方が変更
                result += (f"{before_start_col}列から{before_end_col}列"
                          f"（{before_start_date}～{before_end_date}）にあった"
                          f"Bar（{texts_str}、{color_str}）が"
                          f"{after_start_col}列から{after_end_col}列"
                          f"（{after_start_date}～{after_end_date}）に移動・変更されています。")
            elif position_changed:
                # 位置のみ変更
                date_str = ""
                if before_start_date and before_end_date:
                    date_str = f"（{before_start_date}～{before_end_date}）"
                result += (f"{before_start_col}列から{before_end_col}列{date_str}にあった"
                          f"Bar（{texts_str}、{color_str}）が"
                          f"{after_start_col}列から{after_end_col}列に移動しています。")
            elif date_changed:
                # 日付のみ変更
                result += (f"{before_start_col}列から{before_end_col}列の"
                          f"Bar（{texts_str}、{color_str}）の日付が"
                          f"{before_start_date}～{before_end_date}から"
                          f"{after_start_date}～{after_end_date}に変更されています。")
            else:
                # 変更なし（通常はここには来ない）
                result += f"{before_start_col}列から{before_end_col}列のBar（{texts_str}、{color_str}）に変更があります。"
            
            return result
        
        elif isinstance(before, dict) and after == 'deleted':
            # 削除
            before_start_col, _ = self.parse_position(before['start'])
            before_end_col, _ = self.parse_position(before['end'])
            before_start_date = self.format_date(before.get('start_date'))
            before_end_date = self.format_date(before.get('end_date'))
            
            date_str = ""
            if before_start_date and before_end_date:
                date_str = f"（{before_start_date}～{before_end_date}）"
            
            return (f"・{data1_str}、{row_str}の、"
                   f"{before_start_col}列から{before_end_col}列{date_str}にあった"
                   f"Bar（{texts_str}、{color_str}）が削除されています。")
        
        elif before == 'not_exists' and isinstance(after, dict):
            # 追加
            after_start_col, _ = self.parse_position(after['start'])
            after_end_col, _ = self.parse_position(after['end'])
            after_start_date = self.format_date(after.get('start_date'))
            after_end_date = self.format_date(after.get('end_date'))
            
            date_str = ""
            if after_start_date and after_end_date:
                date_str = f"（{after_start_date}～{after_end_date}）"
            
            return (f"・{data1_str}、{row_str}の、"
                   f"{after_start_col}列から{after_end_col}列{date_str}に"
                   f"Bar（{texts_str}、{color_str}）が追加されています。")
        
        else:
            # 不明な変更
            return f"・{data1_str}、{row_str}で不明な変更が発生しています。"
    
    def generate_report(self, output_path: str):
        """
        Excel形式のレポートを生成
        
        Args:
            output_path: 出力ファイルパス（.xlsx）
        """
        if self.diff_data is None:
            self.load_diff()
        
        # メタデータを取得
        assert self.diff_data is not None, "diff_dataが読み込まれていません"
        metadata = self.diff_data.get('metadata', {})
        file1 = metadata.get('file1', '不明なファイル')
        file2 = metadata.get('file2', '不明なファイル')
        total_changes = metadata.get('total_changes', 0)
        
        # ファイル名から拡張子を削除
        file1_name = Path(file1).stem if file1 != '不明なファイル' else file1
        file2_name = Path(file2).stem if file2 != '不明なファイル' else file2
        
        # Excelワークブックを作成
        wb = Workbook()
        ws = wb.active
        assert ws is not None, "ワークシートの作成に失敗しました"
        ws.title = "変更レポート"
        
        # スタイル定義
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # タイトル行
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']  # type: ignore
        title_cell.value = f"{file1_name}から{file2_name}への変更レポート"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # サマリー行
        ws.merge_cells('A2:H2')
        summary_cell = ws['A2']  # type: ignore
        summary_cell.value = f"合計変更数: {total_changes}件"
        summary_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # ヘッダー行
        headers = ['No.', '作業員情報', '行番号', '変更種類', 'Bar情報', f'変更前\n（{file1_name}）', f'変更後\n（{file2_name}）', 'カテゴリ']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)  # type: ignore
            cell.value = header  # type: ignore
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # プルダウンリストの設定
        dv = DataValidation(type="list", formula1='"A,B,C,D"', allow_blank=True)
        dv.error = '無効な値です'
        dv.errorTitle = '入力エラー'
        dv.prompt = 'A, B, C, D から選択してください'
        dv.promptTitle = 'カテゴリ選択'
        ws.add_data_validation(dv)
        
        # 各変更をExcelに追加
        changes = self.diff_data.get('changes', [])
        row_num = 5
        
        for i, change in enumerate(changes, 1):
            # data1の情報を取得
            data1 = change.get('data1', [])
            data1_str = "、".join(d for d in data1 if d)
            
            # Bar情報を取得
            texts = change.get('texts', [])
            texts_str = "\n".join(f"「{text}」" for text in texts) if texts else "（テキストなし）"
            
            # before と after を取得
            before = change.get('before')
            after = change.get('after')
            
            # 行番号を取得
            row = None
            if isinstance(before, dict) and 'start' in before:
                _, row = self.parse_position(before['start'])
            elif isinstance(after, dict) and 'start' in after:
                _, row = self.parse_position(after['start'])
            
            # 変更種類を判定
            change_type = ""
            before_str = ""
            after_str = ""
            
            if isinstance(before, dict) and isinstance(after, dict):
                # 変更
                before_start_col, _ = self.parse_position(before['start'])
                before_end_col, _ = self.parse_position(before['end'])
                after_start_col, _ = self.parse_position(after['start'])
                after_end_col, _ = self.parse_position(after['end'])
                
                before_start_date = self.format_date(before.get('start_date'))
                before_end_date = self.format_date(before.get('end_date'))
                after_start_date = self.format_date(after.get('start_date'))
                after_end_date = self.format_date(after.get('end_date'))
                
                position_changed = (before_start_col != after_start_col or before_end_col != after_end_col)
                date_changed = (before_start_date != after_start_date or before_end_date != after_end_date)
                
                if position_changed and date_changed:
                    change_type = "位置・日付変更"
                elif position_changed:
                    change_type = "位置変更"
                elif date_changed:
                    change_type = "日付変更"
                else:
                    change_type = "変更"
                
                # 変更前: 日時が上、位置が下
                if before_start_date and before_end_date:
                    before_str = f"{before_start_date}～{before_end_date}\n{before_start_col}～{before_end_col}列"
                else:
                    before_str = f"{before_start_col}～{before_end_col}列"
                
                # 変更後: 日時が上、位置が下
                if after_start_date and after_end_date:
                    after_str = f"{after_start_date}～{after_end_date}\n{after_start_col}～{after_end_col}列"
                else:
                    after_str = f"{after_start_col}～{after_end_col}列"
                
            elif isinstance(before, dict) and after == 'deleted':
                # 削除
                change_type = "削除"
                before_start_col, _ = self.parse_position(before['start'])
                before_end_col, _ = self.parse_position(before['end'])
                before_start_date = self.format_date(before.get('start_date'))
                before_end_date = self.format_date(before.get('end_date'))
                
                # 変更前: 日時が上、位置が下
                if before_start_date and before_end_date:
                    before_str = f"{before_start_date}～{before_end_date}\n{before_start_col}～{before_end_col}列"
                else:
                    before_str = f"{before_start_col}～{before_end_col}列"
                after_str = "削除"
                
            elif before == 'not_exists' and isinstance(after, dict):
                # 追加
                change_type = "追加"
                after_start_col, _ = self.parse_position(after['start'])
                after_end_col, _ = self.parse_position(after['end'])
                after_start_date = self.format_date(after.get('start_date'))
                after_end_date = self.format_date(after.get('end_date'))
                
                before_str = "なし"
                # 変更後: 日時が上、位置が下
                if after_start_date and after_end_date:
                    after_str = f"{after_start_date}～{after_end_date}\n{after_start_col}～{after_end_col}列"
                else:
                    after_str = f"{after_start_col}～{after_end_col}列"
            
            # Excelに書き込み
            ws.cell(row=row_num, column=1).value = i  # type: ignore
            ws.cell(row=row_num, column=2).value = data1_str  # type: ignore
            ws.cell(row=row_num, column=3).value = row if row else ""  # type: ignore
            ws.cell(row=row_num, column=4).value = change_type  # type: ignore
            ws.cell(row=row_num, column=5).value = texts_str  # type: ignore  # Bar情報
            ws.cell(row=row_num, column=6).value = before_str  # type: ignore
            ws.cell(row=row_num, column=7).value = after_str  # type: ignore
            ws.cell(row=row_num, column=8).value = ""  # type: ignore  # カテゴリ列（初期値は空）
            
            # プルダウンリストをカテゴリ列に適用
            dv.add(ws.cell(row=row_num, column=8))  # type: ignore
            
            # セルのスタイル設定
            for col in range(1, 9):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            row_num += 1
        
        # 列幅の調整
        ws.column_dimensions['A'].width = 8   # No.
        ws.column_dimensions['B'].width = 25  # 作業員情報
        ws.column_dimensions['C'].width = 10  # 行番号
        ws.column_dimensions['D'].width = 15  # 変更種類
        ws.column_dimensions['E'].width = 30  # Bar情報
        ws.column_dimensions['F'].width = 25  # 変更前
        ws.column_dimensions['G'].width = 25  # 変更後
        ws.column_dimensions['H'].width = 10  # カテゴリ
        
        # 行の高さ調整（ヘッダー）
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[4].height = 45
        
        # Excelファイルを保存
        wb.save(output_path)
        
        print(f"レポートを {output_path} に出力しました")
        print(f"合計 {total_changes} 件の変更が記録されています")
    
    def print_summary(self):
        """
        サマリーをコンソールに表示
        """
        if self.diff_data is None:
            self.load_diff()
        
        assert self.diff_data is not None, "diff_dataが読み込まれていません"
        metadata = self.diff_data.get('metadata', {})
        changes = self.diff_data.get('changes', [])
        
        # 変更の種類をカウント
        moved = 0
        deleted = 0
        added = 0
        
        for change in changes:
            before = change.get('before')
            after = change.get('after')
            
            if isinstance(before, dict) and isinstance(after, dict):
                moved += 1
            elif isinstance(before, dict) and after == 'deleted':
                deleted += 1
            elif before == 'not_exists' and isinstance(after, dict):
                added += 1
        
        print("\n変更サマリー:")
        print(f"  移動: {moved}件")
        print(f"  削除: {deleted}件")
        print(f"  追加: {added}件")
        print(f"  合計: {metadata.get('total_changes', 0)}件")


def main():
    """
    使用例: diff_semantic.yamlを読み込んでExcelレポートを生成
    """
    # 差分YAMLファイルのパス
    diff_yaml = "diff_semantic.yaml"
    
    # ファイルの存在確認
    if not Path(diff_yaml).exists():
        print(f"エラー: ファイルが見つかりません: {diff_yaml}")
        print("先にmain02.pyを実行して、diff_semantic.yamlを生成してください。")
        return
    
    # レポート生成器を初期化
    reporter = SemanticDiffReporter(diff_yaml)
    
    # Excelレポートを生成
    reporter.generate_report("diff_report.xlsx")
    
    # サマリーを表示
    reporter.print_summary()
    
    print("\nExcelレポートが生成されました: diff_report.xlsx")


if __name__ == "__main__":
    main()
