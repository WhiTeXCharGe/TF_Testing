import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml
from openpyxl import load_workbook


# ============================================================
# Helpers
# ============================================================

def _to_ymd(dt) -> str:
    """Convert pandas/datetime to 'YYYY/MM/DD'."""
    if isinstance(dt, pd.Timestamp):
        return dt.strftime("%Y/%m/%d")
    if isinstance(dt, datetime):
        return dt.strftime("%Y/%m/%d")
    return str(dt)


def _norm(s: str) -> str:
    """Normalize string for matching (trim, collapse spaces, lower)."""
    if not isinstance(s, str):
        return ""
    s = s.replace("　", " ")
    s = s.replace("（", "(").replace("）", ")")
    return " ".join(s.split()).lower()


def load_sheet_as_df(path: str, sheet_name: str) -> pd.DataFrame:
    """
    Load a sheet using openpyxl (read_only=False to avoid ReadOnlyWorksheet bug)
    and convert it to a pandas DataFrame.
    """
    wb = load_workbook(path, data_only=True, read_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    df = pd.DataFrame(rows)
    return df


# ============================================================
# Tool code extraction (e.g. 530N02716A, 830300179A, 852Z00771A)
# ============================================================

TOOLCODE_RE = re.compile(r"\d{3}[A-Z0-9]\d{5}A")


def extract_tool_code(s: str):
    """
    Extract tool code like 530N02716A, 830300179A, 852Z00771A from a string.
    Returns the FIRST match if there are multiple (e.g. '530N01814A_852Z00771A').
    """
    if not isinstance(s, str):
        return None
    m = TOOLCODE_RE.search(s)
    if m:
        return m.group(0)
    return None


# ============================================================
# F22 operation definitions (fixed)
# ============================================================

OPS_DEF = [
    {"id": "f22p1o1", "name": "Power"},
    {"id": "f22p1o2", "name": "Gas"},
    {"id": "f22p1o3", "name": "Exh"},

    {"id": "f22p2o1", "name": "Through Running"},
    {"id": "f22p2o2", "name": "Diw"},
    {"id": "f22p2o3", "name": "Drain Check"},
    {"id": "f22p2o4", "name": "Chemi Sig TEST"},

    {"id": "f22p3o1", "name": "Diw Running"},
    {"id": "f22p3o2", "name": "ISEP"},
    {"id": "f22p3o3", "name": "Chemical"},
    {"id": "f22p3o4", "name": "A1 Port Finish"},

    {"id": "f22p4o1", "name": "Chemical Running"},
    {"id": "f22p4o2", "name": "Acceptance End (T1 End)"},
    {"id": "f22p4o3", "name": "Release VQF (FCCB)"},
]

OPS_BY_ID = {op["id"]: op for op in OPS_DEF}

PHASE1_IDS = ["f22p1o1", "f22p1o2", "f22p1o3"]
PHASE2_IDS = ["f22p2o1", "f22p2o2", "f22p2o3", "f22p2o4"]
PHASE3_IDS = ["f22p3o1", "f22p3o2", "f22p3o3", "f22p3o4"]
PHASE4_IDS = ["f22p4o1", "f22p4o2", "f22p4o3"]

OP_PHASE_INDEX = {}
for op_id in PHASE1_IDS:
    OP_PHASE_INDEX[op_id] = 1
for op_id in PHASE2_IDS:
    OP_PHASE_INDEX[op_id] = 2
for op_id in PHASE3_IDS:
    OP_PHASE_INDEX[op_id] = 3
for op_id in PHASE4_IDS:
    OP_PHASE_INDEX[op_id] = 4

OPS_NAME_TO_ID = {
    _norm("Power"): "f22p1o1",
    _norm("Gas"): "f22p1o2",
    _norm("Exh"): "f22p1o3",

    _norm("Through Running"): "f22p2o1",
    _norm("Diw"): "f22p2o2",
    _norm("Drain Check"): "f22p2o3",
    _norm("Chemi Sig TEST"): "f22p2o4",

    _norm("Diw Running"): "f22p3o1",
    _norm("ISEP"): "f22p3o2",
    _norm("Chemical"): "f22p3o3",
    _norm("A1 Port Finish"): "f22p3o4",

    _norm("Chemical Running"): "f22p4o1",
    _norm("Acceptance End (T1 End)"): "f22p4o2",
    _norm("Acceptance 　　End　 （T1 End）"): "f22p4o2",  # full-width variant
    _norm("Release VQF (FCCB)"): "f22p4o3",
    _norm("Release VQF　（FCCB）"): "f22p4o3",            # full-width variant
}


# ============================================================
# SU_Others: worker info + raw matrix for assignments
# ============================================================

def parse_su_others(path: str, sheet_name: str = "予定表_2024"):
    """
    From 20251201_2 SU_Others.xlsm, build:
      - worker_company_list
      - worker_list (with unavailable_dates + skill_map)
      - plan_range (min/max date in header)
      - worker_date_map: (worker_id, Timestamp) -> raw cell string (non-red only)

    Red cells = "worker unavailable":
      - added to worker.unavailable_dates
      - NEVER produce assignments (even if they have text)
    """
    wb = load_workbook(path, data_only=True, read_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {path}. Available: {wb.sheetnames}")
    ws = wb[sheet_name]

    # Build DataFrame view for header detection, etc.
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    df = pd.DataFrame(rows)
    n_rows, n_cols = df.shape

    # Date header row
    date_row_idx = None
    for r in range(0, min(5, n_rows)):
        row = df.iloc[r]
        if any(isinstance(v, (pd.Timestamp, datetime)) for v in row):
            date_row_idx = r
            break
    if date_row_idx is None:
        raise RuntimeError("Could not find date header row in SU_Others.")

    date_row = df.iloc[date_row_idx]
    date_cols = [c for c, v in enumerate(date_row) if isinstance(v, (pd.Timestamp, datetime))]
    if not date_cols:
        raise RuntimeError("Could not find any date columns in SU_Others.")

    worker_start_row = date_row_idx + 2

    worker_company_map = {}
    worker_company_list = []
    worker_list = []
    worker_row_to_id = {}
    worker_date_map = {}  # (worker_id, Timestamp) -> cell text (non-red)

    def get_worker_company_id(company_name: str) -> str:
        company_name = str(company_name).strip()
        if company_name not in worker_company_map:
            cid = f"wc{len(worker_company_map) + 1}"
            worker_company_map[company_name] = cid
            worker_company_list.append({
                "id": cid,
                "name": company_name,
                "annual_overtime_limit": 360,
                "monthly_overtime_limit": 40,
                "unavailable_dates": [],
            })
        return worker_company_map[company_name]

    # Worker rows + row -> worker_id map
    for r in range(worker_start_row, n_rows):
        company = df.iat[r, 0]
        name = df.iat[r, 1]

        if pd.isna(name):
            continue
        
                # 自由枠 (column E) -> manager flag
        # A=0, B=1, C=2, D=3, E=4
        manager_col_idx = 4
        free_slot = None
        if manager_col_idx < n_cols:
            free_slot = df.iat[r, manager_col_idx]

        is_manager = False
        if isinstance(free_slot, str) and "責" in free_slot:
            # e.g. "M/QC(責)", "M (責)" -> manager
            is_manager = True
            
        company_id = get_worker_company_id(company)
        wid = f"w{len(worker_list) + 1:03d}"

        # collect unavailable dates for this worker (red cells)
        unavailable_set = set()

        for c in date_cols:
            date_val = date_row[c]
            if not isinstance(date_val, (pd.Timestamp, datetime)):
                continue
            dt = pd.to_datetime(date_val)

            cell = ws.cell(row=r + 1, column=c + 1)
            fill = cell.fill
            is_red = (
                fill is not None
                and getattr(fill.fgColor, "type", None) == "rgb"
                and fill.fgColor.rgb == "FFFF0000"
            )

            if is_red:
                unavailable_set.add(_to_ymd(dt))
                # red cell -> no assignment, even if it has text
                continue

            val = cell.value
            if isinstance(val, str) and val.strip():
                worker_date_map[(wid, dt)] = val.strip()

        skill_map = {
            "f22p1": 1,
            "f22p2": 1,
            "f22p3": 1,
            "f22p4": 1,
            # optional: allow misc work too
            "other_op": 1,
        }

        worker_list.append({
            "id": wid,
            "name": str(name).strip(),
            "worker_company": company_id,
            "is_manager": is_manager,
            "skill_map": skill_map,
            "fab_suitability_map": [],
            "unavailable_dates": [{"date": d} for d in sorted(unavailable_set)],
        })
        worker_row_to_id[r] = wid

    # Plan range from date header
    all_dates = list(date_row[date_cols].dropna())
    all_dates.sort()
    plan_range = {
        "start_date": _to_ymd(all_dates[0]),
        "end_date": _to_ymd(all_dates[-1]),
    }

    return {
        "worker_company_list": worker_company_list,
        "worker_list": worker_list,
        "plan_range": plan_range,
        "df": df,
        "date_row_idx": date_row_idx,
        "date_cols": date_cols,
        "worker_start_row": worker_start_row,
        "worker_row_to_id": worker_row_to_id,
        "worker_date_map": worker_date_map,
    }


# ============================================================
# F22_Tool Schedule: 4-phase tool tasks + module/phase meta
# ============================================================

def parse_tool_schedule(path: str, sheet_name: str = "F22_Tool Schedule"):
    """
    From 台湾出張者予定_2025latest.xlsx, 'F22_Tool Schedule':
      - tool_tasks for schedule.workflow_task_list (all modules)
      - date_list for plan_range
      - module_to_phases: module_code -> list of phase meta (for assignments)

    Each tool row is always expanded to 4 phases (P1..P4).
    For each module, phase end_date is (next phase start_date - 1 day).
    """
    df = load_sheet_as_df(path, sheet_name)
    n_rows, n_cols = df.shape

    # ---------- find header row & map columns -> operation ids ----------
    header_row_idx = None
    col_to_op_id = {}
    for r in range(0, min(30, n_rows)):
        row = df.iloc[r]
        tmp_map = {}
        for c, v in enumerate(row):
            if isinstance(v, str) and v.strip():
                key = OPS_NAME_TO_ID.get(_norm(v))
                if key:
                    tmp_map[c] = key
        if len(tmp_map) >= 3:
            header_row_idx = r
            col_to_op_id = tmp_map
            break

    if header_row_idx is None:
        raise RuntimeError("Could not find operation header row in F22_Tool Schedule.")

    tool_tasks = []
    module_to_phases = defaultdict(list)
    all_dates = []
    task_counter = 1

    for r in range(header_row_idx + 2, n_rows):
        location = df.iat[r, 0]
        tsmc_tool = df.iat[r, 2]
        screen_tool = df.iat[r, 3]

        # skip empty rows
        if (pd.isna(location) or str(location).strip() == "") and \
           (pd.isna(tsmc_tool) or str(tsmc_tool).strip() == "") and \
           (pd.isna(screen_tool) or str(screen_tool).strip() == ""):
            continue

        # name = SCREEN tool / TSMC tool
        name_parts = []
        for v in (screen_tool, tsmc_tool):
            if isinstance(v, str) and v.strip():
                name_parts.append(v.strip())
        if not name_parts:
            continue
        task_name = " / ".join(name_parts)

        # module code (for linking with SU_Others; NOT written to YAML)
        module_code = None
        if isinstance(screen_tool, str):
            module_code = extract_tool_code(screen_tool)
        if not module_code and isinstance(tsmc_tool, str):
            module_code = extract_tool_code(tsmc_tool)

        # we only track dates per phase now
        phase_dates = {1: [], 2: [], 3: [], 4: []}
        row_dates = []

        # ---------- parse dates from each operation column ----------
        for c, op_id in col_to_op_id.items():
            if c >= n_cols:
                continue
            cell = df.iat[r, c]

            # blank cell
            if cell is None or cell == "" or (isinstance(cell, float) and pd.isna(cell)):
                continue

            start_dt = end_dt = None

            # Excel real date -> datetime or Timestamp
            if isinstance(cell, (pd.Timestamp, datetime)):
                start_dt = end_dt = cell

            # String cases like "2025/1/6\n＞1/8" or "12/28\n>12/31"
            elif isinstance(cell, str):
                text = cell.strip()
                if not text:
                    continue

                if "\n" in text:
                    first, second = text.split("\n", 1)
                    first = first.strip()
                    second = second.strip().lstrip(">")  # handle "＞1/8" or ">1/8"

                    dt1 = pd.to_datetime(first, errors="coerce")
                    if isinstance(dt1, pd.Timestamp):
                        start_dt = dt1
                        dt2 = pd.to_datetime(second, errors="coerce")
                        if isinstance(dt2, pd.Timestamp):
                            # if month/day only, inherit year from start
                            if dt2.year != dt1.year:
                                dt2 = dt2.replace(year=dt1.year)
                            end_dt = dt2
                        else:
                            end_dt = dt1
                else:
                    dt = pd.to_datetime(text, errors="coerce")
                    if isinstance(dt, pd.Timestamp):
                        start_dt = end_dt = dt

            if start_dt is None:
                continue

            phase_index = OP_PHASE_INDEX.get(op_id)
            if not phase_index:
                continue

            phase_dates[phase_index].append(start_dt)
            phase_dates[phase_index].append(end_dt or start_dt)

            row_dates.append(start_dt)
            row_dates.append(end_dt or start_dt)

            all_dates.append(start_dt)
            all_dates.append(end_dt or start_dt)

        # Default range for the whole row (used as fallback)
        row_dates = [d for d in row_dates if isinstance(d, (pd.Timestamp, datetime))]
        row_start = min(row_dates) if row_dates else None
        row_end = max(row_dates) if row_dates else None

        task_id = f"f22_{task_counter}"
        task_counter += 1

        row_phase_meta = []  # collect per-phase meta for this ONE row

        # ---------- first pass: compute raw start/end per phase ----------
        for ph in (1, 2, 3, 4):
            dates = [d for d in phase_dates[ph] if isinstance(d, (pd.Timestamp, datetime))]
            if dates:
                phase_start = min(dates)
                phase_end = max(dates)
            elif row_start and row_end:
                # no direct dates for this phase -> use row range
                phase_start = row_start
                phase_end = row_end
            else:
                # worst-case fallback
                phase_start = phase_end = pd.Timestamp("2025-01-01")

            phase_id = f"{task_id}_p{ph}"

            row_phase_meta.append({
                "phase_index": ph,
                "phase_id": phase_id,
                "start": phase_start,
                "end": phase_end,
            })

        # ---------- adjust phase end dates: end = next phase start - 1 day ----------
        row_phase_meta_sorted = sorted(row_phase_meta, key=lambda m: m["phase_index"])

        for i in range(len(row_phase_meta_sorted) - 1):
            cur = row_phase_meta_sorted[i]
            nxt = row_phase_meta_sorted[i + 1]

            new_end = nxt["start"] - pd.Timedelta(days=1)
            if new_end < cur["start"]:
                new_end = cur["start"]

            cur["end"] = new_end

        # (last phase keeps its own end)

        # ---------- build phase_task_list with ONE operation per phase ----------
        phase_task_list = []

        for meta in row_phase_meta_sorted:
            ph = meta["phase_index"]
            phase_id = meta["phase_id"]
            phase_start = meta["start"]
            phase_end = meta["end"]

            # inclusive days: (end - start) + 1
            workload_days = int((phase_end - phase_start).days) + 1

            phase_task_list.append({
                "id": phase_id,
                "name": f"Phase {ph}",
                "phase": f"tool_p{ph}",
                "start_date": _to_ymd(phase_start),
                "end_date": _to_ymd(phase_end),
                "operation_task_list": [
                    {
                        "id": phase_id,
                        "name": f"Phase {ph}",
                        "operation": f"f22p{ph}",
                        "workload_days": workload_days,
                    }
                ],
            })

            if module_code:
                module_to_phases[module_code].append({
                    "phase_index": ph,
                    "phase_id": phase_id,
                    "start": phase_start,
                    "end": phase_end,
                    "ops": [],  # not used anymore
                })

        tool_tasks.append({
            "id": task_id,
            "name": task_name,
            "workflow": "wf_tool",
            "fab": "f_tw",
            "phase_task_list": phase_task_list,
            "module_code": module_code,   # internal only, stripped before YAML
        })

    return {
        "tool_tasks": tool_tasks,
        "date_list": all_dates,
        "module_to_phases": module_to_phases,
    }


# ============================================================
# Build assignments (SU_Others x F22 via module code + misc workflow)
# ============================================================

def build_assignments(su_data: dict, tool_data: dict):
    """
    Build:
      - assignments for F22 tools (plan_flexibility = 'Flexible')
      - assignments for "other" work seen only in SU_Others (plan_flexibility = 'Fixed')
      - dummy workflow_task_list entries for all "other" work labels
    """
    module_to_phases = tool_data["module_to_phases"]
    worker_date_map = su_data["worker_date_map"]

    # (worker, phase_id) -> list of dates for F22 tools
    known_assign_map = defaultdict(list)

    # misc label -> set of all dates it appears
    misc_label_dates = defaultdict(set)
    # (worker, label) -> list of dates
    misc_worker_label_dates = defaultdict(list)

    for (wid, dt), text in worker_date_map.items():
        code = extract_tool_code(text)
        if code and code in module_to_phases:
            # map to all phases whose window covers this date
            for phase_meta in module_to_phases[code]:
                ps = phase_meta["start"]
                pe = phase_meta["end"]
                if ps <= dt <= pe:
                    key = (wid, phase_meta["phase_id"])
                    known_assign_map[key].append(dt)
        else:
            # unknown -> misc workflow
            label = text.strip()
            misc_label_dates[label].add(dt)
            misc_worker_label_dates[(wid, label)].append(dt)

    assignments = []

    # ---------- F22 assignments (Flexible) ----------
    for (wid, phase_id), dates in known_assign_map.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        work_date_list = [{"hour": 12, "date": _to_ymd(d)} for d in uniq_dates]
        assignments.append({
            "worker": wid,
            "operation_task": phase_id,
            "start_date": _to_ymd(uniq_dates[0]),
            "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list,
            "plan_flexibility": "Flexible",
        })

    # ---------- Build dummy workflow tasks for misc work ----------
    misc_tasks = []
    misc_label_to_phase = {}  # label -> {"phase_id", "start", "end"}
    misc_counter = 1

    for label, dates in misc_label_dates.items():
        if not dates:
            continue
        start = min(dates)
        end = max(dates)
        task_id = f"misc_{misc_counter}"
        misc_counter += 1
        phase_id = f"{task_id}_p1"
        misc_label_to_phase[label] = {"phase_id": phase_id, "start": start, "end": end}

        workload_days = int((end - start).days) + 1

        misc_tasks.append({
            "id": task_id,
            "name": label,
            "workflow": "wf_other",
            "fab": "f_tw",
            "phase_task_list": [
                {
                    "id": phase_id,
                    "name": "Misc Phase",
                    "phase": "other_p1",
                    "start_date": _to_ymd(start),
                    "end_date": _to_ymd(end),
                    "operation_task_list": [
                        {
                            "id": phase_id,
                            "name": "Misc Phase",
                            "operation": "other_op",
                            "workload_days": workload_days,
                        }
                    ],
                }
            ],
        })

    # ---------- Misc assignments (Fixed) ----------
    for (wid, label), dates in misc_worker_label_dates.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        phase_meta = misc_label_to_phase.get(label)
        if not phase_meta:
            continue
        phase_id = phase_meta["phase_id"]
        work_date_list = [{"hour": 8, "date": _to_ymd(d)} for d in uniq_dates]

        assignments.append({
            "worker": wid,
            "operation_task": phase_id,
            "start_date": _to_ymd(uniq_dates[0]),
            "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list,
            "plan_flexibility": "Fixed",
        })

    return assignments, misc_tasks


# ============================================================
# Build EnvConfig + Schedule and dump YAML
# ============================================================

def build_env_and_schedule(
    su_others_path: str,
    tool_schedule_path: str,
    envconfig_out: str = "EnvConfig_from_excel.yaml",
    schedule_out: str = "Schedule_from_excel.yaml",
):
    su_data = parse_su_others(su_others_path)
    tool_data = parse_tool_schedule(tool_schedule_path)

    # ---------- ENVIRONMENT ----------
    wf_tool_phases = []
    for ph in (1, 2, 3, 4):
        wf_tool_phases.append(
            {
                "id": f"tool_p{ph}",
                "name": f"Phase {ph}",
                "operation_list": [
                    {
                        # Aggregated operation per phase: f22p1, f22p2, f22p3, f22p4
                        "id": f"f22p{ph}",
                        "name": f"F22 Phase {ph}",  # you can change names if you want
                        "work_hours": [8],          # or [4, 8, 12] later
                        "min_worker_num": 1,
                        "max_worker_num": 3,
                    }
                ],
            }
        )

    environment = {
        "workflow_list": [
            {
                "id": "wf_tool",
                "name": "Taiwan Tool Install 2025",
                "phase_list": wf_tool_phases,
            },
            {
                # Dummy workflow for "other" work that only appears in SU_Others
                "id": "wf_other",
                "name": "Other work from SU_Others",
                "phase_list": [
                    {
                        "id": "other_p1",
                        "name": "Misc work",
                        "operation_list": [
                            {
                                "id": "other_op",
                                "name": "Other work",
                                "work_hours": [8],
                                "min_worker_num": 1,
                                "max_worker_num": 3,
                            }
                        ],
                    }
                ],
            },
        ],
        "fab_list": [
            {
                "id": "f_tw",
                "name": "Taiwan Fab (generic)",
                "region": "r_tw",
                "customer_company": "c_tsmc",
                "unavailable_dates": [],
            }
        ],
        "region_list": [
            {
                "id": "r_tw",
                "name": "Taiwan",
                "max_stay_on": 90,
                "max_annual_stay": 240,
                "stay_off_interval": 3,
                "unavailable_dates": [
                    {"weekly": {"weekdays": ["sat", "sun"]}},
                ],
            }
        ],
        "customer_company_list": [
            {
                "id": "c_tsmc",
                "name": "TSMC (generic)",
                "unavailable_dates": [],
            }
        ],
        "worker_company_list": su_data["worker_company_list"],
        "worker_list": su_data["worker_list"],
        "transite_day_map": [],
    }


    # ---------- SCHEDULE ----------
    all_dates = []
    all_dates.append(pd.to_datetime(su_data["plan_range"]["start_date"]))
    all_dates.append(pd.to_datetime(su_data["plan_range"]["end_date"]))
    all_dates.extend(tool_data["date_list"])
    all_dates = [d for d in all_dates if isinstance(d, pd.Timestamp)]
    all_dates.sort()

    if all_dates:
        plan_range = {
            "start_date": _to_ymd(all_dates[0]),
            "end_date": _to_ymd(all_dates[-1]),
        }
    else:
        plan_range = su_data["plan_range"]

    # Remove internal module_code from workflow_task_list
    tool_tasks_for_yaml = []
    for t in tool_data["tool_tasks"]:
        t_copy = dict(t)
        t_copy.pop("module_code", None)
        tool_tasks_for_yaml.append(t_copy)

    assignments, misc_tasks = build_assignments(su_data, tool_data)

    schedule = {
        "plan_range": plan_range,
        "workflow_task_list": tool_tasks_for_yaml + misc_tasks,
        "assignment_list": assignments,
    }

    env_root = {"environment": environment}
    sch_root = {"schedule": schedule}

    class NoAliasDumper(yaml.SafeDumper):
        def ignore_aliases(self, data):
            return True

    with open(envconfig_out, "w", encoding="utf-8") as f:
        yaml.dump(
            env_root,
            f,
            Dumper=NoAliasDumper,
            sort_keys=False,
            allow_unicode=True,
            width=4096,
        )

    with open(schedule_out, "w", encoding="utf-8") as f:
        yaml.dump(
            sch_root,
            f,
            Dumper=NoAliasDumper,
            sort_keys=False,
            allow_unicode=True,
            width=4096,
        )

    return env_root, sch_root


# ============================================================
# Entrypoint
# ============================================================

if __name__ == "__main__":
    su_file = "20251201_2 SU_Others.xlsm"
    tool_file = "20260105 台湾出張者予定_2025latest.xlsx"

    su_path = Path(su_file)
    tool_path = Path(tool_file)

    if su_path.exists() and tool_path.exists():
        build_env_and_schedule(str(su_path), str(tool_path))
        print("EnvConfig_from_excel.yaml and Schedule_from_excel.yaml have been written.")
    else:
        print("Please fix su_file / tool_file paths at the bottom of this script.")
