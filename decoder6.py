
# Decoder6.py
# ---------------------------------------------------------------------
# Generates EnvConfig.yaml + Schedule.yaml for the Timefold scheduler.
#
# Decoder6 vs Decoder5 — what changed and why:
#
# 1) INPUTS. The old "新規製番リスト" (task-CSV) and "スキル集計" (skill-level)
#    Excel files are gone. They are replaced by two new files that share the
#    same sheet names but different roles:
#       - 初期データ追加情報.xlsx    (base; mostly a stub/template)
#       - 初期データ追加情報 _r.xlsx (revised; the real, filled-in data)
#    Both have sheets "製番" (per-module planned phase dates/headcounts) and
#    "作業者" (per-worker regular/spot flag). We merge them, preferring the
#    "_r" file's values and falling back to the base file when "_r" is blank.
#
# 2) SU_Others' own internal layout changed (sheet "予定表_2026" replaces/adds
#    to "予定表_2025"): company/name/role columns moved, and a "担当職種" role
#    column replaced the old free-slot column. Column positions are now
#    detected from the header row instead of hardcoded.
#
# 3) Phase 2 (Hardware Setup) now has two independent operations, Mech (M) and
#    Elec (E), each with their own recommended headcount — instead of one
#    combined p2 skill/operation as in decoder5.
#
# 4) Worker regular/spot (R/S), from 作業者 sheet: every worker's real assigned
#    role(s) get worker_type_by_operation set uniformly to "regular"/"spot"
#    for that worker, matching the single R/S flag in the source data.
#
# 5) If a module has no usable planned phase dates in 製番 (blank), or never
#    appears in SU_Others at all, its planned window now DEFAULTS to the
#    configurable plan range (start..end) instead of being skipped.
#
# 6) Output schema follows GanttChartEditor's current types
#    (src/types/schedule.ts, src/types/envConfig.ts, src/services/yamlService.ts):
#       - workload_hours (not workload_days)
#       - phase/operation ids without underscores: e1p2, e1p2o1, ...
#       - misc_task_list: flat entries (no phase_task_list) for "other work"
#         and "personal business"; assignments reference the misc task's own
#         id directly as operation_task.
#
# The core "match actual work to planned modules" engine (SU_Others outlier
# cleanup, phase-window shifting onto real worked days, dummy/flexible
# classification) is carried over from decoder5 largely unchanged, since none
# of that depends on the Excel schema that changed.
# ---------------------------------------------------------------------

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import unicodedata
import math

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX

# ---------------------------------------------------------------------
# CONFIG (edit here; the 3 input files + plan range are CLI args, see main())
# ---------------------------------------------------------------------
DEFAULT_MAX_WORKER = 8
HOURS_PER_WORKDAY = 10  # matches the "hour: 10" used per assignment work day

CUT_DISTANCE_DAYS = 365
SHIFT_USE_WORKED_DAYS = True
# Decoder5 default was True. On the new data this naive proportional
# calendar pre-split (as opposed to cut_final_zero_workload_modules_to_dummy,
# which checks the real *shifted* result) was wrongly discarding good real
# assignment data for several modules whose actual work isn't evenly spread
# across the calendar span (see TransformationLog "phase zero workload"
# entries). Disabled by default for decoder6; the final post-shift check
# still guards against genuinely empty phases.
CUT_MODULE_IF_PHASE_ZERO_WORKLOAD = False

# Decoder5 default was True (skip modules with no SU_Others match entirely).
# Decoder6 default is False: modules with no actual work found still show up
# in Schedule.yaml, using their planned 製番 dates (as long as those dates are
# complete and fit inside the plan range — see parse_seiban_merged).
SKIP_MODULE_IF_NO_SU_MATCH = False

MIN_WORKED_DAYS_FOR_TOOL = 4
MIN_LEFT_DATE_SPAN_RATIO = 0.20

# Disabled: decoder5's "module appears within the first N days of the
# SU_Others plan range -> treat entire module as dummy" rule assumed a
# single-year sheet, where day-1 entries were likely carryover/phantom data.
# Now that the plan range can span years, "first 10 days of the range" no
# longer reliably means that; it wrongly zapped a real, heavily-worked module
# in testing. Set back to True to re-enable.
ENABLE_HEAD_OF_RANGE_CUT = False
DUMMY_HEAD_DAYS_FROM_PLAN_START = 10
ONGOING_TAIL_KEEP_GAP_DAYS = 30

TRANSFORMATION_LOG = "TransformationLog.txt"

EXCEL_EPOCH = datetime(1899, 12, 30)

# ============================================================
# Generic helpers (unchanged from decoder5)
# ============================================================

def _to_ymd(dt) -> str:
    if isinstance(dt, pd.Timestamp):
        return dt.strftime("%Y/%m/%d")
    if isinstance(dt, datetime):
        return pd.Timestamp(dt).strftime("%Y/%m/%d")
    return str(dt)


def _as_timestamp(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, pd.Timestamp):
        return v.normalize()
    if isinstance(v, datetime):
        return pd.Timestamp(v).normalize()
    if isinstance(v, (int, float)):
        try:
            return pd.Timestamp(EXCEL_EPOCH + pd.Timedelta(days=float(v))).normalize()
        except Exception:
            return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.upper() == "N/A":
            return None
        dt = pd.to_datetime(s, errors="coerce")
        if isinstance(dt, pd.Timestamp) and not pd.isna(dt):
            return dt.normalize()
    return None


def _parse_simple_date(s):
    """Parse a CLI plan-range date string like 2026/01/01 or 2026-01-01."""
    if not s:
        return None
    dt = pd.to_datetime(str(s).strip(), errors="coerce")
    if isinstance(dt, pd.Timestamp) and not pd.isna(dt):
        return dt.normalize()
    raise ValueError(f"Could not parse date: {s!r}")


def _overlaps(a_start, a_end, b_start, b_end) -> bool:
    return (a_start <= b_end) and (a_end >= b_start)


_WS_RE = re.compile(r"\s+")
_ZERO_WIDTH = {"​", "‌", "‍", "﻿"}


def _clean_text(s: str) -> str:
    if not isinstance(s, str):
        return ""
    s = unicodedata.normalize("NFKC", s)
    s = s.replace("　", " ")
    for z in _ZERO_WIDTH:
        s = s.replace(z, "")
    return s


def _norm_name(s: str) -> str:
    s = _clean_text(s)
    s = _WS_RE.sub("", s).strip()
    return s


def _planned_actual_gap_days(planned_start, planned_end, actual_start, actual_end) -> int:
    if planned_start is None or planned_end is None or actual_start is None or actual_end is None:
        return 0
    if actual_end < planned_start:
        return int((planned_start - actual_end).days)
    if actual_start > planned_end:
        return int((actual_start - planned_end).days)
    return 0


def _remember_original_text(su_data: dict, wid: str, dt: pd.Timestamp, old_text: str, new_text: str):
    m = su_data.setdefault("su_outlier_original_text", {})
    k = (wid, _to_ymd(dt))
    if k not in m:
        m[k] = {"old": old_text, "new": new_text}


# ============================================================
# Tool-code extraction (unchanged pattern; still \d{3}[A-Z0-9]\d{5}A)
# ============================================================

TOOLCODE_RE = re.compile(r"\d{3}[A-Z0-9]\d{5}A")


def extract_tool_code(s: str):
    if not isinstance(s, str):
        return None
    m = TOOLCODE_RE.search(s)
    return m.group(0) if m else None


# ============================================================
# SU_Others cut/outlier pipeline (unchanged from decoder5)
# ============================================================

def cut_su_short_span_modules_to_dummy(su_data, min_unique_worked_days=4, planned_meta=None):
    if not su_data:
        return []
    worker_date_map = su_data.get("worker_date_map", {})
    if not worker_date_map:
        return []
    code_to_occ = defaultdict(list)
    for (wid, dt), text in list(worker_date_map.items()):
        code = extract_tool_code(text)
        if not code:
            continue
        if planned_meta is not None and code not in planned_meta:
            continue
        code_to_occ[code].append((dt, wid, text))
    corrections = []
    for code, occ in code_to_occ.items():
        uniq_days = sorted(set(dt for dt, _, _ in occ))
        if len(uniq_days) >= int(min_unique_worked_days):
            continue
        broken = _break_tool_code(code)
        for dt, wid, text in occ:
            if not isinstance(text, str):
                continue
            old = text
            new = old.replace(code, broken, 1)
            if new == old:
                continue
            _remember_original_text(su_data, wid, dt, old, new)
            worker_date_map[(wid, dt)] = new
            corrections.append({
                "wid": wid, "date": _to_ymd(dt), "code": code, "text": old,
                "reason": f"short-span cut: unique worked days {len(uniq_days)} < {min_unique_worked_days} (treated as dummy other)",
            })
    su_data["su_short_span_corrections"] = corrections
    return corrections


def cut_module_if_remaining_dates_too_small_vs_planned(su_data, planned_meta, min_left_date_span_ratio=0.20):
    if not su_data:
        return []
    worker_date_map = su_data.get("worker_date_map", {})
    if not worker_date_map:
        return []
    code_to_occ = defaultdict(list)
    for (wid, dt), text in worker_date_map.items():
        code = extract_tool_code(text)
        if not code:
            continue
        if code not in planned_meta:
            continue
        code_to_occ[code].append((dt, wid, text))
    corrections = []
    for code, occ in code_to_occ.items():
        plan = planned_meta.get(code)
        if not plan:
            continue
        planned_total_days = int(plan.get("total_len", 0))
        if planned_total_days <= 0:
            continue
        remaining_unique_days = sorted(set(dt for dt, _, _ in occ))
        remaining_count = len(remaining_unique_days)
        threshold = int(math.ceil(planned_total_days * float(min_left_date_span_ratio)))
        if remaining_count >= threshold:
            continue
        broken = _break_tool_code(code)
        for dt, wid, text in occ:
            if not isinstance(text, str):
                continue
            old = text
            new = old.replace(code, broken, 1)
            if new == old:
                continue
            _remember_original_text(su_data, wid, dt, old, new)
            worker_date_map[(wid, dt)] = new
            corrections.append({
                "wid": wid, "date": _to_ymd(dt), "code": code, "text": old,
                "reason": (f"module cut: remaining unique worked dates {remaining_count} < "
                           f"ceil(planned_total_days {planned_total_days} * ratio {min_left_date_span_ratio:.2f}) = {threshold}"),
            })
    su_data["su_remaining_ratio_corrections"] = corrections
    return corrections


def cut_module_if_phase_zero_workload(su_data, planned_meta):
    if not su_data:
        return []
    worker_date_map = su_data.get("worker_date_map", {})
    if not worker_date_map:
        return []
    code_to_occ = defaultdict(list)
    for (wid, dt), text in worker_date_map.items():
        code = extract_tool_code(text)
        if not code:
            continue
        if code not in planned_meta:
            continue
        code_to_occ[code].append((dt, wid, text))
    corrections = []
    for code, occ in code_to_occ.items():
        worked_days = sorted(set(dt for dt, _, _ in occ))
        if not worked_days:
            continue
        actual_start = worked_days[0]
        actual_end = worked_days[-1]
        actual_total_span = int((actual_end - actual_start).days) + 1
        if actual_total_span <= 0:
            continue
        span_days = [actual_start + pd.Timedelta(days=i) for i in range(actual_total_span)]
        alloc = _allocate_phase_lengths_v5(
            actual_total_span, planned_meta[code]["phase_len"], phase_ids=(2, 3, 4),
            min_one=(actual_total_span >= 3),
        )
        phase_span_days = {}
        idx = 0
        for ph in (2, 3, 4):
            ln = int(alloc.get(ph, 0))
            phase_span_days[ph] = span_days[idx: idx + ln] if ln > 0 else []
            idx += ln
        worked_set = set(worked_days)
        phase_worked_counts = {ph: sum(1 for d in phase_span_days[ph] if d in worked_set) for ph in (2, 3, 4)}
        if any(phase_worked_counts[ph] == 0 for ph in (2, 3, 4)):
            broken = _break_tool_code(code)
            for dt, wid, text in occ:
                if not isinstance(text, str):
                    continue
                old = text
                new = old.replace(code, broken, 1)
                if new == old:
                    continue
                _remember_original_text(su_data, wid, dt, old, new)
                worker_date_map[(wid, dt)] = new
                corrections.append({
                    "wid": wid, "date": _to_ymd(dt), "code": code, "text": old,
                    "reason": f"module cut: phase zero workload under span-split check (counts={phase_worked_counts})",
                })
    su_data["su_phase_zero_corrections"] = corrections
    return corrections


def cut_final_zero_workload_modules_to_dummy(su_data, shifted_meta):
    if not su_data:
        return []
    worker_date_map = su_data.get("worker_date_map", {})
    if not worker_date_map:
        return []
    zero_codes = set()
    for code, meta in shifted_meta.items():
        if not meta.get("had_su_match"):
            continue
        alloc = meta.get("alloc_worked_days") or {}
        if any(int(alloc.get(ph, 0)) <= 0 for ph in (2, 3, 4)):
            zero_codes.add(code)
    if not zero_codes:
        return []
    corrections = []
    for (wid, dt), text in list(worker_date_map.items()):
        code = extract_tool_code(text)
        if code not in zero_codes:
            continue
        if not isinstance(text, str):
            continue
        old = text
        new = old.replace(code, _break_tool_code(code), 1)
        if new == old:
            continue
        _remember_original_text(su_data, wid, dt, old, new)
        worker_date_map[(wid, dt)] = new
        corrections.append({
            "wid": wid, "date": _to_ymd(dt), "code": code, "text": old,
            "reason": "final module cut: final p2/p3/p4 workload contains zero",
        })
    su_data["su_final_zero_phase_corrections"] = corrections
    return corrections


def cut_modules_with_no_qc_to_dummy(su_data, shifted_meta):
    if not su_data:
        return []
    worker_date_map = su_data.get("worker_date_map", {})
    worker_roles = su_data.get("worker_roles", {})
    if not worker_date_map:
        return []
    code_to_wids = defaultdict(set)
    for (wid, dt), text in worker_date_map.items():
        code = extract_tool_code(text)
        if code:
            code_to_wids[code].add(wid)
    no_qc_codes = set()
    for code, meta in shifted_meta.items():
        if not meta.get("had_su_match"):
            continue
        has_qc = False
        for wid in code_to_wids.get(code, set()):
            role_text = _clean_text(worker_roles.get(wid, "")).upper()
            if "QC" in role_text:
                has_qc = True
                break
        if not has_qc:
            no_qc_codes.add(code)
    if not no_qc_codes:
        return []
    corrections = []
    for (wid, dt), text in list(worker_date_map.items()):
        code = extract_tool_code(text)
        if code not in no_qc_codes:
            continue
        if not isinstance(text, str):
            continue
        old = text
        new = old.replace(code, _break_tool_code(code), 1)
        if new == old:
            continue
        _remember_original_text(su_data, wid, dt, old, new)
        worker_date_map[(wid, dt)] = new
        corrections.append({
            "wid": wid, "date": _to_ymd(dt), "code": code, "text": old,
            "reason": "final module cut: no QC worker exists in actual task cells",
        })
    su_data["su_no_qc_corrections"] = corrections
    return corrections


def _break_tool_code(code: str) -> str:
    if not isinstance(code, str) or len(code) < 2:
        return "OUTLIER"
    return code[:-1] + "X"


def _cluster_by_date_gap(sorted_dates, gap_days: int):
    clusters = []
    cur = []
    for d in sorted_dates:
        if not cur:
            cur = [d]
            continue
        if (d - cur[-1]).days <= gap_days:
            cur.append(d)
        else:
            clusters.append(cur)
            cur = [d]
    if cur:
        clusters.append(cur)
    return clusters


def cut_su_outlier_cells(
    su_data, cluster_gap_days=7, far_gap_days=60, small_cluster_max_unique_days=7,
    cut_module_if_total_cells_lt=4, cut_module_if_unique_days_lt=4, planned_meta=None,
    cut_if_far_from_planned_days=90,
):
    if not su_data:
        return []
    plan_range = su_data.get("plan_range", {}) if su_data else {}
    plan_start = _as_timestamp(plan_range.get("start_date"))
    dummy_head_end = None
    if ENABLE_HEAD_OF_RANGE_CUT and plan_start is not None:
        dummy_head_end = plan_start + pd.Timedelta(days=max(0, DUMMY_HEAD_DAYS_FROM_PLAN_START - 1))
    worker_date_map = su_data.get("worker_date_map", {})
    if not worker_date_map:
        return []
    code_to_occ = defaultdict(list)
    for (wid, dt), text in list(worker_date_map.items()):
        code = extract_tool_code(text)
        if not code:
            continue
        code_to_occ[code].append((dt, wid, text))

    def _longest_consecutive_run(dts_sorted):
        if not dts_sorted:
            return 0
        best = 1
        cur = 1
        for i in range(1, len(dts_sorted)):
            if dts_sorted[i] == dts_sorted[i - 1] + pd.Timedelta(days=1):
                cur += 1
                if cur > best:
                    best = cur
            else:
                cur = 1
        return best

    corrections = []

    def _cut_all_occurrences(code, occ_list, reason):
        broken = _break_tool_code(code)
        for dt, wid, text in occ_list:
            if not isinstance(text, str):
                continue
            old = text
            new = old.replace(code, broken, 1)
            if new == old:
                continue
            _remember_original_text(su_data, wid, dt, old, new)
            worker_date_map[(wid, dt)] = new
            corrections.append({"wid": wid, "date": _to_ymd(dt), "code": code, "text": old, "reason": reason})

    def _rebuild_occ_for_code(code):
        out = []
        for (wid, dt), text in worker_date_map.items():
            if extract_tool_code(text) == code:
                out.append((dt, wid, text))
        return sorted(out, key=lambda x: (x[0], x[1]))

    for code, occ in code_to_occ.items():
        if dummy_head_end is not None:
            has_head_hit = any(plan_start <= dt <= dummy_head_end for dt, _, _ in occ)
            if has_head_hit:
                _cut_all_occurrences(code, occ, f"module cut: module appears within first {DUMMY_HEAD_DAYS_FROM_PLAN_START} days of SU_Others plan range")
                continue

        uniq_days_pre = sorted(set(dt for dt, _, _ in occ))
        if len(uniq_days_pre) < int(cut_module_if_unique_days_lt):
            _cut_all_occurrences(code, occ, f"module cut: unique SU_Others work days < {cut_module_if_unique_days_lt} (unique_days={len(uniq_days_pre)})")
            continue
        if len(occ) < int(cut_module_if_total_cells_lt):
            _cut_all_occurrences(code, occ, f"module cut: total SU_Others cells < {cut_module_if_total_cells_lt} (cells={len(occ)})")
            continue
        if len(occ) <= 1:
            _cut_all_occurrences(code, occ, "module cut: only 1 SU_Others cell for this module")
            continue

        occ_by_wid = defaultdict(list)
        for dt, wid, text in occ:
            occ_by_wid[wid].append((dt, text))
        broken = _break_tool_code(code)

        for wid, wid_occ in occ_by_wid.items():
            wid_dates = sorted(set(dt for dt, _ in wid_occ))
            if len(wid_dates) <= 1:
                for dt, text in wid_occ:
                    if not isinstance(text, str):
                        continue
                    old = text
                    new = old.replace(code, broken, 1)
                    if new == old:
                        continue
                    _remember_original_text(su_data, wid, dt, old, new)
                    worker_date_map[(wid, dt)] = new
                    corrections.append({"wid": wid, "date": _to_ymd(dt), "code": code, "text": old,
                                         "reason": "per-worker cut: only 1 day for this worker on this module (treated as outlier)"})
                continue
            wid_clusters = _cluster_by_date_gap(wid_dates, gap_days=cluster_gap_days)
            keep_dates = set()
            for cl in wid_clusters:
                if _longest_consecutive_run(cl) >= 2:
                    keep_dates.update(cl)
            if not keep_dates:
                for dt, text in wid_occ:
                    if not isinstance(text, str):
                        continue
                    old = text
                    new = old.replace(code, broken, 1)
                    if new == old:
                        continue
                    _remember_original_text(su_data, wid, dt, old, new)
                    worker_date_map[(wid, dt)] = new
                    corrections.append({"wid": wid, "date": _to_ymd(dt), "code": code, "text": old,
                                         "reason": "per-worker cut: no cluster has >=2 consecutive days (treated as outliers)"})
                continue
            for dt, text in wid_occ:
                if dt in keep_dates:
                    continue
                if not isinstance(text, str):
                    continue
                old = text
                new = old.replace(code, broken, 1)
                if new == old:
                    continue
                _remember_original_text(su_data, wid, dt, old, new)
                worker_date_map[(wid, dt)] = new
                corrections.append({"wid": wid, "date": _to_ymd(dt), "code": code, "text": old,
                                     "reason": "per-worker cut: cluster is too small / not consecutive (outlier)"})

        clean_occ = _rebuild_occ_for_code(code)
        if not clean_occ:
            continue
        clean_uniq_days = sorted(set(dt for dt, _, _ in clean_occ))
        if len(clean_uniq_days) < int(cut_module_if_unique_days_lt) or len(clean_occ) < int(cut_module_if_total_cells_lt):
            _cut_all_occurrences(code, clean_occ, "module cut: cleaned evidence became too small after per-worker cleanup")
            continue

        if planned_meta is not None and code in planned_meta:
            pstart = planned_meta[code].get("overall_start")
            pend = planned_meta[code].get("overall_end")
            if pstart is not None and pend is not None:
                cutoff_day = pend + pd.Timedelta(days=int(cut_if_far_from_planned_days))
                uniq_dates = sorted(set(dt for dt, _, _ in clean_occ))
                clusters = _cluster_by_date_gap(uniq_dates, gap_days=1)
                kept_dates = set()
                latest_kept_day = None
                stop_keeping = False
                for cl in clusters:
                    cl_start = cl[0]
                    cl_end = cl[-1]
                    if stop_keeping:
                        continue
                    if cl_start <= cutoff_day:
                        kept_dates.update(cl)
                        latest_kept_day = cl_end
                        continue
                    if latest_kept_day is not None and (cl_start - latest_kept_day).days <= int(ONGOING_TAIL_KEEP_GAP_DAYS):
                        kept_dates.update(cl)
                        latest_kept_day = cl_end
                        continue
                    stop_keeping = True
                broken = _break_tool_code(code)
                for dt, wid, text in clean_occ:
                    if dt in kept_dates:
                        continue
                    if not isinstance(text, str):
                        continue
                    old = text
                    new = old.replace(code, broken, 1)
                    if new == old:
                        continue
                    _remember_original_text(su_data, wid, dt, old, new)
                    worker_date_map[(wid, dt)] = new
                    corrections.append({"wid": wid, "date": _to_ymd(dt), "code": code, "text": old,
                                         "reason": (f"planned-window far cut with ongoing-tail rule: beyond planned end + {cut_if_far_from_planned_days}d "
                                                    f"and more than {ONGOING_TAIL_KEEP_GAP_DAYS}d from last kept cleaned cluster")})

    def _final_cut_isolated_worker_module_pairs(su_data):
        if not su_data:
            return []
        worker_date_map = su_data.get("worker_date_map", {})
        if not worker_date_map:
            return []
        pair_to_occ = defaultdict(list)
        for (wid, dt), text in list(worker_date_map.items()):
            code = extract_tool_code(text)
            if not code:
                continue
            pair_to_occ[(wid, code)].append((dt, text))
        corrections = []

        def _has_adjacent_pair(sorted_dates):
            for i in range(1, len(sorted_dates)):
                if sorted_dates[i] == sorted_dates[i - 1] + pd.Timedelta(days=1):
                    return True
            return False

        for (wid, code), occ in pair_to_occ.items():
            uniq_dates = sorted(set(dt for dt, _ in occ))
            cut_flag = False
            reason = None
            if len(uniq_dates) <= 1:
                cut_flag = True
                reason = "final safety cut: worker has only 1 day on this module"
            elif not _has_adjacent_pair(uniq_dates):
                cut_flag = True
                reason = "final safety cut: worker has no adjacent-day pair on this module"
            if not cut_flag:
                continue
            broken = _break_tool_code(code)
            for dt, text in occ:
                if not isinstance(text, str):
                    continue
                old = text
                new = old.replace(code, broken, 1)
                if new == old:
                    continue
                _remember_original_text(su_data, wid, dt, old, new)
                worker_date_map[(wid, dt)] = new
                corrections.append({"wid": wid, "date": _to_ymd(dt), "code": code, "text": old, "reason": reason})
        su_data["su_final_isolated_pair_corrections"] = corrections
        return corrections

    final_pair_corrections = _final_cut_isolated_worker_module_pairs(su_data)
    corrections.extend(final_pair_corrections)
    su_data["su_outlier_corrections"] = corrections
    return corrections


# ============================================================
# SU_Others parsing — column layout is now auto-detected from the header
# row, since 予定表_2025 and 予定表_2026 use different layouts.
# ============================================================

GREY_RGB_LAST6 = {"A6A6A6", "BFBFBF", "D9D9D9", "808080"}
RED_RGB_LAST6 = {"FF0000"}
GREY_INDEXED = {15, 22, 23, 24, 25, 26, 27, 28, 29}
RED_INDEXED = {10}
IGNORE_WHITE_TEXT = {"FI", "FO"}


def _color_to_rgb6(color):
    if color is None:
        return None
    ctype = getattr(color, "type", None)
    if ctype == "rgb":
        rgb = getattr(color, "rgb", None)
        return str(rgb).upper()[-6:] if rgb else None
    if ctype == "indexed":
        idx = getattr(color, "indexed", None)
        if idx is None:
            return None
        try:
            rgb = COLOR_INDEX[idx]
            return str(rgb).upper()[-6:] if rgb else None
        except Exception:
            return None
    return None


def _cell_fill_rgb6(cell):
    fill = getattr(cell, "fill", None)
    if fill is None:
        return None
    candidates = [getattr(fill, "fgColor", None), getattr(fill, "start_color", None),
                  getattr(fill, "bgColor", None), getattr(fill, "end_color", None)]
    for col in candidates:
        rgb6 = _color_to_rgb6(col)
        if rgb6:
            return rgb6
    return None


def _cell_fill_indexed(cell):
    fill = getattr(cell, "fill", None)
    if fill is None:
        return None
    for col in [getattr(fill, "fgColor", None), getattr(fill, "start_color", None),
                getattr(fill, "bgColor", None), getattr(fill, "end_color", None)]:
        if col is None:
            continue
        if getattr(col, "type", None) == "indexed":
            return getattr(col, "indexed", None)
    return None


def _is_theme_grey(color) -> bool:
    if color is None:
        return False
    if getattr(color, "type", None) != "theme":
        return False
    tint = getattr(color, "tint", None)
    theme = getattr(color, "theme", None)
    if tint is not None:
        try:
            t = float(tint)
            if -0.6 <= t <= 0.6:
                return True
        except Exception:
            pass
    if theme is not None:
        return True
    return False


def _is_red_cell(cell) -> bool:
    rgb6 = _cell_fill_rgb6(cell)
    if rgb6 and rgb6 in RED_RGB_LAST6:
        return True
    idx = _cell_fill_indexed(cell)
    if idx is not None and idx in RED_INDEXED:
        return True
    return False


def _is_grey_cell(cell) -> bool:
    rgb6 = _cell_fill_rgb6(cell)
    if rgb6 and rgb6 in GREY_RGB_LAST6:
        return True
    idx = _cell_fill_indexed(cell)
    if idx is not None and idx in GREY_INDEXED:
        return True
    fill = getattr(cell, "fill", None)
    if fill is not None:
        for col in [getattr(fill, "fgColor", None), getattr(fill, "start_color", None),
                    getattr(fill, "bgColor", None), getattr(fill, "end_color", None)]:
            if _is_theme_grey(col):
                return True
    return False


def _cell_rgb_last6(cell):
    fill = getattr(cell, "fill", None)
    if fill is None:
        return None
    fg = getattr(fill, "fgColor", None)
    if fg is None:
        return None
    if getattr(fg, "type", None) != "rgb":
        return None
    rgb = getattr(fg, "rgb", None)
    if not rgb:
        return None
    return str(rgb).upper()[-6:]


def _find_date_header_ws(ws, max_scan_rows=12):
    for r in range(1, max_scan_rows + 1):
        row_vals = [cell.value for cell in ws[r]]
        if any(isinstance(v, (pd.Timestamp, datetime)) for v in row_vals):
            date_cols = [c for c, v in enumerate(row_vals, start=1) if isinstance(v, (pd.Timestamp, datetime))]
            if not date_cols:
                continue
            dt_by_col = {c: pd.Timestamp(row_vals[c - 1]).normalize() for c in date_cols}

            # Guard against stray cells that carry a date number-format but hold
            # a leftover/blank numeric value (e.g. a "1900-01-09" glitch from a
            # near-zero serial number) — such outliers would otherwise poison
            # min()/max() plan-range computation. Drop columns far (>3y) from
            # the median date of this header row.
            if len(dt_by_col) >= 3:
                sorted_dates = sorted(dt_by_col.values())
                median_dt = sorted_dates[len(sorted_dates) // 2]
                good_cols = [c for c, d in dt_by_col.items() if abs((d - median_dt).days) <= 1095]
                if len(good_cols) < len(date_cols):
                    dropped = sorted(set(date_cols) - set(good_cols))
                    print(f"WARNING: dropping {len(dropped)} SU_Others date-header column(s) far from median "
                          f"(likely formatting artifacts, not real dates): columns={dropped}", file=sys.stderr)
                    date_cols = sorted(good_cols)
                    dt_by_col = {c: dt_by_col[c] for c in good_cols}

            return r, date_cols, dt_by_col
    raise RuntimeError("Could not find date header row in SU_Others.")


def _detect_su_columns(ws, label_row_idx):
    """
    Detect company/name/role/manager columns from the label row, so both the
    old (予定表_2025: 企業名,姓名,業務形態,自由枠,自由枠,2出張先,...) and new
    (予定表_2026: 企業名,ID,姓名,業務形態,2出張先,担当地域,担当職種,作業責任者,...)
    layouts work without hardcoded positions. Returns 0-indexed columns.
    """
    row_vals = [c.value for c in ws[label_row_idx]]
    labels = [str(v).strip() if v is not None else "" for v in row_vals]

    def first_index(label):
        for i, l in enumerate(labels):
            if l == label:
                return i
        return None

    def nth_index(label, n):
        seen = 0
        for i, l in enumerate(labels):
            if l == label:
                seen += 1
                if seen == n:
                    return i
        return None

    company_col = first_index("企業名")
    company_col = 0 if company_col is None else company_col

    name_col = first_index("姓名")
    name_col = 1 if name_col is None else name_col

    manager_col = first_index("作業責任者")

    role_col = first_index("担当職種")
    if role_col is None:
        role_col = nth_index("自由枠", 2)
    if role_col is None:
        role_col = 4

    # Description-block source columns (only present on 予定表_2026's layout;
    # None on the old 予定表_2025 layout, which just leaves description blank).
    gyoumu_col = first_index("業務形態")
    visa1_col = first_index("VISA1")
    visa2_col = first_index("VISA2")
    kaigai_col = first_index("海外運転")
    ojt_col = first_index("OJT")

    return {
        "company": company_col, "name": name_col, "role": role_col, "manager": manager_col,
        "gyoumu": gyoumu_col, "visa1": visa1_col, "visa2": visa2_col,
        "kaigai": kaigai_col, "ojt": ojt_col,
    }


def parse_su_others(path: str, sheet_names=("予定表_2026",), date_filter=None):
    wb = load_workbook(path, data_only=True, read_only=False)
    used_sheets = [s for s in sheet_names if s in wb.sheetnames]
    if not used_sheets:
        raise ValueError(f"None of {sheet_names} found in {path}. Available: {wb.sheetnames}")

    f_start, f_end = date_filter if date_filter else (None, None)

    worker_company_map = {}
    worker_company_list = []

    def get_worker_company_id(company_name: str) -> str:
        company_name = str(company_name).strip()
        if company_name not in worker_company_map:
            cid = f"wc{len(worker_company_map) + 1}"
            worker_company_map[company_name] = cid
            worker_company_list.append({
                "id": cid, "name": company_name,
                "annual_overtime_limit": 10000, "monthly_overtime_limit": 10000,
                "unavailable_dates": [],
            })
        return worker_company_map[company_name]

    worker_key_to_id = {}
    worker_acc = {}
    worker_date_map = {}
    worker_personal_map = {}
    worker_roles = {}
    worker_description = {}

    plan_start = None
    plan_end = None

    for sheet_name in used_sheets:
        ws = wb[sheet_name]
        date_row_idx, date_cols, dt_by_col = _find_date_header_ws(ws)
        label_row_idx = date_row_idx + 1
        cols = _detect_su_columns(ws, label_row_idx)
        company_c, name_c, role_c, manager_c = cols["company"], cols["name"], cols["role"], cols["manager"]

        if f_start is not None and f_end is not None:
            date_cols = [c for c in date_cols if f_start <= dt_by_col[c] <= f_end]
        if not date_cols:
            continue

        worker_start_row = date_row_idx + 2
        s = min(dt_by_col[c] for c in date_cols)
        e = max(dt_by_col[c] for c in date_cols)
        plan_start = s if plan_start is None else min(plan_start, s)
        plan_end = e if plan_end is None else max(plan_end, e)

        blank_streak = 0
        max_col = max(date_cols)
        desc_cols = [c for c in (cols["gyoumu"], cols["visa1"], cols["visa2"], cols["kaigai"], cols["ojt"]) if c is not None]
        min_needed_col = max([company_c, name_c, role_c, (manager_c or 0)] + desc_cols + [0]) + 1

        for r, row_cells in enumerate(
            ws.iter_rows(min_row=worker_start_row, min_col=1, max_col=max(max_col, min_needed_col)),
            start=worker_start_row,
        ):
            company = row_cells[company_c].value if company_c < len(row_cells) else None
            name = row_cells[name_c].value if name_c < len(row_cells) else None

            if name is None or str(name).strip() == "":
                blank_streak += 1
                if blank_streak >= 30:
                    break
                continue
            blank_streak = 0

            company_str = "" if company is None else str(company).strip()
            name_str = str(name).strip()

            # Non-worker note rows (e.g. a "特殊検診" section marker) have no company.
            if not company_str:
                continue

            role_val = row_cells[role_c].value if role_c < len(row_cells) else None
            role_text = str(role_val).strip() if role_val is not None else ""

            if manager_c is not None and manager_c < len(row_cells):
                mgr_val = row_cells[manager_c].value
                is_manager = bool(isinstance(mgr_val, str) and mgr_val.strip() != "")
            else:
                is_manager = bool("責" in role_text)

            key = _norm_name(name_str)
            if key not in worker_key_to_id:
                wid = f"w{len(worker_key_to_id) + 1:03d}"
                worker_key_to_id[key] = wid
                company_id = get_worker_company_id(company_str)
                worker_acc[key] = {
                    "id": wid, "name": name_str, "worker_company": company_id,
                    "is_manager": is_manager, "unavailable_set": set(),
                }
            else:
                if is_manager:
                    worker_acc[key]["is_manager"] = True

            wid = worker_key_to_id[key]
            if role_text:
                prev_role = worker_roles.get(wid, "")
                if len(role_text) >= len(prev_role):
                    worker_roles[wid] = role_text

            if wid not in worker_description:

                def _col_text(col_idx):
                    if col_idx is None or col_idx >= len(row_cells):
                        return ""
                    v = row_cells[col_idx].value
                    return str(v).strip() if v is not None else ""

                gyoumu = _col_text(cols["gyoumu"])
                visa1 = _col_text(cols["visa1"])
                visa2 = _col_text(cols["visa2"])
                kaigai = _col_text(cols["kaigai"])
                ojt_val = _col_text(cols["ojt"])
                visa = " ".join(v for v in (visa1, visa2) if v).strip()

                if gyoumu or visa or kaigai or ojt_val:
                    desc = {"業務形態": gyoumu, "VISA": visa, "海外運転": kaigai}
                    if ojt_val:
                        desc["備考"] = "OJT"
                    worker_description[wid] = desc

            for c in date_cols:
                dt = dt_by_col[c]
                cell = row_cells[c - 1]

                if _is_red_cell(cell):
                    worker_acc[key]["unavailable_set"].add(_to_ymd(dt))
                    continue

                val = cell.value
                text = val.strip() if isinstance(val, str) else ""

                if text.upper() in IGNORE_WHITE_TEXT:
                    continue

                if _is_grey_cell(cell):
                    if text and extract_tool_code(text):
                        worker_date_map[(wid, dt)] = text
                    else:
                        worker_personal_map[(wid, dt)] = text
                    continue

                if text == "":
                    continue

                worker_date_map[(wid, dt)] = text

    worker_list = []
    for acc in worker_acc.values():
        wid = acc["id"]
        worker_list.append({
            "id": wid, "name": acc["name"], "worker_company": acc["worker_company"],
            "is_manager": acc["is_manager"], "role": worker_roles.get(wid, ""),
            "skill_map": {}, "worker_type_by_operation": {}, "fab_suitability_map": [],
            "description": worker_description.get(wid),
            "unavailable_dates": [{"date": d} for d in sorted(acc["unavailable_set"])],
        })

    plan_range = {
        "start_date": _to_ymd(plan_start) if plan_start is not None else "2025/01/01",
        "end_date": _to_ymd(plan_end) if plan_end is not None else "2025/01/01",
    }

    return {
        "worker_company_list": worker_company_list, "worker_company_map": worker_company_map,
        "worker_list": worker_list, "plan_range": plan_range,
        "worker_date_map": worker_date_map, "worker_personal_map": worker_personal_map,
        "worker_roles": worker_roles, "worker_description": worker_description,
    }


# ============================================================
# 製番 sheet parsing (replaces 新規製番リスト "CSV" sheet)
# ============================================================
# Columns (1-based), confirmed against both new Excel files:
#  1 製番 | 2 ユーザー | 3 Fab | 4 地域
#  5 第2工程 開始日 | 6 作業種別(M) | 7 工数(M) | 8 推奨人数(M)
#  9 作業種別(E) | 10 工数(E) | 11 推奨人数(E)
# 12 第3工程 開始日 | 13 作業種別(QC) | 14 工数 | 15 推奨人数
# 16 第4工程 開始日 | 17 作業種別(QC) | 18 工数 | 19 推奨人数
# 20 希望納期

_SEIBAN_COL = {
    "code": 1, "customer": 2, "fab": 3, "region": 4,
    "p2_start": 5, "p2m_headcount": 8,
    "p2e_headcount": 11,
    "p3_start": 12, "p3_headcount": 15,
    "p4_start": 16, "p4_headcount": 19,
    "delivery": 20,
}


def _num_or_none(v):
    if v is None:
        return None
    if isinstance(v, (int, float)) and not (isinstance(v, float) and pd.isna(v)):
        return v
    if isinstance(v, str) and v.strip():
        try:
            return float(v.strip())
        except Exception:
            return None
    return None


def _read_seiban_sheet(path):
    if not path:
        return {}
    wb = load_workbook(path, data_only=True)
    if "製番" not in wb.sheetnames:
        return {}
    ws = wb["製番"]
    rows = {}
    for r in range(3, ws.max_row + 1):
        code_raw = ws.cell(row=r, column=_SEIBAN_COL["code"]).value
        if not isinstance(code_raw, str) or not code_raw.strip():
            continue
        code = code_raw.strip()

        def cell(key):
            return ws.cell(row=r, column=_SEIBAN_COL[key]).value

        rows[code] = {
            "customer": cell("customer"),
            "fab": cell("fab"),
            "region": cell("region"),
            "p2_start": _as_timestamp(cell("p2_start")),
            "p2m_headcount": _num_or_none(cell("p2m_headcount")),
            "p2e_headcount": _num_or_none(cell("p2e_headcount")),
            "p3_start": _as_timestamp(cell("p3_start")),
            "p3_headcount": _num_or_none(cell("p3_headcount")),
            "p4_start": _as_timestamp(cell("p4_start")),
            "p4_headcount": _num_or_none(cell("p4_headcount")),
            "delivery": _as_timestamp(cell("delivery")),
        }
    return rows


def parse_seiban_merged(base_path, r_path, plan_start: pd.Timestamp, plan_end: pd.Timestamp, su_code_span: dict | None = None):
    """
    Returns the same shape as decoder5's parse_tasks_from_csv_v5():
    {"valid_codes", "planned_meta", "cut_rows", "date_list"}, so the rest of
    the pipeline (build_shifted_meta etc.) is unchanged.

    SU_Others is the main source of truth, same philosophy as decoder5's
    "SU_Others provides actual execution span". 製番's p2/p3/p4 start dates
    are only used as a *planned reference* (proportions, ratio checks) — if
    they're missing or out of order, but the module code has real occurrences
    in SU_Others (`su_code_span`), the module is NOT dropped: a nominal
    evenly-split "planned" window is built from SU_Others' own actual span
    instead, and the rest of the pipeline (which shifts onto the real worked
    days regardless) takes it from there. A module is only dropped entirely
    (treated as dummy/not-listed) when 製番 has nothing usable AND SU_Others
    has no occurrences of it either — genuinely no data anywhere.

    There is no plan-range containment check here anymore: a module starting
    before plan_start is not dummied, it's kept and its already-happened
    portion is marked plan_flexibility="Fixed" downstream (build_assignments_v6),
    with only the portion at/after plan_start left "Flexible" for the
    scheduler.

    希望納期 (delivery) is handled separately from p2/p3/p4: in practice it is
    blank in essentially every row of both files (0/60 in the sample data),
    unlike p2/p3/p4 start (26/60 filled) — so treating a missing delivery the
    same as missing start dates would dummy out every module. When missing
    (or earlier than p4_start), it defaults to plan_end.
    """
    su_code_span = su_code_span or {}
    base_rows = _read_seiban_sheet(base_path)
    r_rows = _read_seiban_sheet(r_path)
    all_codes = sorted(set(base_rows) | set(r_rows))

    planned_meta = {}
    cut_rows = []
    all_dates = []

    for code in all_codes:
        b = base_rows.get(code, {})
        rr = r_rows.get(code, {})

        def pick(key):
            v = rr.get(key)
            if v is None or (isinstance(v, str) and not v.strip()):
                v = b.get(key)
            return v

        customer = pick("customer")
        customer = str(customer).strip() if isinstance(customer, str) and str(customer).strip() else "OTHER"
        fab_name = pick("fab")
        fab_name = str(fab_name).strip() if isinstance(fab_name, str) and str(fab_name).strip() else "Other"
        country = pick("region")
        country = str(country).strip() if isinstance(country, str) and str(country).strip() else "Other"

        p2s, p3s, p4s, deliv = pick("p2_start"), pick("p3_start"), pick("p4_start"), pick("delivery")

        complete = all(x is not None for x in (p2s, p3s, p4s))
        ordered = complete and (p2s <= p3s <= p4s)

        if not ordered:
            span = su_code_span.get(code)
            reason = "missing phase start date(s) (p2/p3/p4)" if not complete else "phase start dates out of order"
            if span is None:
                cut_rows.append((code, f"DUMMY: {reason} in both 初期データ追加情報 files, and no SU_Others data found either"))
                continue
            actual_start, actual_end = span
            total_days = int((actual_end - actual_start).days) + 1
            alloc = _allocate_phase_lengths_v5(total_days, {2: 1, 3: 1, 4: 1}, phase_ids=(2, 3, 4), min_one=(total_days >= 3))
            p2s = actual_start
            p3s = p2s + pd.Timedelta(days=int(alloc[2]))
            p4s = p3s + pd.Timedelta(days=int(alloc[3]))
            deliv = actual_end
            cut_rows.append((code, (
                f"NOTE: {reason} in both 初期データ追加情報 files; SU_Others is the main source, so using "
                f"its actual span {_to_ymd(actual_start)} - {_to_ymd(actual_end)} as the planned reference instead"
            )))

        if deliv is None or deliv < p4s:
            reason = "missing" if deliv is None else f"earlier than p4_start ({_to_ymd(deliv)} < {_to_ymd(p4s)})"
            deliv = max(p4s, plan_end)
            cut_rows.append((code, f"NOTE: 希望納期 (delivery) {reason} in both files; defaulted to {_to_ymd(deliv)}"))

        overall_start = p2s
        overall_end = deliv

        starts = {2: p2s, 3: p3s, 4: p4s}
        ends = {
            2: (p3s - pd.Timedelta(days=1)).normalize(),
            3: (p4s - pd.Timedelta(days=1)).normalize(),
            4: deliv,
        }
        for ph in (2, 3, 4):
            if ends[ph] < starts[ph]:
                ends[ph] = starts[ph]

        phase_len = {ph: int((ends[ph] - starts[ph]).days) + 1 for ph in (2, 3, 4)}

        # 希望納期 (delivery) is a customer deadline, not a tight "phase-4 end"
        # like the old sheet's p4終了予定日 was — it can sit hundreds of days
        # past p4's start. Used raw, that blows up phase4's weight and skews
        # every ratio-based heuristic downstream (e.g. a module with a fine,
        # short real worked span gets wrongly judged "too short vs. plan").
        # overall_end / ends[4] still show the true delivery date; only the
        # *proportion* used for splitting/ratio checks is capped.
        cap4 = phase_len[2] + phase_len[3]
        if cap4 <= 0:
            cap4 = phase_len[4]
        if phase_len[4] > cap4:
            phase_len[4] = max(cap4, 1)

        total_len = sum(phase_len.values())
        phase_pct = {ph: (phase_len[ph] / total_len) for ph in (2, 3, 4)}

        planned_meta[code] = {
            "customer": customer, "country": country, "fab_name": fab_name,
            "starts": starts, "ends": ends, "phase_len": phase_len, "phase_pct": phase_pct,
            "overall_start": overall_start, "overall_end": overall_end, "total_len": total_len,
            "p2m_headcount": pick("p2m_headcount"), "p2e_headcount": pick("p2e_headcount"),
            "p3_headcount": pick("p3_headcount"), "p4_headcount": pick("p4_headcount"),
        }
        all_dates.extend([overall_start, overall_end])

    return {
        "valid_codes": sorted(planned_meta.keys()), "planned_meta": planned_meta,
        "cut_rows": cut_rows, "date_list": all_dates,
    }


# ============================================================
# 作業者 sheet parsing (replaces スキル集計) — regular/spot per worker
# ============================================================

def _read_workers_sheet(path):
    if not path:
        return {}
    wb = load_workbook(path, data_only=True)
    if "作業者" not in wb.sheetnames:
        return {}
    ws = wb["作業者"]
    out = {}
    for r in range(1, ws.max_row + 1):
        company = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=3).value
        rs = ws.cell(row=r, column=4).value
        if not isinstance(name, str) or not name.strip():
            continue
        key = _norm_name(name)
        rs_val = str(rs).strip().upper() if isinstance(rs, str) and str(rs).strip() else ""
        out[key] = {
            "name": name.strip(),
            "company": str(company).strip() if isinstance(company, str) and str(company).strip() else "",
            "rs": rs_val if rs_val in ("R", "S") else None,
        }
    return out


def parse_workers_merged(base_path, r_path):
    """
    Merge 作業者 sheets, preferring "_r" (the revised/complete file) over the
    base file. Returns {norm_name: {"name","company","rs"}} where rs is
    "R", "S", or None.
    """
    base = _read_workers_sheet(base_path)
    r_map = _read_workers_sheet(r_path)
    out = {}
    for key in set(base) | set(r_map):
        b = base.get(key, {})
        rr = r_map.get(key, {})
        out[key] = {
            "name": rr.get("name") or b.get("name"),
            "company": rr.get("company") or b.get("company") or "",
            "rs": rr.get("rs") if rr.get("rs") else b.get("rs"),
        }
    return out


# ============================================================
# Shifting / rescaling plan onto SU_Others actual span (unchanged)
# ============================================================

def _allocate_phase_lengths_v5(actual_total_days, planned_phase_len, phase_ids=(2, 3, 4), min_one=True):
    phs = list(phase_ids)
    if actual_total_days <= 0:
        return {ph: 0 for ph in phs}
    total_planned = sum(max(0, int(planned_phase_len.get(ph, 0))) for ph in phs)
    if total_planned <= 0:
        base = actual_total_days // len(phs)
        rem = actual_total_days - base * len(phs)
        out = {ph: base for ph in phs}
        for i in range(rem):
            out[phs[i % len(phs)]] += 1
    else:
        raw = {ph: (actual_total_days * (planned_phase_len.get(ph, 0) / total_planned)) for ph in phs}
        flo = {ph: int(math.floor(raw[ph])) for ph in phs}
        rem = actual_total_days - sum(flo.values())
        frac = sorted(phs, key=lambda ph: (raw[ph] - flo[ph]), reverse=True)
        out = dict(flo)
        for i in range(rem):
            out[frac[i % len(frac)]] += 1
    if min_one and actual_total_days >= len(phs):
        zeros = [ph for ph in phs if out.get(ph, 0) <= 0]
        for z in zeros:
            donors = sorted([ph for ph in phs if out.get(ph, 0) > 1], key=lambda ph: out[ph], reverse=True)
            if not donors:
                break
            d = donors[0]
            out[d] -= 1
            out[z] = 1
    drift = actual_total_days - sum(out.values())
    if phs:
        out[phs[-1]] = out.get(phs[-1], 0) + drift
    return out


def _find_qc_phase3_start_day(code, worked_days, code_occ, su_data, phase34_cap_days):
    worker_roles = su_data.get("worker_roles", {}) if su_data else {}
    worked_index = {d: i for i, d in enumerate(worked_days)}
    PURE_ME_STOP_GAP_TOLERANCE = 3

    def _rt(role_text):
        return _clean_text(role_text).upper()

    def _role_is_pure_qc(role_text):
        rt = _rt(role_text)
        return ("QC" in rt) and ("M" not in rt) and ("E" not in rt)

    def _role_is_pure_m(role_text):
        rt = _rt(role_text)
        return ("M" in rt) and ("QC" not in rt) and ("E" not in rt)

    def _role_is_pure_e(role_text):
        rt = _rt(role_text)
        return ("E" in rt) and ("QC" not in rt) and ("M" not in rt)

    by_wid = defaultdict(list)
    for dt, wid, _disp in sorted(code_occ.get(code, []), key=lambda x: (x[0], x[1])):
        if dt in worked_index:
            by_wid[wid].append(worked_index[dt])

    total_days = len(worked_days)
    latest_p2_end_idx = max(0, total_days - int(phase34_cap_days) - 1) if phase34_cap_days is not None else 0

    pure_qc_first_idx = None
    for wid, idxs in by_wid.items():
        role_text = worker_roles.get(wid, "")
        if not _role_is_pure_qc(role_text):
            continue
        first_idx = min(idxs)
        pure_qc_first_idx = first_idx if pure_qc_first_idx is None else min(pure_qc_first_idx, first_idx)

    pure_me_stop_idx = None
    for wid, idxs in by_wid.items():
        role_text = worker_roles.get(wid, "")
        if not (_role_is_pure_m(role_text) or _role_is_pure_e(role_text)):
            continue
        idxs = sorted(set(idxs))
        if not idxs:
            continue
        cluster_end = idxs[0]
        for i in range(1, len(idxs)):
            gap = idxs[i] - idxs[i - 1]
            if gap <= (PURE_ME_STOP_GAP_TOLERANCE + 1):
                cluster_end = idxs[i]
            else:
                break
        candidate = cluster_end + 1
        if candidate >= total_days:
            continue
        pure_me_stop_idx = candidate if pure_me_stop_idx is None else min(pure_me_stop_idx, candidate)

    if pure_qc_first_idx is not None:
        if pure_qc_first_idx == 0:
            if pure_me_stop_idx is not None:
                earliest_from_cap = max(1, latest_p2_end_idx + 1) if total_days >= 2 else 0
                phase3_start_idx = max(pure_me_stop_idx, earliest_from_cap)
                if total_days >= 2:
                    phase3_start_idx = min(phase3_start_idx, total_days - 1)
                return phase3_start_idx, f"pure QC from start, pure M/E stopped before {_to_ymd(worked_days[phase3_start_idx])}", phase3_start_idx
            return 0, "pure QC exists from first worked day and no pure M/E stop found", pure_qc_first_idx
        else:
            earliest_from_cap = max(1, latest_p2_end_idx + 1) if total_days >= 2 else 0
            phase3_start_idx = max(pure_qc_first_idx, earliest_from_cap)
            if total_days >= 2:
                phase3_start_idx = min(phase3_start_idx, total_days - 1)
            return phase3_start_idx, f"pure QC first joined on {_to_ymd(worked_days[pure_qc_first_idx])}", pure_qc_first_idx

    if pure_me_stop_idx is not None:
        earliest_from_cap = max(1, latest_p2_end_idx + 1) if total_days >= 2 else 0
        phase3_start_idx = max(pure_me_stop_idx, earliest_from_cap)
        if total_days >= 2:
            phase3_start_idx = min(phase3_start_idx, total_days - 1)
        return phase3_start_idx, f"pure M/E stopped before {_to_ymd(worked_days[phase3_start_idx])}", phase3_start_idx

    phase3_start_idx = latest_p2_end_idx + 1 if total_days >= 1 else 0
    if phase3_start_idx < 0:
        phase3_start_idx = 0
    if total_days >= 1:
        phase3_start_idx = min(phase3_start_idx, total_days)
    return phase3_start_idx, "no pure QC and no pure M/E stop found; fallback to cap split", None


def build_shifted_meta(planned_meta: dict, su_data, phase34_cap_days=None):
    orig_map = su_data.get("su_outlier_original_text", {}) if su_data else {}
    shifted_meta = {}
    code_to_shifted_phases = defaultdict(list)

    code_occ = defaultdict(list)
    if su_data is not None:
        for (wid, dt), text in su_data["worker_date_map"].items():
            code = extract_tool_code(text)
            if code:
                disp = text
                k = (wid, _to_ymd(dt))
                if k in orig_map:
                    disp = orig_map[k]["old"]
                code_occ[code].append((dt, wid, disp))

    def _split_tail_equal(tail_days, meta):
        if tail_days <= 0:
            return {3: 0, 4: 0}
        if tail_days == 1:
            return {3: 1, 4: 0}
        half = tail_days // 2
        rem = tail_days % 2
        p3, p4 = half, half
        if rem == 1:
            p3_len = int(meta.get("phase_len", {}).get(3, 0))
            p4_len = int(meta.get("phase_len", {}).get(4, 0))
            if p3_len >= p4_len:
                p3 += 1
            else:
                p4 += 1
        if tail_days >= 2:
            if p3 <= 0:
                p3, p4 = 1, tail_days - 1
            elif p4 <= 0:
                p4, p3 = 1, tail_days - 1
        return {3: p3, 4: p4}

    for code, meta in planned_meta.items():
        occ = code_occ.get(code, [])
        if occ:
            worked_days = sorted(set(x[0] for x in occ))
            actual_start = worked_days[0]
            actual_end = worked_days[-1]
            total_days = len(worked_days)

            phase3_start_idx, trigger_reason, qc_first_idx = _find_qc_phase3_start_day(
                code, worked_days, code_occ, su_data, phase34_cap_days
            )

            p2_days = phase3_start_idx if total_days >= 2 else total_days
            p2_days = max(0, min(p2_days, total_days))
            tail_days = max(0, total_days - p2_days)
            if phase34_cap_days is not None:
                tail_days = min(tail_days, int(phase34_cap_days))
            tail_alloc = _split_tail_equal(tail_days, meta)

            phase_days = {
                2: worked_days[:p2_days],
                3: worked_days[p2_days:p2_days + tail_alloc.get(3, 0)],
                4: worked_days[p2_days + tail_alloc.get(3, 0):p2_days + tail_alloc.get(3, 0) + tail_alloc.get(4, 0)],
            }

            shifted_starts, shifted_ends = {}, {}
            for ph in (2, 3, 4):
                ds = phase_days.get(ph, [])
                if ds:
                    shifted_starts[ph] = ds[0]
                    shifted_ends[ph] = ds[-1]
                else:
                    if ph == 2:
                        shifted_starts[ph] = actual_start
                        shifted_ends[ph] = actual_start
                    else:
                        anchor = shifted_ends.get(ph - 1, actual_start)
                        shifted_starts[ph] = anchor
                        shifted_ends[ph] = anchor

            alloc_worked = {ph: len(phase_days.get(ph, [])) for ph in (2, 3, 4)}

            shifted_meta[code] = {
                "plan": meta, "had_su_match": True,
                "actual_first": actual_start, "actual_last": actual_end, "actual_total": total_days,
                "alloc_span_days": dict(alloc_worked), "alloc_worked_days": alloc_worked,
                "phase_days": phase_days, "shifted_starts": shifted_starts, "shifted_ends": shifted_ends,
                "occ_sample": sorted(occ, key=lambda x: x[0])[:3],
                "occ_last_sample": sorted(occ, key=lambda x: x[0])[-3:],
                "phase3_trigger_reason": trigger_reason,
                "qc_first_join": worked_days[qc_first_idx] if qc_first_idx is not None else None,
                "phase34_cap_days": phase34_cap_days,
            }
        else:
            shifted_meta[code] = {
                "plan": meta, "had_su_match": False,
                "actual_first": None, "actual_last": None, "actual_total": None,
                "alloc_span_days": None, "alloc_worked_days": None, "phase_days": None,
                "shifted_starts": {2: meta["starts"][2], 3: meta["starts"][3], 4: meta["starts"][4]},
                "shifted_ends": {2: meta["ends"][2], 3: meta["ends"][3], 4: meta["ends"][4]},
                "occ_sample": [], "occ_last_sample": [],
                "phase3_trigger_reason": "no SU_Others match", "qc_first_join": None,
                "phase34_cap_days": phase34_cap_days,
            }

        for ph in (2, 3, 4):
            ds = None
            if shifted_meta[code].get("had_su_match") and shifted_meta[code].get("phase_days"):
                ds = set(shifted_meta[code]["phase_days"][ph])
            code_to_shifted_phases[code].append({
                "phase_index": ph, "start": shifted_meta[code]["shifted_starts"][ph],
                "end": shifted_meta[code]["shifted_ends"][ph], "operation": f"p{ph}", "date_set": ds,
            })

    return shifted_meta, code_to_shifted_phases, code_occ


# ============================================================
# Build tool tasks (Schedule.yaml workflow_task_list) + code_to_phases
# ============================================================

def _role_flags(role_text: str):
    rt = _clean_text(role_text).upper()
    return ("M" in rt), ("E" in rt), ("QC" in rt)


def build_tool_tasks(task_meta, shifted_meta):
    """
    New schema: single workflow "wf_tool" per module, phases p2/p3/p4 with
    ids e{n}p2/e{n}p3/e{n}p4; p2 has TWO operations (Mech=o1, Elec=o2), p3/p4
    have one (QC=o1). IDs follow e{n}p{ph}o{k} (no underscore), matching
    GanttChartEditor's current convention.
    """
    tool_tasks = []
    code_to_phases = defaultdict(list)
    all_dates = []
    task_counter = 1

    OP_NAME = {(2, 1): "Mech", (2, 2): "Elec", (3, 1): "QC", (4, 1): "QC"}
    PHASE_NAME = {2: "Hardware Setup", 3: "Function Setup", 4: "Acceptance Inspection"}

    for code in task_meta["valid_codes"]:
        meta = shifted_meta.get(code)
        if not meta:
            continue
        plan = meta["plan"]
        customer, country, fab_name = plan["customer"], plan["country"], plan["fab_name"]

        task_id = f"e{task_counter}"
        task_counter += 1

        phase_task_list = []
        for ph in (2, 3, 4):
            phase_id = f"{task_id}p{ph}"
            start = meta["shifted_starts"][ph]
            end = meta["shifted_ends"][ph]
            if end < start:
                end = start

            op_ks = (1, 2) if ph == 2 else (1,)
            operation_task_list = []
            for k in op_ks:
                operation_task_list.append({
                    "id": f"{phase_id}o{k}", "name": OP_NAME[(ph, k)], "operation": f"p{ph}o{k}",
                    "workload_hours": 0,  # overwritten later from real assignment data
                })

            phase_task_list.append({
                "id": phase_id, "name": PHASE_NAME[ph], "phase": f"p{ph}",
                "start_date": _to_ymd(start), "end_date": _to_ymd(end),
                "operation_task_list": operation_task_list,
            })

            ds = None
            if meta.get("had_su_match") and meta.get("phase_days"):
                ds = set(meta["phase_days"].get(ph, []))
            code_to_phases[code].append({
                "phase_index": ph, "phase_id": phase_id, "start": start, "end": end, "date_set": ds,
            })
            all_dates.extend([start, end])

        tool_tasks.append({
            "id": task_id, "name": code, "workflow": "wf_tool", "fab": None,
            "phase_task_list": phase_task_list,
            "module_code": code, "customer": customer, "country": country, "fab_name": fab_name,
        })

    return tool_tasks, code_to_phases, all_dates


# ============================================================
# Build assignments (Schedule.yaml assignment_list + misc_task_list)
# ============================================================

def build_assignments_v6(su_data, code_to_phases, valid_code_set, plan_start=None, date_filter=None):
    """
    Same "known tool task vs everything else" split as decoder5's
    build_assignments_v5, but:
      - phase 2 attributes each worker's days to Mech (o1) and/or Elec (o2)
        operation based on that worker's role text (M -> o1, E -> o2; if
        neither is present, default to Mech so the day isn't silently lost);
        phase 3/4 always use o1 (QC).
      - "other" SU_Others labels and grey-cell "personal business" both become
        flat misc_task_list entries (no phase/operation wrapper) per the new
        schema; assignments reference the misc task's own id directly.
      - a phase whose real (shifted) start date is before plan_start is
        already underway/complete relative to the plan range: every
        assignment in that phase is marked plan_flexibility="Fixed" (already
        happened, not up for the scheduler to move) instead of "Flexible".
        Phases at/after plan_start stay "Flexible".
    """
    orig_map = su_data.get("su_outlier_original_text", {})
    worker_date_map = su_data["worker_date_map"]
    worker_personal_map = su_data["worker_personal_map"]
    worker_roles = su_data.get("worker_roles", {})
    f_start, f_end = date_filter if date_filter else (None, None)

    # Per-operation Fixed/Flexible: an operation's phase already started
    # before the plan range -> Fixed for every worker on it.
    op_is_fixed = {}
    for phase_list in code_to_phases.values():
        for phase_meta in phase_list:
            phase_id = phase_meta["phase_id"]
            phase_start = phase_meta.get("start")
            is_fixed = bool(plan_start is not None and phase_start is not None and phase_start < plan_start)
            op_ks = (1, 2) if phase_meta["phase_index"] == 2 else (1,)
            for k in op_ks:
                op_is_fixed[f"{phase_id}o{k}"] = is_fixed

    known_assign_map = defaultdict(list)   # (wid, operation_task_id) -> [dt,...]
    misc_label_dates = defaultdict(set)
    misc_worker_label_dates = defaultdict(list)
    personal_label_dates = defaultdict(set)
    personal_worker_label_dates = defaultdict(list)
    pb_norm_to_display = {}
    dummy_tool_labels = defaultdict(set)

    for (wid, dt), raw_text in worker_date_map.items():
        if f_start is not None and f_end is not None and (dt < f_start or dt > f_end):
            continue

        internal_text = raw_text
        display_text = raw_text
        k = (wid, _to_ymd(dt))
        if k in orig_map:
            display_text = orig_map[k]["old"]
            internal_text = orig_map[k]["new"]

        code = extract_tool_code(internal_text)

        if code and (code in valid_code_set) and (code in code_to_phases):
            matched = False
            for phase_meta in code_to_phases[code]:
                ds = phase_meta.get("date_set")
                in_window = (dt in ds) if ds is not None else (
                    phase_meta["start"] is not None and phase_meta["end"] is not None
                    and phase_meta["start"] <= dt <= phase_meta["end"]
                )
                if in_window:
                    phase_id = phase_meta["phase_id"]
                    ph = phase_meta["phase_index"]
                    if ph == 2:
                        has_m, has_e, _ = _role_flags(worker_roles.get(wid, ""))
                        ops = []
                        if has_m:
                            ops.append(f"{phase_id}o1")
                        if has_e:
                            ops.append(f"{phase_id}o2")
                        if not ops:
                            ops = [f"{phase_id}o1"]  # unclassified role -> default to Mech
                    else:
                        ops = [f"{phase_id}o1"]
                    for op_id in ops:
                        known_assign_map[(wid, op_id)].append(dt)
                    matched = True
                    break
            if matched:
                continue

        label = display_text.strip() if isinstance(display_text, str) else ""
        if not label:
            label = "other"
        misc_label_dates[label].add(dt)
        misc_worker_label_dates[(wid, label)].append(dt)

        code_disp = extract_tool_code(display_text)
        if code_disp and code_disp not in valid_code_set:
            dummy_tool_labels[code_disp].add(label)

    for (wid, dt), text in worker_personal_map.items():
        if f_start is not None and f_end is not None and (dt < f_start or dt > f_end):
            continue
        display_label = text if isinstance(text, str) else ""
        norm_label = _clean_text(display_label).strip()
        if norm_label not in pb_norm_to_display:
            pb_norm_to_display[norm_label] = display_label
        personal_label_dates[norm_label].add(dt)
        personal_worker_label_dates[(wid, norm_label)].append(dt)

    assignments = []
    op_workerday_count = defaultdict(int)
    op_date_worker_set = defaultdict(lambda: defaultdict(set))
    tmp_op_to_worker_dates = defaultdict(set)
    for (wid, op_id), dates in known_assign_map.items():
        for d in dates:
            tmp_op_to_worker_dates[(op_id, wid)].add(d)
    for (op_id, wid), dset in tmp_op_to_worker_dates.items():
        op_workerday_count[op_id] += len(dset)
        for d in dset:
            op_date_worker_set[op_id][d].add(wid)

    op_worker_count = {}
    op_assigned_date_count = {}
    for op_id, date_map in op_date_worker_set.items():
        if not date_map:
            op_worker_count[op_id] = 0
            op_assigned_date_count[op_id] = 0
        else:
            op_worker_count[op_id] = max(len(wset) for wset in date_map.values())
            op_assigned_date_count[op_id] = len(date_map)

    for (wid, op_id), dates in known_assign_map.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        work_date_list = [{"hour": HOURS_PER_WORKDAY, "date": _to_ymd(d)} for d in uniq_dates]
        assignments.append({
            "worker": wid, "operation_task": op_id,
            "start_date": _to_ymd(uniq_dates[0]), "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list,
            "plan_flexibility": "Fixed" if op_is_fixed.get(op_id) else "Flexible",
        })

    # misc: "other" SU_Others labels (flat, no phases)
    misc_label_workerday = defaultdict(int)
    tmp_label_wid_dates = defaultdict(set)
    for (wid, label), dates in misc_worker_label_dates.items():
        for d in set(dates):
            tmp_label_wid_dates[label].add((wid, d))
    for label, s in tmp_label_wid_dates.items():
        misc_label_workerday[label] = len(s)

    misc_tasks = []
    misc_label_to_id = {}
    misc_counter = 1
    for label in sorted(misc_label_dates.keys()):
        dates = misc_label_dates[label]
        if not dates:
            continue
        task_id = f"misc_other_{misc_counter}"
        misc_counter += 1
        misc_label_to_id[label] = task_id
        misc_tasks.append({
            "id": task_id, "name": label, "description": None,
            "workflow": "wf_other", "region": "r_other", "color_code": None,
        })

    for (wid, label), dates in misc_worker_label_dates.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        task_id = misc_label_to_id.get(label)
        if not task_id:
            continue
        work_date_list = [{"hour": HOURS_PER_WORKDAY, "date": _to_ymd(d)} for d in uniq_dates]
        assignments.append({
            "worker": wid, "operation_task": task_id,
            "start_date": _to_ymd(uniq_dates[0]), "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list, "plan_flexibility": "Fixed",
        })

    # misc: personal business (grey empty cells) — same flat shape
    pb_label_to_id = {}
    pb_counter = 1
    for norm_label in sorted(personal_label_dates.keys()):
        dates = personal_label_dates[norm_label]
        if not dates:
            continue
        task_id = f"misc_pb_{pb_counter}"
        pb_counter += 1
        pb_label_to_id[norm_label] = task_id
        display_label = pb_norm_to_display.get(norm_label, "") or "Personal Business"
        misc_tasks.append({
            "id": task_id, "name": display_label, "description": None,
            "workflow": "wf_personal_business", "region": "r_other", "color_code": None,
        })

    pb_worker_dates = defaultdict(list)
    for (wid, dt), _ in worker_personal_map.items():
        if f_start is not None and f_end is not None and (dt < f_start or dt > f_end):
            continue
        pb_worker_dates[wid].append(dt)
    for wid in list(pb_worker_dates.keys()):
        pb_worker_dates[wid] = sorted(set(pb_worker_dates[wid]))

    for (wid, norm_label), dates in personal_worker_label_dates.items():
        uniq_dates = sorted(set(dates))
        if not uniq_dates:
            continue
        task_id = pb_label_to_id.get(norm_label)
        if not task_id:
            continue
        work_date_list = [{"hour": HOURS_PER_WORKDAY, "date": _to_ymd(d)} for d in uniq_dates]
        assignments.append({
            "worker": wid, "operation_task": task_id,
            "start_date": _to_ymd(uniq_dates[0]), "end_date": _to_ymd(uniq_dates[-1]),
            "work_date_list": work_date_list, "plan_flexibility": "Fixed",
        })

    return (assignments, misc_tasks, op_workerday_count, op_worker_count,
            op_assigned_date_count, pb_worker_dates, dummy_tool_labels)


# ============================================================
# Transformation log
# ============================================================

def _format_phase_line(ph, start, end, extra=""):
    s = _to_ymd(start) if start is not None else "N/A"
    e = _to_ymd(end) if end is not None else "N/A"
    return f"  - P{ph}: {s} - {e} {extra}" if extra else f"  - P{ph}: {s} - {e}"


def write_transformation_log(out_path, cut_rows, shifted_meta, worker_id_to_name, workload_zero_ops,
                              dummy_tool_labels, su_outlier_corrections, outlier_cut_summary, pb_worker_dates):
    lines = ["Decoder6 Transformation Log", ""]

    lines.append("---------------------- CUT / DEFAULTED (from 製番) ----------------------")
    lines.append("(none)" if not cut_rows else "")
    for code, reason in cut_rows:
        lines.append(f"- {code}: {reason}")
    lines.append("")

    lines.append("---------------------- SU_OTHERS OUTLIER CELLS CUT ----------------------")
    if not su_outlier_corrections:
        lines.append("(none)")
    else:
        by_code = defaultdict(list)
        for rec in su_outlier_corrections:
            by_code[rec.get("code", "?")].append(rec)
        for code in sorted(by_code.keys()):
            lines.append(f"module: {code}")
            recs = sorted(by_code[code], key=lambda r: (r.get("date", ""), r.get("wid", "")))
            for rec in recs[:2000]:
                wid = rec.get("wid", "?")
                nm = worker_id_to_name.get(wid, wid)
                lines.append(f"  - {rec.get('date')} / {wid}({nm})")
                lines.append(f"      {rec.get('text')}")
                lines.append(f"      reason: {rec.get('reason')}")
            if len(recs) > 2000:
                lines.append(f"  ... ({len(recs)-2000} more)")
            lines.append("")
    lines.append("")

    lines.append("---------------------- SHIFTING DATE (plan -> actual) ----------------------")
    for code in sorted(shifted_meta.keys()):
        m = shifted_meta[code]
        plan = m["plan"]
        lines.append(f"module: {code}")
        lines.append("planned (製番):")
        total_len = plan["total_len"]
        for ph in (2, 3, 4):
            pl = plan["phase_len"][ph]
            pct = plan["phase_pct"][ph] * 100.0
            lines.append(_format_phase_line(ph, plan["starts"][ph], plan["ends"][ph], extra=f"(len={pl}d, {pct:.1f}%)"))
        lines.append(f"  planned overall: {_to_ymd(plan['overall_start'])} - {_to_ymd(plan['overall_end'])} (total={total_len}d)")
        if m["had_su_match"]:
            lines.append("shifted result (used in Schedule.yaml):")
            alloc_worked = m.get("alloc_worked_days")
            for ph in (2, 3, 4):
                extra = f"worked_in_phase={alloc_worked.get(ph, 0)}d" if alloc_worked else ""
                lines.append(_format_phase_line(ph, m["shifted_starts"][ph], m["shifted_ends"][ph], extra=extra))
            trig = m.get("phase3_trigger_reason")
            if trig:
                lines.append(f"  phase3 trigger: {trig}")
            lines.append(f"  shifted overall: {_to_ymd(m['actual_first'])} - {_to_ymd(m['actual_last'])} (worked_total={m.get('actual_total')}d)")
        else:
            lines.append("actual (SU_Others): NOT FOUND -> no shift (kept planned/defaulted dates)")
        lines.append("")

    lines.append("---------------------- WORKLOAD WARNING (worker-days == 0) ----------------------")
    if not workload_zero_ops:
        lines.append("(none)")
    else:
        for op_id, mod in workload_zero_ops:
            lines.append(f"- operation_task: {op_id} / module: {mod} (no assigned worker-days in SU_Others after shifting)")
    lines.append("")

    lines.append("---------------------- DUMMY MODULES (SU_Others tool-code not in 製番) ----------------------")
    if not dummy_tool_labels:
        lines.append("(none)")
    else:
        for code in sorted(dummy_tool_labels.keys()):
            labels = sorted(dummy_tool_labels[code])
            lines.append(f"- {code}:")
            for lb in labels[:50]:
                lines.append(f"    - {lb}")
            if len(labels) > 50:
                lines.append(f"    ... ({len(labels)-50} more)")
    lines.append("")

    lines.append("---------------------- SU_Others outlier-cut modules ----------------------")
    if not outlier_cut_summary:
        lines.append("(none)")
    else:
        for code in sorted(outlier_cut_summary.keys()):
            lines.append(f"- {code}:")
            for txt in outlier_cut_summary[code][:20]:
                lines.append(f"    - {txt}")
    lines.append("")

    lines.append("---------------------- PERSONAL BUSINESS (grey empty cells) ----------------------")
    if not pb_worker_dates:
        lines.append("(none)")
    else:
        for wid in sorted(pb_worker_dates.keys()):
            nm = worker_id_to_name.get(wid, wid)
            dts = pb_worker_dates[wid]
            lines.append(f"- {wid}({nm}) : {len(dts)} days")
    lines.append("")

    Path(out_path).write_text("\n".join(lines), encoding="utf-8")


# ============================================================
# Main build
# ============================================================

def build_env_and_schedule_decoder6(
    su_others_path, seiban_info_path, seiban_info_r_path,
    envconfig_out="EnvConfig.yaml", schedule_out="Schedule.yaml", log_out=TRANSFORMATION_LOG,
    plan_start=None, plan_end=None, su_sheet_names=("予定表_2026",),
    phase34_cap_days=None,
):
    # 1) SU_Others actual work data (need this first to get a natural plan range fallback)
    su_data = parse_su_others(su_others_path, sheet_names=su_sheet_names)

    natural_start = _as_timestamp(su_data["plan_range"]["start_date"])
    natural_end = _as_timestamp(su_data["plan_range"]["end_date"])
    resolved_plan_start = plan_start or natural_start
    resolved_plan_end = plan_end or natural_end
    if resolved_plan_start is None or resolved_plan_end is None:
        raise RuntimeError("Could not resolve a plan range: pass --plan-start/--plan-end explicitly.")
    if resolved_plan_end < resolved_plan_start:
        resolved_plan_start, resolved_plan_end = resolved_plan_end, resolved_plan_start

    # 1.5) raw code -> (earliest, latest) occurrence in SU_Others, BEFORE any
    # outlier cleanup — used so parse_seiban_merged can recognize "this
    # module has no usable 製番 dates, but SU_Others is the main source and
    # it does have real data for it" instead of dummying it.
    su_code_dates = defaultdict(list)
    for (_wid, _dt), _text in su_data["worker_date_map"].items():
        _code = extract_tool_code(_text)
        if _code:
            su_code_dates[_code].append(_dt)
    su_code_span = {code: (min(dts), max(dts)) for code, dts in su_code_dates.items()}

    # 2) 製番 (planned modules); SU_Others fills in for modules missing 製番 dates
    task_meta = parse_seiban_merged(seiban_info_path, seiban_info_r_path, resolved_plan_start, resolved_plan_end, su_code_span=su_code_span)
    planned_meta = task_meta["planned_meta"]
    cut_rows = list(task_meta["cut_rows"])
    valid_code_set = set(planned_meta.keys())

    # 3) 作業者 (regular/spot per worker)
    workers_rs = parse_workers_merged(seiban_info_path, seiban_info_r_path)

    # 4) SU_Others outlier cleanup pipeline (unchanged from decoder5)
    su_outlier_corrections = cut_su_outlier_cells(
        su_data, cluster_gap_days=7, far_gap_days=60, small_cluster_max_unique_days=7,
        cut_module_if_unique_days_lt=4, cut_module_if_total_cells_lt=4,
        planned_meta=planned_meta, cut_if_far_from_planned_days=90,
    )
    su_short_span_corrections = cut_su_short_span_modules_to_dummy(
        su_data, min_unique_worked_days=MIN_WORKED_DAYS_FOR_TOOL, planned_meta=planned_meta,
    )
    su_remaining_ratio_corrections = cut_module_if_remaining_dates_too_small_vs_planned(
        su_data, planned_meta=planned_meta, min_left_date_span_ratio=MIN_LEFT_DATE_SPAN_RATIO,
    )
    su_phase_zero_corrections = []
    if CUT_MODULE_IF_PHASE_ZERO_WORKLOAD:
        su_phase_zero_corrections = cut_module_if_phase_zero_workload(su_data, planned_meta=planned_meta)

    shifted_meta_placeholder = {}
    su_final_zero_phase_corrections = cut_final_zero_workload_modules_to_dummy(su_data, shifted_meta_placeholder)

    all_su_corrections = (
        su_outlier_corrections + su_short_span_corrections + su_remaining_ratio_corrections
        + su_phase_zero_corrections + su_final_zero_phase_corrections
    )
    outlier_cut_summary = defaultdict(list)
    for rec in all_su_corrections:
        code = rec.get("code")
        txt = rec.get("text")
        if code and txt and len(outlier_cut_summary[code]) < 20:
            outlier_cut_summary[code].append(txt)
    for code in sorted(outlier_cut_summary.keys()):
        cut_rows.append((code, "SU_Others OUTLIER CUT: occurrences were converted to dummy 'other'"))

    short_span_codes = {rec.get("code") for rec in su_short_span_corrections if rec.get("code")}
    if short_span_codes:
        for code in sorted(short_span_codes):
            cut_rows.append((code, f"DUMMY: SU_Others unique worked days < {MIN_WORKED_DAYS_FOR_TOOL} (converted to wf_other)"))
            planned_meta.pop(code, None)
        valid_code_set = set(planned_meta.keys())
        task_meta["valid_codes"] = [c for c in task_meta["valid_codes"] if c in valid_code_set]

    worker_company_list = su_data["worker_company_list"]
    worker_list = su_data["worker_list"]

    # 5) merge in workers from 作業者 sheets that never appear in SU_Others
    wc_name_to_id = {wc["name"]: wc["id"] for wc in worker_company_list}

    def get_or_create_wc_id(company_name):
        company_name = str(company_name).strip() if isinstance(company_name, str) else ""
        if company_name not in wc_name_to_id:
            cid = f"wc{len(wc_name_to_id) + 1}"
            wc_name_to_id[company_name] = cid
            worker_company_list.append({
                "id": cid, "name": company_name, "annual_overtime_limit": 10000,
                "monthly_overtime_limit": 10000, "unavailable_dates": [],
            })
        return wc_name_to_id[company_name]

    worker_by_key = {_norm_name(w["name"]): w for w in worker_list}
    next_worker_num = 1
    if worker_list:
        try:
            next_worker_num = max(int(w["id"][1:]) for w in worker_list) + 1
        except Exception:
            next_worker_num = len(worker_list) + 1

    for key, meta in workers_rs.items():
        if key in worker_by_key or not meta.get("name"):
            continue
        cid = get_or_create_wc_id(meta.get("company", ""))
        new_w = {
            "id": f"w{next_worker_num:03d}", "name": meta["name"], "worker_company": cid,
            "is_manager": False, "skill_map": {}, "worker_type_by_operation": {},
            "fab_suitability_map": [], "unavailable_dates": [],
        }
        next_worker_num += 1
        worker_list.append(new_w)
        worker_by_key[key] = new_w

    worker_id_to_name = {w["id"]: w["name"] for w in worker_list}

    # 6) shifted meta (real work -> phase windows) using SU_Others
    shifted_meta, code_to_shifted_phases, code_occ = build_shifted_meta(
        planned_meta, su_data, phase34_cap_days=phase34_cap_days,
    )

    su_no_qc_corrections = cut_modules_with_no_qc_to_dummy(su_data, shifted_meta)
    if su_no_qc_corrections:
        no_qc_codes = {rec.get("code") for rec in su_no_qc_corrections if rec.get("code")}
        for code in sorted(no_qc_codes):
            cut_rows.append((code, "DUMMY: no QC worker exists in actual task cells"))
            planned_meta.pop(code, None)
            shifted_meta.pop(code, None)
        valid_code_set = set(planned_meta.keys())
        task_meta["valid_codes"] = [c for c in task_meta["valid_codes"] if c in valid_code_set]

    su_final_zero_phase_corrections = cut_final_zero_workload_modules_to_dummy(su_data, shifted_meta)
    if su_final_zero_phase_corrections:
        zero_codes = {rec.get("code") for rec in su_final_zero_phase_corrections if rec.get("code")}
        for code in sorted(zero_codes):
            cut_rows.append((code, "DUMMY: final p2/p3/p4 workload contains zero"))
            planned_meta.pop(code, None)
            shifted_meta.pop(code, None)
        valid_code_set = set(planned_meta.keys())
        task_meta["valid_codes"] = [c for c in task_meta["valid_codes"] if c in valid_code_set]

    if SKIP_MODULE_IF_NO_SU_MATCH:
        no_su_codes = [code for code, m in shifted_meta.items() if not m.get("had_su_match")]
        for code in no_su_codes:
            cut_rows.append((code, "SKIPPED: SU_Others NOT FOUND; module omitted from Schedule.yaml"))
            planned_meta.pop(code, None)
            shifted_meta.pop(code, None)
        valid_code_set = set(planned_meta.keys())
        task_meta["valid_codes"] = [c for c in task_meta["valid_codes"] if c in valid_code_set]

    cut_due_to_distance = []
    for code, m in list(shifted_meta.items()):
        if not m.get("had_su_match"):
            continue
        plan = m["plan"]
        gap = _planned_actual_gap_days(plan["overall_start"], plan["overall_end"], m["actual_first"], m["actual_last"])
        if gap > CUT_DISTANCE_DAYS:
            cut_due_to_distance.append((code, gap))
    if cut_due_to_distance:
        for code, gap in cut_due_to_distance:
            cut_rows.append((code, f"actual span is outside planned window by {gap} days (> {CUT_DISTANCE_DAYS})"))
            planned_meta.pop(code, None)
            shifted_meta.pop(code, None)
        valid_code_set = set(planned_meta.keys())
        task_meta["valid_codes"] = [c for c in task_meta["valid_codes"] if c in valid_code_set]

    # 7) build tool tasks (Schedule.yaml workflow_task_list) with new schema
    tool_tasks, code_to_phases, all_dates = build_tool_tasks(task_meta, shifted_meta)

    # 8) customer/region/fab lists
    customer_name_to_id = {"OTHER": "c_other"}
    region_name_to_id = {"Other": "r_other"}
    fab_name_to_id = {"Other": "f_other"}
    customer_company_list = [{"id": "c_other", "name": "OTHER", "unavailable_dates": []}]
    region_list = [{"id": "r_other", "name": "Other", "max_stay_on": 10000, "max_annual_stay": 10000,
                     "stay_off_interval": 3, "unavailable_dates": []}]
    fab_list = [{"id": "f_other", "name": "Other", "region": "r_other", "customer_company": "c_other", "unavailable_dates": []}]

    def get_customer_id(name):
        nm = name.strip() if isinstance(name, str) and name.strip() else "OTHER"
        if nm not in customer_name_to_id:
            cid = f"c{len(customer_name_to_id)}"
            customer_name_to_id[nm] = cid
            customer_company_list.append({"id": cid, "name": nm, "unavailable_dates": []})
        return customer_name_to_id[nm]

    def get_region_id(country):
        nm = country.strip() if isinstance(country, str) and country.strip() else "Other"
        if nm not in region_name_to_id:
            rid = f"r{len(region_name_to_id)}"
            region_name_to_id[nm] = rid
            region_list.append({"id": rid, "name": nm, "max_stay_on": 10000, "max_annual_stay": 10000,
                                 "stay_off_interval": 3, "unavailable_dates": []})
        return region_name_to_id[nm]

    def get_fab_id(fab_name, country, customer):
        nm = fab_name.strip() if isinstance(fab_name, str) and fab_name.strip() else "Other"
        if nm not in fab_name_to_id:
            fid = f"f{len(fab_name_to_id)}"
            fab_name_to_id[nm] = fid
            fab_list.append({"id": fid, "name": nm, "region": get_region_id(country),
                              "customer_company": get_customer_id(customer), "unavailable_dates": []})
        return fab_name_to_id[nm]

    for t in tool_tasks:
        t["fab"] = get_fab_id(t.get("fab_name"), t.get("country"), t.get("customer"))

    def build_transite_day_map(region_list, days_default=1):
        region_ids = [r["id"] for r in region_list if r.get("id")]
        out = []
        for fr in region_ids:
            for to in region_ids:
                if fr == to:
                    continue
                d = 0 if (fr == "r_other" and to == "r_other") else days_default
                out.append({"from": fr, "to": to, "days": d})
        return out

    transite_day_map = build_transite_day_map(region_list, days_default=1)

    # 9) environment
    environment = {
        "workflow_list": [
            {
                "id": "wf_tool", "name": "Tool Install",
                "phase_list": [
                    {"id": "p2", "name": "Hardware Setup", "operation_list": [
                        {"id": "p2o1", "name": "Mech", "work_hours": [HOURS_PER_WORKDAY], "min_worker_num": 1, "max_worker_num": DEFAULT_MAX_WORKER},
                        {"id": "p2o2", "name": "Elec", "work_hours": [HOURS_PER_WORKDAY], "min_worker_num": 1, "max_worker_num": DEFAULT_MAX_WORKER},
                    ]},
                    {"id": "p3", "name": "Function Setup", "operation_list": [
                        {"id": "p3o1", "name": "QC", "work_hours": [HOURS_PER_WORKDAY], "min_worker_num": 1, "max_worker_num": DEFAULT_MAX_WORKER},
                    ]},
                    {"id": "p4", "name": "Acceptance Inspection", "operation_list": [
                        {"id": "p4o1", "name": "QC", "work_hours": [HOURS_PER_WORKDAY], "min_worker_num": 1, "max_worker_num": DEFAULT_MAX_WORKER},
                    ]},
                ],
            },
        ],
        "fab_list": fab_list, "region_list": region_list, "customer_company_list": customer_company_list,
        "worker_company_list": worker_company_list, "transite_day_map": transite_day_map,
        "worker_list": worker_list,
    }

    # 10) assignments
    (assignments, misc_tasks, op_workerday_count, op_worker_count, op_assigned_date_count,
     pb_worker_dates, dummy_tool_labels) = build_assignments_v6(su_data, code_to_phases, valid_code_set, plan_start=resolved_plan_start)

    # 11) skills: base (other_op/pb always ok) + inferred from SU_Others role text
    for w in environment["worker_list"]:
        w["skill_map"] = {"p2o1": 0, "p2o2": 0, "p3o1": 0, "p4o1": 0, "other_op": 1, "personal_business_op": 1}
        w.setdefault("worker_type_by_operation", {})

    worker_by_id = {w["id"]: w for w in environment["worker_list"]}
    worker_roles = su_data.get("worker_roles", {})

    for wid, role_text in worker_roles.items():
        w = worker_by_id.get(wid)
        if w is None:
            continue
        has_m, has_e, has_qc = _role_flags(role_text)
        if has_m:
            w["skill_map"]["p2o1"] = max(int(w["skill_map"].get("p2o1", 0)), 1)
        if has_e:
            w["skill_map"]["p2o2"] = max(int(w["skill_map"].get("p2o2", 0)), 1)
        if has_qc:
            w["skill_map"]["p3o1"] = max(int(w["skill_map"].get("p3o1", 0)), 1)
            w["skill_map"]["p4o1"] = max(int(w["skill_map"].get("p4o1", 0)), 1)

    # worker_type_by_operation from 作業者 R/S, applied uniformly to that
    # worker's real task-skill operations (p2o1/p2o2/p3o1/p4o1), per:
    # "if a worker is R or S, all of that worker's task skills are R or S".
    RS_TO_TYPE = {"R": "regular", "S": "spot"}
    for w in environment["worker_list"]:
        key = _norm_name(w["name"])
        meta = workers_rs.get(key)
        if not meta or not meta.get("rs"):
            continue
        wtype = RS_TO_TYPE[meta["rs"]]
        for op in ("p2o1", "p2o2", "p3o1", "p4o1"):
            if int(w["skill_map"].get(op, 0)) > 0:
                w["worker_type_by_operation"][op] = wtype

    # 12) plan_range (explicit / resolved above)
    plan_range = {"start_date": _to_ymd(resolved_plan_start), "end_date": _to_ymd(resolved_plan_end)}

    tool_tasks_for_yaml = []
    for t in tool_tasks:
        t2 = dict(t)
        for k in ("module_code", "customer", "country", "fab_name"):
            t2.pop(k, None)
        tool_tasks_for_yaml.append(t2)

    op_id_to_module_code = {}
    for t in tool_tasks_for_yaml:
        mod = t.get("name", "")
        for pt in t.get("phase_task_list", []):
            for ot in pt.get("operation_task_list", []):
                op_id_to_module_code[ot["id"]] = mod

    # 13) workload_hours + recommends_worker_min/max, from real assignment data
    #     (falling back to 製番's 推奨人数 headcount where SU_Others gave nothing)
    headcount_by_op = {}
    for code, meta in shifted_meta.items():
        plan = meta.get("plan", {})
        task = next((t for t in tool_tasks if t.get("module_code") == code), None)
        if not task:
            continue
        eid = task["id"]
        headcount_by_op[f"{eid}p2o1"] = plan.get("p2m_headcount")
        headcount_by_op[f"{eid}p2o2"] = plan.get("p2e_headcount")
        headcount_by_op[f"{eid}p3o1"] = plan.get("p3_headcount")
        headcount_by_op[f"{eid}p4o1"] = plan.get("p4_headcount")

    workload_zero_ops = []
    for t in tool_tasks_for_yaml:
        for pt in t.get("phase_task_list", []):
            for ot in pt.get("operation_task_list", []):
                op_id = ot["id"]
                worked_days = int(op_workerday_count.get(op_id, 0))
                if worked_days == 0:
                    workload_zero_ops.append((op_id, op_id_to_module_code.get(op_id, "")))
                ot["workload_hours"] = worked_days * HOURS_PER_WORKDAY

                peak_w = int(op_worker_count.get(op_id, 0))
                assigned_date_count = int(op_assigned_date_count.get(op_id, 0))
                excel_headcount = headcount_by_op.get(op_id)

                if assigned_date_count > 0:
                    recommend_avg = worked_days / assigned_date_count
                    rec_min = int(math.floor(recommend_avg))
                    rec_max = int(math.ceil(recommend_avg))
                else:
                    rec_min, rec_max = 0, 0
                if worked_days > 0:
                    rec_min = max(1, rec_min)
                    rec_max = max(1, rec_max)
                if excel_headcount:
                    rec_max = max(rec_max, int(excel_headcount))
                    rec_min = max(min(rec_min, rec_max), 1 if rec_max > 0 else 0)

                ot["recommends_worker_min"] = rec_min
                ot["recommends_worker_max"] = rec_max
                ot["description"] = None
                ot["color_code"] = None

                if peak_w > DEFAULT_MAX_WORKER or (excel_headcount and excel_headcount > DEFAULT_MAX_WORKER):
                    pass  # environment max_worker_num stays a generic global default; per-task caps live in recommends_*

    for t in tool_tasks_for_yaml:
        t["description"] = None
        for pt in t["phase_task_list"]:
            pt["description"] = None

    schedule = {
        "plan_range": plan_range,
        "workflow_task_list": tool_tasks_for_yaml,
        "misc_task_list": misc_tasks,
        "assignment_list": assignments,
    }

    env_root = {"environment": environment}
    sch_root = {"schedule": schedule}

    _write_env_config_yaml(envconfig_out, environment)
    _write_schedule_yaml(schedule_out, schedule)

    write_transformation_log(
        log_out, cut_rows=cut_rows, shifted_meta=shifted_meta, worker_id_to_name=worker_id_to_name,
        workload_zero_ops=sorted(set(workload_zero_ops)), dummy_tool_labels=dummy_tool_labels,
        su_outlier_corrections=all_su_corrections, outlier_cut_summary=outlier_cut_summary,
        pb_worker_dates=pb_worker_dates,
    )

    return env_root, sch_root, shifted_meta


# ============================================================
# YAML writers, matching GanttChartEditor/src/services/yamlService.ts exactly
# (snake_case keys; workload_hours; e{n}p{ph}o{k} ids; misc_task_list flat).
# ============================================================

def _yd(s):
    return s  # dates are already "YYYY/MM/DD" strings from _to_ymd


_YAML_RISKY_CHARS_RE = re.compile(r'[:#\[\]{}&*!|>\'"%@`,]|^[\-?]')
_YAML_KEYWORD_RE = re.compile(r'^(true|false|null|yes|no|on|off|~)$', re.IGNORECASE)
_YAML_NUMLIKE_RE = re.compile(r'^[+-]?(\d+\.?\d*|\.\d+)([eE][+-]?\d+)?$')


def _ys(v):
    """
    YAML-safe scalar string. Free-text labels straight from SU_Others cells
    (misc task names etc.) can be almost anything — a bare "," or "-" alone,
    something containing ": ", a lone number-looking token — all of which
    break unquoted plain-scalar syntax. Quote (with proper JSON/YAML-safe
    escaping) whenever the value isn't unambiguously plain; otherwise leave
    it bare for readability.
    """
    if v is None:
        return ""
    s = str(v)
    if s == "":
        return '""'
    if (
        _YAML_RISKY_CHARS_RE.search(s)
        or s != s.strip()
        or "\n" in s
        or _YAML_KEYWORD_RE.match(s)
        or _YAML_NUMLIKE_RE.match(s)
    ):
        return json.dumps(s, ensure_ascii=False)
    return s


def _flow_arr(items):
    return "[" + ", ".join(str(i) for i in items) + "]"


def _flow_num_map(d):
    return "{" + ", ".join(f"{k}: {v}" for k, v in d.items()) + "}"


def _emit_unavail_dates(lines, dates, indent):
    if not dates:
        lines.append(f"{indent}unavailable_dates: []")
        return
    lines.append(f"{indent}unavailable_dates:")
    for ud in dates:
        if "date" in ud:  # decoder-internal shape: {"date": "YYYY/MM/DD"}
            lines.append(f"{indent}- single:")
            lines.append(f"{indent}    days:")
            lines.append(f"{indent}    - {ud['date']}")
        elif "weekly" in ud:
            lines.append(f"{indent}- weekly:")
            lines.append(f"{indent}    weekdays: {_flow_arr(ud['weekly']['weekdays'])}")
        elif "single" in ud:
            lines.append(f"{indent}- single:")
            lines.append(f"{indent}    days:")
            for d in ud["single"]["days"]:
                lines.append(f"{indent}    - {d}")


def _write_env_config_yaml(path, env):
    L = []
    p = L.append
    p("environment:")

    p("  workflow_list:")
    for wf in env["workflow_list"]:
        p(f"  - id: {wf['id']}")
        p(f"    name: {_ys(wf.get('name'))}")
        p("    phase_list:")
        for ph in wf.get("phase_list", []):
            p(f"    - id: {ph['id']}")
            p(f"      name: {_ys(ph.get('name'))}")
            p("      operation_list:")
            for op in ph.get("operation_list", []):
                p(f"      - id: {op['id']}")
                p(f"        name: {_ys(op.get('name'))}")
                p(f"        work_hours: {_flow_arr(op.get('work_hours', []))}")
                p(f"        min_worker_num: {op.get('min_worker_num', 0)}")
                p(f"        max_worker_num: {op.get('max_worker_num', 0)}")

    p("  fab_list:")
    for f in env["fab_list"]:
        p(f"  - id: {f['id']}")
        p(f"    name: {_ys(f.get('name'))}")
        p(f"    region: {f.get('region')}")
        p(f"    customer_company: {f.get('customer_company')}")
        _emit_unavail_dates(L, f.get("unavailable_dates", []), "    ")

    p("  region_list:")
    for r in env["region_list"]:
        p(f"  - id: {r['id']}")
        p(f"    name: {_ys(r.get('name'))}")
        p(f"    max_stay_on: {r.get('max_stay_on', 0)}")
        p(f"    max_annual_stay: {r.get('max_annual_stay', 0)}")
        p(f"    stay_off_interval: {r.get('stay_off_interval', 0)}")
        _emit_unavail_dates(L, r.get("unavailable_dates", []), "    ")

    p("  customer_company_list:")
    for c in env["customer_company_list"]:
        p(f"  - id: {c['id']}")
        p(f"    name: {_ys(c.get('name'))}")
        _emit_unavail_dates(L, c.get("unavailable_dates", []), "    ")

    p("  worker_company_list:")
    for wc in env["worker_company_list"]:
        p(f"  - id: {wc['id']}")
        p(f"    name: {_ys(wc.get('name'))}")
        p(f"    annual_overtime_limit: {wc.get('annual_overtime_limit', 0)}")
        p(f"    monthly_overtime_limit: {wc.get('monthly_overtime_limit', 0)}")
        _emit_unavail_dates(L, wc.get("unavailable_dates", []), "    ")

    p("  transite_day_map:")
    for t in env["transite_day_map"]:
        p(f"  - from: {t['from']}")
        p(f"    to: {t['to']}")
        p(f"    days: {t['days']}")

    p("  worker_list:")
    for w in env["worker_list"]:
        p(f"  - id: {w['id']}")
        p(f"    name: {_ys(w.get('name'))}")
        p(f"    worker_company: {w.get('worker_company')}")
        p(f"    is_manager: {'true' if w.get('is_manager') else 'false'}")
        sm = w.get("skill_map") or {}
        p(f"    skill_map: {_flow_num_map(sm) if sm else '{}'}")
        wto = w.get("worker_type_by_operation") or {}
        if wto:
            p(f"    worker_type_by_operation: {{{', '.join(f'{k}: {v}' for k, v in wto.items())}}}")
        else:
            p("    worker_type_by_operation: {}")
        fsm = w.get("fab_suitability_map") or []
        if fsm:
            p("    fab_suitability_map:")
            for entry in fsm:
                p(f"    - kind: {entry['kind']}")
                p(f"      suitability: {_flow_num_map(entry['suitability'])}")
        else:
            p("    fab_suitability_map: []")
        _emit_unavail_dates(L, w.get("unavailable_dates", []), "    ")
        desc = w.get("description")
        if desc is not None:
            p("    description:")
            p(f"      業務形態: {_ys(desc.get('業務形態', ''))}")
            p(f"      VISA: {_ys(desc.get('VISA', ''))}")
            p(f"      海外運転: {_ys(desc.get('海外運転', ''))}")
            if "備考" in desc:
                p(f"      備考: {_ys(desc['備考'])}")

    Path(path).write_text("\n".join(L) + "\n", encoding="utf-8")


def _write_schedule_yaml(path, sch):
    L = []
    p = L.append
    p("schedule:")
    p("  plan_range:")
    p(f"    start_date: {sch['plan_range']['start_date']}")
    p(f"    end_date: {sch['plan_range']['end_date']}")

    p("  workflow_task_list:")
    for wt in sch["workflow_task_list"]:
        p(f"  - id: {wt['id']}")
        p(f"    name: {_ys(wt.get('name'))}")
        p(f"    description: {_ys(wt.get('description'))}")
        p(f"    workflow: {wt['workflow']}")
        if wt.get("fab") is not None:
            p(f"    fab: {wt['fab']}")
        p("    phase_task_list:")
        for pt in wt["phase_task_list"]:
            p(f"    - id: {pt['id']}")
            p(f"      name: {_ys(pt.get('name'))}")
            p(f"      description: {_ys(pt.get('description'))}")
            p(f"      phase: {pt['phase']}")
            p(f"      start_date: {pt['start_date']}")
            p(f"      end_date: {pt['end_date']}")
            p("      operation_task_list:")
            for ot in pt["operation_task_list"]:
                p(f"      - id: {ot['id']}")
                p(f"        name: {_ys(ot.get('name'))}")
                p(f"        description: {_ys(ot.get('description'))}")
                p(f"        operation: {ot['operation']}")
                p(f"        workload_hours: {ot.get('workload_hours', 0)}")
                if "recommends_worker_min" in ot:
                    p(f"        recommends_worker_min: {ot['recommends_worker_min']}")
                    p(f"        recommends_worker_max: {ot['recommends_worker_max']}")
                p(f"        color_code: {_ys(ot.get('color_code')) if ot.get('color_code') else ''}")

    p("  misc_task_list:")
    for mt in sch["misc_task_list"]:
        p(f"  - id: {mt['id']}")
        p(f"    name: {_ys(mt.get('name'))}")
        p(f"    description: {_ys(mt.get('description'))}")
        p(f"    workflow: {mt['workflow']}")
        if mt.get("region") is not None:
            p(f"    region: {mt['region']}")
        if mt.get("color_code"):
            p(f"    color_code: {mt['color_code']}")

    p("  assignment_list:")
    for a in sch["assignment_list"]:
        p(f"  - worker: {a['worker']}")
        p(f"    operation_task: {a['operation_task']}")
        p(f"    start_date: {a['start_date']}")
        p(f"    end_date: {a['end_date']}")
        p("    work_date_list:")
        for w in a["work_date_list"]:
            p(f"    - date: {w['date']}")
            p(f"      hour: {w['hour']}")
        p(f"    plan_flexibility: {a['plan_flexibility']}")
        p(f"    description: {_ys(a.get('description'))}")

    Path(path).write_text("\n".join(L) + "\n", encoding="utf-8")


# ============================================================
# CLI
# ============================================================

def main():
    ap = argparse.ArgumentParser(description="Decoder6: SU_Others + 初期データ追加情報(_r) -> EnvConfig.yaml + Schedule.yaml")
    ap.add_argument("--su-others", required=True, help="Path to the SU_Others .xlsm file")
    ap.add_argument("--seiban-info", required=True, help="Path to 初期データ追加情報.xlsx (base)")
    ap.add_argument("--seiban-info-r", required=True, help="Path to 初期データ追加情報 _r.xlsx (revised; authoritative)")
    ap.add_argument("--plan-start", default=None, help="Plan range start date, e.g. 2025/09/01. Default: earliest date found in SU_Others.")
    ap.add_argument("--plan-end", default=None, help="Plan range end date, e.g. 2026/03/31. Default: latest date found in SU_Others.")
    ap.add_argument("--su-sheets", default="予定表_2026", help="Comma-separated SU_Others sheet names to read, in order. 予定表_2025 is ignored by default per current test scope.")
    ap.add_argument("--envconfig-out", default="EnvConfig.yaml")
    ap.add_argument("--schedule-out", default="Schedule.yaml")
    ap.add_argument("--log-out", default=TRANSFORMATION_LOG)
    args = ap.parse_args()

    plan_start = _parse_simple_date(args.plan_start) if args.plan_start else None
    plan_end = _parse_simple_date(args.plan_end) if args.plan_end else None
    su_sheets = tuple(s.strip() for s in args.su_sheets.split(",") if s.strip())

    for label, path in [("--su-others", args.su_others), ("--seiban-info", args.seiban_info), ("--seiban-info-r", args.seiban_info_r)]:
        if not Path(path).exists():
            print(f"ERROR: {label} file not found: {path}", file=sys.stderr)
            sys.exit(1)

    build_env_and_schedule_decoder6(
        args.su_others, args.seiban_info, args.seiban_info_r,
        envconfig_out=args.envconfig_out, schedule_out=args.schedule_out, log_out=args.log_out,
        plan_start=plan_start, plan_end=plan_end, su_sheet_names=su_sheets,
    )
    print(f"{args.envconfig_out}, {args.schedule_out}, and {args.log_out} have been written.")


if __name__ == "__main__":
    main()
