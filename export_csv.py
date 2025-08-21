# export_csv.py
import os
from argparse import ArgumentParser
from collections import defaultdict
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

from employee_scheduler_second import solve_from_config

LIGHT_BLUE = "ADD8E6"  # deadline highlight color

def _req_date(r):
    """Return the Python date for a RequirementSlot, supporting both r.day and r.bucket.day."""
    # Newer per-day-slot model
    if hasattr(r, "day") and r.day is not None and hasattr(r.day, "d"):
        return r.day.d
    # Legacy bucket model
    if hasattr(r, "bucket") and r.bucket is not None and hasattr(r.bucket, "day") and r.bucket.day is not None:
        return r.bucket.day.d
    return None

def add_task_sheet(wb, solution, start_day, deadline_by_task, dates):
    """Sheet 1: rows = task_code, cols = dates, cells = comma-joined employee names, deadline highlighted."""
    ws = wb.active
    ws.title = "Schedule by Task"

    # map (task_code, date) -> list of employee names
    cell = defaultdict(list)
    for r in solution.reqs:
        if r.employee:
            d = _req_date(r)
            if d is not None:
                cell[(r.skill.code(), d)].append(r.employee.name)

    # deadline date per task_code
    dl_date_for = {task_code: start_day + timedelta(days=idx) for task_code, idx in deadline_by_task.items()}

    # task list ordered by process then letter
    tasks = sorted(
        {r.skill.code() for r in solution.reqs},
        key=lambda code: (int(code.split("-")[0][1:]), code.split("-")[1])
    )

    # styles
    deadline_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    # layout settings
    LABEL_W = 10
    DATE_W  = 22
    ROW_H   = 36

    # header
    ws.cell(row=1, column=1, value="task \\ date").font = bold
    for j, d in enumerate(dates, start=2):
        c = ws.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws.column_dimensions[get_column_letter(j)].width = DATE_W
    ws.column_dimensions["A"].width = LABEL_W
    ws.freeze_panes = "B2"

    # body
    for i, code in enumerate(tasks, start=2):
        ws.cell(row=i, column=1, value=code).font = bold
        ws.row_dimensions[i].height = ROW_H
        dl_date = dl_date_for.get(code, None)

        for j, d in enumerate(dates, start=2):
            names = cell.get((code, d), [])
            text = ", ".join(sorted(names))
            c = ws.cell(row=i, column=j, value=text)
            c.alignment = center
            if dl_date is not None and d == dl_date:
                c.fill = deadline_fill

def add_employee_sheet(wb, solution, dates):
    """Sheet 2: rows = employee, cols = dates, cells = comma-joined task codes (no deadline highlight)."""
    ws = wb.create_sheet(title="Schedule by Employee")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    # Build (employee -> {date -> set(task_codes)})
    emp_day_tasks = defaultdict(lambda: defaultdict(set))

    for r in solution.reqs:
        if not r.employee:
            continue
        d = _req_date(r)
        if d is None:
            continue
        emp_day_tasks[r.employee.name][d].add(r.skill.code())

    # Sort employees by name
    employee_names = sorted({e.name for e in solution.employees})

    # Header
    ws.cell(row=1, column=1, value="employee \\ date").font = bold
    for j, d in enumerate(dates, start=2):
        c = ws.cell(row=1, column=j, value=d.strftime("%Y-%m-%d"))
        c.font = bold
        ws.column_dimensions[get_column_letter(j)].width = 18
    ws.column_dimensions["A"].width = 18
    ws.freeze_panes = "B2"

    # Body
    for i, name in enumerate(employee_names, start=2):
        ws.cell(row=i, column=1, value=name).font = bold
        ws.row_dimensions[i].height = 24
        day_map = emp_day_tasks.get(name, {})
        for j, d in enumerate(dates, start=2):
            tasks_today = sorted(day_map.get(d, []))
            text = ", ".join(tasks_today)
            c = ws.cell(row=i, column=j, value=text)
            c.alignment = center

def write_excel_matrix(solution, start_day, deadline_by_task, out_path: str = "schedule_matrix.xlsx"):
    dates = [d.d for d in solution.days]  # list[date]

    wb = Workbook()
    # Sheet 1: by task (with deadline highlight)
    add_task_sheet(wb, solution, start_day, deadline_by_task, dates)
    # Sheet 2: by employee (no highlight)
    add_employee_sheet(wb, solution, dates)

    wb.save(out_path)
    print(f"Wrote Excel: {os.path.abspath(out_path)}")

def main():
    ap = ArgumentParser(description="Export schedule matrix with task + employee sheets")
    ap.add_argument("--config", default="config.yaml", help="Path to YAML config")
    ap.add_argument("--out", default="schedule_matrix.xlsx", help="Output .xlsx path")
    args = ap.parse_args()

    solution, start_day, deadline_by_task = solve_from_config(args.config)
    print(f"Best score: {solution.score}")
    write_excel_matrix(solution, start_day, deadline_by_task, out_path=args.out)

if __name__ == "__main__":
    main()
