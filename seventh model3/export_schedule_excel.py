# export_schedule_excel.py
# ---------------------------------------------------------------------
# Builds an Excel with four sheets from Schedule.yaml + EnvConfig.yaml
# SHEET 1 -> "Tasks x Dates" (meta cols + required/assigned per task,
#             per-day cells highlight: start/lightblue, deadline/red,
#             window breach/purple, ordering breach/blue, staffing minmax/pink.
#             assigned_hours cell yellow if under-assigned)
# SHEET 2 -> "Employees x Dates" (company | employee | skills,
#             per-day cells highlight: skill mismatch/orange)
# SHEET 3 -> "Dashboard" (KPIs + tables + CHARTS)
# SHEET 4 -> "Breaches" (window / ordering / skill mismatch / minmax)
# ---------------------------------------------------------------------

import os
import math
import yaml
from argparse import ArgumentParser
from datetime import date, datetime, timedelta
from collections import defaultdict, Counter

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

# ------------------------------- COLORS --------------------------------
LIGHT_BLUE   = "ADD8E6"  # start marker (not a violation)
RED          = "FF9999"  # deadline marker (not a violation)

PURPLE_WIN   = "E6B8F7"  # phase window breach (violation)
BLUE_ORDER   = "9DC3E6"  # phase ordering breach (violation)
PINK_MINMAX  = "F6B5C9"  # staffing min/max breach (violation)
YELLOW_UNDER = "FFF2CC"  # under-assigned task (violation) -> assigned_hours cell
ORANGE_SKILL = "F8CBAD"  # skill mismatch (violation) -> sheet2 per-day cells

# ------------------------------- CONFIG --------------------------------
# Required hours from Schedule.yaml only:
#   required_task = workload_days * 8
REQUIRED_HOURS_MODE = True

OT_THRESHOLD = 8
CAP_HOURS    = 12

# ------------------------------ UTILITIES ------------------------------
def _d(s):
    """Parse 'YYYY/MM/DD' or 'YYYY-MM-DD' to date."""
    if isinstance(s, date):
        return s
    s = str(s).replace("-", "/")
    return datetime.strptime(s, "%Y/%m/%d").date()

def _hfmt(h):
    return f"{int(h)}H"

def _safe(dct, key):
    try:
        return dct[key]
    except Exception:
        return None

# ------------------------------- LOADERS -------------------------------
def load_env(env_path):
    with open(env_path, "r", encoding="utf-8") as f:
        env = yaml.safe_load(f)
    env = env.get("environment", {})

    workflows = {w["id"]: w for w in env.get("workflow_list", [])}
    fabs      = {f["id"]: f for f in env.get("fab_list", [])}
    regions   = {r["id"]: r for r in env.get("region_list", [])}
    customers = {c["id"]: c for c in env.get("customer_company_list", [])}
    wcompanies= {c["id"]: c for c in env.get("worker_company_list", [])}
    workers   = {w["id"]: w for w in env.get("worker_list", [])}

    # operation meta: workflow_id -> phase_id -> op_id -> meta
    op_meta = defaultdict(lambda: defaultdict(dict))
    for wf in env.get("workflow_list", []):
        for ph in wf.get("phase_list", []):
            for op in ph.get("operation_list", []):
                op_meta[wf["id"]][ph["id"]][op["id"]] = {
                    "name": op.get("name", op["id"]),
                    "work_hours": op.get("work_hours", [8,10,12]),
                    "min_worker_num": op.get("min_worker_num", 1),
                    "max_worker_num": op.get("max_worker_num", 99),
                }
    return {
        "workflows": workflows,
        "fabs": fabs,
        "regions": regions,
        "customers": customers,
        "workers": workers,
        "worker_companies": wcompanies,
        "op_meta": op_meta,
    }

def load_schedule(path):
    with open(path, "r", encoding="utf-8") as f:
        s = yaml.safe_load(f)
    s = s.get("schedule", s)

    plan_start = _d(s["plan_range"]["start_date"])
    plan_end   = _d(s["plan_range"]["end_date"])
    modules    = s.get("workflow_task_list", [])
    asg_raw    = s.get("assignment_list", [])

    # Normalize assignments and expand per-day rows
    assignments = []
    # --- also capture per-assignment blocks for utilization ---
    assignment_blocks = []
    for a in asg_raw:
        wd_key = "work_date_lsit" if "work_date_lsit" in a else "work_date_list"
        work_dates = []
        for ditem in a.get(wd_key, []):
            d = _d(ditem["date"])
            h = int(ditem["hour"])
            assignments.append({
                "worker": a["worker"],
                "operation_task": a["operation_task"],  # e.g. e1p3o2
                "date": d,
                "hours": h,
            })
            work_dates.append(d)
        # prefer explicit start/end; fallback to min/max of work_dates
        sd = a.get("start_date")
        ed = a.get("end_date")
        start_date = _d(sd) if sd else (min(work_dates) if work_dates else None)
        end_date   = _d(ed) if ed else (max(work_dates) if work_dates else None)
        assignment_blocks.append({
            "worker": a["worker"],
            "operation_task": a["operation_task"],
            "start_date": start_date,
            "end_date": end_date,
            "work_dates": sorted(set(work_dates)),
        })
    return plan_start, plan_end, modules, assignments, assignment_blocks

# ---------------------------- AGGREGATIONS -----------------------------
def build_maps(env, modules, assignments):
    workers   = env["workers"]
    fabs      = env["fabs"]
    regions   = env["regions"]
    customers = env["customers"]
    wcompanies= env["worker_companies"]

    mod_map = {m["id"]: m for m in modules}

    # module start & phase end
    module_start = {}
    phase_end = {}  # (module_id, phase_id) -> date
    op_phase_of_module = {}  # (module_id, op_id) -> phase_id

    for m in modules:
        starts = []
        for ph in m.get("phase_task_list", []):
            p_start = _d(ph["start_date"])
            p_end   = _d(ph["end_date"])
            starts.append(p_start)
            phase_end[(m["id"], ph["phase"])] = p_end
            for ot in ph.get("operation_task_list", []):
                op_phase_of_module[(m["id"], ot["operation"])] = ph["phase"]
        module_start[m["id"]] = min(starts) if starts else None

    # worker info
    worker_name = {wid: workers[wid].get("name", wid) for wid in workers}
    worker_company_name = {
        wid: wcompanies.get(workers[wid].get("worker_company"), {}).get(
            "name", workers[wid].get("worker_company")
        )
        for wid in workers
    }
    worker_skills = {wid: workers[wid].get("skill_map", {}) for wid in workers}

    # module metadata for sheet 1
    mod_meta_cols = {}
    for m in modules:
        fab_id = m.get("fab", "")
        fab    = fabs.get(fab_id, {})
        region = regions.get(fab.get("region"), {})
        cust   = customers.get(fab.get("customer_company"), {})
        mod_meta_cols[m["id"]] = {
            "module": m["id"],
            "module_name": m.get("name",""),
            "fab_id": fab_id,
            "fab_name": fab.get("name",""),
            "region": region.get("name",""),
            "customer": cust.get("name",""),
        }

    # catalog rows: (module, op)
    module_ops = []
    for m in modules:
        for ph in m.get("phase_task_list", []):
            for ot in ph.get("operation_task_list", []):
                module_ops.append((m["id"], ot["operation"], ot.get("name","")))

    # map: per-day strings for Tasks x Dates
    tde = defaultdict(list)  # (module_id, op_id, date) -> ["AA(12H)" ...]
    # map: per-day for Employees x Dates
    edt = defaultdict(list)  # (company, worker_name, date) -> ["e1p1o2 (12H)"]
    # stats
    emp_total_hours = Counter()
    emp_workdays    = Counter()
    per_day_total   = Counter()
    day_op_heads    = defaultdict(int)  # (module_id, op_id, date) -> heads

    # also keep a per-cell list of assignments to evaluate violations later
    per_cell_assigns_task = defaultdict(list)     # key: (m_id, op_id, date) -> list of dicts
    per_cell_assigns_emp  = defaultdict(list)     # key: (comp, wname, date) -> list of dicts

    for a in assignments:
        wid   = a["worker"]
        wname = worker_name.get(wid, wid)
        comp  = worker_company_name.get(wid, "")
        idx   = a["operation_task"].find("p")
        m_id  = a["operation_task"][:idx] if idx > 0 else a["operation_task"]
        op_id = a["operation_task"][idx:] if idx > 0 else ""

        # sheet1 strings
        tde[(m_id, op_id, a["date"])].append(
            f'{wname}({_hfmt(a["hours"])})'
        )
        per_cell_assigns_task[(m_id, op_id, a["date"])].append({
            "worker": wid, "wname": wname, "hours": a["hours"]
        })

        # sheet2 strings
        edt[(comp, wname, a["date"])].append(f'{m_id}{op_id} ({_hfmt(a["hours"])})')
        per_cell_assigns_emp[(comp, wname, a["date"])].append({
            "worker": wid, "m_id": m_id, "op_id": op_id, "hours": a["hours"]
        })

        emp_total_hours[wname] += a["hours"]
        per_day_total[a["date"]] += a["hours"]
        day_op_heads[(m_id, op_id, a["date"])] += 1

    # employee workdays
    for (comp, wname, d), items in edt.items():
        if items:
            emp_workdays[wname] += 1

    return {
        "mod_map": mod_map,
        "module_start": module_start,
        "phase_end": phase_end,
        "op_phase_of_module": op_phase_of_module,
        "mod_meta_cols": mod_meta_cols,
        "module_ops": module_ops,
        "tde": tde,
        "edt": edt,
        "worker_name": worker_name,
        "worker_company_name": worker_company_name,
        "worker_skills": worker_skills,
        "emp_total_hours": emp_total_hours,
        "emp_workdays": emp_workdays,
        "per_day_total": per_day_total,
        "day_op_heads": day_op_heads,
        "per_cell_assigns_task": per_cell_assigns_task,
        "per_cell_assigns_emp": per_cell_assigns_emp,
        "worker_name": worker_name,
    }

# ----------------------- REQUIRED HOURS (task/module) ------------------
def compute_required_hours_task_module(modules):
    """Return:
       - req_task[(m_id, op_id)] = workload_days * 8
       - req_module[m_id]        = sum of its tasks
    """
    req_task   = defaultdict(int)
    req_module = defaultdict(int)
    for m in modules:
        m_id = m["id"]
        total_days = 0
        for ph in m.get("phase_task_list", []):
            for ot in ph.get("operation_task_list", []):
                days = int(ot.get("workload_days", 0))
                req_task[(m_id, ot["operation"])] += days * 8
                total_days += days
        req_module[m_id] = total_days * 8
    return req_task, req_module

# --------------------------- VIOLATION DETECTION -----------------------
def detect_violations(env, modules, assignments, maps):
    """Build lookups for coloring and Sheet 4 tables."""
    # Phase windows for each (m_id, phase) from Schedule
    phase_window = {}  # (m_id, phase_id) -> (start, end)
    for m in modules:
        for ph in m.get("phase_task_list", []):
            phase_window[(m["id"], ph["phase"])] = (_d(ph["start_date"]), _d(ph["end_date"]))

    # Phase order: for each module, determine last assignment date per phase (from data)
    last_assigned_in_phase = defaultdict(lambda: defaultdict(lambda: None))  # m_id -> phase -> last_date
    # Also per-phase set of ops (for mapping op_id->phase quickly)
    # maps["op_phase_of_module"] already provides mapping.

    # assigned_by_task / by_module
    assigned_task  = defaultdict(int)  # (m_id, op_id)
    assigned_mod   = defaultdict(int)  # m_id
    for a in assignments:
        op_task = a["operation_task"]
        idx = op_task.find("p")
        m_id = op_task[:idx] if idx > 0 else op_task
        op_id= op_task[idx:] if idx > 0 else ""
        assigned_task[(m_id, op_id)] += a["hours"]
        assigned_mod[m_id]           += a["hours"]
        # update last assigned in its phase
        ph = maps["op_phase_of_module"].get((m_id, op_id))
        d  = a["date"]
        prev = last_assigned_in_phase[m_id].get(ph)
        if prev is None or d > prev:
            last_assigned_in_phase[m_id][ph] = d

    # Staffing min/max (by op/date)
    op_meta = env["op_meta"]
    wf_id = modules[0].get("workflow") if modules else None

    # Build violation sets for quick coloring
    # Sheet 1 per-day keys: (m_id, op_id, date)
    win_breach_cells   = set()
    order_breach_cells = set()
    minmax_breach_cells= set()

    # Sheet 2 per-day keys: (comp, wname, date)
    skill_mismatch_cells = set()

    # Tables for sheet 4
    tbl_window  = []  # [date, module, phase, op, worker, reason, phase_start, phase_end]
    tbl_order   = []  # [date, module, later_phase, op, worker, required_prev_phase_last_date]
    tbl_skill   = []  # [date, worker, company, module, op]
    tbl_minmax  = []  # [date, module, op, heads, min, max, status]

    # --- window & ordering & minmax (sheet1) ---
    for (m_id, op_id, dt), heads in maps["day_op_heads"].items():
        ph = maps["op_phase_of_module"].get((m_id, op_id))
        if ph:
            start, end = phase_window.get((m_id, ph), (None, None))
            # window
            if (start and dt < start) or (end and dt > end):
                win_breach_cells.add((m_id, op_id, dt))
                # we don't know which specific worker in this cell; include rows for each assignment at that cell
                for A in maps["per_cell_assigns_task"].get((m_id, op_id, dt), []):
                    tbl_window.append([dt.isoformat(), m_id, ph, op_id, A["wname"],
                                       "early" if (start and dt < start) else "late",
                                       start.isoformat() if start else "",
                                       end.isoformat() if end else ""])

            # ordering (phase k used before k-1 finished)
            # Get last assigned date in previous phase
            try:
                prev_phase_num = int(ph[1:]) - 1
                if prev_phase_num >= 1:
                    prev_phase = f"p{prev_phase_num}"
                    prev_last = last_assigned_in_phase[m_id].get(prev_phase)
                    if prev_last and dt <= prev_last:
                        order_breach_cells.add((m_id, op_id, dt))
                        for A in maps["per_cell_assigns_task"].get((m_id, op_id, dt), []):
                            tbl_order.append([dt.isoformat(), m_id, ph, op_id, A["wname"],
                                              prev_last.isoformat()])
            except Exception:
                pass

        # min/max staffing for that op/date
        meta = _safe(_safe(op_meta, wf_id), ph) if wf_id else None
        meta = _safe(meta, op_id) or {}
        min_w = int(meta.get("min_worker_num", 1))
        max_w = int(meta.get("max_worker_num", 999))
        status = ""
        if heads < min_w:
            status = f"below min ({heads}<{min_w})"
        elif heads > max_w:
            status = f"above max ({heads}>{max_w})"
        if status:
            minmax_breach_cells.add((m_id, op_id, dt))
            tbl_minmax.append([dt.isoformat(), m_id, op_id, heads, min_w, max_w, status])

    # --- skill mismatch (sheet2 only) ---
    for (comp, wname, dt), lst in maps["per_cell_assigns_emp"].items():
        mismatch = False
        for info in lst:
            wid   = None
            # find wid by name quickly (reverse map once)
            # we stored wid in info already
            wid = info["worker"]
            skills = maps["worker_skills"].get(wid, {})
            if not skills.get(info["op_id"]):
                mismatch = True
                tbl_skill.append([dt.isoformat(), wname, comp, info["m_id"], info["op_id"]])
        if mismatch:
            skill_mismatch_cells.add((comp, wname, dt))

    return {
        "assigned_task": assigned_task,
        "assigned_mod": assigned_mod,
        "last_assigned_in_phase": last_assigned_in_phase,
        "win_breach_cells": win_breach_cells,
        "order_breach_cells": order_breach_cells,
        "minmax_breach_cells": minmax_breach_cells,
        "skill_mismatch_cells": skill_mismatch_cells,
        "tbl_window": tbl_window,
        "tbl_order": tbl_order,
        "tbl_skill": tbl_skill,
        "tbl_minmax": tbl_minmax,
    }

# ------------------------------- WRITERS -------------------------------
def write_sheet_tasks_dates(wb, plan_start, plan_end, env, maps, modules, vios, req_task):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="999999")

    ws = wb.create_sheet("Tasks x Dates")

    # headers incl. required/assigned per task
    headers = ["module", "module_name", "fab_id", "fab_name", "region", "customer", "task",
               "required_hours", "assigned_hours"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h); c.font = bold; c.alignment = center

    # date columns
    dates = []
    d = plan_start
    while d <= plan_end:
        dates.append(d); d += timedelta(days=1)

    for j, dt in enumerate(dates, start=len(headers)+1):
        c = ws.cell(row=1, column=j, value=dt.isoformat())
        c.font = bold; c.alignment = center
        ws.column_dimensions[get_column_letter(j)].width = 26

    widths = [8, 18, 8, 14, 12, 10, 18, 16, 16]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = get_column_letter(len(headers)+1) + "2"

    # coloring fills
    fill_start   = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
    fill_deadln  = PatternFill(start_color=RED,        end_color=RED,        fill_type="solid")
    fill_win     = PatternFill(start_color=PURPLE_WIN, end_color=PURPLE_WIN, fill_type="solid")
    fill_order   = PatternFill(start_color=BLUE_ORDER, end_color=BLUE_ORDER, fill_type="solid")
    fill_minmax  = PatternFill(start_color=PINK_MINMAX,end_color=PINK_MINMAX,fill_type="solid")
    fill_under   = PatternFill(start_color=YELLOW_UNDER,end_color=YELLOW_UNDER,fill_type="solid")

    # sort rows
    def mod_sort_key(m_id):
        try: return int(m_id[1:])
        except: return 999
    rows = sorted(maps["module_ops"], key=lambda x: (mod_sort_key(x[0]), x[1]))

    # helper for module start & phase end
    module_start = maps["module_start"]
    phase_end    = maps["phase_end"]
    op_phase     = maps["op_phase_of_module"]

    # write rows
    for r_idx, (m_id, op_id, op_name) in enumerate(rows, start=2):
        meta = maps["mod_meta_cols"][m_id]
        ws.cell(row=r_idx, column=1, value=meta["module"]).font = bold
        ws.cell(row=r_idx, column=2, value=meta["module_name"]).font = bold
        ws.cell(row=r_idx, column=3, value=meta["fab_id"]).font = bold
        ws.cell(row=r_idx, column=4, value=meta["fab_name"]).font = bold
        ws.cell(row=r_idx, column=5, value=meta["region"]).font = bold
        ws.cell(row=r_idx, column=6, value=meta["customer"]).font = bold
        ws.cell(row=r_idx, column=7, value=f"{op_id} {op_name}").font = bold

        # required/assigned per task
        req = req_task.get((m_id, op_id), 0)
        asg = vios["assigned_task"].get((m_id, op_id), 0)
        ws.cell(row=r_idx, column=8, value=req).alignment = center
        c_asg = ws.cell(row=r_idx, column=9, value=asg); c_asg.alignment = center
        if asg < req:
            c_asg.fill = fill_under  # highlight assigned_hours cell only

        # date cells
        m_start = module_start.get(m_id)
        ph_id   = op_phase.get((m_id, op_id))
        p_end   = phase_end.get((m_id, ph_id))

        for j, dt in enumerate(dates, start=len(headers)+1):
            txt = " | ".join(sorted(maps["tde"].get((m_id, op_id, dt), [])))
            c = ws.cell(row=r_idx, column=j, value=txt)
            c.alignment = center
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)

            # decor (non-violation)
            if m_start and dt == m_start:
                c.fill = fill_start
            if p_end and dt == p_end:
                c.fill = fill_deadln

            # violations overlays (priority: minmax > window > ordering)
            key = (m_id, op_id, dt)
            if key in vios["minmax_breach_cells"]:
                c.fill = fill_minmax
            elif key in vios["win_breach_cells"]:
                c.fill = fill_win
            elif key in vios["order_breach_cells"]:
                c.fill = fill_order

    # row heights
    ws.row_dimensions[1].height = 22
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 36

    # legend at bottom
    legend_row = ws.max_row + 2
    ws.cell(row=legend_row,   column=1, value="Legend").font = bold
    def legend(label, fill, col):
        cell = ws.cell(row=legend_row+1, column=col, value=label)
        cell.fill = fill; cell.alignment = center
    legend("Start",   fill_start,   1)
    legend("Deadline",fill_deadln,  2)
    legend("Window breach", fill_win, 3)
    legend("Ordering breach", fill_order, 4)
    legend("Staffing min/max breach", fill_minmax, 5)
    legend("Under-assigned (assigned_hours)", fill_under, 6)

def write_sheet_employees_dates(wb, plan_start, plan_end, env, maps, vios):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws = wb.create_sheet("Employees x Dates")

    base_headers = ["company", "employee", "skills"]
    for i, h in enumerate(base_headers, start=1):
        c = ws.cell(row=1, column=i, value=h); c.font = bold

    # dates
    dates = []
    d = plan_start
    while d <= plan_end:
        dates.append(d); d += timedelta(days=1)

    for j, dt in enumerate(dates, start=len(base_headers)+1):
        c = ws.cell(row=1, column=j, value=dt.isoformat()); c.font = bold
        ws.column_dimensions[get_column_letter(j)].width = 30

    workdays_col = len(base_headers) + len(dates) + 1
    workhours_col= len(base_headers) + len(dates) + 2
    ws.cell(row=1, column=workdays_col, value="Workdays").font = bold
    ws.cell(row=1, column=workhours_col, value="WorkHours").font = bold

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 54
    ws.column_dimensions[get_column_letter(workdays_col)].width = 11
    ws.column_dimensions[get_column_letter(workhours_col)].width = 11
    ws.freeze_panes = "D2"

    # skills text
    def skills_text(wname):
        wid = None
        for k, v in env["workers"].items():
            if v.get("name") == wname:
                wid = k; break
        if wid is None:
            return ""
        items = list(maps["worker_skills"].get(wid, {}).items())
        def keyer(k):
            try:
                p, o = k.split("o")
                return (int(p[1:]), int(o))
            except Exception:
                return (999, 999)
        items.sort(key=lambda kv: keyer(kv[0]))
        return ", ".join(f"{k}:{v}" for k, v in items)

    # roster
    roster = []
    seen = set()
    for (comp, wname, d), _ in maps["edt"].items():
        if (comp, wname) not in seen:
            roster.append((comp, wname)); seen.add((comp, wname))
    roster.sort(key=lambda t: (t[0] or "", t[1] or ""))

    fill_skill = PatternFill(start_color=ORANGE_SKILL, end_color=ORANGE_SKILL, fill_type="solid")

    for i, (comp, wname) in enumerate(roster, start=2):
        ws.row_dimensions[i].height = 34
        ws.cell(row=i, column=1, value=comp or "").alignment = left
        ws.cell(row=i, column=2, value=wname or "").alignment = left
        ws.cell(row=i, column=3, value=skills_text(wname)).alignment = left

        for j, dt in enumerate(dates, start=len(base_headers)+1):
            txt = " | ".join(sorted(maps["edt"].get((comp, wname, dt), [])))
            c = ws.cell(row=i, column=j, value=txt); c.alignment = center
            if (comp, wname, dt) in vios["skill_mismatch_cells"]:
                c.fill = fill_skill

        ws.cell(row=i, column=workdays_col, value=maps["emp_workdays"].get(wname, 0)).alignment = center
        ws.cell(row=i, column=workhours_col, value=maps["emp_total_hours"].get(wname, 0)).alignment = center

    # legend
    legend_row = ws.max_row + 2
    ws.cell(row=legend_row, column=1, value="Legend").font = bold
    cell = ws.cell(row=legend_row+1, column=1, value="Skill mismatch")
    cell.fill = fill_skill; cell.alignment = center

# ----------------------- DASHBOARD (unchanged layout) ------------------
def write_sheet_dashboard(wb, plan_start, plan_end, env, maps, modules, assignments, assignment_blocks, req_module):
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws = wb.create_sheet("Dashboard")
    for col in range(1, 80):
        ws.column_dimensions[get_column_letter(col)].width = 14

    # KPIs
    unique_workers = len({a["worker"] for a in assignments})

    avg_hours = 0.0
    if maps["emp_workdays"]:
        total_hours = sum(maps["emp_total_hours"].values())
        total_days  = sum(maps["emp_workdays"].values())
        if total_days:
            avg_hours = total_hours / total_days

    ot_hours = 0
    per_emp_day = defaultdict(int)
    for (comp, wname, d), items in maps["edt"].items():
        day_hours = 0
        for s in items:
            h = int(s.split("(")[-1].rstrip("H)"))
            day_hours += h
        per_emp_day[(wname, d)] += day_hours
    for (_, _dte), h in per_emp_day.items():
        if h > OT_THRESHOLD:
            ot_hours += (h - OT_THRESHOLD)
    cap_breach_cnt = sum(1 for h in per_emp_day.values() if h > CAP_HOURS)
    cap_breach_pct = (100.0 * cap_breach_cnt / max(1, len(per_emp_day))) if per_emp_day else 0.0

    # Completion %
    assigned_by_module = defaultdict(int)
    for a in assignments:
        idx = a["operation_task"].find("p")
        m_id = a["operation_task"][:idx] if idx > 0 else a["operation_task"]
        assigned_by_module[m_id] += a["hours"]

    total_req = sum(req_module.values())
    completion_pct = (100.0 * sum(assigned_by_module.values()) / total_req) if total_req else 100.0

    ws["A1"].value = "KPIs"; ws["A1"].font = bold
    kpi_rows = [
        ("Unique workers", unique_workers),
        ("Avg hours/worker-day", round(avg_hours, 2)),
        ("Overtime hours(>8)", ot_hours),
        ("Cap breach(>12h) days", cap_breach_cnt),
        ("Cap breach(%)", f"{cap_breach_pct:.1f}%"),
        ("Completion", f"{completion_pct:.1f}%"),
    ]
    for i, (k, v) in enumerate(kpi_rows, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

    # Progress by module
    start_row = 10
    ws.cell(row=start_row, column=1, value="Progress by module").font = bold
    hdr = ["module", "required_hours", "assigned_hours", "%complete", "planned_end", "last_assigned", "delay(vs plan)"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(row=start_row+1, column=j, value=h).font = bold

    # last_assigned per module (exact prefix split)
    last_assigned_by_module = {}
    for a in assignments:
        op_task = a["operation_task"]
        idx = op_task.find("p")
        m_id_a = op_task[:idx] if idx > 0 else op_task
        prev = last_assigned_by_module.get(m_id_a)
        if prev is None or a["date"] > prev:
            last_assigned_by_module[m_id_a] = a["date"]

    fill_under = PatternFill(start_color=YELLOW_UNDER, end_color=YELLOW_UNDER, fill_type="solid")

    prog_rows = []
    for m in modules:
        m_id = m["id"]
        req = req_module.get(m_id, 0)
        asg = assigned_by_module.get(m_id, 0)
        pct = (100.0 * asg / req) if req else 100.0
        planned_end = max(_d(ph["end_date"]) for ph in m.get("phase_task_list", [])) if m.get("phase_task_list") else None
        last_asg = last_assigned_by_module.get(m_id)
        delay = None
        if planned_end and last_asg:
            delay = (last_asg - planned_end).days
        prog_rows.append((m_id, req, asg, pct,
                          planned_end.isoformat() if planned_end else "",
                          last_asg.isoformat() if last_asg else "",
                          delay if delay is not None else ""))

    for i, row in enumerate(prog_rows, start=start_row+2):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            # module-level under-assignment: shade assigned_hours & %complete
            if j in (3,4):  # assigned_hours, %complete
                req = row[1]; asg = row[2]
                if isinstance(req,(int,float)) and isinstance(asg,(int,float)) and asg < req:
                    c.fill = fill_under

    # ---------------- Assignment Utilization (employee-wide earliest start/latest end) ----------------
    util_start = start_row + 2 + len(prog_rows) + 2
    ws.cell(row=util_start, column=1, value="Assignment Utilization").font = bold
    util_hdr = ["employee", "start_date", "end_date", "utilization%"]
    for j, h in enumerate(util_hdr, start=1):
        ws.cell(row=util_start+1, column=j, value=h).font = bold

    # Aggregate per employee across all assignment blocks
    # - earliest start over all their tasks
    # - latest end over all their tasks
    # - worked_days = union of all dates they actually worked
    worker_name = maps["worker_name"]  # id -> display name
    agg = {}  # wid -> {"name":..., "start":date, "end":date, "worked": set(date)}
    for blk in assignment_blocks:
        wid = blk["worker"]
        nm = worker_name.get(wid, wid)
        sd = blk.get("start_date")
        ed = blk.get("end_date")
        wds = set(blk.get("work_dates", []))
        if wid not in agg:
            agg[wid] = {"name": nm, "start": sd, "end": ed, "worked": set(wds)}
        else:
            if sd and (agg[wid]["start"] is None or sd < agg[wid]["start"]):
                agg[wid]["start"] = sd
            if ed and (agg[wid]["end"] is None or ed > agg[wid]["end"]):
                agg[wid]["end"] = ed
            agg[wid]["worked"].update(wds)

    # Build rows (sorted by employee name)
    util_rows = []
    for wid, info in agg.items():
        nm = info["name"]
        sd = info["start"]
        ed = info["end"]
        worked_days = len(info["worked"])
        if sd and ed:
            planned_days = (ed - sd).days + 1
            pct = (worked_days / planned_days * 100.0) if planned_days > 0 else 0.0
        else:
            planned_days = 0
            pct = 0.0
        util_rows.append((
            nm,
            sd.isoformat() if sd else "",
            ed.isoformat() if ed else "",
            f"{pct:.1f}%"
        ))

    util_rows.sort(key=lambda r: (r[0] or ""))

    for i, row in enumerate(util_rows, start=util_start+2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    # Overtime & capacity tables
    b0 = util_start + 2 + len(util_rows) + 2
    ws.cell(row=b0, column=1, value="Overtime & capacity").font = bold
    ws.cell(row=b0+1, column=1, value="employee").font = bold
    ws.cell(row=b0+1, column=2, value="total_ot_hours").font = bold

    ot_by_emp = Counter()
    for (wname, d), h in per_emp_day.items():
        if h > OT_THRESHOLD:
            ot_by_emp[wname] += (h - OT_THRESHOLD)
    ot_table = sorted(ot_by_emp.items(), key=lambda kv: kv[1], reverse=True)

    for i, (name, hrs) in enumerate(ot_table, start=b0+2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=int(hrs))

    ws.cell(row=b0+1, column=5, value="date").font = bold
    ws.cell(row=b0+1, column=6, value="total_hours").font = bold

    day_rows = []
    d = plan_start
    while d <= plan_end:
        day_rows.append((d.strftime("%m/%d"), maps["per_day_total"].get(d, 0)))
        d += timedelta(days=1)

    for i, (dd, hrs) in enumerate(day_rows, start=b0+2):
        ws.cell(row=i, column=5, value=dd)
        ws.cell(row=i, column=6, value=int(hrs))

    # Workload balance by employee
    d0 = b0 + 2 + max(len(ot_table), len(day_rows)) + 2
    ws.cell(row=d0, column=1, value="Workload balance (per employee)").font = bold
    hdr3 = ["employee", "workdays", "total_hours", "avg/day", "stdev/day", "CoV"]
    for j, h in enumerate(hdr3, start=1):
        ws.cell(row=d0+1, column=j, value=h).font = bold

    per_emp_day_list = defaultdict(list)
    for (w, d), h in per_emp_day.items():
        per_emp_day_list[w].append(h)

    balance_rows = []
    for wname in sorted(maps["emp_total_hours"].keys()):
        days = maps["emp_workdays"].get(wname, 0)
        tot  = maps["emp_total_hours"].get(wname, 0)
        avg  = (tot / days) if days else 0.0
        lst  = per_emp_day_list.get(wname, [])
        stdev= 0.0
        if len(lst) >= 2:
            m = sum(lst) / len(lst)
            stdev = math.sqrt(sum((x - m) ** 2 for x in lst) / (len(lst) - 1))
        cov = (stdev / avg * 100.0) if avg else 0.0
        balance_rows.append((wname, days, int(tot), round(avg,2), round(stdev,2), f"{cov:.1f}%"))

    for i, row in enumerate(balance_rows, start=d0+2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    # helper top-20 (for chart)
    top_col = 10
    ws.cell(row=d0, column=top_col, value="Top 20 total hours").font = bold
    ws.cell(row=d0+1, column=top_col,   value="employee").font = bold
    ws.cell(row=d0+1, column=top_col+1, value="total_hours").font = bold
    balance_rows_top20 = sorted(balance_rows, key=lambda r: r[2], reverse=True)[:20]
    for i, row in enumerate(balance_rows_top20, start=d0+2):
        ws.cell(row=i, column=top_col,   value=row[0])
        ws.cell(row=i, column=top_col+1, value=row[2])

    # charts
    anchor_line = "N2"
    anchor_ot20 = "N22"
    anchor_tot20= "N44"

    if day_rows:
        chart2 = LineChart()
        chart2.title = "Total hours per day"
        chart2.y_axis.title = "hours"
        chart2.x_axis.title = "date"
        chart2.x_axis.tickLblSkip = 9
        dr0 = b0 + 1
        mrows = len(day_rows)
        cats = Reference(ws, min_col=5, min_row=dr0+1, max_row=dr0+mrows)
        vals = Reference(ws, min_col=6, min_row=dr0,   max_row=dr0+mrows)
        chart2.add_data(vals, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.height = 12; chart2.width = 24
        ws.add_chart(chart2, anchor_line)

    if balance_rows_top20:
        chart4 = BarChart()
        chart4.type = "bar"
        chart4.title = "Total hours by employee (Top 20)"
        chart4.x_axis.title = "hours"
        chart4.y_axis.title = "employee"
        bal0 = d0 + 1
        mrows = len(balance_rows_top20)
        cats = Reference(ws, min_col=top_col,   min_row=bal0+1, max_row=bal0+mrows)
        vals = Reference(ws, min_col=top_col+1, min_row=bal0,   max_row=bal0+mrows)
        chart4.add_data(vals, titles_from_data=True)
        chart4.set_categories(cats)
        chart4.height = 12; chart4.width = 24
        ws.add_chart(chart4, anchor_tot20)

def write_sheet_breaches(wb, vios):
    bold = Font(bold=True)
    ws = wb.create_sheet("Breaches")

    row = 1
    def write_table(title, headers, rows):
        nonlocal row
        ws.cell(row=row, column=1, value=title).font = bold
        row += 1
        for j, h in enumerate(headers, start=1):
            ws.cell(row=row, column=j, value=h).font = bold
        row += 1
        for r in rows:
            for j, v in enumerate(r, start=1):
                ws.cell(row=row, column=j, value=v)
            row += 1
        row += 2

    write_table(
        "Phase window breaches",
        ["date", "module", "phase", "op_id", "worker", "reason", "phase_start", "phase_end"],
        sorted(vios["tbl_window"], key=lambda x:(x[1],x[2],x[0]))
    )

    write_table(
        "Phase ordering breaches",
        ["date", "module", "phase(later)", "op_id", "worker", "required_prev_phase_last_date"],
        sorted(vios["tbl_order"], key=lambda x:(x[1],x[2],x[0]))
    )

    write_table(
        "Skill mismatches",
        ["date", "worker", "company", "module", "op_id"],
        sorted(vios["tbl_skill"], key=lambda x:(x[1],x[2],x[0]))
    )

    write_table(
        "Staffing min/max breaches",
        ["date", "module", "op_id", "heads", "min", "max", "status"],
        sorted(vios["tbl_minmax"], key=lambda x:(x[0],x[1],x[2]))
    )

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18

# -------------------------------- MAIN ---------------------------------
def main():
    ap = ArgumentParser(description="Read Schedule.yaml + EnvConfig.yaml and export Excel.")
    ap.add_argument("--schedule", default="Schedule.yaml")
    ap.add_argument("--env",      default="EnvConfig.yaml")
    ap.add_argument("--out",      default="schedule_export.xlsx")
    args = ap.parse_args()

    plan_start, plan_end, modules, assignments, assignment_blocks = load_schedule(args.schedule)
    env  = load_env(args.env)
    maps = build_maps(env, modules, assignments)

    req_task, req_module = compute_required_hours_task_module(modules)
    vios = detect_violations(env, modules, assignments, maps)

    wb = Workbook()
    # remove default
    del wb[wb.sheetnames[0]]

    # ===== SHEET 1 =====
    write_sheet_tasks_dates(wb, plan_start, plan_end, env, maps, modules, vios, req_task)
    # ===== SHEET 2 =====
    write_sheet_employees_dates(wb, plan_start, plan_end, env, maps, vios)
    # ===== SHEET 3 =====
    write_sheet_dashboard(wb, plan_start, plan_end, env, maps, modules, assignments, assignment_blocks, req_module)
    # ===== SHEET 4 =====
    write_sheet_breaches(wb, vios)

    wb.save(args.out)
    print(f"Wrote Excel: {os.path.abspath(args.out)}")

if __name__ == "__main__":
    main()
